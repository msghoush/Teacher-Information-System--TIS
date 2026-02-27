import csv
import io
import json
import logging
import os
import re
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Dict, Iterator
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

AUDIT_LOG_DIR = Path(os.getenv("AUDIT_LOG_DIR", "logs"))
AUDIT_LOG_NAME = os.getenv("AUDIT_LOG_NAME", "system_audit.log")
AUDIT_LOG_MAX_BYTES = int(os.getenv("AUDIT_LOG_MAX_BYTES", str(5 * 1024 * 1024)))
AUDIT_LOG_BACKUP_COUNT = int(os.getenv("AUDIT_LOG_BACKUP_COUNT", "7"))

AUDIT_CSV_FIELDS = [
    "Date (UTC)",
    "Time (UTC)",
    "User ID",
    "Username",
    "Role",
    "Action",
    "Action Details",
    "HTTP Method",
    "Endpoint",
    "Status",
    "Outcome",
    "Branch Scope",
    "Academic Year Scope",
    "Client IP",
    "Duration (ms)",
    "Error",
    "User Agent",
]


def get_audit_log_path() -> Path:
    return AUDIT_LOG_DIR / AUDIT_LOG_NAME


def get_audit_csv_filename() -> str:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    base_name = Path(AUDIT_LOG_NAME).stem
    return f"{base_name}_{timestamp}.csv"


def get_audit_xlsx_filename() -> str:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    base_name = Path(AUDIT_LOG_NAME).stem
    return f"{base_name}_{timestamp}.xlsx"


def get_audit_logger() -> logging.Logger:
    logger = logging.getLogger("tis.audit")
    if logger.handlers:
        return logger

    AUDIT_LOG_DIR.mkdir(parents=True, exist_ok=True)
    handler = RotatingFileHandler(
        filename=get_audit_log_path(),
        maxBytes=AUDIT_LOG_MAX_BYTES,
        backupCount=AUDIT_LOG_BACKUP_COUNT,
        encoding="utf-8",
    )
    handler.setFormatter(logging.Formatter("%(message)s"))

    logger.setLevel(logging.INFO)
    logger.addHandler(handler)
    logger.propagate = False
    return logger


def write_audit_event(event: Dict[str, Any]) -> None:
    payload = dict(event)
    payload["timestamp_utc"] = datetime.now(timezone.utc).isoformat()
    get_audit_logger().info(
        json.dumps(payload, separators=(",", ":"), ensure_ascii=True)
    )


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _split_utc_timestamp(timestamp_value: Any) -> tuple[str, str]:
    timestamp_text = _to_text(timestamp_value)
    if not timestamp_text:
        return "", ""

    normalized = timestamp_text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return "", timestamp_text

    parsed_utc = parsed.astimezone(timezone.utc)
    return (
        parsed_utc.strftime("%Y-%m-%d"),
        parsed_utc.strftime("%H:%M:%S"),
    )


def _extract_numeric_id(path: str) -> str:
    match = re.search(r"/(\d+)$", path)
    return match.group(1) if match else ""


def _classify_action(method: str, path: str) -> str:
    if method == "POST" and path == "/login":
        return "User Login"
    if method == "GET" and path == "/logout":
        return "User Logout"
    if method == "GET" and path == "/dashboard":
        return "View Dashboard"
    if method == "POST" and path == "/admin/current-year":
        return "Set Current Academic Year"
    if method == "POST" and path == "/developer/open-academic-year":
        return "Open New Academic Year"
    if method == "POST" and path == "/scope/branch":
        return "Switch Branch Scope"
    if method == "POST" and path == "/scope/academic-year":
        return "Switch Academic Year Scope"
    if method == "GET" and path == "/admin/audit-log":
        return "Download Audit Log"
    if method == "GET" and path == "/reports/allocation-plan.xlsx":
        return "Download Allocation Plan"

    if method == "GET" and path in {"/subjects", "/subjects/"}:
        return "View Subjects"
    if method == "POST" and path in {"/subjects", "/subjects/"}:
        return "Create Subject"
    if method == "POST" and path == "/subjects/import":
        return "Import Subjects"
    if method == "GET" and path == "/subjects/template":
        return "Download Subject Template"
    if method == "GET" and path.startswith("/subjects/edit/"):
        return "Open Subject Edit"
    if method == "POST" and path.startswith("/subjects/edit/"):
        return "Update Subject"
    if method == "GET" and path.startswith("/subjects/delete/"):
        return "Delete Subject"

    if method == "GET" and path in {"/users", "/users/"}:
        return "View Users"
    if method == "POST" and path in {"/users", "/users/"}:
        return "Create User"
    if method == "GET" and path.startswith("/users/edit/"):
        return "Open User Edit"
    if method == "POST" and path.startswith("/users/edit/"):
        return "Update User"
    if method == "GET" and path.startswith("/users/delete/"):
        return "Delete User"
    if method == "POST" and path == "/users/delete-bulk":
        return "Bulk Delete Users"

    if method == "GET" and path in {"/teachers", "/teachers/"}:
        return "View Teachers"
    if method == "POST" and path in {"/teachers", "/teachers/"}:
        return "Create Teacher"
    if method == "GET" and path.startswith("/teachers/edit/"):
        return "Open Teacher Edit"
    if method == "POST" and path.startswith("/teachers/edit/"):
        return "Update Teacher"
    if method == "GET" and path.startswith("/teachers/delete/"):
        return "Delete Teacher"

    return "System Action"


def _resolve_outcome(status_code: Any, error_name: str) -> str:
    if error_name:
        return "Error"

    try:
        code = int(status_code)
    except (TypeError, ValueError):
        return "Unknown"

    if code >= 500:
        return "Server Error"
    if code >= 400:
        return "Denied/Failed"
    if code >= 300:
        return "Redirect"
    return "Success"


def _build_action_details(event: Dict[str, Any], action: str) -> str:
    path = _to_text(event.get("path"))
    details = []
    target_id = _extract_numeric_id(path)
    if target_id and action in {
        "Delete Subject",
        "Update Subject",
        "Open Subject Edit",
        "Delete User",
        "Update User",
        "Open User Edit",
    }:
        details.append(f"Target ID: {target_id}")

    query = _to_text(event.get("query"))
    if query:
        details.append(f"Query: {query}")

    error_name = _to_text(event.get("error"))
    if error_name:
        details.append(f"Error: {error_name}")

    return " | ".join(details)


def _event_to_csv_row(event: Dict[str, Any]) -> Dict[str, str]:
    method = _to_text(event.get("method"))
    path = _to_text(event.get("path"))
    action = _classify_action(method, path)
    action_details = _build_action_details(event, action)
    date_text, time_text = _split_utc_timestamp(event.get("timestamp_utc"))
    status_text = _to_text(event.get("status_code"))
    error_text = _to_text(event.get("error"))

    return {
        "Date (UTC)": date_text,
        "Time (UTC)": time_text,
        "User ID": _to_text(event.get("actor_user_id")),
        "Username": _to_text(event.get("actor_username")),
        "Role": _to_text(event.get("actor_role")),
        "Action": action,
        "Action Details": action_details,
        "HTTP Method": method,
        "Endpoint": path,
        "Status": status_text,
        "Outcome": _resolve_outcome(status_text, error_text),
        "Branch Scope": _to_text(event.get("scope_branch_id")),
        "Academic Year Scope": _to_text(event.get("scope_academic_year_id")),
        "Client IP": _to_text(event.get("client_ip")),
        "Duration (ms)": _to_text(event.get("duration_ms")),
        "Error": error_text,
        "User Agent": _to_text(event.get("user_agent")),
    }


def iter_audit_rows(log_path: Path) -> Iterator[Dict[str, str]]:
    with log_path.open("r", encoding="utf-8", errors="replace") as source:
        for raw_line in source:
            line = raw_line.strip()
            if not line:
                continue

            try:
                event = json.loads(line)
                yield _event_to_csv_row(event)
            except json.JSONDecodeError:
                row = {
                    field: "" for field in AUDIT_CSV_FIELDS
                }
                row["Action"] = "Unparsed Log Entry"
                row["Action Details"] = line[:240]
                yield row


def iter_audit_csv_bytes(log_path: Path) -> Iterator[bytes]:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=AUDIT_CSV_FIELDS)
    writer.writeheader()
    yield buffer.getvalue().encode("utf-8")
    buffer.seek(0)
    buffer.truncate(0)

    for row in iter_audit_rows(log_path):
        writer.writerow(row)
        yield buffer.getvalue().encode("utf-8")
        buffer.seek(0)
        buffer.truncate(0)


def build_audit_xlsx_bytes(log_path: Path) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Audit Log"

    sheet.append(AUDIT_CSV_FIELDS)
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_index, header in enumerate(AUDIT_CSV_FIELDS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in iter_audit_rows(log_path):
        sheet.append([row.get(field, "") for field in AUDIT_CSV_FIELDS])

    sheet.freeze_panes = "A2"

    preferred_widths = {
        "A": 13,  # Date
        "B": 11,  # Time
        "C": 16,  # User ID
        "D": 18,  # Username
        "E": 14,  # Role
        "F": 28,  # Action
        "G": 46,  # Action Details
        "H": 12,  # Method
        "I": 34,  # Endpoint
        "J": 10,  # Status
        "K": 14,  # Outcome
        "L": 14,  # Branch Scope
        "M": 20,  # Academic Year Scope
        "N": 16,  # Client IP
        "O": 13,  # Duration
        "P": 20,  # Error
        "Q": 42,  # User Agent
    }
    for column_key, width in preferred_widths.items():
        sheet.column_dimensions[column_key].width = width

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
