from fastapi import FastAPI, Request, Form, Depends, Query, File, UploadFile
from fastapi.responses import RedirectResponse, HTMLResponse, PlainTextResponse, StreamingResponse, FileResponse
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import inspect, text, func, or_
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
import csv
import html
import io
import json
import logging
import math
import os
import re
import secrets
import smtplib
import ssl
import time
from typing import Optional, Any
from urllib.parse import quote, quote_plus, urlencode
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from database import engine, SessionLocal
import models
import auth
import authorization
import branding_storage
import db_migrations
import email_service
import email_templates
import location_service
import permission_registry
from visual_design import (
    VISUAL_COMPONENT_MAP,
    build_visual_design_config,
    is_custom_component_key,
    normalize_visual_payload,
    rows_to_visual_settings,
)
from dependencies import get_db
from routers import subjects, users, teachers, planning, timetable, academic_calendar, observations
from auth import get_password_hash
from models import User, Branch, AcademicYear
from teacher_capacity import (
    get_teacher_capacity_breakdown,
)
from ui_shell import DEFAULT_SCHOOL_LOGO_SLOTS, build_shell_context, get_school_logo_slots
from audit import (
    get_audit_log_path,
    get_audit_logger,
    write_audit_event,
    iter_audit_csv_bytes,
    get_audit_csv_filename,
    build_audit_xlsx_bytes,
    get_audit_xlsx_filename,
)
from homeroom_defaults import (
    get_effective_subject_count,
    get_homeroom_bundle_subject_labels,
    is_default_homeroom_subject,
    is_homeroom_bundle_subject,
    is_lower_primary_homeroom_grade,
)
from subject_colors import (
    build_subject_theme,
    generate_subject_color_by_code,
    normalize_hex_color,
    resolve_subject_color,
    to_excel_hex,
)
from teacher_qualifications import (
    QUALIFICATION_KIND_DEGREE,
    QUALIFICATION_KIND_SPECIALIZATION,
    build_qualification_key,
    ensure_qualification_options_seeded,
    get_subject_alignment_group_keys,
)
from timetable_logic import (
    ALL_DAY_KEY,
    BLOCK_TYPE_OPTIONS,
    WORKING_DAY_OPTIONS,
    build_time_slots,
    get_timetable_setting_row,
    get_timetable_settings_payload,
    normalize_non_teaching_block_values,
    normalize_timetable_settings_values,
    validate_non_teaching_block_overlap,
)

# ---------------------------------------
# Create Tables
# ---------------------------------------
models.Base.metadata.create_all(bind=engine)
db_migrations.run_pending_migrations(engine)

# ---------------------------------------
# App Initialization
# ---------------------------------------
app = FastAPI(title="Teacher Information System")
app.mount(
    "/static",
    branding_storage.ProtectedBrandingStaticFiles(directory="static"),
    name="static",
)
app.mount(
    "/landing-public",
    StaticFiles(directory="tis-landing-website/public"),
    name="landing_public",
)

templates = Jinja2Templates(directory="templates")
ACADEMIC_YEAR_NAME_PATTERN = re.compile(r"^\d{4}-\d{4}$")
PUBLIC_LANDING_HOSTS = {"tisplatform.com", "www.tisplatform.com"}
DEMO_REQUEST_EMAIL_TO = os.getenv("TIS_DEMO_REQUEST_EMAIL_TO", "info@tisplatform.com")
DEMO_REQUEST_FIELDS = (
    ("school_name", "School Name"),
    ("full_name", "Contact Name"),
    ("email", "Email"),
    ("phone", "Phone"),
    ("country", "Country"),
    ("school_type", "School Type"),
    ("teachers", "Number of Teachers"),
    ("students", "Number of Students"),
    ("branches", "Number of Branches"),
    ("interested_plan", "Interested Plan"),
    ("message", "Message"),
)
DEMO_REQUEST_STATUSES = (
    "New",
    "Contacted",
    "Demo Scheduled",
    "Converted",
    "Not a Fit",
)


def _get_positive_int_env(name: str, default: int) -> int:
    raw_value = str(os.getenv(name, "") or "").strip()
    if not raw_value:
        return default
    try:
        parsed_value = int(raw_value)
    except ValueError:
        return default
    return parsed_value if parsed_value > 0 else default


def _get_bool_env(name: str, default: bool = False) -> bool:
    raw_value = str(os.getenv(name, "") or "").strip().lower()
    if not raw_value:
        return default
    if raw_value in {"1", "true", "yes", "on"}:
        return True
    if raw_value in {"0", "false", "no", "off"}:
        return False
    return default


IDLE_TIMEOUT_COOKIE_KEY = "last_activity_ts"
IDLE_TIMEOUT_MINUTES = _get_positive_int_env("TIS_IDLE_TIMEOUT_MINUTES", 30)
IDLE_TIMEOUT_SECONDS = IDLE_TIMEOUT_MINUTES * 60
IDLE_TIMEOUT_LOGIN_MESSAGE = "Session timed out due to inactivity. Please log in again."
AUTH_SESSION_COOKIE_KEYS = (
    auth.SESSION_COOKIE_KEY,
    "user_id",
    "branch_id",
    "school_group_id",
    "academic_year_id",
    IDLE_TIMEOUT_COOKIE_KEY,
)
IDLE_TIMEOUT_EXEMPT_PATHS = {
    "/",
    "/login",
    "/logout",
    "/forgot-password",
    "/request-demo",
    "/favicon.ico",
}
IDLE_TIMEOUT_EXEMPT_PREFIXES = (
    "/static/",
)


def _clear_auth_session_cookies(response):
    for cookie_key in AUTH_SESSION_COOKIE_KEYS:
        response.delete_cookie(cookie_key)
    return response


def _request_hostname(request: Request) -> str:
    host = str(request.headers.get("host") or request.url.hostname or "").strip().lower()
    if not host:
        return ""
    if host.startswith("["):
        return host.split("]", 1)[0].lstrip("[")
    return host.split(":", 1)[0].rstrip(".")


def _is_public_landing_host(request: Request) -> bool:
    return _request_hostname(request) in PUBLIC_LANDING_HOSTS


REPORT_STANDARD_MAX_HOURS = 24
# Version 21: General Science auto matching keeps Science as the lead family
# before Biology, Chemistry, and ICT, including saved/effective plan payloads.
HIRING_PLAN_POOL_LOGIC_VERSION = 21
CROSS_SUBJECT_SUPPORT_RULES = {
    "english": {"social studies english"},
    "arabic": {"social studies ksa"},
    "arbic": {"social studies ksa"},
}
PRIORITY_STAFFING_SUBJECT_PREFIXES = {
    "english": ("english",),
    "arabic": ("arabic", "arbic"),
    "math": ("mathematics", "math", "maths"),
    "science": ("science", "general science", "sce", "biology", "chemistry", "physics"),
    "islamic": ("islamic studies", "islamic", "quran", "holy quran"),
}
HIRING_FAMILY_PRIORITY = {
    "english": 0,
    "social_english": 1,
    "social": 1,
    "wellbeing": 2,
    "reflection": 3,
    "performing_arts": 4,
    "art": 5,
    "arabic": 6,
    "islamic": 7,
    "quran": 7,
    "social_arabic": 8,
    "math": 9,
    "mental_math": 9,
    "physics": 10,
    "science": 11,
    "biology": 12,
    "chemistry": 13,
    "ict": 14,
    "pe": 15,
    "homeroom": 16,
}
HIRING_COMPATIBILITY_GROUPS = {
    "english": "english_pool",
    "social_english": "english_pool",
    "social": "english_pool",
    "wellbeing": "english_pool",
    "reflection": "english_pool",
    "performing_arts": "english_pool",
    "art": "english_pool",
    "math": "math_pool",
    "mental_math": "math_pool",
    "physics": "math_pool",
    "science": "general_science_pool",
    "biology": "general_science_pool",
    "chemistry": "general_science_pool",
    "ict": "general_science_pool",
    "arabic": "arabic_pool",
    "islamic": "arabic_pool",
    "quran": "arabic_pool",
    "social_arabic": "arabic_pool",
    "pe": "physical_education",
    "homeroom": "homeroom_pool",
}
HIRING_GROUP_LABELS = {
    "english_pool": "English Pool",
    "arabic_pool": "Arabic Pool",
    "math_pool": "Math Pool",
    "general_science_pool": "General Science Pool",
    "physical_education": "Physical Education Pool",
    "homeroom_pool": "Homeroom Pool",
}
HIRING_POOL_ACCENT_COLORS = {
    "english_pool": "#2563EB",
    "arabic_pool": "#0F766E",
    "math_pool": "#7C3AED",
    "general_science_pool": "#1D4ED8",
    "physical_education": "#EA580C",
    "homeroom_pool": "#B45309",
}
HIRING_POOL_FAMILY_ORDER = {
    "english_pool": ["english", "social_english", "social", "wellbeing", "reflection", "performing_arts", "art"],
    "arabic_pool": ["arabic", "islamic", "quran", "social_arabic"],
    "math_pool": ["math", "mental_math", "physics"],
    "general_science_pool": ["science", "biology", "chemistry", "ict"],
    "physical_education": ["pe"],
    "homeroom_pool": ["homeroom"],
}
HIRING_NAMED_POOL_KEYS = {
    "english_pool",
    "math_pool",
    "arabic_pool",
    "general_science_pool",
    "physical_education",
    "homeroom_pool",
}
HIRING_PROFILE_GROUP_LABEL_KEYS = HIRING_NAMED_POOL_KEYS
HIRING_POOL_ALLOWED_FAMILIES = {
    "english_pool": {"english", "social_english", "social", "wellbeing", "reflection", "performing_arts", "art"},
    "math_pool": {"math", "mental_math", "physics"},
    "general_science_pool": {"science", "biology", "chemistry", "ict"},
    "arabic_pool": {"arabic", "islamic", "quran", "social_arabic"},
    "physical_education": {"pe"},
    "homeroom_pool": {"homeroom"},
}
HIRING_GENERAL_SCIENCE_FAMILIES = {"science", "biology", "chemistry", "ict"}
HIRING_FAMILY_LABELS = {
    "english": "English",
    "arabic": "Arabic",
    "math": "Mathematics",
    "mental_math": "Mental Math",
    "physics": "Physics",
    "science": "Science",
    "biology": "Biology",
    "chemistry": "Chemistry",
    "ict": "ICT",
    "islamic": "Islamic",
    "quran": "Quran",
    "social_arabic": "Social Studies Arabic",
    "social_english": "Social Studies English",
    "social": "Social Studies",
    "pe": "Physical Education",
    "homeroom": "Homeroom",
    "wellbeing": "Well Being",
    "art": "Art",
    "performing_arts": "Performing Arts",
    "reflection": "Reflection",
}
REPORT_EXPORT_SUBJECT_FILL_PALETTE = [
    "E8F1FF",
    "EAF8F4",
    "FFF4E6",
    "FDECF3",
    "EEEAFE",
    "E9F8FF",
    "F2F7E8",
    "FCEBEB",
]
get_audit_logger()

FAVICON_IMAGE_PATH = str(
    branding_storage.tis_logo_absolute_path(theme="light", compact=True)
)
FAVICON_CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}
PROFILE_PHOTO_UPLOAD_DIR = os.path.join("static", "uploads", "profile_photos")
PROFILE_PHOTO_RELATIVE_DIR = "uploads/profile_photos"
PROFILE_PHOTO_MAX_BYTES = 3 * 1024 * 1024
DEFAULT_SCHOOL_GROUP_NAME = "Al-Andalus Schools"
MAJOR_ALIGNMENT_STOPWORDS = {
    "a",
    "an",
    "and",
    "b",
    "ba",
    "bed",
    "bs",
    "bsc",
    "certificate",
    "degree",
    "diploma",
    "ed",
    "education",
    "for",
    "in",
    "ma",
    "major",
    "masters",
    "minor",
    "msc",
    "of",
    "phd",
    "teaching",
    "the",
    "with",
}


def _ensure_profile_photo_upload_dir():
    os.makedirs(PROFILE_PHOTO_UPLOAD_DIR, exist_ok=True)


def _ensure_school_group_schema():
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())
    if "school_groups" not in table_names:
        models.SchoolGroup.__table__.create(bind=engine, checkfirst=True)
    if "school_group_logos" not in table_names:
        models.SchoolGroupLogo.__table__.create(bind=engine, checkfirst=True)


def _ensure_subject_color_schema():
    # Subject color schema is managed by db_migrations.
    return


def _ensure_role_permission_schema():
    models.RolePermission.__table__.create(bind=engine, checkfirst=True)


def _backfill_subject_colors(db: Session):
    subjects = db.query(models.Subject).all()
    changes_made = False
    for subject in subjects:
        subject_code = getattr(subject, "subject_code", "")
        subject_name = getattr(subject, "subject_name", "")
        stored_color = normalize_hex_color(getattr(subject, "color", ""))
        expected_color = resolve_subject_color(
            subject_code,
            subject_name=subject_name,
        )
        legacy_color = generate_subject_color_by_code(subject_code)
        normalized_current_color = normalize_hex_color(getattr(subject, "color", ""))
        resolved_color = stored_color
        if stored_color is None or stored_color == legacy_color:
            resolved_color = expected_color
        if normalized_current_color != resolved_color:
            subject.color = resolved_color
            changes_made = True

    if changes_made:
        db.commit()


def _detect_profile_photo_extension(file_bytes: bytes) -> str:
    if file_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if file_bytes.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if file_bytes.startswith((b"GIF87a", b"GIF89a")):
        return ".gif"
    if (
        len(file_bytes) >= 12
        and file_bytes[:4] == b"RIFF"
        and file_bytes[8:12] == b"WEBP"
    ):
        return ".webp"
    return ""


def _profile_photo_media_type_from_extension(extension: str) -> str:
    normalized_extension = str(extension or "").strip().lower()
    if normalized_extension == ".png":
        return "image/png"
    if normalized_extension in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if normalized_extension == ".gif":
        return "image/gif"
    if normalized_extension == ".webp":
        return "image/webp"
    return "application/octet-stream"


def _normalize_profile_photo_relative_path(relative_path: str) -> str:
    return str(relative_path or "").replace("\\", "/").lstrip("/")


def _delete_profile_photo_file(relative_path: str):
    normalized_relative_path = _normalize_profile_photo_relative_path(relative_path)
    if not normalized_relative_path.startswith(f"{PROFILE_PHOTO_RELATIVE_DIR}/"):
        return

    absolute_path = os.path.abspath(
        os.path.join("static", *normalized_relative_path.split("/"))
    )
    upload_root = os.path.abspath(PROFILE_PHOTO_UPLOAD_DIR)
    if not absolute_path.startswith(upload_root):
        return
    if os.path.exists(absolute_path):
        try:
            os.remove(absolute_path)
        except OSError:
            return


def _migrate_profile_photos_to_database(db: Session):
    users_needing_migration = db.query(User).filter(
        User.profile_image_path.isnot(None),
        User.profile_image_path != "",
        User.profile_image_data.is_(None),
    ).all()

    if not users_needing_migration:
        return

    has_changes = False
    for user_row in users_needing_migration:
        normalized_relative_path = _normalize_profile_photo_relative_path(
            getattr(user_row, "profile_image_path", "") or ""
        )
        if not normalized_relative_path.startswith(f"{PROFILE_PHOTO_RELATIVE_DIR}/"):
            continue

        absolute_path = os.path.abspath(
            os.path.join("static", *normalized_relative_path.split("/"))
        )
        upload_root = os.path.abspath(PROFILE_PHOTO_UPLOAD_DIR)
        if not absolute_path.startswith(upload_root):
            continue
        if not os.path.exists(absolute_path):
            continue

        try:
            with open(absolute_path, "rb") as image_file:
                file_bytes = image_file.read()
        except OSError:
            continue

        if not file_bytes:
            continue

        detected_extension = _detect_profile_photo_extension(file_bytes)
        if not detected_extension:
            _, detected_extension = os.path.splitext(absolute_path)

        user_row.profile_image_content_type = _profile_photo_media_type_from_extension(
            detected_extension
        )
        user_row.profile_image_data = file_bytes
        has_changes = True

    if has_changes:
        db.commit()


def _safe_redirect_path(path: str) -> str:
    cleaned = str(path or "").strip()
    if not cleaned.startswith("/") or cleaned.startswith("//"):
        return "/dashboard"
    return cleaned


def _redirect_with_notice(path: str, notice: str):
    safe_path = _safe_redirect_path(path)
    base_path, _, fragment = safe_path.partition("#")
    separator = "&" if "?" in base_path else "?"
    url = f"{base_path}{separator}notice={quote_plus(str(notice or '').strip())}"
    if fragment:
        url = f"{url}#{fragment}"
    return RedirectResponse(url=url, status_code=302)


def _redirect_with_error(path: str, error: str):
    safe_path = _safe_redirect_path(path)
    base_path, _, fragment = safe_path.partition("#")
    separator = "&" if "?" in base_path else "?"
    url = f"{base_path}{separator}error={quote_plus(str(error or '').strip())}"
    if fragment:
        url = f"{url}#{fragment}"
    return RedirectResponse(url=url, status_code=302)


SAUDI_REGIONS = (
    "Riyadh Region",
    "Makkah Region",
    "Madinah Region",
    "Eastern Province",
    "Al Qassim Region",
    "Hail Region",
    "Tabuk Region",
    "Northern Borders Region",
    "Al Jawf Region",
    "Jazan Region",
    "Najran Region",
    "Al Bahah Region",
    "Asir Region",
)
SAUDI_REGION_LOOKUP = {
    region.casefold(): region
    for region in SAUDI_REGIONS
}


def _normalize_branch_region(value: str) -> str | None:
    cleaned = " ".join(str(value or "").split())
    if not cleaned:
        return None
    return SAUDI_REGION_LOOKUP.get(cleaned.casefold())


def _resolve_submitted_location(
    *,
    country_code: str,
    region_id: str,
    region_manual: str,
    city_id: str,
    city_manual: str,
    legacy_region: str = "",
    required: bool = True,
) -> location_service.ResolvedLocation | None:
    modern_values = (country_code, region_id, region_manual, city_id, city_manual)
    if any(str(value or "").strip() for value in modern_values):
        if not required and not str(region_id or "").strip() and not str(region_manual or "").strip():
            return None
        return location_service.resolve_location(
            country_code=country_code,
            region_id=region_id,
            region_manual=region_manual,
            city_id=city_id,
            city_manual=city_manual,
            require_city=required,
        )

    legacy_location = location_service.infer_legacy_saudi_location(legacy_region)
    if legacy_location:
        return legacy_location
    if required:
        raise location_service.LocationValidationError(
            "Select a country, region/state/province, and city."
        )
    return None


def _normalize_location_detail(value: object, label: str) -> str:
    cleaned = " ".join(str(value or "").split())
    if any(ord(character) < 32 or ord(character) == 127 for character in cleaned):
        raise location_service.LocationValidationError(
            f"{label} contains unsupported control characters."
        )
    if len(cleaned) > 160:
        raise location_service.LocationValidationError(
            f"{label} must be 160 characters or fewer."
        )
    return cleaned


def _bulk_json_error(errors: list[dict[str, object]], status_code: int = 422):
    return JSONResponse(
        {
            "ok": False,
            "errors": errors,
        },
        status_code=status_code,
    )


def _bulk_payload_items(payload: object) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    if not isinstance(payload, dict) or not isinstance(payload.get("items"), list):
        return [], [
            {
                "id": None,
                "label": "Changes",
                "field": "items",
                "message": "Submit a valid list of changed items.",
            }
        ]
    raw_items = payload.get("items") or []
    if not raw_items:
        return [], [
            {
                "id": None,
                "label": "Changes",
                "field": "items",
                "message": "No modified items were submitted.",
            }
        ]
    if len(raw_items) > 200:
        return [], [
            {
                "id": None,
                "label": "Changes",
                "field": "items",
                "message": "Save no more than 200 modified items at once.",
            }
        ]
    if not all(isinstance(item, dict) for item in raw_items):
        return [], [
            {
                "id": None,
                "label": "Changes",
                "field": "items",
                "message": "Each modified item must be a valid record.",
            }
        ]
    return list(raw_items), []


def _bulk_item_error(
    *,
    item_id: object,
    label: str,
    field: str,
    message: str,
) -> dict[str, object]:
    return {
        "id": item_id,
        "label": str(label or "Record").strip() or "Record",
        "field": field,
        "message": message,
    }


def _branch_usage_counts(db: Session, branch_id: int) -> dict[str, int]:
    return {
        "users_count": db.query(models.User).filter(
            models.User.branch_id == branch_id
        ).count(),
        "subjects_count": db.query(models.Subject).filter(
            models.Subject.branch_id == branch_id
        ).count(),
        "teachers_count": db.query(models.Teacher).filter(
            models.Teacher.branch_id == branch_id
        ).count(),
        "planning_sections_count": db.query(models.PlanningSection).filter(
            models.PlanningSection.branch_id == branch_id
        ).count(),
    }


ACADEMIC_YEAR_USAGE_MODELS = (
    ("users_count", "users", models.User),
    ("notifications_count", "notifications", models.SystemNotification),
    ("calendar_event_types_count", "calendar event types", models.CalendarEventType),
    ("calendar_events_count", "calendar events", models.CalendarEvent),
    ("subjects_count", "subjects", models.Subject),
    ("teachers_count", "teachers", models.Teacher),
    ("observations_count", "observations", models.Observation),
    ("planning_sections_count", "planning sections", models.PlanningSection),
    ("timetable_settings_count", "timetable settings", models.TimetableSetting),
    ("timetable_entries_count", "timetable entries", models.TimetableEntry),
    ("hiring_plan_drafts_count", "hiring plan drafts", models.HiringPlanDraft),
)


def _academic_year_usage_counts(db: Session, academic_year_id: int) -> dict[str, int]:
    return {
        key: db.query(model).filter(
            model.academic_year_id == academic_year_id
        ).count()
        for key, _label, model in ACADEMIC_YEAR_USAGE_MODELS
    }


def _build_academic_year_configuration_rows(
    db: Session,
    school_group_id: int,
) -> list[dict[str, object]]:
    rows = []
    academic_years = db.query(models.AcademicYear).filter(
        models.AcademicYear.school_group_id == school_group_id
    ).order_by(models.AcademicYear.year_name.desc()).all()
    for academic_year in academic_years:
        usage_counts = _academic_year_usage_counts(db, academic_year.id)
        linked_records_count = sum(int(count or 0) for count in usage_counts.values())
        usage_summary = ", ".join(
            f"{usage_counts[key]} {label}"
            for key, label, _model in ACADEMIC_YEAR_USAGE_MODELS
            if usage_counts[key]
        )
        rows.append(
            {
                "id": academic_year.id,
                "year_name": academic_year.year_name,
                "is_active": bool(academic_year.is_active),
                "usage_counts": usage_counts,
                "linked_records_count": linked_records_count,
                "usage_summary": usage_summary,
                "can_delete": linked_records_count == 0,
            }
        )
    return rows


def _build_branch_configuration_rows(
    db: Session,
    *,
    scoped_branch_id: int | None = None,
) -> list[dict[str, object]]:
    branch_rows = []
    branches = db.query(models.Branch).order_by(
        models.Branch.status.desc(),
        models.Branch.name.asc(),
    ).all()
    active_branch_count = sum(1 for branch in branches if bool(branch.status))

    for branch in branches:
        usage_counts = _branch_usage_counts(db, branch.id)
        linked_records_count = sum(int(value or 0) for value in usage_counts.values())
        can_delete = linked_records_count == 0 and (
            not bool(branch.status) or active_branch_count > 1
        )
        can_deactivate = not bool(branch.status) or active_branch_count > 1
        saved_region = str(branch.location or "").strip()
        normalized_region = _normalize_branch_region(saved_region)
        legacy_location = location_service.infer_legacy_saudi_location(saved_region)

        branch_rows.append(
            {
                "id": branch.id,
                "school_group_id": getattr(branch, "school_group_id", None),
                "name": str(branch.name or "").strip(),
                "country_code": str(getattr(branch, "country_code", "") or "").strip()
                or (legacy_location.country_code if legacy_location else ""),
                "country_name": str(getattr(branch, "country_name", "") or "").strip()
                or (legacy_location.country_name if legacy_location else ""),
                "region": legacy_location.region_name if legacy_location else (
                    str(getattr(branch, "region_name", "") or "").strip()
                    or normalized_region
                    or saved_region
                ),
                "city": str(getattr(branch, "city_name", "") or "").strip(),
                "district": str(getattr(branch, "district_name", "") or "").strip(),
                "neighborhood": str(
                    getattr(branch, "neighborhood_name", "") or ""
                ).strip(),
                "status": bool(branch.status),
                "is_current_scope": scoped_branch_id == branch.id,
                "usage_counts": usage_counts,
                "linked_records_count": linked_records_count,
                "can_delete": can_delete,
                "can_deactivate": can_deactivate,
            }
        )

    return branch_rows


def _build_school_delete_summary(db: Session, school_group_id: int | None) -> dict[str, object]:
    if not school_group_id:
        return {
            "can_delete": False,
            "reason": "No school selected.",
            "branch_count": 0,
            "academic_years_count": 0,
            "school_logos_count": 0,
            "branch_logos_count": 0,
            "linked_records_count": 0,
        }

    branch_ids = [
        branch_id
        for (branch_id,) in db.query(models.Branch.id).filter(
            models.Branch.school_group_id == school_group_id
        ).all()
    ]
    branch_usage_counts = {
        branch_id: _branch_usage_counts(db, branch_id)
        for branch_id in branch_ids
    }
    linked_records_count = sum(
        int(count or 0)
        for usage_counts in branch_usage_counts.values()
        for count in usage_counts.values()
    )
    academic_years_count = db.query(models.AcademicYear).filter(
        models.AcademicYear.school_group_id == school_group_id
    ).count()
    school_logos_count = db.query(models.SchoolGroupLogo).filter(
        models.SchoolGroupLogo.school_group_id == school_group_id
    ).count()
    branch_logos_count = 0
    if branch_ids:
        branch_logos_count = db.query(models.BranchLogo).filter(
            models.BranchLogo.branch_id.in_(branch_ids)
        ).count()
    school_count = db.query(models.SchoolGroup).count()

    blockers = []
    if school_count <= 1:
        blockers.append("at least one school must remain")
    if linked_records_count:
        blockers.append(f"{linked_records_count} linked branch records")
    if academic_years_count:
        blockers.append(f"{academic_years_count} academic years")
    if school_logos_count or branch_logos_count:
        blockers.append(f"{school_logos_count + branch_logos_count} logo records")

    return {
        "can_delete": not blockers,
        "reason": "; ".join(blockers),
        "branch_count": len(branch_ids),
        "academic_years_count": academic_years_count,
        "school_logos_count": school_logos_count,
        "branch_logos_count": branch_logos_count,
        "linked_records_count": linked_records_count,
    }


def _get_role_permission_rows(
    db: Session,
    role: str,
    school_group_id: int | None = None,
) -> list[models.RolePermission]:
    normalized_role = permission_registry.normalize_managed_role(role)
    if not normalized_role:
        return []

    query = db.query(models.RolePermission).filter(
        models.RolePermission.role == normalized_role
    )
    if school_group_id is None:
        query = query.filter(models.RolePermission.school_group_id.is_(None))
    else:
        query = query.filter(models.RolePermission.school_group_id == school_group_id)
    return query.all()


def _get_allowed_permission_keys(
    db: Session,
    role: str,
    school_group_id: int | None = None,
) -> set[str]:
    normalized_role = permission_registry.normalize_managed_role(role)
    if not normalized_role:
        return set()
    allowed_keys = permission_registry.get_default_permissions_for_role(normalized_role)
    for permission_row in _get_role_permission_rows(db, normalized_role, None):
        if permission_row.permission_key in permission_registry.PERMISSION_LABELS:
            if permission_row.is_allowed:
                allowed_keys.add(permission_row.permission_key)
            else:
                allowed_keys.discard(permission_row.permission_key)
    if school_group_id:
        for permission_row in _get_role_permission_rows(db, normalized_role, school_group_id):
            if permission_row.permission_key in permission_registry.PERMISSION_LABELS:
                if permission_row.is_allowed:
                    allowed_keys.add(permission_row.permission_key)
            else:
                allowed_keys.discard(permission_row.permission_key)
    return permission_registry.constrain_role_permissions(normalized_role, allowed_keys)


def _build_role_permission_payload(
    db: Session,
    role: str,
    school_group_id: int | None = None,
) -> dict:
    allowed_keys = _get_allowed_permission_keys(db, role, school_group_id)
    return permission_registry.build_role_permission_payload(role, allowed_keys)


def _build_role_permission_summary_map(
    db: Session,
    school_group_id: int | None = None,
) -> dict[str, dict[str, object]]:
    return {
        role: _build_role_permission_payload(db, role, school_group_id)
        for role in permission_registry.MANAGED_ROLES
    }


def _set_role_permission_rows(
    db: Session,
    *,
    role: str,
    allowed_keys: set[str],
    school_group_id: int | None,
    updated_by_user_id: str | None,
):
    normalized_role = permission_registry.normalize_managed_role(role)
    if not normalized_role:
        return

    valid_keys = set(permission_registry.ALL_PERMISSION_KEYS)
    allowed_keys = permission_registry.constrain_role_permissions(
        normalized_role,
        allowed_keys & valid_keys,
    )
    existing_rows = {
        row.permission_key: row
        for row in _get_role_permission_rows(db, normalized_role, school_group_id)
    }
    now = datetime.utcnow()
    for permission_key in valid_keys:
        is_allowed = permission_key in allowed_keys
        row = existing_rows.get(permission_key)
        if row:
            row.is_allowed = is_allowed
            row.updated_by_user_id = updated_by_user_id
            row.updated_at = now
        else:
            db.add(
                models.RolePermission(
                    school_group_id=school_group_id,
                    role=normalized_role,
                    permission_key=permission_key,
                    is_allowed=is_allowed,
                    updated_by_user_id=updated_by_user_id,
                    created_at=now,
                    updated_at=now,
                )
            )


def _seed_global_role_permissions(db: Session):
    for role in permission_registry.MANAGED_ROLES:
        if _get_role_permission_rows(db, role, None):
            continue
        _set_role_permission_rows(
            db,
            role=role,
            allowed_keys=permission_registry.get_default_permissions_for_role(role),
            school_group_id=None,
            updated_by_user_id="system",
        )
    db.commit()


def _ensure_default_school_group(db: Session):
    school_group = db.query(models.SchoolGroup).filter(
        models.SchoolGroup.name == DEFAULT_SCHOOL_GROUP_NAME
    ).first()
    if not school_group:
        legacy_group = db.query(models.SchoolGroup).filter(
            models.SchoolGroup.name.in_(["Al Andalus Schools", "Andalus Schools"])
        ).first()
        if legacy_group:
            legacy_group.name = DEFAULT_SCHOOL_GROUP_NAME
            legacy_group.updated_at = datetime.utcnow()
            school_group = legacy_group
    if not school_group:
        school_group = db.query(models.SchoolGroup).order_by(
            models.SchoolGroup.id.asc()
        ).first()
    if not school_group:
        school_group = models.SchoolGroup(
            name=DEFAULT_SCHOOL_GROUP_NAME,
            status=True,
            country_code="SA",
            country_name="Saudi Arabia",
        )
        db.add(school_group)
        db.flush()

    changed = False
    if (
        school_group.name == DEFAULT_SCHOOL_GROUP_NAME
        and not str(getattr(school_group, "country_code", "") or "").strip()
    ):
        school_group.country_code = "SA"
        school_group.country_name = "Saudi Arabia"
        changed = True
    branches_without_group = db.query(models.Branch).filter(
        models.Branch.school_group_id.is_(None)
    ).all()
    for branch in branches_without_group:
        branch.school_group_id = school_group.id
        changed = True

    academic_years_without_group = db.query(models.AcademicYear).filter(
        models.AcademicYear.school_group_id.is_(None)
    ).all()
    for academic_year in academic_years_without_group:
        academic_year.school_group_id = school_group.id
        changed = True

    all_school_groups = db.query(models.SchoolGroup).all()
    for item in all_school_groups:
        has_branch = db.query(models.Branch).filter(
            models.Branch.school_group_id == item.id
        ).first()
        if not has_branch:
            db.add(
                models.Branch(
                    school_group_id=item.id,
                    name=item.name,
                    location="Makkah",
                    country_code="SA",
                    country_name="Saudi Arabia",
                    region_name="Makkah",
                    status=True,
                )
            )
            changed = True

    if changed:
        school_group.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(school_group)
    return school_group


AL_ANDALUS_LEGACY_LOGOS = (
    {
        "slot_key": "primary",
        "label": "Little Andalus International Schools",
        "source": branding_storage.STATIC_ROOT / "images" / "andalus-logo.png",
        "filename": "primary.png",
    },
    {
        "slot_key": "accreditation",
        "label": "Cognia",
        "source": branding_storage.STATIC_ROOT / "images" / "cognia-logo.png",
        "filename": "accreditation.png",
    },
    {
        "slot_key": "secondary",
        "label": "Andalus International Schools",
        "source": branding_storage.STATIC_ROOT / "images" / "andalus-logo-main.png",
        "filename": "secondary.png",
    },
)
AL_ANDALUS_GROUP_NAMES = {
    DEFAULT_SCHOOL_GROUP_NAME,
    "Al Andalus Schools",
    "Andalus Schools",
}


def _ensure_organization_branding_storage(db: Session):
    logger = logging.getLogger(__name__)
    school_groups = db.query(models.SchoolGroup).all()
    for school_group in school_groups:
        branding_storage.ensure_organization_logo_dir(school_group.id)

    changed = False
    for logo in db.query(models.SchoolGroupLogo).all():
        try:
            migrated_path = branding_storage.migrate_legacy_logo_file(
                logo.image_path,
                school_group_id=logo.school_group_id,
            )
        except (branding_storage.BrandingStorageError, FileNotFoundError) as exc:
            logger.warning(
                "Unable to migrate school-group logo %s: %s",
                getattr(logo, "id", ""),
                exc,
            )
            continue
        if migrated_path != logo.image_path:
            logo.image_path = migrated_path
            logo.updated_at = datetime.utcnow()
            changed = True

    branch_rows = {
        branch.id: branch for branch in db.query(models.Branch).all()
    }
    for logo in db.query(models.BranchLogo).all():
        branch = branch_rows.get(logo.branch_id)
        school_group_id = getattr(branch, "school_group_id", None)
        if not school_group_id:
            continue
        try:
            migrated_path = branding_storage.migrate_legacy_logo_file(
                logo.image_path,
                school_group_id=school_group_id,
                branch_id=branch.id,
            )
        except (branding_storage.BrandingStorageError, FileNotFoundError) as exc:
            logger.warning(
                "Unable to migrate branch logo %s: %s",
                getattr(logo, "id", ""),
                exc,
            )
            continue
        if migrated_path != logo.image_path:
            logo.image_path = migrated_path
            logo.updated_at = datetime.utcnow()
            changed = True

    al_andalus = next(
        (
            group
            for group in school_groups
            if str(group.name or "").strip() in AL_ANDALUS_GROUP_NAMES
        ),
        None,
    )
    if al_andalus:
        existing_slots = {
            row.slot_key
            for row in db.query(models.SchoolGroupLogo).filter(
                models.SchoolGroupLogo.school_group_id == al_andalus.id
            ).all()
        }
        slot_definitions = {
            slot["slot_key"]: slot for slot in DEFAULT_SCHOOL_LOGO_SLOTS
        }
        for legacy_logo in AL_ANDALUS_LEGACY_LOGOS:
            if legacy_logo["slot_key"] in existing_slots:
                continue
            try:
                relative_path = branding_storage.copy_legacy_default_logo(
                    legacy_logo["source"],
                    school_group_id=al_andalus.id,
                    filename=legacy_logo["filename"],
                )
            except FileNotFoundError as exc:
                logger.warning("Unable to copy Al-Andalus legacy logo: %s", exc)
                continue
            slot = slot_definitions[legacy_logo["slot_key"]]
            db.add(
                models.SchoolGroupLogo(
                    school_group_id=al_andalus.id,
                    slot_key=legacy_logo["slot_key"],
                    label=legacy_logo["label"],
                    image_path=relative_path,
                    content_type="image/png",
                    sort_order=slot["sort_order"],
                    updated_by_user_id="system",
                )
            )
            changed = True

    if changed:
        db.commit()


def _normalize_qualification_label(value: str) -> str:
    return " ".join(str(value or "").split()).strip()


def _normalize_qualification_kind(value: str) -> str:
    normalized_value = str(value or "").strip().lower()
    if normalized_value == QUALIFICATION_KIND_DEGREE:
        return QUALIFICATION_KIND_DEGREE
    return QUALIFICATION_KIND_SPECIALIZATION


def _build_qualification_configuration_rows(
    db: Session,
) -> list[dict[str, object]]:
    qualification_rows = []
    for option in ensure_qualification_options_seeded(db):
        usage_count = db.query(models.TeacherQualificationSelection).filter(
            models.TeacherQualificationSelection.qualification_key == option["key"]
        ).count()
        qualification_rows.append(
            {
                "key": option["key"],
                "label": option["label"],
                "kind": option["kind"],
                "group_label": (
                    "Degrees"
                    if option["kind"] == QUALIFICATION_KIND_DEGREE
                    else "Majors / Teaching Specializations"
                ),
                "alignment_keys": ", ".join(option["alignment_keys"]),
                "legacy_aliases": ", ".join(option["legacy_aliases"]),
                "sort_order": int(option.get("sort_order", 0) or 0),
                "usage_count": usage_count,
                "can_delete": usage_count == 0,
            }
        )
    return qualification_rows


CONFIGURATION_MODULES = (
    {
        "key": "overview",
        "label": "Overview",
        "href": "/system-configuration",
        "icon": "settings",
        "description": "Open the configuration hub.",
        "permission_keys": (
            "configuration.view",
            "schools.view",
            "branches.view",
            "academic_years.view",
            "branding.view",
            "configuration.manage_permissions",
            "configuration.manage_degrees",
            "configuration.manage_specializations",
            "timetable.manage_settings",
            "timetable.manage_blocks",
            "calendar.manage_event_types",
        ),
        "permission_mode": "any",
    },
    {
        "key": "school-management",
        "label": "School Management",
        "href": "/system-configuration/schools",
        "icon": "branch",
        "description": "Manage schools, branches, years, branding, and access.",
        "permission_keys": (
            "schools.view",
            "branches.view",
            "academic_years.view",
            "branding.view",
        ),
        "permission_mode": "any",
    },
    {
        "key": "users",
        "label": "Users",
        "href": "/users",
        "icon": "users",
        "description": "Manage user accounts, roles, and active status.",
        "permission_keys": ("users.view",),
    },
    {
        "key": "role-permissions",
        "label": "Role Permissions",
        "href": "/system-configuration/role-permissions",
        "icon": "shield",
        "description": "Assign detailed TIS permissions to each access role.",
        "permission_keys": ("configuration.manage_permissions",),
    },
    {
        "key": "degrees",
        "label": "Degrees",
        "href": "/system-configuration/degrees",
        "icon": "copy",
        "description": "Manage academic degree options.",
        "permission_keys": ("configuration.manage_degrees",),
    },
    {
        "key": "specializations",
        "label": "Specializations",
        "href": "/system-configuration/specializations",
        "icon": "subjects",
        "description": "Manage majors and teaching specializations.",
        "permission_keys": ("configuration.manage_specializations",),
    },
    {
        "key": "timetable-settings",
        "label": "Timetable Settings",
        "href": "/system-configuration/timetable-settings",
        "icon": "timetable",
        "description": "Configure school days, periods, and blocked times.",
        "permission_keys": ("timetable.manage_settings", "timetable.manage_blocks"),
        "permission_mode": "any",
    },
    {
        "key": "academic-calendar",
        "label": "Academic Calendar",
        "href": "/system-configuration/calendar",
        "icon": "calendar",
        "description": "Configure calendar event types for the active scope.",
        "permission_keys": ("calendar.manage_event_types",),
    },
)


def _get_configuration_modules(
    active_key: str,
    db: Session | None = None,
    current_user=None,
) -> list[dict[str, object]]:
    modules = []
    for module in CONFIGURATION_MODULES:
        permission_keys = tuple(module.get("permission_keys", ()))
        permission_mode = str(module.get("permission_mode", "all")).strip().lower()
        if permission_keys and db is not None and current_user is not None:
            if permission_mode == "any":
                allowed = any(
                    auth.has_permission(db, current_user, permission_key)
                    for permission_key in permission_keys
                )
            else:
                allowed = all(
                    auth.has_permission(db, current_user, permission_key)
                    for permission_key in permission_keys
                )
            if not allowed:
                continue
        modules.append(
            {
                key: value
                for key, value in module.items()
                if key not in {"permission_keys", "permission_mode"}
            }
            | {"active": module["key"] == active_key}
        )
    return modules


def _build_configuration_hub_stats(
    school_group_rows,
    branch_rows,
    academic_year_rows,
    degree_rows,
    specialization_rows,
    active_year,
    timetable_settings_count,
    calendar_event_type_count,
):
    return [
        {
            "label": "Schools",
            "icon": "branch",
            "value": len(school_group_rows),
            "note": "Organization containers",
        },
        {
            "label": "Branches",
            "icon": "branch",
            "value": len(branch_rows),
            "note": f"{sum(1 for row in branch_rows if row['status'])} active, {sum(1 for row in branch_rows if not row['status'])} inactive",
        },
        {
            "label": "Academic Years",
            "icon": "year",
            "value": len(academic_year_rows),
            "note": (
                f"Current: {active_year.year_name}"
                if active_year
                else "No live academic year yet"
            ),
        },
        {
            "label": "Degrees",
            "icon": "copy",
            "value": len(degree_rows),
            "note": "Teacher form options",
        },
        {
            "label": "Specializations",
            "icon": "subjects",
            "value": len(specialization_rows),
            "note": "Teacher form options",
        },
        {
            "label": "Timetable Settings",
            "icon": "timetable",
            "value": timetable_settings_count,
            "note": "Saved school-day profiles",
        },
        {
            "label": "Calendar Event Types",
            "icon": "calendar",
            "value": calendar_event_type_count,
            "note": "Calendar type definitions",
        },
    ]


def _build_configuration_context(request: Request, db: Session, current_user):
    scoped_branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    scoped_academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    user_school_group_id = _get_user_school_group_id(db, current_user)
    can_manage_all = _can_manage_all_school_scopes(db, current_user)
    branch_rows = _build_branch_configuration_rows(
        db,
        scoped_branch_id=scoped_branch_id,
    )
    if not can_manage_all and user_school_group_id:
        branch_rows = [
            row for row in branch_rows
            if row.get("school_group_id") == user_school_group_id
        ]

    school_group_query = db.query(models.SchoolGroup).order_by(
        models.SchoolGroup.status.desc(),
        models.SchoolGroup.name.asc(),
    )
    if not can_manage_all and user_school_group_id:
        school_group_query = school_group_query.filter(
            models.SchoolGroup.id == user_school_group_id
        )
    school_group_rows = school_group_query.all()

    academic_year_query = db.query(models.AcademicYear)
    if not can_manage_all and user_school_group_id:
        academic_year_query = academic_year_query.filter(
            models.AcademicYear.school_group_id == user_school_group_id
        )
    academic_year_rows = academic_year_query.order_by(
        models.AcademicYear.year_name.desc()
    ).all()
    qualification_rows = _build_qualification_configuration_rows(db)
    degree_rows = [
        row for row in qualification_rows
        if row["kind"] == QUALIFICATION_KIND_DEGREE
    ]
    specialization_rows = [
        row for row in qualification_rows
        if row["kind"] == QUALIFICATION_KIND_SPECIALIZATION
    ]
    active_year = next(
        (year for year in academic_year_rows if bool(year.is_active)),
        None,
    )
    timetable_settings_query = db.query(models.TimetableSetting)
    calendar_event_type_query = db.query(models.CalendarEventType)
    if not can_manage_all and user_school_group_id:
        scoped_branch_ids = [
            branch_row["id"]
            for branch_row in branch_rows
        ]
        if scoped_branch_ids:
            timetable_settings_query = timetable_settings_query.filter(
                models.TimetableSetting.branch_id.in_(scoped_branch_ids)
            )
            calendar_event_type_query = calendar_event_type_query.filter(
                models.CalendarEventType.branch_id.in_(scoped_branch_ids)
            )
        else:
            timetable_settings_query = timetable_settings_query.filter(
                models.TimetableSetting.id == -1
            )
            calendar_event_type_query = calendar_event_type_query.filter(
                models.CalendarEventType.id == -1
            )
    timetable_settings_count = timetable_settings_query.count()
    calendar_event_type_count = calendar_event_type_query.count()
    return {
        "branch_rows": branch_rows,
        "school_group_rows": school_group_rows,
        "school_group_count": len(school_group_rows),
        "branch_count": len(branch_rows),
        "active_branch_count": sum(1 for row in branch_rows if row["status"]),
        "inactive_branch_count": sum(1 for row in branch_rows if not row["status"]),
        "academic_year_rows": academic_year_rows,
        "active_year": active_year,
        "degree_rows": degree_rows,
        "specialization_rows": specialization_rows,
        "timetable_settings_count": timetable_settings_count,
        "calendar_event_type_count": calendar_event_type_count,
        "configuration_modules": _get_configuration_modules(
            "overview",
            db,
            current_user,
        ),
        "configuration_stats": _build_configuration_hub_stats(
            school_group_rows,
            branch_rows,
            academic_year_rows,
            degree_rows,
            specialization_rows,
            active_year,
            timetable_settings_count,
            calendar_event_type_count,
        ),
        "error_message": str(request.query_params.get("error", "") or "").strip(),
        "scoped_academic_year_id": scoped_academic_year_id,
    }


def _resolve_client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    if request.client:
        return request.client.host or ""
    return ""


def _resolve_audit_actor(request: Request):
    actor_user_id = getattr(request.state, "audit_actor_user_id", None)
    actor_username = getattr(request.state, "audit_actor_username", None)
    actor_role = getattr(request.state, "audit_actor_role", None)
    actor_branch_id = getattr(request.state, "audit_actor_branch_id", None)

    if actor_user_id:
        return {
            "actor_user_id": actor_user_id,
            "actor_username": actor_username or "",
            "actor_role": actor_role or "",
            "actor_branch_id": actor_branch_id,
        }

    session_user_id = auth.get_session_user_id(request)
    if session_user_id:
        return {
            "actor_user_id": session_user_id,
            "actor_username": "",
            "actor_role": "Unknown",
            "actor_branch_id": None,
        }

    return {
        "actor_user_id": "anonymous",
        "actor_username": "",
        "actor_role": "Anonymous",
        "actor_branch_id": None,
    }


def _write_request_audit_log(
    request: Request,
    status_code: int,
    duration_ms: float,
    error_name: str = "",
):
    try:
        actor = _resolve_audit_actor(request)
        request_query = str(request.url.query)
        if request.url.path == "/platform/account/verify-email":
            request_query = "[redacted]"
        write_audit_event(
            {
                "event_type": "http_request",
                "method": request.method,
                "path": request.url.path,
                "query": request_query,
                "status_code": status_code,
                "duration_ms": duration_ms,
                "client_ip": _resolve_client_ip(request),
                "user_agent": request.headers.get("user-agent", ""),
                "scope_branch_id": request.cookies.get("branch_id"),
                "scope_academic_year_id": request.cookies.get("academic_year_id"),
                "error": error_name,
                **actor,
            }
        )
    except Exception:
        # Audit logging must not block business operations.
        pass


def _normalize_grade_label(value) -> str:
    cleaned = str(value).strip().upper()
    if cleaned in {"0", "K", "KG", "KINDERGARTEN"}:
        return "KG"
    try:
        parsed = int(cleaned)
    except (TypeError, ValueError):
        return ""
    if 1 <= parsed <= 12:
        return str(parsed)
    return ""


def _grade_sort_key(grade_label: str) -> int:
    if grade_label == "KG":
        return 0
    try:
        return int(grade_label)
    except (TypeError, ValueError):
        return 99


def _build_subject_identity(subject_name: str, fallback_code: str = ""):
    cleaned_name = " ".join(str(subject_name or "").split())
    if cleaned_name:
        return cleaned_name.lower(), cleaned_name

    cleaned_code = " ".join(str(fallback_code or "").split()).upper()
    if cleaned_code:
        return cleaned_code.lower(), cleaned_code

    return "", ""


def _build_subject_display_labels(
    subject_name: str,
    subject_code: str = "",
    weekly_hours: int = 0,
    grade_label=None,
):
    bundle_subject_labels = get_homeroom_bundle_subject_labels(
        subject_code=subject_code,
        subject_name=subject_name,
        weekly_hours=weekly_hours,
        grade_label=grade_label,
    )
    if bundle_subject_labels:
        return [
            f"{bundle_subject} ({subject_name}, {weekly_hours}h bundle)"
            for bundle_subject in bundle_subject_labels
        ]
    return [subject_name]


def _build_teacher_display_name(teacher) -> str:
    if teacher is None:
        return "Unknown Teacher"
    name_parts = [
        str(getattr(teacher, "first_name", "") or "").strip(),
        str(getattr(teacher, "middle_name", "") or "").strip(),
        str(getattr(teacher, "last_name", "") or "").strip(),
    ]
    full_name = " ".join(part for part in name_parts if part).strip()
    if full_name:
        return full_name
    teacher_pk = getattr(teacher, "id", None)
    return f"Teacher #{teacher_pk}" if teacher_pk is not None else "Unknown Teacher"


def _normalize_subject_family_key(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())
    return " ".join(normalized.split())


def _normalize_alignment_text(value: str) -> str:
    return _normalize_subject_family_key(value)


def _build_alignment_token_set(value: str):
    return {
        token
        for token in _normalize_alignment_text(value).split()
        if len(token) > 1 and token not in MAJOR_ALIGNMENT_STOPWORDS
    }


def _subject_starts_with_any(normalized_value: str, prefixes) -> bool:
    normalized_text = _normalize_subject_family_key(normalized_value)
    if not normalized_text:
        return False

    for prefix in prefixes:
        normalized_prefix = _normalize_subject_family_key(prefix)
        if not normalized_prefix:
            continue
        if normalized_text == normalized_prefix or normalized_text.startswith(
            normalized_prefix + " "
        ):
            return True
    return False


def _get_priority_staffing_subject_family(subject_key: str = "", subject_name: str = "") -> str:
    candidates = [
        _normalize_subject_family_key(subject_name),
        _normalize_subject_family_key(subject_key),
    ]
    for family, prefixes in PRIORITY_STAFFING_SUBJECT_PREFIXES.items():
        if any(_subject_starts_with_any(candidate, prefixes) for candidate in candidates):
            return family
    return ""


def _is_priority_staffing_subject(subject_key: str = "", subject_name: str = "") -> bool:
    return bool(_get_priority_staffing_subject_family(subject_key, subject_name))


def _subject_matches_degree_major(
    degree_major: str,
    subject_name: str,
    subject_key: str = "",
) -> bool:
    normalized_major = _normalize_alignment_text(degree_major)
    if not normalized_major:
        return False

    major_tokens = _build_alignment_token_set(normalized_major)
    subject_texts = [
        _normalize_alignment_text(subject_name),
        _normalize_alignment_text(subject_key),
    ]
    for subject_text in subject_texts:
        if not subject_text:
            continue
        if len(subject_text) >= 4 and subject_text in normalized_major:
            return True
        if len(normalized_major) >= 4 and normalized_major in subject_text:
            return True
        if major_tokens and (_build_alignment_token_set(subject_text) & major_tokens):
            return True

    return False


def _subject_matches_teacher_major(teacher, subject_name: str, subject_key: str = "") -> bool:
    return _subject_matches_degree_major(
        getattr(teacher, "degree_major", ""),
        subject_name,
        subject_key,
    )


def _text_contains_any_keyword(normalized_text: str, keywords) -> bool:
    if not normalized_text:
        return False
    for keyword in keywords:
        normalized_keyword = _normalize_alignment_text(keyword)
        if not normalized_keyword:
            continue
        if (
            normalized_text == normalized_keyword
            or f" {normalized_keyword} " in f" {normalized_text} "
        ):
            return True
    return False


def _resolve_teacher_major_priority_rule(major_text: str) -> dict:
    normalized_major = _normalize_alignment_text(major_text)
    default_rule = {
        "rule_key": "generic_major_match",
        "label": "Generic major alignment",
        "pool_key": "",
        "priority_families": [],
        "allowed_families": [],
        "search_all_pools": False,
    }
    if not normalized_major:
        return default_rule

    is_arabic_major = _text_contains_any_keyword(
        normalized_major,
        ("arabic", "arabic language", "arabic literature"),
    )
    if is_arabic_major:
        return {
            "rule_key": "arabic_major",
            "label": "Arabic major",
            "pool_key": "arabic_pool",
            "priority_families": ["arabic"],
            "allowed_families": ["arabic", "islamic", "quran", "social_arabic"],
            "search_all_pools": False,
        }

    is_english_major = _text_contains_any_keyword(
        normalized_major,
        ("english", "english literature", "language arts", "ela"),
    )
    if is_english_major:
        return {
            "rule_key": "english_major",
            "label": "English major",
            "pool_key": "english_pool",
            "priority_families": ["english"],
            "allowed_families": [
                "english",
                "social_english",
                "social",
                "wellbeing",
                "reflection",
                "performing_arts",
                "art",
            ],
            "search_all_pools": False,
        }

    is_art_major = _text_contains_any_keyword(
        normalized_major,
        ("art", "art education", "visual art", "visual arts", "drawing", "painting"),
    )
    if is_art_major:
        return {
            "rule_key": "art_major",
            "label": "Art major",
            "pool_key": "english_pool",
            "priority_families": ["art"],
            "allowed_families": ["art", "performing_arts", "reflection", "wellbeing"],
            "search_all_pools": False,
        }

    is_performing_arts_major = _text_contains_any_keyword(
        normalized_major,
        (
            "performing arts",
            "performance",
            "music",
            "music education",
            "drama",
            "theatre",
            "theater",
            "dance",
        ),
    )
    if is_performing_arts_major:
        return {
            "rule_key": "performing_arts_major",
            "label": "Performing Arts major",
            "pool_key": "english_pool",
            "priority_families": ["performing_arts"],
            "allowed_families": ["performing_arts", "art", "reflection", "wellbeing"],
            "search_all_pools": False,
        }

    is_reflection_major = _text_contains_any_keyword(
        normalized_major,
        ("reflection", "reflective", "advisory", "character education"),
    )
    if is_reflection_major:
        return {
            "rule_key": "reflection_major",
            "label": "Reflection major",
            "pool_key": "english_pool",
            "priority_families": ["reflection"],
            "allowed_families": ["reflection", "wellbeing", "performing_arts", "art"],
            "search_all_pools": False,
        }

    is_wellbeing_major = _text_contains_any_keyword(
        normalized_major,
        ("wellbeing", "well being", "health", "sel", "life skills"),
    )
    if is_wellbeing_major:
        return {
            "rule_key": "wellbeing_major",
            "label": "Well Being major",
            "pool_key": "english_pool",
            "priority_families": ["wellbeing"],
            "allowed_families": ["wellbeing", "reflection", "performing_arts", "art"],
            "search_all_pools": False,
        }

    is_ict_major = _text_contains_any_keyword(
        normalized_major,
        (
            "ict",
            "computer",
            "computer science",
            "software engineering",
            "information technology",
            "informatics",
            "programming",
            "technology",
            "computing",
            "digital",
        ),
    )
    if is_ict_major:
        return {
            "rule_key": "ict_major",
            "label": "ICT / Computer Science major",
            "pool_key": "general_science_pool",
            "priority_families": ["ict"],
            "allowed_families": ["ict"],
            "search_all_pools": True,
        }

    is_physics_major = _text_contains_any_keyword(
        normalized_major,
        ("physics", "physical science"),
    )
    if is_physics_major:
        return {
            "rule_key": "physics_major",
            "label": "Physics major",
            "pool_key": "math_pool",
            "priority_families": ["math", "physics"],
            "allowed_families": ["math", "physics"],
            "search_all_pools": False,
        }

    is_math_major = _text_contains_any_keyword(
        normalized_major,
        ("math", "mathematics", "maths", "algebra", "geometry", "calculus", "statistics"),
    )
    if is_math_major:
        return {
            "rule_key": "math_major",
            "label": "Math major",
            "pool_key": "math_pool",
            "priority_families": ["math"],
            "allowed_families": ["math", "physics"],
            "search_all_pools": False,
        }

    is_islamic_major = _text_contains_any_keyword(
        normalized_major,
        (
            "islamic",
            "islamic studies",
            "quran",
            "quranic",
            "hadith",
            "fiqh",
            "tawheed",
            "religion",
            "noorani",
            "noraniah",
            "qaadah",
        ),
    )
    if is_islamic_major:
        return {
            "rule_key": "islamic_major",
            "label": "Islamic / Quran major",
            "pool_key": "arabic_pool",
            "priority_families": ["islamic", "quran"],
            "allowed_families": ["islamic", "quran", "social_arabic", "arabic"],
            "search_all_pools": False,
        }

    is_social_studies_major = _text_contains_any_keyword(
        normalized_major,
        (
            "social studies",
            "social",
            "history",
            "geography",
            "economics",
            "economic",
            "humanities",
            "civics",
        ),
    )
    if is_social_studies_major:
        return {
            "rule_key": "social_studies_major",
            "label": "Social Studies major",
            "pool_key": "english_pool",
            "priority_families": ["social_english", "social"],
            "allowed_families": ["social_english", "social"],
            "search_all_pools": True,
        }

    is_science_major = _text_contains_any_keyword(
        normalized_major,
        (
            "science",
            "general science",
            "biology",
            "chemistry",
            "biochemistry",
            "life science",
            "laboratory",
            "lab",
        ),
    )
    if is_science_major and not is_physics_major:
        return {
            "rule_key": "general_science_major",
            "label": "General Science major",
            "pool_key": "general_science_pool",
            "priority_families": ["science"],
            "allowed_families": ["science", "biology", "chemistry", "ict"],
            "search_all_pools": False,
        }

    return default_rule


def _build_subject_priority_family(subject_name: str, subject_key: str = "", subject_code: str = "") -> str:
    return _detect_hiring_subject_family(
        {
            "subject_name": subject_name,
            "subject_key": subject_key,
            "subject_code": subject_code,
        }
    )


def _build_profile_subject_compatibility(profile, subject_key: str, subject_name: str):
    override_subject_keys = set(profile.get("override_subject_keys", []))
    eligible_subject_keys = (
        set(profile.get("subject_keys", []))
        | set(profile.get("secondary_subject_keys", []))
        | set(profile.get("support_subject_keys", []))
        | set(profile.get("homeroom_subject_keys", []))
    )
    if subject_key in override_subject_keys:
        return {
            "priority": 2,
            "basis": "Admin override",
        }

    if subject_key in eligible_subject_keys:
        return {
            "priority": 0,
            "basis": "Eligible subject",
        }

    if _subject_matches_degree_major(
        profile.get("degree_major", ""),
        subject_name,
        subject_key,
    ):
        return {
            "priority": 1,
            "basis": "Major match",
        }

    return None


def _build_explicit_section_subject_keys(section_assignments):
    explicit_keys = set()
    for assignment in section_assignments or []:
        planning_section_id = getattr(assignment, "planning_section_id", None)
        subject_code = str(getattr(assignment, "subject_code", "") or "").strip().upper()
        if planning_section_id and subject_code:
            explicit_keys.add((planning_section_id, subject_code))
    return explicit_keys


def _build_reporting_context_from_section_assignments(
    db: Session,
    subjects,
    planning_sections,
    teachers,
    section_assignments,
):
    sections_by_grade = {}
    current_sections_by_grade = {}
    new_sections_by_grade = {}

    for section in planning_sections:
        grade_label = _normalize_grade_label(section.grade_level)
        if not grade_label:
            continue

        sections_by_grade[grade_label] = sections_by_grade.get(grade_label, 0) + 1
        status = str(section.class_status or "").strip().lower()
        if status == "current":
            current_sections_by_grade[grade_label] = (
                current_sections_by_grade.get(grade_label, 0) + 1
            )
        elif status == "new":
            new_sections_by_grade[grade_label] = (
                new_sections_by_grade.get(grade_label, 0) + 1
            )

    scoped_subjects_by_code = {
        subject.subject_code: subject
        for subject in subjects
        if subject.subject_code
    }
    subject_demand_map = {}
    required_hours_by_grade = {}
    required_current_hours_by_grade = {}
    required_new_hours_by_grade = {}

    for subject in subjects:
        grade_label = _normalize_grade_label(subject.grade)
        if not grade_label:
            continue

        weekly_hours = int(subject.weekly_hours or 0)
        if weekly_hours <= 0:
            continue

        sections_count = sections_by_grade.get(grade_label, 0)
        if sections_count <= 0:
            continue

        current_sections_count = current_sections_by_grade.get(grade_label, 0)
        new_sections_count = new_sections_by_grade.get(grade_label, 0)

        subject_key, subject_label = _build_subject_identity(
            subject_name=subject.subject_name,
            fallback_code=subject.subject_code or "",
        )
        if not subject_key:
            continue

        required_hours = weekly_hours * sections_count
        required_current_hours = weekly_hours * current_sections_count
        required_new_hours = weekly_hours * new_sections_count

        required_hours_by_grade[grade_label] = (
            required_hours_by_grade.get(grade_label, 0) + required_hours
        )
        required_current_hours_by_grade[grade_label] = (
            required_current_hours_by_grade.get(grade_label, 0) + required_current_hours
        )
        required_new_hours_by_grade[grade_label] = (
            required_new_hours_by_grade.get(grade_label, 0) + required_new_hours
        )

        if subject_key not in subject_demand_map:
            subject_demand_map[subject_key] = {
                "subject_name": subject_label,
                "subject_code": subject.subject_code or "",
                "subject_color": resolve_subject_color(
                    subject.subject_code or subject_key,
                    getattr(subject, "color", ""),
                    subject_name=subject.subject_name,
                ),
                "weekly_hours": weekly_hours,
                "primary_grade_label": grade_label,
                "bundle_subject_labels": list(
                    get_homeroom_bundle_subject_labels(
                        subject_code=subject.subject_code or "",
                        subject_name=subject.subject_name or "",
                        weekly_hours=weekly_hours,
                        grade_label=grade_label,
                    )
                ),
                "required_hours": 0,
                "required_current_hours": 0,
                "required_new_hours": 0,
                "grades": set(),
            }

        entry = subject_demand_map[subject_key]
        entry["required_hours"] += required_hours
        entry["required_current_hours"] += required_current_hours
        entry["required_new_hours"] += required_new_hours
        entry["grades"].add(grade_label)

    teacher_ids = sorted(
        teacher.id
        for teacher in teachers
        if getattr(teacher, "id", None)
    )
    planning_sections_by_id = {
        section.id: section
        for section in planning_sections
        if getattr(section, "id", None)
    }
    teachers_by_id = {
        teacher.id: teacher
        for teacher in teachers
        if getattr(teacher, "id", None)
    }
    teacher_subject_map = {
        teacher_id: set()
        for teacher_id in teacher_ids
    }
    if teacher_ids:
        teacher_allocations = db.query(models.TeacherSubjectAllocation).filter(
            models.TeacherSubjectAllocation.teacher_id.in_(teacher_ids)
        ).all()
    else:
        teacher_allocations = []

    for allocation in teacher_allocations:
        subject = scoped_subjects_by_code.get(allocation.subject_code)
        if not subject:
            continue
        subject_key, _ = _build_subject_identity(
            subject_name=subject.subject_name,
            fallback_code=subject.subject_code or "",
        )
        if subject_key and subject_key in subject_demand_map:
            teacher_subject_map.setdefault(allocation.teacher_id, set()).add(subject_key)

    for teacher in teachers:
        subject_key_set = teacher_subject_map.setdefault(getattr(teacher, "id", None), set())
        if subject_key_set:
            continue
        fallback_code = str(teacher.subject_code or "").strip().upper()
        if not fallback_code:
            continue
        fallback_subject = scoped_subjects_by_code.get(fallback_code)
        if not fallback_subject:
            continue
        subject_key, _ = _build_subject_identity(
            subject_name=fallback_subject.subject_name,
            fallback_code=fallback_subject.subject_code or "",
        )
        if subject_key and subject_key in subject_demand_map:
            subject_key_set.add(subject_key)

    actual_hours_by_teacher = {}
    actual_hours_by_teacher_subject = {}
    actual_hours_by_subject = {}
    actual_teacher_contributors_by_subject = {}
    actual_homeroom_breakdown_by_teacher = {}
    actual_homeroom_section_labels_by_teacher = {}
    actual_homeroom_allocations_by_teacher = {}
    explicit_section_subject_keys = _build_explicit_section_subject_keys(
        section_assignments
    )
    homeroom_assignments_by_teacher = _build_homeroom_assignments_by_teacher(
        subjects=subjects,
        planning_sections=planning_sections,
        explicit_section_subject_keys=explicit_section_subject_keys,
        valid_teacher_ids=set(teachers_by_id.keys()),
    )

    for assignment in section_assignments:
        teacher = teachers_by_id.get(getattr(assignment, "teacher_id", None))
        if not teacher:
            continue
        subject = scoped_subjects_by_code.get(assignment.subject_code)
        if not subject:
            continue

        subject_key, _ = _build_subject_identity(
            subject_name=subject.subject_name,
            fallback_code=subject.subject_code or "",
        )
        if not subject_key or subject_key not in subject_demand_map:
            continue

        subject_hours = int(subject.weekly_hours or 0)
        teacher_id = getattr(teacher, "id", None)
        actual_hours_by_teacher[teacher_id] = (
            actual_hours_by_teacher.get(teacher_id, 0) + subject_hours
        )
        teacher_subject_hours = actual_hours_by_teacher_subject.setdefault(
            teacher_id,
            {},
        )
        teacher_subject_hours[subject_key] = (
            teacher_subject_hours.get(subject_key, 0) + subject_hours
        )
        actual_hours_by_subject[subject_key] = (
            actual_hours_by_subject.get(subject_key, 0) + subject_hours
        )
        subject_contributors = actual_teacher_contributors_by_subject.setdefault(
            subject_key,
            {},
        )
        teacher_name = _build_teacher_display_name(teacher)
        teacher_contributor = subject_contributors.setdefault(
            teacher_id,
            {
                "teacher_id": teacher_id,
                "teacher_name": teacher_name,
                "allocated_hours": 0,
                "section_labels": set(),
                "section_hours": {},
            },
        )
        teacher_contributor["allocated_hours"] += subject_hours
        planning_section = planning_sections_by_id.get(assignment.planning_section_id)
        if planning_section:
            grade_label = _normalize_grade_label(planning_section.grade_level)
            section_name = str(planning_section.section_name or "").strip().upper()
            if grade_label and section_name:
                display_grade = "KG" if grade_label == "KG" else f"G{grade_label}"
                section_label = f"{display_grade}-{section_name}"
                teacher_contributor["section_labels"].add(section_label)
                teacher_contributor.setdefault("section_hours", {})
                teacher_contributor["section_hours"][section_label] = (
                    int(teacher_contributor["section_hours"].get(section_label, 0))
                    + subject_hours
                )

    for teacher_id, homeroom_items in homeroom_assignments_by_teacher.items():
        for item in homeroom_items:
            subject_key = item["subject_key"]
            subject_hours = int(item["required_hours"])
            if subject_hours <= 0:
                continue

            actual_hours_by_teacher[teacher_id] = (
                actual_hours_by_teacher.get(teacher_id, 0) + subject_hours
            )
            teacher_subject_hours = actual_hours_by_teacher_subject.setdefault(
                teacher_id,
                {},
            )
            teacher_subject_hours[subject_key] = (
                teacher_subject_hours.get(subject_key, 0) + subject_hours
            )
            actual_hours_by_subject[subject_key] = (
                actual_hours_by_subject.get(subject_key, 0) + subject_hours
            )
            actual_homeroom_breakdown_by_teacher.setdefault(teacher_id, {})
            actual_homeroom_breakdown_by_teacher[teacher_id][subject_key] = (
                actual_homeroom_breakdown_by_teacher[teacher_id].get(subject_key, 0)
                + subject_hours
            )
            actual_homeroom_section_labels_by_teacher.setdefault(teacher_id, [])
            if item["class_label"] not in actual_homeroom_section_labels_by_teacher[teacher_id]:
                actual_homeroom_section_labels_by_teacher[teacher_id].append(
                    item["class_label"]
                )
            actual_homeroom_allocations_by_teacher.setdefault(teacher_id, []).append(
                {
                    **item,
                    "allocated_hours": subject_hours,
                }
            )
            subject_contributors = actual_teacher_contributors_by_subject.setdefault(
                subject_key,
                {},
            )
            teacher_name = _build_teacher_display_name(teachers_by_id.get(teacher_id))
            teacher_contributor = subject_contributors.setdefault(
                teacher_id,
                {
                    "teacher_id": teacher_id,
                    "teacher_name": teacher_name,
                    "allocated_hours": 0,
                    "section_labels": set(),
                    "section_hours": {},
                },
            )
            teacher_contributor["allocated_hours"] += subject_hours
            teacher_contributor["section_labels"].add(item["class_label"])
            teacher_contributor.setdefault("section_hours", {})
            teacher_contributor["section_hours"][item["class_label"]] = (
                int(teacher_contributor["section_hours"].get(item["class_label"], 0))
                + subject_hours
            )

    teacher_profiles = []
    total_existing_capacity_hours = 0
    for teacher in teachers:
        teacher_id = getattr(teacher, "id", None)
        allocation_breakdown = dict(
            actual_hours_by_teacher_subject.get(teacher_id, {})
        )
        subject_keys = sorted(
            allocation_breakdown.keys(),
            key=lambda key: subject_demand_map.get(key, {}).get("subject_name", key),
        )
        allocated_hours = sum(allocation_breakdown.values())
        teacher_capacity_breakdown = get_teacher_capacity_breakdown(
            teacher,
            default_max_hours=REPORT_STANDARD_MAX_HOURS,
        )
        teacher_capacity = teacher_capacity_breakdown["international_capacity_hours"]
        total_existing_capacity_hours += teacher_capacity

        teacher_profiles.append(
            {
                "teacher": teacher,
                "name": _build_teacher_display_name(teacher),
                "subject_keys": subject_keys,
                "support_subject_keys": [],
                "subject_count": len(subject_keys),
                "primary_subject_basis_hours": max(allocation_breakdown.values(), default=0),
                "homeroom_allocated_hours": sum(
                    actual_homeroom_breakdown_by_teacher.get(teacher_id, {}).values()
                ),
                "homeroom_subject_keys": sorted(
                    actual_homeroom_breakdown_by_teacher.get(teacher_id, {}).keys(),
                    key=lambda key: subject_demand_map.get(key, {}).get("subject_name", key),
                ),
                "homeroom_section_labels": list(
                    actual_homeroom_section_labels_by_teacher.get(teacher_id, [])
                ),
                "homeroom_class_allocations": [
                    dict(item)
                    for item in actual_homeroom_allocations_by_teacher.get(teacher_id, [])
                ],
                "homeroom_allocation_breakdown": dict(
                    actual_homeroom_breakdown_by_teacher.get(teacher_id, {})
                ),
                "allocated_hours": allocated_hours,
                "remaining_capacity_hours": max(teacher_capacity - allocated_hours, 0),
                "allocation_breakdown": allocation_breakdown,
                "capacity_hours": teacher_capacity,
                "total_capacity_hours": teacher_capacity_breakdown[
                    "total_capacity_hours"
                ],
                "extra_hours_count": teacher_capacity_breakdown["extra_hours"],
                "national_section_hours": teacher_capacity_breakdown[
                    "national_section_hours"
                ],
                "primary_allocated_hours": allocated_hours,
                "support_allocated_hours": 0,
            }
        )

    teachers_per_subject = {}
    for teacher_id, subject_keys in teacher_subject_map.items():
        covered_subject_keys = set(subject_keys) | set(
            actual_homeroom_breakdown_by_teacher.get(teacher_id, {}).keys()
        )
        for subject_key in covered_subject_keys:
            teachers_per_subject[subject_key] = (
                teachers_per_subject.get(subject_key, 0) + 1
            )

    report_subject_rows = []
    for subject_key, demand in subject_demand_map.items():
        required_hours = demand["required_hours"]
        allocated_hours = min(actual_hours_by_subject.get(subject_key, 0), required_hours)
        remaining_hours = max(required_hours - allocated_hours, 0)
        subject_contributors = []
        for contributor in actual_teacher_contributors_by_subject.get(subject_key, {}).values():
            contributor_allocated_hours = int(contributor.get("allocated_hours", 0))
            contributor_teacher_id = contributor.get("teacher_id")
            total_teacher_allocated_hours = int(
                actual_hours_by_teacher.get(
                    contributor_teacher_id,
                    contributor_allocated_hours,
                )
            )
            other_subject_hours = max(
                total_teacher_allocated_hours - contributor_allocated_hours,
                0,
            )
            current_subject_share = (
                round(
                    (contributor_allocated_hours / REPORT_STANDARD_MAX_HOURS)
                    * 100
                )
                if REPORT_STANDARD_MAX_HOURS > 0
                else 0
            )
            other_subject_share = (
                round((other_subject_hours / REPORT_STANDARD_MAX_HOURS) * 100)
                if REPORT_STANDARD_MAX_HOURS > 0
                else 0
            )
            total_load_share = (
                round(
                    (total_teacher_allocated_hours / REPORT_STANDARD_MAX_HOURS)
                    * 100
                )
                if REPORT_STANDARD_MAX_HOURS > 0
                else 0
            )
            section_hours = dict(contributor.get("section_hours", {}))
            section_labels = sorted(
                label for label in contributor.get("section_labels", set()) if label
            )
            section_details = [
                f"{label} ({int(section_hours.get(label, 0))}h)"
                if int(section_hours.get(label, 0)) > 0
                else label
                for label in section_labels
            ]
            subject_contributors.append(
                {
                    "teacher_id": contributor_teacher_id,
                    "teacher_name": contributor.get("teacher_name", "-"),
                    "allocated_hours": contributor_allocated_hours,
                    "current_subject_hours": contributor_allocated_hours,
                    "other_allocated_hours": other_subject_hours,
                    "total_allocated_hours": total_teacher_allocated_hours,
                    "share_percentage": min(total_load_share, 100),
                    "current_share_percentage": min(current_subject_share, 100),
                    "other_share_percentage": min(other_subject_share, 100),
                    "total_share_percentage": total_load_share,
                    "capacity_hours": REPORT_STANDARD_MAX_HOURS,
                    "remaining_capacity_hours": max(
                        REPORT_STANDARD_MAX_HOURS - total_teacher_allocated_hours,
                        0,
                    ),
                    "section_labels": section_details,
                    "section_count": len(section_labels),
                }
            )
        subject_contributors.sort(
            key=lambda item: (
                -item["allocated_hours"],
                item["teacher_name"],
            )
        )
        teacher_requirement_blocks = (
            math.ceil(remaining_hours / REPORT_STANDARD_MAX_HOURS)
            if remaining_hours > 0
            else 0
        )
        coverage_percentage = (
            round((allocated_hours / required_hours) * 100)
            if required_hours > 0
            else 0
        )
        grades = sorted(demand["grades"], key=_grade_sort_key)

        report_subject_rows.append(
            {
                "subject_key": subject_key,
                "subject_name": demand["subject_name"],
                "subject_code": demand.get("subject_code", subject_key),
                "subject_color": resolve_subject_color(
                    demand.get("subject_code", subject_key),
                    demand.get("subject_color", ""),
                    subject_name=demand.get("subject_name", ""),
                ),
                "grades": grades,
                "required_hours": required_hours,
                "required_current_hours": demand["required_current_hours"],
                "required_new_hours": demand["required_new_hours"],
                "allocated_hours": allocated_hours,
                "remaining_hours": remaining_hours,
                "coverage_percentage": coverage_percentage,
                "teachers_with_subject": len(subject_contributors),
                "assigned_teacher_count": len(subject_contributors),
                "assigned_teacher_names": [
                    contributor["teacher_name"] for contributor in subject_contributors
                ],
                "assigned_teacher_labels": [
                    f"{contributor['teacher_name']} ({contributor['allocated_hours']}h)"
                    for contributor in subject_contributors
                ],
                "assigned_teacher_contributors": subject_contributors,
                "effective_subject_count": max(
                    len(demand.get("bundle_subject_labels", [])),
                    1,
                ),
                "teacher_requirement_blocks": teacher_requirement_blocks,
                "additional_teachers_needed": teacher_requirement_blocks,
                "additional_teachers_note": "",
                "priority_staffing_subject": _is_priority_staffing_subject(
                    subject_key,
                    demand["subject_name"],
                ),
                "internal_absorption_recommended": False,
            }
        )

    report_subject_rows.sort(
        key=lambda row: (
            -row["remaining_hours"],
            row["subject_name"],
        )
    )

    report_gap_rows = [
        dict(row)
        for row in report_subject_rows
        if row["remaining_hours"] > 0
    ]
    max_remaining_hours = max(
        (row["remaining_hours"] for row in report_gap_rows),
        default=0,
    )
    for row in report_gap_rows:
        row["gap_chart_pct"] = (
            round((row["remaining_hours"] / max_remaining_hours) * 100, 1)
            if max_remaining_hours > 0
            else 0
        )
    report_gap_rows = report_gap_rows[:8]

    report_teacher_rows = []
    for profile in teacher_profiles:
        subject_labels = [
            subject_demand_map[subject_key]["subject_name"]
            for subject_key in profile["subject_keys"]
            if subject_key in subject_demand_map
        ]
        allocation_labels = [
            f"{subject_demand_map[subject_key]['subject_name']} ({hours}h)"
            for subject_key, hours in sorted(
                profile["allocation_breakdown"].items(),
                key=lambda item: (-item[1], subject_demand_map.get(item[0], {}).get("subject_name", item[0])),
            )
            if subject_key in subject_demand_map
        ]
        subject_chips = [
            {
                "name": subject_demand_map[subject_key]["subject_name"],
                "color": subject_demand_map[subject_key].get("subject_color", "#0A4EA3"),
                "hours": hours,
            }
            for subject_key, hours in sorted(
                profile["allocation_breakdown"].items(),
                key=lambda item: (-item[1], subject_demand_map.get(item[0], {}).get("subject_name", item[0])),
            )
            if subject_key in subject_demand_map
        ]

        teacher = profile["teacher"]
        report_teacher_rows.append(
            {
                "teacher_pk": teacher.id,
                "teacher_id": teacher.teacher_id or "-",
                "teacher_name": profile["name"],
                "degree_major": str(getattr(teacher, "degree_major", "") or "").strip(),
                "subject_labels": subject_labels,
                "subject_chips": subject_chips,
                "support_subject_labels": [],
                "homeroom_subject_labels": [
                    subject_demand_map[subject_key]["subject_name"]
                    for subject_key in profile.get("homeroom_subject_keys", [])
                    if subject_key in subject_demand_map
                ],
                "homeroom_section_labels": list(
                    profile.get("homeroom_section_labels", [])
                ),
                "homeroom_allocation_labels": [
                    f"{item['class_label']}: {item['subject_name']} ({item['allocated_hours']}h)"
                    for item in profile.get("homeroom_class_allocations", [])
                ],
                "allocation_labels": allocation_labels,
                "expected_allocated_hours": profile["allocated_hours"],
                "homeroom_allocated_hours": profile.get("homeroom_allocated_hours", 0),
                "primary_allocated_hours": profile.get("primary_allocated_hours", 0),
                "support_allocated_hours": profile.get("support_allocated_hours", 0),
                "primary_subject_basis_hours": profile.get(
                    "primary_subject_basis_hours",
                    0,
                ),
                "capacity_hours": profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS),
                "total_capacity_hours": profile.get(
                    "total_capacity_hours",
                    REPORT_STANDARD_MAX_HOURS,
                ),
                "extra_hours_count": profile.get("extra_hours_count", 0),
                "national_section_hours": profile.get("national_section_hours", 0),
                "remaining_capacity_hours": profile["remaining_capacity_hours"],
            }
        )

    report_teacher_rows.sort(
        key=lambda row: (
            -row["expected_allocated_hours"],
            row["teacher_name"],
        )
    )

    report_grade_rows = []
    for grade_label, total_sections in sections_by_grade.items():
        report_grade_rows.append(
            {
                "grade_label": grade_label,
                "sections_total": total_sections,
                "sections_current": current_sections_by_grade.get(grade_label, 0),
                "sections_new": new_sections_by_grade.get(grade_label, 0),
                "required_hours_total": required_hours_by_grade.get(grade_label, 0),
                "required_hours_current": required_current_hours_by_grade.get(
                    grade_label, 0
                ),
                "required_hours_new": required_new_hours_by_grade.get(grade_label, 0),
            }
        )
    report_grade_rows.sort(
        key=lambda row: _grade_sort_key(row["grade_label"])
    )

    total_required_hours = sum(
        row["required_hours"] for row in report_subject_rows
    )
    total_remaining_hours = sum(
        row["remaining_hours"] for row in report_subject_rows
    )
    total_allocated_hours = total_required_hours - total_remaining_hours
    total_staffing_requirement_blocks = (
        math.ceil(total_remaining_hours / REPORT_STANDARD_MAX_HOURS)
        if total_remaining_hours > 0
        else 0
    )
    total_existing_teachers = len(teachers)
    coverage_percentage = (
        round((total_allocated_hours / total_required_hours) * 100)
        if total_required_hours > 0
        else 0
    )
    teachers_with_subject_alignment = sum(
        1 for subject_keys in teacher_subject_map.values() if subject_keys
    )
    teachers_utilized = sum(
        1 for profile in teacher_profiles if profile["allocated_hours"] > 0
    )
    teachers_full_load = sum(
        1
        for profile in teacher_profiles
        if profile["allocated_hours"] >= profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
    )
    unused_existing_capacity_hours = max(
        total_existing_capacity_hours - total_allocated_hours,
        0,
    )
    total_required_current_hours = sum(
        required_current_hours_by_grade.values()
    )
    total_required_new_hours = sum(
        required_new_hours_by_grade.values()
    )
    total_new_sections_planned = sum(new_sections_by_grade.values())
    largest_gap_row = report_gap_rows[0] if report_gap_rows else None

    report_summary = {
        "total_required_hours": total_required_hours,
        "total_required_current_hours": total_required_current_hours,
        "total_required_new_hours": total_required_new_hours,
        "total_allocated_hours": total_allocated_hours,
        "total_remaining_hours": total_remaining_hours,
        "coverage_percentage": coverage_percentage,
        "total_additional_teachers_needed": total_staffing_requirement_blocks,
        "total_staffing_requirement_blocks": total_staffing_requirement_blocks,
        "total_existing_teachers": total_existing_teachers,
        "total_existing_capacity_hours": total_existing_capacity_hours,
        "unused_existing_capacity_hours": unused_existing_capacity_hours,
        "teachers_with_subject_alignment": teachers_with_subject_alignment,
        "teachers_utilized": teachers_utilized,
        "teachers_full_load": teachers_full_load,
        "teachers_idle": max(total_existing_teachers - teachers_utilized, 0),
        "total_new_sections_planned": total_new_sections_planned,
        "subjects_with_gaps": len(report_gap_rows),
        "homeroom_default_coverage_hours": sum(
            int(profile.get("homeroom_allocated_hours", 0))
            for profile in teacher_profiles
        ),
        "largest_gap_subject_name": (
            largest_gap_row["subject_name"] if largest_gap_row else ""
        ),
        "largest_gap_hours": (
            int(largest_gap_row["remaining_hours"]) if largest_gap_row else 0
        ),
        "largest_gap_teachers_needed": (
            int(largest_gap_row["teacher_requirement_blocks"])
            if largest_gap_row
            else 0
        ),
    }

    teacher_profiles_export = []
    for profile in teacher_profiles:
        teacher = profile["teacher"]
        teacher_profiles_export.append(
            {
                "teacher_pk": teacher.id,
                "teacher_id": teacher.teacher_id or "-",
                "teacher_name": profile["name"],
                "subject_keys": list(profile["subject_keys"]),
                "support_subject_keys": list(profile["support_subject_keys"]),
                "allocation_breakdown": dict(profile["allocation_breakdown"]),
                "allocated_hours": int(profile["allocated_hours"]),
                "remaining_capacity_hours": int(profile["remaining_capacity_hours"]),
                "capacity_hours": int(
                    profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
                ),
                "total_capacity_hours": int(
                    profile.get("total_capacity_hours", REPORT_STANDARD_MAX_HOURS)
                ),
                "national_section_hours": int(
                    profile.get("national_section_hours", 0)
                ),
                "primary_allocated_hours": int(profile.get("primary_allocated_hours", 0)),
                "support_allocated_hours": int(profile.get("support_allocated_hours", 0)),
                "primary_subject_basis_hours": int(
                    profile.get("primary_subject_basis_hours", 0)
                ),
            }
        )

    return {
        "summary": report_summary,
        "subject_rows": report_subject_rows,
        "gap_rows": report_gap_rows,
        "teacher_rows": report_teacher_rows,
        "grade_rows": report_grade_rows,
        "teacher_profiles": teacher_profiles_export,
    }


def _build_report_class_allocation_data_from_section_assignments(
    db: Session,
    subjects,
    planning_sections,
    teachers,
    reporting_context,
    section_assignments,
):
    class_rows = _build_report_class_rows(planning_sections)
    subjects_by_grade, subject_name_by_key = _build_report_subject_catalog(subjects)
    teacher_by_id = {
        teacher.id: teacher
        for teacher in teachers
        if getattr(teacher, "id", None)
    }
    teacher_subject_name_map = {}
    teacher_primary_hours_map = {}
    teacher_class_allocations = {}
    teacher_homeroom_allocations = {}
    teacher_homeroom_subject_keys = {}
    teacher_homeroom_section_labels = {}
    teacher_homeroom_hours = {}
    teacher_total_allocated_hours = {}
    teacher_class_fill_subject_keys = {}

    demand_items_by_subject = {}
    demand_items_lookup = {}
    demand_items_by_section_subject_code = {}
    for class_row in class_rows:
        grade_subjects = subjects_by_grade.get(class_row["grade_label"], [])
        for subject_item in grade_subjects:
            demand_item = {
                "planning_section_id": class_row["planning_section_id"],
                "class_key": class_row["class_key"],
                "class_label": class_row["class_label"],
                "class_status": class_row["class_status"],
                "grade_label": class_row["grade_label"],
                "section_name": class_row["section_name"],
                "subject_key": subject_item["subject_key"],
                "subject_code": subject_item["subject_code"],
                "subject_name": subject_item["subject_name"],
                "required_hours": subject_item["weekly_hours"],
                "remaining_hours": subject_item["weekly_hours"],
                "allocated_hours": 0,
                "teacher_id": None,
                "teacher_name": "",
                "coverage_type": "",
            }
            demand_items_by_subject.setdefault(subject_item["subject_key"], []).append(
                demand_item
            )
            demand_items_lookup[
                (class_row["class_key"], subject_item["subject_key"])
            ] = demand_item
            demand_items_by_section_subject_code[
                (
                    class_row["planning_section_id"],
                    str(subject_item["subject_code"] or "").strip().upper(),
                )
            ] = demand_item

    for subject_key in demand_items_by_subject:
        demand_items_by_subject[subject_key].sort(
            key=lambda item: (
                0 if item["class_status"] == "Current" else 1,
                _grade_sort_key(item["grade_label"]),
                item["section_name"],
                item["class_label"],
            )
        )

    explicit_section_subject_keys = _build_explicit_section_subject_keys(
        section_assignments
    )
    homeroom_assignments_by_teacher = _build_homeroom_assignments_by_teacher(
        subjects=subjects,
        planning_sections=planning_sections,
        explicit_section_subject_keys=explicit_section_subject_keys,
        valid_teacher_ids=set(teacher_by_id.keys())
    )

    assignment_rows = []

    def _register_assignment(
        teacher_id,
        teacher_name,
        demand_item,
        allocated_hours,
        coverage_type,
    ):
        if allocated_hours <= 0 or not demand_item:
            return

        demand_item["allocated_hours"] = min(
            int(demand_item.get("required_hours", 0)),
            int(demand_item.get("allocated_hours", 0)) + allocated_hours,
        )
        demand_item["remaining_hours"] = max(
            int(demand_item.get("required_hours", 0)) - int(demand_item["allocated_hours"]),
            0,
        )
        demand_item["teacher_id"] = teacher_id
        demand_item["teacher_name"] = teacher_name
        demand_item["coverage_type"] = coverage_type

        teacher_total_allocated_hours[teacher_id] = (
            teacher_total_allocated_hours.get(teacher_id, 0) + allocated_hours
        )
        teacher_primary_hours_map.setdefault(teacher_id, {})
        subject_key = demand_item["subject_key"]
        teacher_primary_hours_map[teacher_id][subject_key] = (
            teacher_primary_hours_map[teacher_id].get(subject_key, 0) + allocated_hours
        )
        teacher_subject_name_map.setdefault(
            teacher_id,
            set(),
        ).add(subject_key)

        class_key = demand_item["class_key"]
        teacher_class_allocations.setdefault(teacher_id, {}).setdefault(class_key, []).append(
            {
                "subject_key": subject_key,
                "subject_code": demand_item["subject_code"],
                "subject_name": demand_item["subject_name"],
                "allocated_hours": allocated_hours,
                "class_status": demand_item["class_status"],
            }
        )
        teacher_class_fill_subject_keys.setdefault(teacher_id, {})
        teacher_class_fill_subject_keys[teacher_id].setdefault(class_key, subject_key)

        assignment_rows.append(
            {
                "teacher_id": (
                    getattr(teacher_by_id.get(teacher_id), "teacher_id", None)
                    or "-"
                ),
                "teacher_name": teacher_name or "-",
                "class_label": demand_item["class_label"],
                "class_status": demand_item["class_status"],
                "subject_code": demand_item["subject_code"],
                "subject_name": demand_item["subject_name"],
                "allocated_hours": allocated_hours,
                "coverage_type": coverage_type,
            }
        )

    for assignment in section_assignments:
        subject_code = str(getattr(assignment, "subject_code", "") or "").strip().upper()
        demand_item = demand_items_by_section_subject_code.get(
            (getattr(assignment, "planning_section_id", None), subject_code)
        )
        if not demand_item:
            continue
        teacher = teacher_by_id.get(getattr(assignment, "teacher_id", None))
        if not teacher:
            continue
        teacher_name = _build_teacher_display_name(teacher)
        _register_assignment(
            teacher_id=teacher.id,
            teacher_name=teacher_name,
            demand_item=demand_item,
            allocated_hours=int(demand_item["required_hours"]),
            coverage_type="Manual",
        )

    for teacher_id, homeroom_items in homeroom_assignments_by_teacher.items():
        teacher = teacher_by_id.get(teacher_id)
        if not teacher:
            continue
        teacher_name = _build_teacher_display_name(teacher)
        for item in homeroom_items:
            demand_item = demand_items_lookup.get(
                (item.get("class_key"), item.get("subject_key"))
            )
            if not demand_item:
                continue
            teacher_homeroom_hours[teacher_id] = (
                teacher_homeroom_hours.get(teacher_id, 0) + int(item["required_hours"])
            )
            teacher_homeroom_subject_keys.setdefault(teacher_id, set()).add(
                item["subject_key"]
            )
            teacher_homeroom_section_labels.setdefault(teacher_id, [])
            if item["class_label"] not in teacher_homeroom_section_labels[teacher_id]:
                teacher_homeroom_section_labels[teacher_id].append(item["class_label"])
            teacher_homeroom_allocations.setdefault(teacher_id, []).append(
                {
                    **item,
                    "allocated_hours": int(item["required_hours"]),
                }
            )
            _register_assignment(
                teacher_id=teacher_id,
                teacher_name=teacher_name,
                demand_item=demand_item,
                allocated_hours=int(item["required_hours"]),
                coverage_type="Homeroom",
            )

    assignment_rows.sort(
        key=lambda row: (
            row["teacher_name"],
            row["class_label"],
            row["subject_code"],
        )
    )

    teacher_matrix_rows = []
    underloaded_teacher_rows = []
    teacher_profiles = reporting_context.get("teacher_profiles", [])
    profile_by_teacher_pk = {
        profile.get("teacher_pk"): profile
        for profile in teacher_profiles
        if profile.get("teacher_pk")
    }
    teacher_rows_by_pk = {
        row.get("teacher_pk"): row
        for row in reporting_context.get("teacher_rows", [])
        if row.get("teacher_pk")
    }

    sorted_teachers = sorted(
        teachers,
        key=lambda teacher: (
            -teacher_total_allocated_hours.get(getattr(teacher, "id", None), 0),
            _build_teacher_display_name(teacher),
            getattr(teacher, "id", 0) or 0,
        ),
    )

    for teacher in sorted_teachers:
        teacher_id = getattr(teacher, "id", None)
        profile = profile_by_teacher_pk.get(teacher_id, {})
        report_teacher_row = teacher_rows_by_pk.get(teacher_id, {})
        class_cells = {}
        class_fill_subject_keys = {}
        for class_key, allocation_items in teacher_class_allocations.get(teacher_id, {}).items():
            allocation_items.sort(
                key=lambda item: (-item["allocated_hours"], item["subject_code"])
            )
            class_cells[class_key] = "\n".join(
                f"{item['subject_code']} ({item['allocated_hours']}h)"
                for item in allocation_items
            )
            class_fill_subject_keys[class_key] = allocation_items[0]["subject_key"]

        teacher_matrix_row = {
            "teacher_pk": teacher_id,
            "teacher_id": teacher.teacher_id or "-",
            "teacher_name": _build_teacher_display_name(teacher),
            "degree_major": str(getattr(teacher, "degree_major", "") or "").strip(),
            "expected_allocated_hours": int(
                teacher_total_allocated_hours.get(teacher_id, 0)
            ),
            "capacity_hours": int(
                report_teacher_row.get(
                    "capacity_hours",
                    get_teacher_capacity_breakdown(
                        teacher,
                        default_max_hours=REPORT_STANDARD_MAX_HOURS,
                    )["international_capacity_hours"],
                )
            ),
            "remaining_capacity_hours": int(
                report_teacher_row.get("remaining_capacity_hours", 0)
            ),
            "recommended_absorption_hours": 0,
            "recommended_assignment_labels": [],
            "primary_subject_label": ", ".join(
                subject_name_by_key.get(subject_key, subject_key.title())
                for subject_key in sorted(
                    teacher_subject_name_map.get(teacher_id, set()),
                    key=lambda key: subject_name_by_key.get(key, key),
                )
            )
            or "-",
            "support_subject_label": "-",
            "class_cells": class_cells,
            "class_fill_subject_keys": class_fill_subject_keys,
        }
        teacher_matrix_rows.append(teacher_matrix_row)

        if int(teacher_matrix_row["remaining_capacity_hours"]) > 0:
            underloaded_teacher_rows.append(
                {
                    "teacher_pk": teacher_id,
                    "teacher_id": teacher.teacher_id or "-",
                    "teacher_name": _build_teacher_display_name(teacher),
                    "degree_major": teacher_matrix_row["degree_major"],
                    "current_load_hours": int(
                        teacher_matrix_row["expected_allocated_hours"]
                    ),
                    "capacity_hours": int(teacher_matrix_row["capacity_hours"]),
                    "remaining_capacity_hours": int(
                        teacher_matrix_row["remaining_capacity_hours"]
                    ),
                    "projected_allocated_hours": int(
                        teacher_matrix_row["expected_allocated_hours"]
                    ),
                    "projected_remaining_capacity_hours": int(
                        teacher_matrix_row["remaining_capacity_hours"]
                    ),
                    "recommended_absorption_hours": 0,
                    "recommended_assignment_labels": [],
                }
            )

    unassigned_rows = []
    subject_section_rows = []
    subject_section_map = {}
    for subject_key, subject_items in demand_items_by_subject.items():
        covered_section_labels = []
        partial_section_labels = []
        uncovered_section_labels = []
        for demand_item in subject_items:
            required_hours = int(demand_item["required_hours"])
            allocated_hours = int(demand_item["allocated_hours"])
            remaining_hours = int(demand_item["remaining_hours"])
            class_label = demand_item["class_label"]
            if remaining_hours <= 0:
                covered_section_labels.append(class_label)
            elif allocated_hours > 0:
                partial_section_labels.append(
                    f"{class_label} / {remaining_hours}h"
                )
            else:
                uncovered_section_labels.append(f"{class_label} / {remaining_hours}h")
                unassigned_rows.append(
                    {
                        "class_label": class_label,
                        "class_status": demand_item["class_status"],
                        "subject_code": demand_item["subject_code"],
                        "subject_name": demand_item["subject_name"],
                        "remaining_hours": remaining_hours,
                    }
                )

        section_row = {
            "subject_key": subject_key,
            "total_sections": len(subject_items),
            "covered_sections_count": len(covered_section_labels),
            "partial_sections_count": len(partial_section_labels),
            "recommended_sections_count": 0,
            "uncovered_sections_count": len(uncovered_section_labels),
            "covered_section_labels": covered_section_labels,
            "partial_section_labels": partial_section_labels,
            "recommended_section_labels": [],
            "uncovered_section_labels": uncovered_section_labels,
            "recommended_hours": 0,
        }
        subject_section_rows.append(section_row)
        subject_section_map[subject_key] = section_row

    unassigned_rows.sort(
        key=lambda row: (
            row["class_label"],
            row["subject_code"],
        )
    )
    underloaded_teacher_rows.sort(
        key=lambda row: (
            -row["remaining_capacity_hours"],
            row["teacher_name"],
        )
    )

    return {
        "class_rows": class_rows,
        "teacher_matrix_rows": teacher_matrix_rows,
        "assignment_rows": assignment_rows,
        "unassigned_rows": unassigned_rows,
        "subject_section_rows": subject_section_rows,
        "subject_section_map": subject_section_map,
        "underloaded_teacher_rows": underloaded_teacher_rows,
        "recommendation_rows": [],
    }


def _build_reporting_context(
    db: Session,
    subjects,
    planning_sections,
    teachers,
):
    sections_by_grade = {}
    current_sections_by_grade = {}
    new_sections_by_grade = {}

    for section in planning_sections:
        grade_label = _normalize_grade_label(section.grade_level)
        if not grade_label:
            continue

        sections_by_grade[grade_label] = sections_by_grade.get(grade_label, 0) + 1
        status = str(section.class_status or "").strip().lower()
        if status == "current":
            current_sections_by_grade[grade_label] = (
                current_sections_by_grade.get(grade_label, 0) + 1
            )
        elif status == "new":
            new_sections_by_grade[grade_label] = (
                new_sections_by_grade.get(grade_label, 0) + 1
            )

    scoped_subjects_by_code = {
        subject.subject_code: subject
        for subject in subjects
        if subject.subject_code
    }
    subject_demand_map = {}
    required_hours_by_grade = {}
    required_current_hours_by_grade = {}
    required_new_hours_by_grade = {}

    for subject in subjects:
        grade_label = _normalize_grade_label(subject.grade)
        if not grade_label:
            continue

        weekly_hours = int(subject.weekly_hours or 0)
        if weekly_hours <= 0:
            continue

        sections_count = sections_by_grade.get(grade_label, 0)
        if sections_count <= 0:
            continue

        current_sections_count = current_sections_by_grade.get(grade_label, 0)
        new_sections_count = new_sections_by_grade.get(grade_label, 0)

        subject_key, subject_label = _build_subject_identity(
            subject_name=subject.subject_name,
            fallback_code=subject.subject_code or "",
        )
        if not subject_key:
            continue

        required_hours = weekly_hours * sections_count
        required_current_hours = weekly_hours * current_sections_count
        required_new_hours = weekly_hours * new_sections_count

        required_hours_by_grade[grade_label] = (
            required_hours_by_grade.get(grade_label, 0) + required_hours
        )
        required_current_hours_by_grade[grade_label] = (
            required_current_hours_by_grade.get(grade_label, 0) + required_current_hours
        )
        required_new_hours_by_grade[grade_label] = (
            required_new_hours_by_grade.get(grade_label, 0) + required_new_hours
        )

        if subject_key not in subject_demand_map:
            subject_demand_map[subject_key] = {
                "subject_name": subject_label,
                "subject_code": subject.subject_code or "",
                "subject_color": resolve_subject_color(
                    subject.subject_code or subject_key,
                    getattr(subject, "color", ""),
                    subject_name=subject.subject_name,
                ),
                "weekly_hours": weekly_hours,
                "primary_grade_label": grade_label,
                "bundle_subject_labels": list(
                    get_homeroom_bundle_subject_labels(
                        subject_code=subject.subject_code or "",
                        subject_name=subject.subject_name or "",
                        weekly_hours=weekly_hours,
                        grade_label=grade_label,
                    )
                ),
                "required_hours": 0,
                "required_current_hours": 0,
                "required_new_hours": 0,
                "grades": set(),
            }

        entry = subject_demand_map[subject_key]
        entry["required_hours"] += required_hours
        entry["required_current_hours"] += required_current_hours
        entry["required_new_hours"] += required_new_hours
        entry["grades"].add(grade_label)

    teacher_subject_map = {
        teacher.id: set()
        for teacher in teachers
        if getattr(teacher, "id", None)
    }
    teacher_subject_override_map = {
        teacher.id: set()
        for teacher in teachers
        if getattr(teacher, "id", None)
    }
    teacher_subject_effective_count_map = {
        teacher.id: 0
        for teacher in teachers
        if getattr(teacher, "id", None)
    }
    teacher_subject_hours_map = {
        teacher.id: {}
        for teacher in teachers
        if getattr(teacher, "id", None)
    }
    teacher_ids = sorted(teacher_subject_map.keys())

    if teacher_ids:
        teacher_allocations = db.query(models.TeacherSubjectAllocation).filter(
            models.TeacherSubjectAllocation.teacher_id.in_(teacher_ids)
        ).all()
    else:
        teacher_allocations = []

    for allocation in teacher_allocations:
        subject_key_set = teacher_subject_map.get(allocation.teacher_id)
        if subject_key_set is None:
            continue

        subject = scoped_subjects_by_code.get(allocation.subject_code)
        if not subject:
            continue

        subject_key, _ = _build_subject_identity(
            subject_name=subject.subject_name,
            fallback_code=subject.subject_code or "",
        )
        if not subject_key or subject_key not in subject_demand_map:
            continue
        subject_key_set.add(subject_key)
        teacher_subject_effective_count_map[allocation.teacher_id] = (
            teacher_subject_effective_count_map.get(allocation.teacher_id, 0)
            + get_effective_subject_count(
                subject_code=subject.subject_code or "",
                subject_name=subject.subject_name or "",
                weekly_hours=subject.weekly_hours,
                grade_label=_normalize_grade_label(subject.grade),
            )
        )
        if allocation.compatibility_override:
            teacher_subject_override_map.setdefault(allocation.teacher_id, set()).add(
                subject_key
            )
        subject_hours = int(subject.weekly_hours or 0)
        subject_hours_map = teacher_subject_hours_map.get(allocation.teacher_id, {})
        subject_hours_map[subject_key] = (
            subject_hours_map.get(subject_key, 0) + max(subject_hours, 0)
        )

    for teacher in teachers:
        subject_key_set = teacher_subject_map.get(getattr(teacher, "id", None))
        if subject_key_set is None or subject_key_set:
            continue

        fallback_code = str(teacher.subject_code or "").strip().upper()
        if not fallback_code:
            continue

        fallback_subject = scoped_subjects_by_code.get(fallback_code)
        if not fallback_subject:
            continue

        subject_key, _ = _build_subject_identity(
            subject_name=fallback_subject.subject_name,
            fallback_code=fallback_subject.subject_code or "",
        )
        if subject_key and subject_key in subject_demand_map:
            subject_key_set.add(subject_key)
            teacher_subject_effective_count_map[getattr(teacher, "id", None)] = (
                teacher_subject_effective_count_map.get(getattr(teacher, "id", None), 0)
                + get_effective_subject_count(
                    subject_code=fallback_subject.subject_code or "",
                    subject_name=fallback_subject.subject_name or "",
                    weekly_hours=fallback_subject.weekly_hours,
                    grade_label=_normalize_grade_label(fallback_subject.grade),
                )
            )
            fallback_subject_hours_map = teacher_subject_hours_map.get(
                getattr(teacher, "id", None),
                {},
            )
            fallback_subject_hours_map[subject_key] = max(
                fallback_subject_hours_map.get(subject_key, 0),
                int(fallback_subject.weekly_hours or 0),
            )

    teacher_profiles = []
    for teacher in teachers:
        teacher_id = getattr(teacher, "id", None)
        teacher_major_text = str(getattr(teacher, "degree_major", "") or "").strip()
        candidate_subject_keys = sorted(
            teacher_subject_map.get(teacher_id, set()),
            key=lambda key: subject_demand_map[key]["subject_name"],
        )
        override_subject_keys = set(teacher_subject_override_map.get(teacher_id, set()))
        subject_hours_map = teacher_subject_hours_map.get(teacher_id, {})
        ranked_subject_keys = []
        major_aligned_subject_keys = []
        primary_subject_keys = []
        secondary_subject_keys = []
        primary_subject_key = None
        major_priority_rule = _resolve_teacher_major_priority_rule(teacher_major_text)
        if candidate_subject_keys:
            ranked_subject_keys = sorted(
                candidate_subject_keys,
                key=lambda key: (
                    -int(
                        _subject_matches_teacher_major(
                            teacher,
                            subject_demand_map[key]["subject_name"],
                            key,
                        )
                    ),
                    -subject_hours_map.get(key, 0),
                    -subject_demand_map[key]["required_hours"],
                    subject_demand_map[key]["subject_name"],
                ),
            )
            major_aligned_subject_keys = [
                key
                for key in ranked_subject_keys
                if _subject_matches_teacher_major(
                    teacher,
                    subject_demand_map[key]["subject_name"],
                    key,
                )
            ]
            ranked_index_by_key = {
                subject_key: index
                for index, subject_key in enumerate(ranked_subject_keys)
            }
            rule_pool_key = str(major_priority_rule.get("pool_key", "") or "")
            rule_priority_families = list(
                major_priority_rule.get("priority_families", []) or []
            )
            rule_allowed_families = list(
                major_priority_rule.get("allowed_families", []) or []
            )
            candidate_subject_family_map = {
                key: _build_subject_priority_family(
                    subject_name=subject_demand_map[key]["subject_name"],
                    subject_key=key,
                    subject_code=subject_demand_map[key].get("subject_code", ""),
                )
                for key in ranked_subject_keys
            }
            rule_subject_keys = []
            if rule_pool_key and rule_allowed_families:
                allowed_rule_families = set(rule_allowed_families)
                rule_candidate_subject_keys = list(ranked_subject_keys)
                if major_priority_rule.get("rule_key") == "general_science_major":
                    seen_rule_candidate_keys = set(rule_candidate_subject_keys)
                    for demand_subject_key, demand_subject in subject_demand_map.items():
                        if demand_subject_key in seen_rule_candidate_keys:
                            continue
                        demand_subject_family = _build_subject_priority_family(
                            subject_name=demand_subject["subject_name"],
                            subject_key=demand_subject_key,
                            subject_code=demand_subject.get("subject_code", ""),
                        )
                        if (
                            HIRING_COMPATIBILITY_GROUPS.get(demand_subject_family)
                            == rule_pool_key
                            and demand_subject_family in allowed_rule_families
                        ):
                            rule_candidate_subject_keys.append(demand_subject_key)
                            seen_rule_candidate_keys.add(demand_subject_key)
                            candidate_subject_family_map[demand_subject_key] = (
                                demand_subject_family
                            )
                pool_subject_keys = [
                    key
                    for key in rule_candidate_subject_keys
                    if HIRING_COMPATIBILITY_GROUPS.get(
                        candidate_subject_family_map.get(key, "")
                    )
                    == rule_pool_key
                    and candidate_subject_family_map.get(key, "") in allowed_rule_families
                ]
                family_order_map = {
                    family: index
                    for index, family in enumerate(rule_allowed_families)
                }
                rule_subject_keys = sorted(
                    pool_subject_keys,
                    key=lambda key: (
                        family_order_map.get(
                            candidate_subject_family_map.get(key, ""),
                            len(family_order_map) + 1,
                        ),
                        ranked_index_by_key.get(key, 999),
                        -subject_demand_map[key]["required_hours"],
                        subject_demand_map[key]["subject_name"],
                    ),
                )

            if rule_subject_keys:
                primary_subject_keys = list(rule_subject_keys)
                secondary_subject_keys = []
            elif major_aligned_subject_keys:
                primary_subject_keys = list(major_aligned_subject_keys)
                secondary_subject_keys = [
                    key for key in ranked_subject_keys if key not in primary_subject_keys
                ]
            else:
                primary_subject_keys = list(ranked_subject_keys)

            primary_subject_key = ranked_subject_keys[0]

        support_subject_keys = list(secondary_subject_keys)
        if major_priority_rule.get("pool_key") and primary_subject_keys:
            support_subject_keys = []
        seen_support_subject_keys = set(support_subject_keys)
        for base_subject_key in ranked_subject_keys:
            for support_subject_key in CROSS_SUBJECT_SUPPORT_RULES.get(
                base_subject_key,
                set(),
            ):
                normalized_support_key = _normalize_subject_family_key(
                    support_subject_key
                )
                if (
                    normalized_support_key
                    and normalized_support_key in subject_demand_map
                    and normalized_support_key not in ranked_subject_keys
                    and normalized_support_key not in seen_support_subject_keys
                ):
                    support_subject_keys.append(normalized_support_key)
                    seen_support_subject_keys.add(normalized_support_key)

        logging.getLogger("uvicorn.error").debug(
            "Smart auto-plan teacher rule: teacher=%s major=%s rule=%s pool=%s primary=%s secondary=%s support=%s",
            _build_teacher_display_name(teacher),
            teacher_major_text or "-",
            major_priority_rule.get("rule_key", "generic_major_match"),
            major_priority_rule.get("pool_key", "") or "-",
            primary_subject_keys,
            secondary_subject_keys,
            support_subject_keys,
        )

        teacher_capacity_breakdown = get_teacher_capacity_breakdown(
            teacher,
            default_max_hours=REPORT_STANDARD_MAX_HOURS,
        )
        teacher_capacity = teacher_capacity_breakdown["international_capacity_hours"]

        teacher_profiles.append(
            {
                "teacher": teacher,
                "name": _build_teacher_display_name(teacher),
                "degree_major": teacher_major_text,
                "subject_keys": primary_subject_keys,
                "secondary_subject_keys": secondary_subject_keys,
                "support_subject_keys": support_subject_keys,
                "eligible_subject_keys": primary_subject_keys + support_subject_keys,
                "override_subject_keys": sorted(override_subject_keys),
                "subject_count": max(
                    teacher_subject_effective_count_map.get(teacher_id, 0),
                    len(ranked_subject_keys),
                ),
                "primary_subject_basis_hours": (
                    subject_hours_map.get(primary_subject_key, 0)
                    if primary_subject_key
                    else 0
                ),
                "homeroom_allocated_hours": 0,
                "homeroom_subject_keys": [],
                "homeroom_section_labels": [],
                "homeroom_class_allocations": [],
                "homeroom_allocation_breakdown": {},
                "allocated_hours": 0,
                "primary_allocated_hours": 0,
                "support_allocated_hours": 0,
                "remaining_capacity_hours": teacher_capacity,
                "capacity_hours": teacher_capacity,
                "total_capacity_hours": teacher_capacity_breakdown[
                    "total_capacity_hours"
                ],
                "extra_hours_count": teacher_capacity_breakdown["extra_hours"],
                "major_priority_rule_key": major_priority_rule.get("rule_key", ""),
                "major_priority_rule_label": major_priority_rule.get("label", ""),
                "major_priority_pool_key": major_priority_rule.get("pool_key", ""),
                "major_priority_family_order": list(
                    major_priority_rule.get("allowed_families", []) or []
                ),
                "major_priority_primary_families": list(
                    major_priority_rule.get("priority_families", []) or []
                ),
                "national_section_hours": teacher_capacity_breakdown[
                    "national_section_hours"
                ],
                "allocation_breakdown": {},
            }
        )

    remaining_hours_by_subject = {
        subject_key: data["required_hours"]
        for subject_key, data in subject_demand_map.items()
    }
    planning_section_ids = [
        section.id
        for section in planning_sections
        if getattr(section, "id", None)
    ]
    explicit_section_assignments = []
    if planning_section_ids:
        explicit_section_assignments = db.query(models.TeacherSectionAssignment).filter(
            models.TeacherSectionAssignment.planning_section_id.in_(planning_section_ids)
        ).all()
    explicit_section_subject_keys = _build_explicit_section_subject_keys(
        explicit_section_assignments
    )
    # Reserve homeroom-owned class subjects before pooled branchwide allocation.
    homeroom_assignments_by_teacher = _build_homeroom_assignments_by_teacher(
    subjects=subjects,
    planning_sections=planning_sections,
    explicit_section_subject_keys=explicit_section_subject_keys,
    valid_teacher_ids={teacher.id for teacher in teachers if getattr(teacher, "id", None)},
    )

    for profile in teacher_profiles:
        teacher_id = getattr(profile["teacher"], "id", None)
        homeroom_items = homeroom_assignments_by_teacher.get(teacher_id, [])
        homeroom_remaining_capacity = int(
            profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
        )
        homeroom_allocation_breakdown = {}
        homeroom_class_allocations = []
        homeroom_section_labels = []

        for item in homeroom_items:
            if homeroom_remaining_capacity <= 0:
                break

            subject_key = item["subject_key"]
            subject_remaining_hours = remaining_hours_by_subject.get(subject_key, 0)
            if subject_remaining_hours <= 0:
                continue

            allocated_hours = min(
                homeroom_remaining_capacity,
                int(item["required_hours"]),
                subject_remaining_hours,
            )
            if allocated_hours <= 0:
                continue

            remaining_hours_by_subject[subject_key] = (
                subject_remaining_hours - allocated_hours
            )
            homeroom_remaining_capacity -= allocated_hours
            homeroom_allocation_breakdown[subject_key] = (
                homeroom_allocation_breakdown.get(subject_key, 0) + allocated_hours
            )
            homeroom_class_allocations.append(
                {
                    **item,
                    "allocated_hours": allocated_hours,
                }
            )
            if item["class_label"] not in homeroom_section_labels:
                homeroom_section_labels.append(item["class_label"])

        homeroom_allocated_hours = sum(homeroom_allocation_breakdown.values())
        profile["homeroom_allocated_hours"] = homeroom_allocated_hours
        profile["homeroom_subject_keys"] = sorted(
            homeroom_allocation_breakdown.keys(),
            key=lambda key: subject_demand_map[key]["subject_name"],
        )
        profile["homeroom_section_labels"] = homeroom_section_labels
        profile["homeroom_class_allocations"] = homeroom_class_allocations
        profile["homeroom_allocation_breakdown"] = homeroom_allocation_breakdown
        profile["allocated_hours"] = homeroom_allocated_hours
        profile["remaining_capacity_hours"] = max(
            int(profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS))
            - homeroom_allocated_hours,
            0,
        )

    allocation_sequence = sorted(
        teacher_profiles,
        key=lambda profile: (
            profile["subject_count"] if profile["subject_count"] else 999,
            profile["name"],
            profile["teacher"].id or 0,
        ),
    )

    for profile in allocation_sequence:
        if not profile["eligible_subject_keys"]:
            continue

        remaining_capacity = int(
            profile.get("remaining_capacity_hours", REPORT_STANDARD_MAX_HOURS)
        )
        allocation_breakdown = {}
        major_family_order = list(profile.get("major_priority_family_order", []) or [])
        major_family_order_map = {
            family: index for index, family in enumerate(major_family_order)
        }

        while remaining_capacity > 0:
            primary_candidate_subject_keys = [
                subject_key
                for subject_key in profile["subject_keys"]
                if remaining_hours_by_subject.get(subject_key, 0) > 0
            ]
            secondary_candidate_subject_keys = [
                subject_key
                for subject_key in profile.get("secondary_subject_keys", [])
                if remaining_hours_by_subject.get(subject_key, 0) > 0
            ]
            support_candidate_subject_keys = [
                subject_key
                for subject_key in profile["support_subject_keys"]
                if (
                    subject_key not in profile.get("secondary_subject_keys", [])
                    and remaining_hours_by_subject.get(subject_key, 0) > 0
                )
            ]

            if primary_candidate_subject_keys:
                candidate_subject_keys = primary_candidate_subject_keys
            elif secondary_candidate_subject_keys:
                candidate_subject_keys = secondary_candidate_subject_keys
            else:
                candidate_subject_keys = support_candidate_subject_keys

            if not candidate_subject_keys:
                logging.getLogger("uvicorn.error").debug(
                    "Smart auto-plan stop: teacher=%s rule=%s remaining_capacity=%sh reason=no compatible uncovered subjects",
                    profile.get("name", "Unknown Teacher"),
                    profile.get("major_priority_rule_key", "generic_major_match"),
                    remaining_capacity,
                )
                break

            candidate_subject_keys.sort(
                key=lambda subject_key: (
                    major_family_order_map.get(
                        _build_subject_priority_family(
                            subject_name=subject_demand_map[subject_key]["subject_name"],
                            subject_key=subject_key,
                            subject_code=subject_demand_map[subject_key].get(
                                "subject_code", ""
                            ),
                        ),
                        len(major_family_order_map) + 1,
                    ),
                    -remaining_hours_by_subject.get(subject_key, 0),
                    subject_demand_map[subject_key]["subject_name"],
                )
            )
            selected_subject_key = candidate_subject_keys[0]
            subject_remaining_hours = remaining_hours_by_subject.get(
                selected_subject_key, 0
            )
            allocated_hours = min(remaining_capacity, subject_remaining_hours)
            if allocated_hours <= 0:
                logging.getLogger("uvicorn.error").debug(
                    "Smart auto-plan stop: teacher=%s subject=%s remaining_capacity=%sh reason=allocated_hours_zero",
                    profile.get("name", "Unknown Teacher"),
                    selected_subject_key,
                    remaining_capacity,
                )
                break

            selected_family = _build_subject_priority_family(
                subject_name=subject_demand_map[selected_subject_key]["subject_name"],
                subject_key=selected_subject_key,
                subject_code=subject_demand_map[selected_subject_key].get(
                    "subject_code", ""
                ),
            )
            skipped_subject_details = [
                (
                    subject_key,
                    _build_subject_priority_family(
                        subject_name=subject_demand_map[subject_key]["subject_name"],
                        subject_key=subject_key,
                        subject_code=subject_demand_map[subject_key].get(
                            "subject_code", ""
                        ),
                    ),
                    remaining_hours_by_subject.get(subject_key, 0),
                )
                for subject_key in candidate_subject_keys[1:]
            ]
            logging.getLogger("uvicorn.error").debug(
                "Smart auto-plan pick: teacher=%s major_rule=%s pool=%s selected=%s family=%s allocated=%sh remaining_before=%sh skipped=%s",
                profile.get("name", "Unknown Teacher"),
                profile.get("major_priority_rule_key", "generic_major_match"),
                profile.get("major_priority_pool_key", "") or "-",
                selected_subject_key,
                selected_family or "-",
                allocated_hours,
                remaining_capacity,
                skipped_subject_details,
            )

            allocation_breakdown[selected_subject_key] = (
                allocation_breakdown.get(selected_subject_key, 0) + allocated_hours
            )
            remaining_hours_by_subject[selected_subject_key] = (
                subject_remaining_hours - allocated_hours
            )
            remaining_capacity -= allocated_hours

        homeroom_allocated_hours = int(profile.get("homeroom_allocated_hours", 0))
        allocated_hours_total = sum(allocation_breakdown.values())
        teacher_capacity_hours = int(
            profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
        )
        if homeroom_allocated_hours + allocated_hours_total > teacher_capacity_hours:
            overflow_hours = (
                homeroom_allocated_hours + allocated_hours_total - teacher_capacity_hours
            )
            reduction_order = (
                list(profile["support_subject_keys"]) + list(profile["subject_keys"])
            )
            for subject_key in reduction_order:
                if overflow_hours <= 0:
                    break
                current_hours = allocation_breakdown.get(subject_key, 0)
                if current_hours <= 0:
                    continue
                reduce_hours = min(current_hours, overflow_hours)
                updated_hours = current_hours - reduce_hours
                if updated_hours > 0:
                    allocation_breakdown[subject_key] = updated_hours
                else:
                    allocation_breakdown.pop(subject_key, None)
                remaining_hours_by_subject[subject_key] = (
                    remaining_hours_by_subject.get(subject_key, 0) + reduce_hours
                )
                overflow_hours -= reduce_hours

        primary_allocated_hours = sum(
            allocation_breakdown.get(subject_key, 0)
            for subject_key in profile["subject_keys"]
        )
        support_allocated_hours = sum(
            allocation_breakdown.get(subject_key, 0)
            for subject_key in profile["support_subject_keys"]
        )
        total_allocated_hours = min(
            homeroom_allocated_hours + sum(allocation_breakdown.values()),
            teacher_capacity_hours,
        )
        logging.getLogger("uvicorn.error").debug(
            "Smart auto-plan final: teacher=%s major=%s rule=%s pool=%s allocated=%sh remaining_capacity=%sh breakdown=%s",
            profile.get("name", "Unknown Teacher"),
            profile.get("degree_major", "") or "-",
            profile.get("major_priority_rule_key", "generic_major_match"),
            profile.get("major_priority_pool_key", "") or "-",
            total_allocated_hours,
            max(teacher_capacity_hours - total_allocated_hours, 0),
            allocation_breakdown,
        )
        profile["allocation_breakdown"] = allocation_breakdown
        profile["allocated_hours"] = total_allocated_hours
        profile["primary_allocated_hours"] = primary_allocated_hours
        profile["support_allocated_hours"] = support_allocated_hours
        profile["remaining_capacity_hours"] = max(
            teacher_capacity_hours - total_allocated_hours,
            0,
        )

    teachers_per_subject = {}
    for profile in teacher_profiles:
        covered_subject_keys = set(profile["eligible_subject_keys"]) | set(
            profile.get("homeroom_subject_keys", [])
        )
        for subject_key in covered_subject_keys:
            teachers_per_subject[subject_key] = (
                teachers_per_subject.get(subject_key, 0) + 1
            )

    report_subject_rows = []
    for subject_key, demand in subject_demand_map.items():
        required_hours = demand["required_hours"]
        remaining_hours = remaining_hours_by_subject.get(subject_key, 0)
        allocated_hours = max(required_hours - remaining_hours, 0)
        teacher_requirement_blocks = (
            math.ceil(remaining_hours / REPORT_STANDARD_MAX_HOURS)
            if remaining_hours > 0
            else 0
        )
        grades = sorted(demand["grades"], key=_grade_sort_key)
        coverage_percentage = (
            round((allocated_hours / required_hours) * 100)
            if required_hours > 0
            else 0
        )

        report_subject_rows.append(
            {
                "subject_key": subject_key,
                "subject_name": demand["subject_name"],
                "subject_code": demand.get("subject_code", subject_key),
                "subject_color": resolve_subject_color(
                    demand.get("subject_code", subject_key),
                    demand.get("subject_color", ""),
                    subject_name=demand.get("subject_name", ""),
                ),
                "grades": grades,
                "required_hours": required_hours,
                "required_current_hours": demand["required_current_hours"],
                "required_new_hours": demand["required_new_hours"],
                "allocated_hours": allocated_hours,
                "remaining_hours": remaining_hours,
                "coverage_percentage": coverage_percentage,
                "teachers_with_subject": teachers_per_subject.get(subject_key, 0),
                "effective_subject_count": max(
                    len(demand.get("bundle_subject_labels", [])),
                    1,
                ),
                "teacher_requirement_blocks": teacher_requirement_blocks,
                "additional_teachers_needed": teacher_requirement_blocks,
                "additional_teachers_note": "",
                "priority_staffing_subject": _is_priority_staffing_subject(
                    subject_key,
                    demand["subject_name"],
                ),
                "internal_absorption_recommended": False,
            }
        )

    report_subject_rows.sort(
        key=lambda row: (
            -row["remaining_hours"],
            row["subject_name"],
        )
    )

    report_gap_rows = [
        dict(row)
        for row in report_subject_rows
        if row["remaining_hours"] > 0
    ]
    max_remaining_hours = max(
        (row["remaining_hours"] for row in report_gap_rows),
        default=0,
    )
    for row in report_gap_rows:
        row["gap_chart_pct"] = (
            round((row["remaining_hours"] / max_remaining_hours) * 100, 1)
            if max_remaining_hours > 0
            else 0
        )
    report_gap_rows = report_gap_rows[:8]

    report_teacher_rows = []
    for profile in teacher_profiles:
        subject_labels = [
            label
            for subject_key in profile["subject_keys"]
            for label in _build_subject_display_labels(
                subject_name=subject_demand_map[subject_key]["subject_name"],
                subject_code=subject_demand_map[subject_key].get("subject_code", ""),
                weekly_hours=subject_demand_map[subject_key].get("weekly_hours", 0),
                grade_label=subject_demand_map[subject_key].get("primary_grade_label"),
            )
        ]
        support_subject_labels = [
            label
            for subject_key in profile["support_subject_keys"]
            for label in _build_subject_display_labels(
                subject_name=subject_demand_map[subject_key]["subject_name"],
                subject_code=subject_demand_map[subject_key].get("subject_code", ""),
                weekly_hours=subject_demand_map[subject_key].get("weekly_hours", 0),
                grade_label=subject_demand_map[subject_key].get("primary_grade_label"),
            )
        ]
        homeroom_subject_labels = [
            label
            for subject_key in profile.get("homeroom_subject_keys", [])
            for label in _build_subject_display_labels(
                subject_name=subject_demand_map[subject_key]["subject_name"],
                subject_code=subject_demand_map[subject_key].get("subject_code", ""),
                weekly_hours=subject_demand_map[subject_key].get("weekly_hours", 0),
                grade_label=subject_demand_map[subject_key].get("primary_grade_label"),
            )
        ]
        homeroom_section_labels = list(profile.get("homeroom_section_labels", []))
        allocation_labels = [
            f"{subject_demand_map[subject_key]['subject_name']} ({hours}h)"
            for subject_key, hours in sorted(
                profile["allocation_breakdown"].items(),
                key=lambda item: (-item[1], subject_demand_map[item[0]]["subject_name"]),
            )
        ]
        homeroom_labels_by_class = {}
        for item in profile.get("homeroom_class_allocations", []):
            homeroom_labels_by_class.setdefault(item["class_label"], []).append(item)
        homeroom_allocation_labels = []
        for class_label in sorted(homeroom_labels_by_class.keys()):
            allocation_items = homeroom_labels_by_class[class_label]
            allocation_items.sort(
                key=lambda item: (
                    -item["allocated_hours"],
                    item["subject_name"],
                )
            )
            homeroom_allocation_labels.append(
                f"{class_label}: "
                + ", ".join(
                    (
                        f"{item['subject_name']} ({item['allocated_hours']}h)"
                        if not get_homeroom_bundle_subject_labels(
                            subject_code=item.get("subject_code", ""),
                            subject_name=item.get("subject_name", ""),
                            weekly_hours=item.get("allocated_hours", 0),
                            grade_label=item.get("grade_label"),
                        )
                        else (
                            f"{item['subject_name']} ({item['allocated_hours']}h) "
                            f"| Includes {', '.join(get_homeroom_bundle_subject_labels(item.get('subject_code', ''), item.get('subject_name', ''), item.get('allocated_hours', 0), item.get('grade_label')))}"
                        )
                    )
                    for item in allocation_items
                )
            )
        subject_chips = [
            {
                "name": subject_demand_map[subject_key]["subject_name"],
                "color": subject_demand_map[subject_key].get("subject_color", "#0A4EA3"),
                "hours": hours,
            }
            for subject_key, hours in sorted(
                profile["allocation_breakdown"].items(),
                key=lambda item: (-item[1], subject_demand_map.get(item[0], {}).get("subject_name", item[0])),
            )
            if subject_key in subject_demand_map
        ]

        teacher = profile["teacher"]
        report_teacher_rows.append(
            {
                "teacher_pk": teacher.id,
                "teacher_id": teacher.teacher_id or "-",
                "teacher_name": profile["name"],
                "degree_major": str(getattr(teacher, "degree_major", "") or "").strip(),
                "subject_labels": subject_labels,
                "subject_chips": subject_chips,
                "support_subject_labels": support_subject_labels,
                "homeroom_subject_labels": homeroom_subject_labels,
                "homeroom_section_labels": homeroom_section_labels,
                "homeroom_allocation_labels": homeroom_allocation_labels,
                "allocation_labels": allocation_labels,
                "expected_allocated_hours": profile["allocated_hours"],
                "homeroom_allocated_hours": profile.get("homeroom_allocated_hours", 0),
                "primary_allocated_hours": profile.get("primary_allocated_hours", 0),
                "support_allocated_hours": profile.get("support_allocated_hours", 0),
                "primary_subject_basis_hours": profile.get(
                    "primary_subject_basis_hours",
                    0,
                ),
                "capacity_hours": profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS),
                "total_capacity_hours": profile.get(
                    "total_capacity_hours",
                    REPORT_STANDARD_MAX_HOURS,
                ),
                "extra_hours_count": profile.get("extra_hours_count", 0),
                "national_section_hours": profile.get("national_section_hours", 0),
                "remaining_capacity_hours": profile["remaining_capacity_hours"],
            }
        )

    report_teacher_rows.sort(
        key=lambda row: (
            -row["expected_allocated_hours"],
            row["teacher_name"],
        )
    )

    report_grade_rows = []
    for grade_label, total_sections in sections_by_grade.items():
        report_grade_rows.append(
            {
                "grade_label": grade_label,
                "sections_total": total_sections,
                "sections_current": current_sections_by_grade.get(grade_label, 0),
                "sections_new": new_sections_by_grade.get(grade_label, 0),
                "required_hours_total": required_hours_by_grade.get(grade_label, 0),
                "required_hours_current": required_current_hours_by_grade.get(
                    grade_label, 0
                ),
                "required_hours_new": required_new_hours_by_grade.get(grade_label, 0),
            }
        )
    report_grade_rows.sort(
        key=lambda row: _grade_sort_key(row["grade_label"])
    )

    total_required_hours = sum(
        row["required_hours"] for row in report_subject_rows
    )
    total_remaining_hours = sum(
        row["remaining_hours"] for row in report_subject_rows
    )
    total_allocated_hours = total_required_hours - total_remaining_hours
    total_staffing_requirement_blocks = (
        math.ceil(total_remaining_hours / REPORT_STANDARD_MAX_HOURS)
        if total_remaining_hours > 0
        else 0
    )
    total_existing_teachers = len(teachers)
    total_existing_capacity_hours = sum(
        int(profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS))
        for profile in teacher_profiles
    )
    coverage_percentage = (
        round((total_allocated_hours / total_required_hours) * 100)
        if total_required_hours > 0
        else 0
    )
    teachers_with_subject_alignment = sum(
        1
        for profile in teacher_profiles
        if profile["subject_count"] > 0 or profile.get("homeroom_section_labels")
    )
    teachers_utilized = sum(
        1 for profile in teacher_profiles if profile["allocated_hours"] > 0
    )
    teachers_full_load = sum(
        1
        for profile in teacher_profiles
        if profile["allocated_hours"] >= profile.get(
            "capacity_hours",
            REPORT_STANDARD_MAX_HOURS,
        )
    )
    unused_existing_capacity_hours = max(
        total_existing_capacity_hours - total_allocated_hours,
        0,
    )
    total_required_current_hours = sum(
        required_current_hours_by_grade.values()
    )
    total_required_new_hours = sum(
        required_new_hours_by_grade.values()
    )
    total_new_sections_planned = sum(new_sections_by_grade.values())
    largest_gap_row = report_gap_rows[0] if report_gap_rows else None

    report_summary = {
        "total_required_hours": total_required_hours,
        "total_required_current_hours": total_required_current_hours,
        "total_required_new_hours": total_required_new_hours,
        "total_allocated_hours": total_allocated_hours,
        "total_remaining_hours": total_remaining_hours,
        "coverage_percentage": coverage_percentage,
        "total_additional_teachers_needed": total_staffing_requirement_blocks,
        "total_staffing_requirement_blocks": total_staffing_requirement_blocks,
        "total_existing_teachers": total_existing_teachers,
        "total_existing_capacity_hours": total_existing_capacity_hours,
        "unused_existing_capacity_hours": unused_existing_capacity_hours,
        "teachers_with_subject_alignment": teachers_with_subject_alignment,
        "teachers_utilized": teachers_utilized,
        "teachers_full_load": teachers_full_load,
        "teachers_idle": max(total_existing_teachers - teachers_utilized, 0),
        "total_new_sections_planned": total_new_sections_planned,
        "subjects_with_gaps": len(report_gap_rows),
        "homeroom_default_coverage_hours": sum(
            int(profile.get("homeroom_allocated_hours", 0))
            for profile in teacher_profiles
        ),
        "largest_gap_subject_name": (
            largest_gap_row["subject_name"] if largest_gap_row else ""
        ),
        "largest_gap_hours": (
            int(largest_gap_row["remaining_hours"]) if largest_gap_row else 0
        ),
        "largest_gap_teachers_needed": (
            int(largest_gap_row["teacher_requirement_blocks"])
            if largest_gap_row
            else 0
        ),
    }

    teacher_profiles_export = []
    for profile in teacher_profiles:
        teacher = profile["teacher"]
        teacher_profiles_export.append(
            {
                "teacher_pk": teacher.id,
                "teacher_id": teacher.teacher_id or "-",
                "teacher_name": profile["name"],
                "degree_major": str(getattr(teacher, "degree_major", "") or "").strip(),
                "subject_keys": list(profile["subject_keys"]),
                "secondary_subject_keys": list(profile.get("secondary_subject_keys", [])),
                "support_subject_keys": list(profile["support_subject_keys"]),
                "homeroom_subject_keys": list(profile.get("homeroom_subject_keys", [])),
                "homeroom_section_labels": list(
                    profile.get("homeroom_section_labels", [])
                ),
                "homeroom_class_allocations": [
                    dict(item)
                    for item in profile.get("homeroom_class_allocations", [])
                ],
                "homeroom_allocation_breakdown": dict(
                    profile.get("homeroom_allocation_breakdown", {})
                ),
                "allocation_breakdown": dict(profile["allocation_breakdown"]),
                "allocated_hours": int(profile["allocated_hours"]),
                "remaining_capacity_hours": int(profile["remaining_capacity_hours"]),
                "homeroom_allocated_hours": int(
                    profile.get("homeroom_allocated_hours", 0)
                ),
                "capacity_hours": int(
                    profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
                ),
                "total_capacity_hours": int(
                    profile.get("total_capacity_hours", REPORT_STANDARD_MAX_HOURS)
                ),
                "national_section_hours": int(
                    profile.get("national_section_hours", 0)
                ),
                "primary_allocated_hours": int(profile.get("primary_allocated_hours", 0)),
                "support_allocated_hours": int(profile.get("support_allocated_hours", 0)),
                "primary_subject_basis_hours": int(
                    profile.get("primary_subject_basis_hours", 0)
                ),
            }
        )

    return {
        "summary": report_summary,
        "subject_rows": report_subject_rows,
        "gap_rows": report_gap_rows,
        "teacher_rows": report_teacher_rows,
        "grade_rows": report_grade_rows,
        "teacher_profiles": teacher_profiles_export,
    }


def _build_dashboard_report_visuals(
    report_summary,
    report_subject_rows,
    report_grade_rows,
    planning_current_sections_count,
    planning_new_sections_count,
):
    coverage_pct = max(
        0,
        min(100, int(report_summary.get("coverage_percentage", 0) or 0)),
    )
    uncovered_pct = max(0, 100 - coverage_pct)

    total_sections = planning_current_sections_count + planning_new_sections_count
    current_section_pct = (
        round((planning_current_sections_count / total_sections) * 100)
        if total_sections > 0
        else 0
    )
    new_section_pct = max(0, 100 - current_section_pct) if total_sections > 0 else 0

    existing_teachers = int(report_summary.get("total_existing_teachers", 0) or 0)
    new_teachers_required = int(
        report_summary.get("total_new_teachers_required", 0) or 0
    )
    total_teacher_mix = existing_teachers + new_teachers_required
    existing_teacher_pct = (
        round((existing_teachers / total_teacher_mix) * 100)
        if total_teacher_mix > 0
        else 0
    )
    new_teacher_pct = (
        max(0, 100 - existing_teacher_pct)
        if total_teacher_mix > 0
        else 0
    )

    top_subject_gap = [
        {
            "label": row["subject_name"],
            "value": int(row["remaining_hours"]),
            "coverage": int(row["coverage_percentage"]),
        }
        for row in sorted(
            (
                row
                for row in report_subject_rows
                if int(row.get("remaining_hours", 0) or 0) > 0
            ),
            key=lambda row: (-row["remaining_hours"], row["subject_name"]),
        )
    ]
    max_subject_gap = max((item["value"] for item in top_subject_gap), default=0)
    for item in top_subject_gap:
        item["width_pct"] = (
            round((item["value"] / max_subject_gap) * 100, 1)
            if max_subject_gap > 0
            else 0
        )

    grade_mix = [
        {
            "label": row["grade_label"],
            "total_hours": int(row["required_hours_total"]),
            "new_hours": int(row["required_hours_new"]),
        }
        for row in report_grade_rows
        if int(row.get("required_hours_total", 0) or 0) > 0
    ]
    max_grade_hours = max((item["total_hours"] for item in grade_mix), default=0)
    for item in grade_mix:
        item["width_pct"] = (
            round((item["total_hours"] / max_grade_hours) * 100, 1)
            if max_grade_hours > 0
            else 0
        )

    return {
        "coverage_pct": coverage_pct,
        "uncovered_pct": uncovered_pct,
        "current_section_pct": current_section_pct,
        "new_section_pct": new_section_pct,
        "existing_teacher_pct": existing_teacher_pct,
        "new_teacher_pct": new_teacher_pct,
        "top_subject_gap": top_subject_gap,
        "grade_mix": grade_mix,
    }


def _enrich_report_summary_hiring_metrics(report_summary):
    summary = dict(report_summary or {})
    uncovered_hours = int(summary.get("total_remaining_hours", 0) or 0)
    whole_new_hires = uncovered_hours // REPORT_STANDARD_MAX_HOURS
    remaining_uncovered_hours_after_hires = (
        uncovered_hours % REPORT_STANDARD_MAX_HOURS
        if uncovered_hours > 0
        else 0
    )
    total_existing_teachers = int(summary.get("total_existing_teachers", 0) or 0)
    total_teachers_needed_branch = total_existing_teachers + whole_new_hires
    summary["total_uncovered_hours"] = uncovered_hours
    summary["total_new_teachers_required"] = whole_new_hires
    summary["remaining_uncovered_hours_after_hires"] = (
        remaining_uncovered_hours_after_hires
    )
    summary["hireable_covered_hours"] = (
        whole_new_hires * REPORT_STANDARD_MAX_HOURS
    )
    summary["hiring_plan_equivalent_full_teacher_count"] = whole_new_hires
    summary["hiring_plan_equivalent_remaining_hours"] = (
        remaining_uncovered_hours_after_hires
    )
    summary["total_teachers_needed_branch"] = total_teachers_needed_branch
    summary["new_hire_capacity_hours"] = whole_new_hires * REPORT_STANDARD_MAX_HOURS
    summary["staffing_remainder_has_gap"] = remaining_uncovered_hours_after_hires > 0
    return summary


def _detect_hiring_subject_family(subject_row: dict) -> str:
    subject_name = str(subject_row.get("subject_name", "") or "")
    subject_code = str(subject_row.get("subject_code", "") or "")
    subject_key = str(subject_row.get("subject_key", "") or "")
    normalized_text = _normalize_subject_family_key(
        f"{subject_code} {subject_name} {subject_key}"
    )
    alignment_groups = set(
        get_subject_alignment_group_keys(subject_name, subject_code)
    )

    if re.search(r"\b(qur|quran|qur an|qno|qaad|qaadah|noraniah|noorani)\b", normalized_text):
        return "quran"
    if re.search(r"\b(home\s*room|homeroom)\b", normalized_text):
        return "homeroom"
    if re.search(r"\b(reflection|reflective|advisory|character education)\b", normalized_text):
        return "reflection"
    if re.search(r"\b(social studies english|social english|sse)\b", normalized_text):
        return "social_english"
    if re.search(
        r"\b(social studies arabic|social arabic|social studies ksa|ksa social|social studies saudi|saudi social|ssa)\b",
        normalized_text,
    ):
        return "social_arabic"
    if re.search(r"\b(social studies english|social english|sse|social studies|social|humanities|civics)\b", normalized_text):
        return "social_english"
    if re.search(r"\b(english|ela|phonics|reading|writing|literacy|language arts)\b", normalized_text):
        return "english"
    if re.search(r"\b(arabic|arbic|lang ar|arab)\b", normalized_text):
        return "arabic"
    if re.search(r"\b(math|maths|mathematics|algebra|geometry|calculus)\b", normalized_text):
        return "math"
    if re.search(r"\b(physics|physical science|phy)\b", normalized_text):
        return "physics"
    if (
        re.search(r"\b(science|sciences|general science|general sciences|integrated science|integrated sciences|steam)\b|\b(?:sci|sce)(?:\b|\d)", normalized_text)
        and not re.search(
            r"\b(computer science|computer sciences|life science|life sciences|chemical science|chemical sciences|physical science|physical sciences)\b",
            normalized_text,
        )
    ):
        return "science"
    if re.search(r"\b(biology|life science|life sciences)\b|\bbio(?:\b|\d)", normalized_text):
        return "biology"
    if re.search(r"\b(chemistry|chemical science|chemical sciences)\b|\bchem(?:\b|\d)", normalized_text):
        return "chemistry"
    if re.search(r"\b(ict|information communication technology|computer|computing|technology|coding|robotics)\b|\bcs(?:\b|\d)", normalized_text):
        return "ict"
    if re.search(r"\b(pe|physical education|sport|fitness)\b", normalized_text):
        return "pe"
    if re.search(r"\b(well being|wellbeing|health|sel|life skills)\b", normalized_text):
        return "wellbeing"
    if re.search(r"\b(performing|performance|drama|theatre|theater|dance|music)\b", normalized_text):
        return "performing_arts"
    if re.search(r"\b(art|drawing|painting|visual art)\b", normalized_text):
        return "art"

    if "islamic" in alignment_groups or re.search(
        r"\b(islamic|hadith|fiqh|tawheed|religion)\b",
        normalized_text,
    ):
        return "islamic"
    if re.search(r"\b(mental math|mental|abacus|mmt)\b", normalized_text):
        return "mental_math"
    if "math" in alignment_groups:
        return "math"
    if "physics" in alignment_groups:
        return "physics"
    if "biology" in alignment_groups:
        return "biology"
    if "chemistry" in alignment_groups:
        return "chemistry"
    if "science" in alignment_groups:
        return "science"
    if "computer" in alignment_groups:
        return "ict"
    if "arabic" in alignment_groups:
        return "arabic"
    if "english" in alignment_groups:
        return "english"
    if alignment_groups.intersection({"social", "history", "geography"}):
        if re.search(r"\b(ar|arabic|ksa|saudi)\b", normalized_text):
            return "social_arabic"
        return "social_english"
    if "pe" in alignment_groups:
        return "pe"
    if "music" in alignment_groups or re.search(
        r"\b(performing|performance|drama|theatre|theater|dance|music)\b",
        normalized_text,
    ):
        return "performing_arts"
    if "art" in alignment_groups:
        return "art"
    return "other"


def _get_hiring_subject_sort_key(item: dict):
    return (
        HIRING_FAMILY_PRIORITY.get(item.get("family"), 99),
        -int(item.get("remaining_hours", 0) or 0),
        str(item.get("subject_name", "") or "").lower(),
    )


def _build_hiring_coverage_label(coverage_items: list[dict]) -> str:
    return " + ".join(
        f"{int(item.get('hours', 0) or 0)}h {item.get('subject_name', 'Subject')}"
        for item in coverage_items
        if int(item.get("hours", 0) or 0) > 0
    )


def _build_hiring_profile_label(
    coverage_items: list[dict],
    *,
    group_key: str = "",
    dedicated: bool = False,
) -> str:
    if group_key in HIRING_PROFILE_GROUP_LABEL_KEYS and group_key in HIRING_GROUP_LABELS:
        return HIRING_GROUP_LABELS[group_key]

    if dedicated and len(coverage_items) == 1:
        return f"{coverage_items[0].get('subject_name', 'Subject')} Pool"

    family_labels = []
    for item in coverage_items:
        family = item.get("family", "")
        label = HIRING_FAMILY_LABELS.get(family)
        if label and label not in family_labels:
            family_labels.append(label)

    if len(family_labels) == 1:
        return f"{family_labels[0]} Pool"
    if group_key and group_key in HIRING_GROUP_LABELS:
        return HIRING_GROUP_LABELS[group_key]
    if family_labels:
        return f"{' / '.join(family_labels[:3])} Pool"
    return "Specialist Pool"


def _build_hiring_profile_reason(
    profile_type: str,
    coverage_items: list[dict],
    total_hours: int,
    *,
    group_key: str = "",
    teacher_count: int = 1,
) -> str:
    if profile_type == "dedicated":
        subject_name = coverage_items[0].get("subject_name", "this subject")
        return (
            f"{total_hours}h in {subject_name} creates "
            f"{teacher_count} full 24h specialist load"
            f"{'s' if teacher_count != 1 else ''}, so dedicated hiring is recommended."
        )

    coverage_label = _build_hiring_coverage_label(coverage_items)
    group_label = HIRING_GROUP_LABELS.get(group_key, "the same specialization")
    if total_hours >= REPORT_STANDARD_MAX_HOURS:
        return (
            f"{coverage_label} forms one full 24h load. These subjects are grouped "
            f"because they belong to {group_label} compatibility."
        )

    if len(coverage_items) > 1:
        return (
            f"{coverage_label} is compatible under {group_label}, but totals "
            f"{total_hours}h, below a full 24h teacher load."
        )

    subject_name = coverage_items[0].get("subject_name", "this subject")
    return (
        f"{subject_name} has {total_hours}h uncovered and no compatible uncovered "
        "subject remains to complete a full 24h load."
    )


def _normalize_hiring_pool_group_key(group_key: str = "", family: str = "") -> str:
    normalized_group = re.sub(
        r"[^a-z0-9]+",
        "_",
        str(group_key or "").strip().lower(),
    ).strip("_")
    normalized_family = re.sub(
        r"[^a-z0-9]+",
        "_",
        str(family or "").strip().lower(),
    ).strip("_")
    if normalized_group in {"math", "math_pool", "single_math", "single_mental_math", "single_physics"}:
        return "math_pool"
    if normalized_group in {
        "biology",
        "biology_pool",
        "chemistry",
        "chemistry_pool",
        "ict",
        "ict_pool",
        "science",
        "science_pool",
        "science_teacher",
        "general_science_pool",
        "general_science",
        "general_science_teacher",
        "general_science_related",
        "general_science_related_pool",
        "general_science_related_subjects",
        "science_related",
        "science_related_pool",
        "science_related_subjects",
        "single_science",
        "single_biology",
        "single_chemistry",
        "single_ict",
    }:
        return "general_science_pool"
    if normalized_group in {"english", "english_humanities", "english_remainder", "single_english", "single_social_english", "single_social"}:
        return "english_pool"
    if normalized_group in {"arabic", "arabic_related", "single_arabic", "single_islamic", "single_quran", "single_social_arabic"}:
        return "arabic_pool"
    if normalized_group in {"pe", "student_life", "single_pe"}:
        return "physical_education"
    if normalized_group in {"homeroom", "home_room", "homeroom_pool", "homeroom_teacher", "single_homeroom"}:
        return "homeroom_pool"

    if normalized_family in {"math", "mental_math", "physics"}:
        return "math_pool"
    if normalized_family in HIRING_GENERAL_SCIENCE_FAMILIES:
        return "general_science_pool"
    if normalized_family in {"english", "social_english", "social", "wellbeing", "reflection", "performing_arts", "art"}:
        return "english_pool"
    if normalized_family in {"arabic", "islamic", "quran", "social_arabic"}:
        return "arabic_pool"
    if normalized_family == "pe":
        return "physical_education"
    if normalized_family == "homeroom":
        return "homeroom_pool"

    return normalized_group


def _recalculate_hiring_editor_profile_capacity(profile: dict) -> dict:
    total_hours = sum(int(item.get("hours", 0) or 0) for item in profile.get("items", []) or [])
    block_size_hours = max(
        1,
        int(profile.get("block_size_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS),
    )
    recommended_capacity = (
        max(
            block_size_hours,
            ((total_hours + block_size_hours - 1) // block_size_hours) * block_size_hours,
        )
        if total_hours > 0
        else block_size_hours
    )
    profile["block_size_hours"] = block_size_hours
    profile["max_hours"] = max(
        int(profile.get("max_hours", block_size_hours) or block_size_hours),
        recommended_capacity,
    )
    return profile


def _sort_hiring_editor_profile_items(profile: dict) -> dict:
    group_key = _normalize_hiring_pool_group_key(str(profile.get("group_key", "") or ""))
    family_order = HIRING_POOL_FAMILY_ORDER.get(group_key, [])
    if not family_order:
        return profile

    family_order_map = {family: index for index, family in enumerate(family_order)}
    profile["items"] = sorted(
        profile.get("items", []) or [],
        key=lambda item: (
            family_order_map.get(
                str(item.get("family", "") or "").strip().lower(),
                len(family_order_map) + 1,
            ),
            str(item.get("subject_name", "") or "").strip().lower(),
            str(item.get("subject_code", "") or "").strip().upper(),
        ),
    )
    return profile


def _apply_general_science_editor_rule(
    profiles: list[dict],
    unassigned_items: list[dict],
) -> tuple[list[dict], list[dict]]:
    normalized_profiles: list[dict] = []
    general_science_profile: dict | None = None

    def ensure_general_science_profile() -> dict:
        nonlocal general_science_profile
        if general_science_profile is None:
            general_science_profile = {
                "id": "profile-general-science-pool",
                "name": HIRING_GROUP_LABELS["general_science_pool"],
                "group_key": "general_science_pool",
                "assignment_note": "Science first, with Biology, Chemistry, and ICT merged into one pool",
                "accent_color": HIRING_POOL_ACCENT_COLORS["general_science_pool"],
                "max_hours": REPORT_STANDARD_MAX_HOURS,
                "block_size_hours": REPORT_STANDARD_MAX_HOURS,
                "is_manual": False,
                "allow_over_capacity": False,
                "override_compatibility": False,
                "items": [],
            }
            normalized_profiles.insert(0, general_science_profile)
        return general_science_profile

    for profile in profiles:
        if profile.get("group_key") == "general_science_pool":
            general_science_profile = ensure_general_science_profile()
            general_science_profile["id"] = str(profile.get("id", "") or general_science_profile["id"])
            general_science_profile["name"] = HIRING_GROUP_LABELS["general_science_pool"]
            general_science_profile["accent_color"] = HIRING_POOL_ACCENT_COLORS["general_science_pool"]
            general_science_profile["block_size_hours"] = int(profile.get("block_size_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS)
            general_science_profile["max_hours"] = int(profile.get("max_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS)
            general_science_profile["allow_over_capacity"] = bool(profile.get("allow_over_capacity", False))
            general_science_profile["override_compatibility"] = bool(profile.get("override_compatibility", False))
            general_science_profile["is_manual"] = False
            general_science_profile["items"].extend(profile.get("items", []) or [])
        else:
            normalized_profiles.append(profile)

    retained_unassigned_items: list[dict] = []
    for item in unassigned_items:
        family = str(item.get("family", "") or "").strip().lower()
        retained_unassigned_items.append(
            {
                **item,
                "family": family,
            }
        )

    seen_named_keys: dict[str, dict] = {}
    deduped_profiles: list[dict] = []
    for profile in normalized_profiles:
        gk = profile.get("group_key", "")
        if gk and gk in HIRING_NAMED_POOL_KEYS:
            if gk in seen_named_keys:
                existing = seen_named_keys[gk]
                existing["items"].extend(profile.get("items", []))
                existing["max_hours"] = max(
                    int(existing.get("max_hours", REPORT_STANDARD_MAX_HOURS)),
                    int(profile.get("max_hours", REPORT_STANDARD_MAX_HOURS)),
                )
            else:
                if gk == "general_science_pool":
                    profile["name"] = HIRING_GROUP_LABELS["general_science_pool"]
                    profile["is_manual"] = False
                seen_named_keys[gk] = profile
                deduped_profiles.append(profile)
        else:
            deduped_profiles.append(profile)

    return (
        [
            _recalculate_hiring_editor_profile_capacity(
                _sort_hiring_editor_profile_items(profile)
            )
            for profile in deduped_profiles
        ],
        retained_unassigned_items,
    )


def _build_hiring_pool_reason(
    group_key: str,
    coverage_items: list[dict],
    total_hours: int,
    full_teacher_count: int,
    remaining_hours: int,
) -> str:
    group_label = HIRING_GROUP_LABELS.get(group_key, "this specialization")
    if group_key == "arabic_pool":
        base_reason = (
            "Arabic Pool keeps Arabic first and groups Islamic, Quran, Qaadah Nooraniah, and Social Studies Arabic "
            "as compatible Arabic/identity-related coverage."
        )
    elif group_key == "english_pool":
        base_reason = (
            "English Pool keeps English first and groups Social Studies English, Well Being, Reflection, "
            "Performing Arts, and Art directly in the same pool."
        )
    elif group_key == "math_pool":
        base_reason = (
            "Math Pool groups Mathematics / Math with Physics as compatible coverage."
        )
    elif group_key == "general_science_pool":
        base_reason = (
            "General Science Pool is opened whenever Science has uncovered hours and groups Science first, "
            "then Biology, Chemistry, and ICT in the same recommended pool."
        )
    elif group_key == "physical_education":
        base_reason = (
            "Physical Education is kept in its own pool and is not merged with unrelated subjects."
        )
    elif group_key == "homeroom_pool":
        base_reason = (
            "Homeroom Pool groups uncovered homeroom-owned load into its own recommendation so branches with"
            " homeroom teaching demand do not fall back into Unassigned."
        )
    else:
        base_reason = f"These subjects are grouped under the {group_label} specialization."

    if full_teacher_count > 0:
        return (
            f"{base_reason} This pool contains {total_hours}h uncovered, which converts to "
            f"{full_teacher_count} full 24h teacher block{'s' if full_teacher_count != 1 else ''}"
            f"{f' with {remaining_hours}h still remaining in the pool.' if remaining_hours > 0 else '.'}"
        )

    return (
        f"{base_reason} This pool currently contains {total_hours}h uncovered, which is below one full 24h teacher block"
        f" and leaves {remaining_hours}h waiting for either compatible absorption or the next full hire threshold."
    )


def _build_hiring_coverage_recommendation(report_subject_rows: list[dict]) -> dict:
    uncovered_items = []
    for row in report_subject_rows or []:
        remaining_hours = int(row.get("remaining_hours", 0) or 0)
        if remaining_hours <= 0:
            continue
        family = _detect_hiring_subject_family(row)
        group_key = HIRING_COMPATIBILITY_GROUPS.get(family, f"single_{family}")
        uncovered_items.append(
            {
                "subject_key": row.get("subject_key", ""),
                "subject_name": row.get("subject_name", "Subject"),
                "subject_code": row.get("subject_code", ""),
                "subject_color": row.get("subject_color", "#0A4EA3"),
                "remaining_hours": remaining_hours,
                "family": family,
                "group_key": group_key,
                "priority_staffing_subject": bool(row.get("priority_staffing_subject")),
            }
        )

    uncovered_items.sort(key=_get_hiring_subject_sort_key)
    total_uncovered_hours = sum(int(item.get("remaining_hours", 0) or 0) for item in uncovered_items)
    profiles = []
    profile_counter = 1
    family_buckets = {}
    for item in uncovered_items:
        family_buckets.setdefault(item.get("family", "other"), []).append(dict(item))

    for family_items in family_buckets.values():
        family_items.sort(key=_get_hiring_subject_sort_key)

    def consume_family_hours(family: str, limit: int | None = None) -> list[dict]:
        remaining_limit = None if limit is None else max(int(limit), 0)
        consumed = []
        for item in family_buckets.get(family, []):
            open_hours = int(item.get("remaining_hours", 0) or 0)
            if open_hours <= 0:
                continue
            if remaining_limit is not None and remaining_limit <= 0:
                break
            take_hours = open_hours if remaining_limit is None else min(open_hours, remaining_limit)
            if take_hours <= 0:
                continue
            item["remaining_hours"] = open_hours - take_hours
            if remaining_limit is not None:
                remaining_limit -= take_hours
            consumed.append(
                {
                    **item,
                    "hours": take_hours,
                }
            )
        return consumed

    def consume_families_in_order(families: list[str], limit: int | None = None) -> list[dict]:
        consumed = []
        remaining_limit = None if limit is None else max(int(limit), 0)
        for family in families:
            if remaining_limit is not None and remaining_limit <= 0:
                break
            family_consumed = consume_family_hours(family, remaining_limit)
            consumed.extend(family_consumed)
            if remaining_limit is not None:
                remaining_limit -= sum(
                    int(item.get("hours", 0) or 0)
                    for item in family_consumed
                )
        return consumed

    def build_pool_profile(group_key: str, pool_items: list[dict]):
        nonlocal profile_counter
        if not pool_items:
            return None

        total_hours = sum(
            int(item.get("hours", item.get("remaining_hours", 0)) or 0)
            for item in pool_items
        )
        if total_hours <= 0:
            return None
        group_key = _normalize_hiring_pool_group_key(group_key)

        full_teacher_count = total_hours // REPORT_STANDARD_MAX_HOURS
        remaining_hours = total_hours % REPORT_STANDARD_MAX_HOURS
        unique_families = []
        for item in pool_items:
            family = item.get("family", "")
            if family and family not in unique_families:
                unique_families.append(family)

        coverage_items = [
            {
                **item,
                "hours": int(item.get("hours", item.get("remaining_hours", 0)) or 0),
            }
            for item in pool_items
        ]
        capacity_to_next_block = (
            REPORT_STANDARD_MAX_HOURS - remaining_hours
            if remaining_hours > 0
            else 0
        )
        recommended_capacity_hours = (
            full_teacher_count * REPORT_STANDARD_MAX_HOURS
            if remaining_hours == 0
            else (full_teacher_count + 1) * REPORT_STANDARD_MAX_HOURS
        )
        recommended_capacity_hours = max(
            REPORT_STANDARD_MAX_HOURS,
            recommended_capacity_hours,
        )
        progress_width_pct = (
            round((remaining_hours / REPORT_STANDARD_MAX_HOURS) * 100, 1)
            if remaining_hours > 0
            else 100.0
        )
        is_multi_subject_pool = len(unique_families) > 1 or len(coverage_items) > 1

        profile = {
            "id": f"hire-plan-{profile_counter}",
            "group_key": group_key,
            "profile_label": _build_hiring_profile_label(
                coverage_items,
                group_key=group_key,
                dedicated=not is_multi_subject_pool,
            ),
            "teacher_count": full_teacher_count,
            "full_teacher_count": full_teacher_count,
            "profile_type": "combined_pool" if is_multi_subject_pool else "specialist_pool",
            "status_label": (
                "Recommended subject pool"
                if group_key in HIRING_NAMED_POOL_KEYS
                else (
                    "Combined specialization pool"
                    if is_multi_subject_pool
                    else "Specialist priority pool"
                )
            ),
            "total_hours": total_hours,
            "group_total_hours": total_hours,
            "capacity_hours": recommended_capacity_hours,
            "block_size_hours": REPORT_STANDARD_MAX_HOURS,
            "remaining_hours": remaining_hours,
            "remaining_capacity_hours": capacity_to_next_block,
            "hours_to_next_block": capacity_to_next_block,
            "is_full_load": full_teacher_count > 0,
            "coverage_items": coverage_items,
            "coverage_label": _build_hiring_coverage_label(coverage_items),
            "reason": _build_hiring_pool_reason(
                group_key,
                coverage_items,
                total_hours,
                full_teacher_count,
                remaining_hours,
            ),
            "accent_color": HIRING_POOL_ACCENT_COLORS.get(
                group_key,
                coverage_items[0].get("subject_color", "#0A4EA3"),
            ),
            "progress_width_pct": progress_width_pct,
        }

        coverage_families = {item.get("family", "") for item in coverage_items}
        if group_key == "english_pool":
            if coverage_families == {"english"}:
                profile["assignment_note"] = "English-first pool"
            else:
                profile["assignment_note"] = "English Pool compatible subjects grouped together"
        elif group_key == "math_pool":
            if "physics" in coverage_families:
                profile["assignment_note"] = "Physics grouped with Math"
            else:
                profile["assignment_note"] = "Math priority pool"
        elif group_key == "general_science_pool":
            profile["assignment_note"] = "Science first, with Biology, Chemistry, and ICT merged into one pool"
        elif group_key == "arabic_pool":
            profile["assignment_note"] = "Arabic / identity-related pool"
        elif group_key == "physical_education":
            profile["assignment_note"] = "Separate Physical Education pool"
        elif group_key == "homeroom_pool":
            profile["assignment_note"] = "Homeroom-owned load grouped into its own pool"
        else:
            profile["assignment_note"] = "Compatible pool grouping"

        profile_counter += 1
        return profile

    pool_definitions = [
        (
            "english_pool",
            ["english", "social_english", "social", "wellbeing", "reflection", "performing_arts", "art"],
        ),
        (
            "general_science_pool",
            ["science", "biology", "chemistry", "ict"],
        ),
        (
            "arabic_pool",
            ["arabic", "islamic", "quran", "social_arabic"],
        ),
        (
            "math_pool",
            ["math", "mental_math", "physics"],
        ),
        (
            "physical_education",
            ["pe"],
        ),
        (
            "homeroom_pool",
            ["homeroom"],
        ),
    ]

    for group_key, families in pool_definitions:
        pool_items = consume_families_in_order(families)
        profile = build_pool_profile(group_key, pool_items)
        if profile:
            profiles.append(profile)

    remaining_other_items = []
    for family, items in family_buckets.items():
        for item in items:
            open_hours = int(item.get("remaining_hours", 0) or 0)
            if open_hours <= 0:
                continue
            remaining_other_items.append(
                {
                    **item,
                    "hours": open_hours,
                }
            )
            item["remaining_hours"] = 0

    if remaining_other_items:
        remaining_other_items.sort(key=_get_hiring_subject_sort_key)
        # Group remaining items: named-pool leftovers get merged into existing profiles.
        # Truly unrecognized items remain unallocated and will show in editor unassigned.
        named_pool_leftovers: dict[str, list[dict]] = {}
        truly_other_items: list[dict] = []
        for item in remaining_other_items:
            resolved_gk = _normalize_hiring_pool_group_key(
                item.get("group_key", f"single_{item.get('family', 'other')}"),
                item.get("family", ""),
            )
            if resolved_gk in HIRING_NAMED_POOL_KEYS:
                named_pool_leftovers.setdefault(resolved_gk, []).append(item)
            else:
                truly_other_items.append(item)

        # Merge named-pool leftovers into their existing profile if one exists,
        # otherwise build a new profile for that pool.
        for pool_gk, leftover_items in named_pool_leftovers.items():
            existing_profile = next(
                (p for p in profiles if p.get("group_key") == pool_gk), None
            )
            if existing_profile is not None:
                for item in leftover_items:
                    item_hours = int(item.get("hours", item.get("remaining_hours", 0)) or 0)
                    if item_hours <= 0:
                        continue
                    existing_profile["coverage_items"].append({
                        **item,
                        "hours": item_hours,
                    })
                    existing_profile["total_hours"] = existing_profile.get("total_hours", 0) + item_hours
                    existing_profile["group_total_hours"] = existing_profile["total_hours"]
                    family = item.get("family", "")
                    if family and family not in existing_profile.get("unique_families", []):
                        existing_profile.setdefault("unique_families", []).append(family)
                # Recalculate derived fields
                total_hours = existing_profile["total_hours"]
                full_teacher_count = total_hours // REPORT_STANDARD_MAX_HOURS
                remaining_hours = total_hours % REPORT_STANDARD_MAX_HOURS
                existing_profile["full_teacher_count"] = full_teacher_count
                existing_profile["teacher_count"] = full_teacher_count
                existing_profile["remaining_hours"] = remaining_hours
                existing_profile["is_full_load"] = full_teacher_count > 0
                existing_profile["capacity_hours"] = max(
                    REPORT_STANDARD_MAX_HOURS,
                    (full_teacher_count * REPORT_STANDARD_MAX_HOURS
                     if remaining_hours == 0
                     else (full_teacher_count + 1) * REPORT_STANDARD_MAX_HOURS),
                )
                existing_profile["remaining_capacity_hours"] = (
                    REPORT_STANDARD_MAX_HOURS - remaining_hours if remaining_hours > 0 else 0
                )
                existing_profile["coverage_label"] = _build_hiring_coverage_label(
                    existing_profile["coverage_items"]
                )
                coverage_families_now = {ci.get("family", "") for ci in existing_profile["coverage_items"]}
            else:
                profile = build_pool_profile(pool_gk, leftover_items)
                if profile:
                    profiles.append(profile)

        if truly_other_items:
            logging.info(
                "Leaving %s item(s) (%sh) unassigned because no compatible named pool exists.",
                len(truly_other_items),
                sum(int(item.get("hours", 0) or 0) for item in truly_other_items),
            )

    full_teacher_count = sum(int(profile.get("full_teacher_count", 0) or 0) for profile in profiles)
    partial_profile_count = sum(1 for profile in profiles if int(profile.get("remaining_hours", 0) or 0) > 0)
    covered_hours = sum(int(profile.get("total_hours", 0) or 0) for profile in profiles)
    remaining_capacity_hours = sum(
        int(profile.get("remaining_capacity_hours", 0) or 0)
        for profile in profiles
        if int(profile.get("remaining_hours", 0) or 0) > 0
    )
    equivalent_full_teacher_count = covered_hours // REPORT_STANDARD_MAX_HOURS
    equivalent_remaining_hours = covered_hours % REPORT_STANDARD_MAX_HOURS
    unallocated_hours = max(total_uncovered_hours - covered_hours, 0)

    pool_debug_parts = [
        (
            f"{profile.get('profile_label', 'Profile')}:"
            f"{int(profile.get('total_hours', 0) or 0)}h="
            f"{int(profile.get('full_teacher_count', 0) or 0)}x24+"
            f"{int(profile.get('remaining_hours', 0) or 0)}"
        )
        for profile in profiles
    ]
    logging.info(
        "Hiring plan debug: pools=%s total_uncovered=%s covered_in_pools=%s equivalent=%sx24+%s details=%s",
        len(profiles),
        total_uncovered_hours,
        covered_hours,
        equivalent_full_teacher_count,
        equivalent_remaining_hours,
        " | ".join(pool_debug_parts),
    )
    if unallocated_hours > 0:
        logging.warning(
            "Hiring plan has %sh uncovered but not assigned to a visible pool.",
            unallocated_hours,
        )

    return {
        "hiring_plan_profiles": profiles,
        "hiring_plan_profile_count": len(profiles),
        "hiring_plan_full_teacher_count": full_teacher_count,
        "hiring_plan_partial_profile_count": partial_profile_count,
        "hiring_plan_covered_hours": covered_hours,
        "hiring_plan_unallocated_hours": unallocated_hours,
        "hiring_plan_remaining_capacity_hours": remaining_capacity_hours,
        "hiring_plan_equivalent_full_teacher_count": equivalent_full_teacher_count,
        "hiring_plan_equivalent_remaining_hours": equivalent_remaining_hours,
    }


def _build_hiring_plan_editor_payload(report_summary: dict, report_subject_rows: list[dict]) -> dict:
    profiles_source = list((report_summary or {}).get("hiring_plan_profiles", []) or [])
    subject_remaining: dict[str, dict[str, Any]] = {}

    for row in report_subject_rows or []:
        remaining_hours = int(row.get("remaining_hours", 0) or 0)
        if remaining_hours <= 0:
            continue
        subject_key = str(row.get("subject_key", "") or "")
        if not subject_key:
            continue
        subject_remaining[subject_key] = {
            "subject_key": subject_key,
            "subject_name": str(row.get("subject_name", "Subject") or "Subject"),
            "subject_code": str(row.get("subject_code", "") or ""),
            "family": _detect_hiring_subject_family(row),
            "subject_color": str(row.get("subject_color", "#0A4EA3") or "#0A4EA3"),
            "hours": remaining_hours,
        }

    profile_items = []
    for profile in profiles_source:
        items = []
        for item in profile.get("coverage_items", []) or []:
            subject_key = str(item.get("subject_key", "") or "")
            item_hours = int(item.get("hours", 0) or 0)
            if item_hours <= 0:
                continue
            subject_name = str(item.get("subject_name", "Subject") or "Subject")
            if subject_key and subject_key in subject_remaining:
                available = int(subject_remaining[subject_key]["hours"])
                take = min(available, item_hours)
                if take <= 0:
                    continue
                subject_remaining[subject_key]["hours"] = max(available - take, 0)
                item_hours = take

            item_family = str(item.get("family", "") or "")
            if not item_family:
                item_family = _detect_hiring_subject_family(item)

            items.append(
                {
                    "id": f"chip-{len(items)+1}-{subject_key or subject_name.lower().replace(' ', '-')}",
                    "subject_key": subject_key,
                    "subject_name": subject_name,
                    "subject_code": str(item.get("subject_code", "") or ""),
                    "family": item_family,
                    "subject_color": str(item.get("subject_color", profile.get("accent_color", "#0A4EA3")) or "#0A4EA3"),
                    "hours": item_hours,
                }
            )

        profile_group_key = _normalize_hiring_pool_group_key(str(profile.get("group_key", "") or ""))
        profile_items.append(
            {
                "id": str(profile.get("id", f"profile-{len(profile_items)+1}")),
                "name": HIRING_GROUP_LABELS.get(
                    profile_group_key,
                    str(profile.get("profile_label", "Proposed Pool") or "Proposed Pool"),
                ),
                "group_key": profile_group_key,
                "assignment_note": str(profile.get("assignment_note", "Compatible pool grouping") or "Compatible pool grouping"),
                "accent_color": str(
                    HIRING_POOL_ACCENT_COLORS.get(
                        profile_group_key,
                        profile.get("accent_color", "#0A4EA3"),
                    )
                    or "#0A4EA3"
                ),
                "max_hours": int(profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS),
                "block_size_hours": int(profile.get("block_size_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS),
                "is_manual": False,
                "allow_over_capacity": False,
                "override_compatibility": False,
                "items": items,
            }
        )

    unassigned_items = []
    for remaining in subject_remaining.values():
        remaining_hours = int(remaining.get("hours", 0) or 0)
        if remaining_hours <= 0:
            continue
        unassigned_items.append(
            {
                "id": f"chip-unassigned-{len(unassigned_items)+1}-{remaining['subject_key']}",
                "subject_key": remaining["subject_key"],
                "subject_name": remaining["subject_name"],
                "subject_code": remaining["subject_code"],
                "family": remaining["family"],
                "subject_color": remaining["subject_color"],
                "hours": remaining_hours,
            }
        )

    return {
        "pool_logic_version": HIRING_PLAN_POOL_LOGIC_VERSION,
        "profiles": profile_items,
        "unassigned_items": unassigned_items,
        "locked": True,
        "summary": {
            "total_uncovered_hours": int((report_summary or {}).get("total_uncovered_hours", 0) or 0),
            "total_new_teachers_required": int((report_summary or {}).get("total_new_teachers_required", 0) or 0),
            "remaining_uncovered_hours_after_hires": int((report_summary or {}).get("remaining_uncovered_hours_after_hires", 0) or 0),
        },
    }


def _normalize_hiring_plan_payload(raw_payload: dict) -> dict:
    payload = dict(raw_payload or {})
    profiles = []
    for raw_profile in payload.get("profiles", []) or []:
        items = []
        is_manual_profile = bool(raw_profile.get("is_manual", False))
        for raw_item in raw_profile.get("items", []) or []:
            hours = int(raw_item.get("hours", 0) or 0)
            if hours <= 0:
                continue
            family = str(raw_item.get("family", "") or "").strip().lower()
            if not family or family == "other":
                family = _detect_hiring_subject_family(raw_item)
            items.append(
                {
                    "id": str(raw_item.get("id", "") or ""),
                    "subject_key": str(raw_item.get("subject_key", "") or ""),
                    "subject_name": str(raw_item.get("subject_name", "Subject") or "Subject"),
                    "subject_code": str(raw_item.get("subject_code", "") or ""),
                    "family": family,
                    "subject_color": str(raw_item.get("subject_color", "#0A4EA3") or "#0A4EA3"),
                    "hours": hours,
                }
            )

        raw_profile_group_key = str(raw_profile.get("group_key", "") or "")
        profile_group_key = _normalize_hiring_pool_group_key(raw_profile_group_key)
        if not is_manual_profile and not raw_profile_group_key.strip():
            profile_name_group_key = _normalize_hiring_pool_group_key(
                str(raw_profile.get("name", "") or "")
            )
            if profile_name_group_key == "general_science_pool":
                profile_group_key = profile_name_group_key
        if not is_manual_profile and profile_group_key not in HIRING_NAMED_POOL_KEYS and items:
            item_group_keys = {
                _normalize_hiring_pool_group_key(family=str(item.get("family", "") or ""))
                for item in items
            }
            if item_group_keys == {"general_science_pool"}:
                profile_group_key = "general_science_pool"
        profiles.append(
            {
                "id": str(raw_profile.get("id", "") or ""),
                "name": HIRING_GROUP_LABELS.get(
                    profile_group_key,
                    str(raw_profile.get("name", "Proposed Pool") or "Proposed Pool"),
                ),
                "group_key": profile_group_key,
                "assignment_note": str(raw_profile.get("assignment_note", "") or ""),
                "accent_color": str(
                    HIRING_POOL_ACCENT_COLORS.get(
                        profile_group_key,
                        raw_profile.get("accent_color", "#0A4EA3"),
                    )
                    or "#0A4EA3"
                ),
                "max_hours": int(raw_profile.get("max_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS),
                "block_size_hours": int(raw_profile.get("block_size_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS),
                "is_manual": is_manual_profile,
                "allow_over_capacity": bool(raw_profile.get("allow_over_capacity", False)),
                "override_compatibility": bool(raw_profile.get("override_compatibility", False)),
                "items": items,
            }
        )

    unassigned_items = []
    for raw_item in payload.get("unassigned_items", []) or []:
        hours = int(raw_item.get("hours", 0) or 0)
        if hours <= 0:
            continue
        family = str(raw_item.get("family", "") or "").strip().lower()
        if not family or family == "other":
            family = _detect_hiring_subject_family(raw_item)
        unassigned_items.append(
            {
                "id": str(raw_item.get("id", "") or ""),
                "subject_key": str(raw_item.get("subject_key", "") or ""),
                "subject_name": str(raw_item.get("subject_name", "Subject") or "Subject"),
                "subject_code": str(raw_item.get("subject_code", "") or ""),
                "family": family,
                "subject_color": str(raw_item.get("subject_color", "#0A4EA3") or "#0A4EA3"),
                "hours": hours,
            }
        )

    profiles, unassigned_items = _apply_general_science_editor_rule(
        profiles,
        unassigned_items,
    )

    return {
        "pool_logic_version": HIRING_PLAN_POOL_LOGIC_VERSION,
        "profiles": profiles,
        "unassigned_items": unassigned_items,
        "locked": bool(payload["locked"]) if "locked" in payload else True,
        "summary": dict(payload.get("summary", {}) or {}),
    }


def _collect_hiring_plan_warnings(plan_payload: dict) -> list[str]:
    warnings = []
    for profile in plan_payload.get("profiles", []) or []:
        profile_name = str(profile.get("name", "Proposed profile") or "Proposed profile")
        max_hours = int(profile.get("max_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS)
        profile_hours = sum(int(item.get("hours", 0) or 0) for item in profile.get("items", []) or [])

        if profile_hours > max_hours and not profile.get("allow_over_capacity", False):
            warnings.append(
                f"{profile_name} is over capacity ({profile_hours}h / {max_hours}h). Enable override to allow this."
            )

        profile_group_key = _normalize_hiring_pool_group_key(str(profile.get("group_key", "") or ""))
        allowed_families = HIRING_POOL_ALLOWED_FAMILIES.get(profile_group_key)
        incompatible_families = []
        group_keys = set()
        for item in profile.get("items", []) or []:
            family = str(item.get("family", "") or "")
            if not family:
                family = _detect_hiring_subject_family(item)
            if allowed_families is not None:
                if family not in allowed_families:
                    incompatible_families.append(HIRING_FAMILY_LABELS.get(family, family or "Subject"))
            else:
                group_keys.add(HIRING_COMPATIBILITY_GROUPS.get(family, f"single_{family}"))

        if incompatible_families and not profile.get("override_compatibility", False):
            warnings.append(
                f"{profile_name} includes incompatible subject families ({', '.join(sorted(set(incompatible_families)))})."
            )

        mixed_groups = {key for key in group_keys if key}
        if allowed_families is None and len(mixed_groups) > 1 and not profile.get("override_compatibility", False):
            warnings.append(
                f"{profile_name} mixes normally incompatible subject groups ({', '.join(sorted(mixed_groups))})."
            )

    return warnings


def _build_report_class_rows(planning_sections):
    class_rows = []
    seen_class_keys = set()

    for section in planning_sections:
        grade_label = _normalize_grade_label(section.grade_level)
        if not grade_label:
            continue

        section_name = str(section.section_name or "").strip().upper()
        if not section_name:
            continue

        class_key = f"{grade_label}-{section_name}"
        if class_key in seen_class_keys:
            continue
        seen_class_keys.add(class_key)

        raw_status = str(section.class_status or "").strip().lower()
        class_status = "New" if raw_status == "new" else "Current"
        display_grade = "KG" if grade_label == "KG" else f"G{grade_label}"

        class_rows.append(
            {
                "planning_section_id": section.id,
                "class_key": class_key,
                "class_label": f"{display_grade}-{section_name}",
                "grade_label": grade_label,
                "section_name": section_name,
                "class_status": class_status,
                "homeroom_teacher_id": section.homeroom_teacher_id,
            }
        )

    class_rows.sort(
        key=lambda row: (
            _grade_sort_key(row["grade_label"]),
            row["section_name"],
            0 if row["class_status"] == "Current" else 1,
        )
    )
    return class_rows


def _build_report_subject_catalog(subjects):
    subjects_by_grade = {}
    subject_name_by_key = {}

    for subject in subjects:
        grade_label = _normalize_grade_label(subject.grade)
        if not grade_label:
            continue

        weekly_hours = int(subject.weekly_hours or 0)
        if weekly_hours <= 0:
            continue

        subject_code = str(subject.subject_code or "").strip().upper()
        subject_key, subject_name = _build_subject_identity(
            subject_name=subject.subject_name,
            fallback_code=subject_code,
        )
        if not subject_key:
            continue

        if subject_key not in subject_name_by_key:
            subject_name_by_key[subject_key] = subject_name

        subject_color = resolve_subject_color(
            subject_code or subject_key,
            getattr(subject, "color", ""),
            subject_name=subject_name,
        )
        theme = build_subject_theme(subject_color)

        subjects_by_grade.setdefault(grade_label, []).append(
            {
                "subject_key": subject_key,
                "subject_code": subject_code or subject_name,
                "subject_name": subject_name,
                "subject_color": subject_color,
                "subject_color_soft": theme["soft"],
                "subject_color_text": theme["text"],
                "subject_color_border": theme["border"],
                "weekly_hours": weekly_hours,
            }
        )

    for grade_label in subjects_by_grade:
        subjects_by_grade[grade_label].sort(
            key=lambda item: (item["subject_name"], item["subject_code"])
        )

    return subjects_by_grade, subject_name_by_key


def _build_homeroom_assignments_by_teacher(
    subjects,
    planning_sections,
    explicit_section_subject_keys=None,
    valid_teacher_ids=None,
):
    class_rows = _build_report_class_rows(planning_sections)
    subjects_by_grade, _ = _build_report_subject_catalog(subjects)
    assignments_by_teacher = {}
    explicit_section_subject_keys = explicit_section_subject_keys or set()
    has_teacher_filter = valid_teacher_ids is not None
    valid_teacher_ids = set(valid_teacher_ids or [])

    for class_row in class_rows:
        homeroom_teacher_id = class_row.get("homeroom_teacher_id")
        if not homeroom_teacher_id:
            continue
        if has_teacher_filter and homeroom_teacher_id not in valid_teacher_ids:
            continue

        bundle_subject_items = []
        default_subject_items = []
        for subject_item in subjects_by_grade.get(class_row["grade_label"], []):
            explicit_section_key = (
                class_row.get("planning_section_id"),
                str(subject_item["subject_code"] or "").strip().upper(),
            )
            if explicit_section_key in explicit_section_subject_keys:
                continue

            if is_homeroom_bundle_subject(
                subject_code=subject_item["subject_code"],
                subject_name=subject_item["subject_name"],
                weekly_hours=subject_item["weekly_hours"],
                grade_label=class_row["grade_label"],
            ):
                bundle_subject_items.append(subject_item)
                continue

            if is_default_homeroom_subject(
                class_row["grade_label"],
                subject_key=subject_item["subject_key"],
                subject_name=subject_item["subject_name"],
                subject_code=subject_item["subject_code"],
            ):
                default_subject_items.append(subject_item)

        for subject_item in bundle_subject_items or default_subject_items:

            assignments_by_teacher.setdefault(homeroom_teacher_id, []).append(
                {
                    "teacher_id": homeroom_teacher_id,
                    "class_key": class_row["class_key"],
                    "class_label": class_row["class_label"],
                    "class_status": class_row["class_status"],
                    "grade_label": class_row["grade_label"],
                    "section_name": class_row["section_name"],
                    "subject_key": subject_item["subject_key"],
                    "subject_code": subject_item["subject_code"],
                    "subject_name": subject_item["subject_name"],
                    "required_hours": subject_item["weekly_hours"],
                }
            )

    for teacher_id in assignments_by_teacher:
        assignments_by_teacher[teacher_id].sort(
            key=lambda item: (
                0 if item["class_status"] == "Current" else 1,
                _grade_sort_key(item["grade_label"]),
                item["section_name"],
                item["subject_name"],
            )
        )

    return assignments_by_teacher


def _build_report_class_allocation_data(subjects, planning_sections, reporting_context):
    class_rows = _build_report_class_rows(planning_sections)
    subjects_by_grade, subject_name_by_key = _build_report_subject_catalog(subjects)
    teacher_profiles = reporting_context.get("teacher_profiles", [])

    demand_items_by_subject = {}
    demand_items_lookup = {}
    for class_row in class_rows:
        grade_subjects = subjects_by_grade.get(class_row["grade_label"], [])
        for subject_item in grade_subjects:
            demand_item = {
                "class_key": class_row["class_key"],
                "class_label": class_row["class_label"],
                "class_status": class_row["class_status"],
                "grade_label": class_row["grade_label"],
                "section_name": class_row["section_name"],
                "subject_key": subject_item["subject_key"],
                "subject_code": subject_item["subject_code"],
                "subject_name": subject_item["subject_name"],
                "required_hours": subject_item["weekly_hours"],
                "remaining_hours": subject_item["weekly_hours"],
                "recommended_hours": 0,
                "recommended_assignments": [],
            }
            demand_items_by_subject.setdefault(subject_item["subject_key"], []).append(
                demand_item
            )
            demand_items_lookup[
                (class_row["class_key"], subject_item["subject_key"])
            ] = demand_item

    for subject_key in demand_items_by_subject:
        demand_items_by_subject[subject_key].sort(
            key=lambda item: (
                0 if item["class_status"] == "Current" else 1,
                _grade_sort_key(item["grade_label"]),
                item["section_name"],
                item["class_label"],
            )
        )

    sorted_profiles = sorted(
        teacher_profiles,
        key=lambda profile: (
            -int(profile.get("allocated_hours", 0)),
            profile.get("teacher_name", ""),
            str(profile.get("teacher_id", "")),
        ),
    )

    teacher_matrix_rows = []
    assignment_rows = []
    teacher_matrix_row_by_pk = {}

    for profile in sorted_profiles:
        allocation_breakdown = profile.get("allocation_breakdown", {}) or {}
        primary_subject_keys = list(profile.get("subject_keys", []))
        support_subject_keys = list(profile.get("support_subject_keys", []))
        homeroom_class_allocations = list(
            profile.get("homeroom_class_allocations", [])
        )

        ordered_subject_keys = []
        for subject_key in primary_subject_keys + support_subject_keys:
            if subject_key not in ordered_subject_keys:
                ordered_subject_keys.append(subject_key)
        for subject_key in sorted(allocation_breakdown.keys()):
            if subject_key not in ordered_subject_keys:
                ordered_subject_keys.append(subject_key)

        class_allocations = {}

        # Seed class allocations with reserved homeroom work so export matches report math.
        for item in homeroom_class_allocations:
            allocated_hours = int(item.get("allocated_hours", 0))
            if allocated_hours <= 0:
                continue

            demand_item = demand_items_lookup.get(
                (item.get("class_key"), item.get("subject_key"))
            )
            if demand_item:
                demand_item["remaining_hours"] = max(
                    int(demand_item["remaining_hours"]) - allocated_hours,
                    0,
                )

            class_key = item.get("class_key")
            class_allocations.setdefault(class_key, []).append(
                {
                    "subject_key": item.get("subject_key"),
                    "subject_code": item.get("subject_code"),
                    "subject_name": item.get("subject_name"),
                    "allocated_hours": allocated_hours,
                    "class_status": item.get("class_status"),
                }
            )

            assignment_rows.append(
                {
                    "teacher_id": profile.get("teacher_id", "-"),
                    "teacher_name": profile.get("teacher_name", "-"),
                    "class_label": item.get("class_label", "-"),
                    "class_status": item.get("class_status", "Current"),
                    "subject_code": item.get("subject_code", "-"),
                    "subject_name": item.get("subject_name", "-"),
                    "allocated_hours": allocated_hours,
                    "coverage_type": "Homeroom",
                }
            )

        for subject_key in ordered_subject_keys:
            subject_hours_quota = int(allocation_breakdown.get(subject_key, 0))
            if subject_hours_quota <= 0:
                continue

            for demand_item in demand_items_by_subject.get(subject_key, []):
                if subject_hours_quota <= 0:
                    break

                subject_hours_remaining = int(demand_item["remaining_hours"])
                if subject_hours_remaining <= 0:
                    continue

                allocated_hours = min(subject_hours_quota, subject_hours_remaining)
                demand_item["remaining_hours"] = subject_hours_remaining - allocated_hours
                subject_hours_quota -= allocated_hours

                class_key = demand_item["class_key"]
                class_allocations.setdefault(class_key, []).append(
                    {
                        "subject_key": subject_key,
                        "subject_code": demand_item["subject_code"],
                        "subject_name": demand_item["subject_name"],
                        "allocated_hours": allocated_hours,
                        "class_status": demand_item["class_status"],
                    }
                )

                assignment_rows.append(
                    {
                        "teacher_id": profile.get("teacher_id", "-"),
                        "teacher_name": profile.get("teacher_name", "-"),
                        "class_label": demand_item["class_label"],
                        "class_status": demand_item["class_status"],
                        "subject_code": demand_item["subject_code"],
                        "subject_name": demand_item["subject_name"],
                        "allocated_hours": allocated_hours,
                        "coverage_type": (
                            "Support"
                            if subject_key in support_subject_keys
                            else "Primary"
                        ),
                    }
                )

        class_cells = {}
        class_fill_subject_keys = {}
        for class_key, allocation_items in class_allocations.items():
            allocation_items.sort(
                key=lambda item: (-item["allocated_hours"], item["subject_code"])
            )
            class_cells[class_key] = "\n".join(
                f"{item['subject_code']} ({item['allocated_hours']}h)"
                for item in allocation_items
            )
            class_fill_subject_keys[class_key] = allocation_items[0]["subject_key"]

        teacher_matrix_rows.append(
            {
                "teacher_pk": profile.get("teacher_pk"),
                "teacher_id": profile.get("teacher_id", "-"),
                "teacher_name": profile.get("teacher_name", "-"),
                "degree_major": profile.get("degree_major", ""),
                "expected_allocated_hours": int(profile.get("allocated_hours", 0)),
                "capacity_hours": int(
                    profile.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
                ),
                "remaining_capacity_hours": int(
                    profile.get("remaining_capacity_hours", REPORT_STANDARD_MAX_HOURS)
                ),
                "recommended_absorption_hours": 0,
                "recommended_assignment_labels": [],
                "primary_subject_label": ", ".join(
                    subject_name_by_key.get(subject_key, subject_key.title())
                    for subject_key in primary_subject_keys
                )
                or "-",
                "support_subject_label": ", ".join(
                    subject_name_by_key.get(subject_key, subject_key.title())
                    for subject_key in support_subject_keys
                )
                or "-",
                "class_cells": class_cells,
                "class_fill_subject_keys": class_fill_subject_keys,
            }
        )
        teacher_matrix_row_by_pk[profile.get("teacher_pk")] = teacher_matrix_rows[-1]

    assignment_rows.sort(
        key=lambda row: (
            row["teacher_name"],
            row["class_label"],
            row["subject_code"],
        )
    )

    recommendation_rows = []

    unassigned_rows = []
    for subject_items in demand_items_by_subject.values():
        for demand_item in subject_items:
            if int(demand_item["remaining_hours"]) <= 0:
                continue
            unassigned_rows.append(
                {
                    "class_label": demand_item["class_label"],
                    "class_status": demand_item["class_status"],
                    "subject_code": demand_item["subject_code"],
                    "subject_name": demand_item["subject_name"],
                    "remaining_hours": int(demand_item["remaining_hours"]),
                }
            )

    unassigned_rows.sort(
        key=lambda row: (
            row["class_label"],
            row["subject_code"],
        )
    )

    subject_section_rows = []
    subject_section_map = {}
    for subject_key, subject_items in demand_items_by_subject.items():
        total_sections = len(subject_items)
        covered_section_labels = []
        partial_section_labels = []
        recommended_section_labels = []
        uncovered_section_labels = []

        for demand_item in subject_items:
            required_hours = int(demand_item["required_hours"])
            remaining_hours = int(demand_item["remaining_hours"])
            recommended_hours = int(demand_item.get("recommended_hours", 0))
            allocated_hours = max(required_hours - remaining_hours, 0)
            class_label = demand_item["class_label"]

            if remaining_hours <= 0:
                covered_section_labels.append(class_label)
            elif allocated_hours > 0:
                partial_section_labels.append(
                    f"{class_label} / {remaining_hours}h"
                )

            if recommended_hours > 0:
                recommended_section_labels.append(
                    f"{class_label}: "
                    + ", ".join(
                        f"{entry['teacher_name']} ({entry['allocated_hours']}h, {entry['match_basis']})"
                        for entry in demand_item.get("recommended_assignments", [])
                    )
                )
            elif remaining_hours > 0 and allocated_hours <= 0:
                uncovered_section_labels.append(f"{class_label} / {remaining_hours}h")

        section_row = {
            "subject_key": subject_key,
            "total_sections": total_sections,
            "covered_sections_count": len(covered_section_labels),
            "partial_sections_count": len(partial_section_labels),
            "recommended_sections_count": len(recommended_section_labels),
            "uncovered_sections_count": len(uncovered_section_labels),
            "covered_section_labels": covered_section_labels,
            "partial_section_labels": partial_section_labels,
            "recommended_section_labels": recommended_section_labels,
            "uncovered_section_labels": uncovered_section_labels,
            "recommended_hours": sum(
                int(item.get("recommended_hours", 0))
                for item in subject_items
            ),
        }
        subject_section_rows.append(section_row)
        subject_section_map[subject_key] = section_row

    underloaded_teacher_rows = []
    for teacher_matrix_row in teacher_matrix_rows:
        current_remaining_capacity = int(
            teacher_matrix_row.get("remaining_capacity_hours", 0)
        )
        if current_remaining_capacity <= 0:
            continue

        recommended_absorption_hours = int(
            teacher_matrix_row.get("recommended_absorption_hours", 0)
        )
        expected_allocated_hours = int(
            teacher_matrix_row.get("expected_allocated_hours", 0)
        )
        capacity_hours = int(
            teacher_matrix_row.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
        )
        projected_allocated_hours = min(
            expected_allocated_hours + recommended_absorption_hours,
            capacity_hours,
        )
        projected_remaining_capacity = max(
            capacity_hours - projected_allocated_hours,
            0,
        )

        underloaded_teacher_rows.append(
            {
                "teacher_pk": teacher_matrix_row.get("teacher_pk"),
                "teacher_id": teacher_matrix_row.get("teacher_id", "-"),
                "teacher_name": teacher_matrix_row.get("teacher_name", "-"),
                "degree_major": teacher_matrix_row.get("degree_major", ""),
                "current_load_hours": expected_allocated_hours,
                "capacity_hours": capacity_hours,
                "remaining_capacity_hours": current_remaining_capacity,
                "projected_allocated_hours": projected_allocated_hours,
                "projected_remaining_capacity_hours": projected_remaining_capacity,
                "recommended_absorption_hours": recommended_absorption_hours,
                "recommended_assignment_labels": list(
                    teacher_matrix_row.get("recommended_assignment_labels", [])
                ),
            }
        )

    underloaded_teacher_rows.sort(
        key=lambda row: (
            -row["recommended_absorption_hours"],
            -row["remaining_capacity_hours"],
            row["teacher_name"],
        )
    )

    return {
        "class_rows": class_rows,
        "teacher_matrix_rows": teacher_matrix_rows,
        "assignment_rows": assignment_rows,
        "unassigned_rows": unassigned_rows,
        "subject_section_rows": subject_section_rows,
        "subject_section_map": subject_section_map,
        "underloaded_teacher_rows": underloaded_teacher_rows,
        "recommendation_rows": recommendation_rows,
    }


def _decorate_staffing_report_rows(report_subject_rows, report_summary):
    def _hex_to_rgb(color_value: str):
        cleaned_value = str(color_value or "").strip().lstrip("#")
        if len(cleaned_value) != 6:
            return (10, 78, 163)
        try:
            return (
                int(cleaned_value[0:2], 16),
                int(cleaned_value[2:4], 16),
                int(cleaned_value[4:6], 16),
            )
        except ValueError:
            return (10, 78, 163)

    def _rgb_to_hex(rgb_value):
        red, green, blue = rgb_value
        safe_red = max(0, min(255, int(round(red))))
        safe_green = max(0, min(255, int(round(green))))
        safe_blue = max(0, min(255, int(round(blue))))
        return f"#{safe_red:02x}{safe_green:02x}{safe_blue:02x}"

    def _blend_hex_colors(start_color: str, end_color: str, ratio: float):
        safe_ratio = max(0.0, min(1.0, float(ratio)))
        start_rgb = _hex_to_rgb(start_color)
        end_rgb = _hex_to_rgb(end_color)
        blended_rgb = (
            start_rgb[0] + (end_rgb[0] - start_rgb[0]) * safe_ratio,
            start_rgb[1] + (end_rgb[1] - start_rgb[1]) * safe_ratio,
            start_rgb[2] + (end_rgb[2] - start_rgb[2]) * safe_ratio,
        )
        return _rgb_to_hex(blended_rgb)

    def _subject_coverage_donut_palette(coverage_percentage: int):
        safe_coverage = max(0, min(100, int(coverage_percentage or 0)))
        theme_red = "#b42318"
        theme_orange = "#d97706"
        theme_amber = "#c79a14"
        theme_green = "#0d7a47"

        if safe_coverage <= 40:
            primary_color = theme_red
            remainder_tint = "#f2c5c8"
            secondary_mix_ratio = 0.38
        elif safe_coverage <= 60:
            interpolation_ratio = (safe_coverage - 40) / 20
            primary_color = _blend_hex_colors(
                theme_red,
                theme_orange,
                interpolation_ratio,
            )
            remainder_tint = "#f6d1cb"
            secondary_mix_ratio = 0.50
        elif safe_coverage <= 80:
            interpolation_ratio = (safe_coverage - 60) / 20
            primary_color = _blend_hex_colors(
                theme_orange,
                theme_amber,
                interpolation_ratio,
            )
            remainder_tint = "#f9e1bf"
            secondary_mix_ratio = 0.64
        else:
            interpolation_ratio = (safe_coverage - 80) / 20
            primary_color = _blend_hex_colors(
                theme_amber,
                theme_green,
                interpolation_ratio,
            )
            remainder_tint = "#d6ecde"
            secondary_mix_ratio = 0.72

        secondary_color = _blend_hex_colors(
            primary_color,
            remainder_tint,
            secondary_mix_ratio,
        )
        return primary_color, secondary_color

    decorated_rows = []
    priority_subjects_with_gaps = 0
    priority_urgent_subjects = 0
    priority_uncovered_sections = 0
    priority_gap_hours = 0
    largest_priority_row = None

    for row in report_subject_rows:
        decorated_row = dict(row)
        remaining_hours = int(decorated_row.get("remaining_hours", 0))
        coverage_percentage = int(decorated_row.get("coverage_percentage", 0))
        donut_primary_color, donut_secondary_color = _subject_coverage_donut_palette(
            coverage_percentage
        )
        teacher_blocks = int(
            decorated_row.get(
                "teacher_requirement_blocks",
                decorated_row.get("additional_teachers_needed", 0),
            )
        )
        uncovered_sections_count = int(decorated_row.get("uncovered_sections_count", 0))
        partial_sections_count = int(decorated_row.get("partial_sections_count", 0))
        open_gap_sections = uncovered_sections_count + partial_sections_count
        priority_subject = bool(
            decorated_row.get("priority_staffing_subject")
            or _is_priority_staffing_subject(
                decorated_row.get("subject_key", ""),
                decorated_row.get("subject_name", ""),
            )
        )
        priority_alert = priority_subject and remaining_hours > 0
        priority_urgent = priority_alert and open_gap_sections >= 2

        decorated_row["priority_staffing_subject"] = priority_subject
        decorated_row["priority_staffing_alert"] = priority_alert
        decorated_row["priority_staffing_urgent"] = priority_urgent
        decorated_row["open_gap_sections_count"] = open_gap_sections
        decorated_row["subject_donut_primary"] = donut_primary_color
        decorated_row["subject_donut_secondary"] = donut_secondary_color
        decorated_row["subject_accent_surface"] = _blend_hex_colors(
            donut_secondary_color,
            "#ffffff",
            0.42,
        )
        decorated_row["subject_status_bg"] = _blend_hex_colors(
            donut_primary_color,
            "#ffffff",
            0.79,
        )
        decorated_row["subject_status_border"] = _blend_hex_colors(
            donut_primary_color,
            "#ffffff",
            0.50,
        )
        decorated_row["subject_status_text"] = _blend_hex_colors(
            donut_primary_color,
            "#0f172a",
            0.06,
        )

        if priority_alert:
            priority_subjects_with_gaps += 1
            priority_uncovered_sections += open_gap_sections
            priority_gap_hours += remaining_hours
            if largest_priority_row is None or (
                remaining_hours,
                open_gap_sections,
                decorated_row.get("subject_name", ""),
            ) > (
                int(largest_priority_row.get("remaining_hours", 0)),
                int(largest_priority_row.get("open_gap_sections_count", 0)),
                largest_priority_row.get("subject_name", ""),
            ):
                largest_priority_row = decorated_row
        if priority_urgent:
            priority_urgent_subjects += 1

        if remaining_hours == 0:
            decorated_row.update(
                {
                    "staffing_status_label": "Fully Covered",
                    "staffing_status_class": "report-status-pill-good",
                    "staffing_item_class": "is-covered",
                    "staffing_alert_class": "is-covered",
                    "staffing_fill_color": "#0f7f7a",
                    "staffing_fill_remainder": "#dbe7f8",
                    "staffing_icon": "shield",
                }
            )
        elif priority_urgent:
            decorated_row.update(
                {
                    "staffing_status_label": "Priority Alert",
                    "staffing_status_class": "report-status-pill-critical",
                    "staffing_item_class": "is-critical",
                    "staffing_alert_class": "is-critical",
                    "staffing_fill_color": "#b42318",
                    "staffing_fill_remainder": "#f5c6ca",
                    "staffing_icon": "alert",
                }
            )
        elif priority_alert:
            decorated_row.update(
                {
                    "staffing_status_label": "Specialist Needed",
                    "staffing_status_class": "report-status-pill-gap",
                    "staffing_item_class": "is-warning",
                    "staffing_alert_class": "is-warning",
                    "staffing_fill_color": "#d97706",
                    "staffing_fill_remainder": "#fde7c2",
                    "staffing_icon": "priority",
                }
            )
        elif teacher_blocks >= 2 or remaining_hours >= 24 or open_gap_sections >= 2:
            decorated_row.update(
                {
                    "staffing_status_label": "Major Gap",
                    "staffing_status_class": "report-status-pill-critical",
                    "staffing_item_class": "is-critical",
                    "staffing_alert_class": "is-critical",
                    "staffing_fill_color": "#b42318",
                    "staffing_fill_remainder": "#f5c6ca",
                    "staffing_icon": "alert",
                }
            )
        elif teacher_blocks >= 1:
            decorated_row.update(
                {
                    "staffing_status_label": "Coverage Gap",
                    "staffing_status_class": "report-status-pill-gap",
                    "staffing_item_class": "is-warning",
                    "staffing_alert_class": "is-warning",
                    "staffing_fill_color": "#d97706",
                    "staffing_fill_remainder": "#fde7c2",
                    "staffing_icon": "hire",
                }
            )
        elif coverage_percentage >= 70:
            decorated_row.update(
                {
                    "staffing_status_label": "Partial Gap",
                    "staffing_status_class": "report-status-pill-partial",
                    "staffing_item_class": "is-warning",
                    "staffing_alert_class": "is-warning",
                    "staffing_fill_color": "#d97706",
                    "staffing_fill_remainder": "#fde7c2",
                    "staffing_icon": "partial",
                }
            )
        else:
            decorated_row.update(
                {
                    "staffing_status_label": "Major Gap",
                    "staffing_status_class": "report-status-pill-gap",
                    "staffing_item_class": "is-warning",
                    "staffing_alert_class": "is-warning",
                    "staffing_fill_color": "#d97706",
                    "staffing_fill_remainder": "#fde7c2",
                    "staffing_icon": "gap",
                }
            )

        if priority_alert:
            decorated_row["staffing_action_label"] = "24h Blocks"
            decorated_row["staffing_action_value"] = str(teacher_blocks)
            decorated_row["staffing_note"] = (
                "Multiple sections are still open in this priority subject."
                if priority_urgent
                else "Priority subject coverage should be reviewed first."
            )
        elif remaining_hours == 0:
            decorated_row["staffing_action_label"] = "24h Blocks"
            decorated_row["staffing_action_value"] = "0"
            decorated_row["staffing_note"] = ""
        else:
            decorated_row["staffing_action_label"] = "24h Blocks"
            decorated_row["staffing_action_value"] = str(teacher_blocks)
            if open_gap_sections > 0:
                decorated_row["staffing_note"] = (
                    f"{open_gap_sections} section(s) still have uncovered hours."
                )
            else:
                decorated_row["staffing_note"] = "Use the uncovered-hour total to decide next staffing steps."

        decorated_rows.append(decorated_row)

    max_remaining_hours = max(
        (int(row.get("remaining_hours", 0)) for row in decorated_rows if int(row.get("remaining_hours", 0)) > 0),
        default=0,
    )
    for row in decorated_rows:
        remaining_hours = int(row.get("remaining_hours", 0))
        row["gap_chart_pct"] = (
            round((remaining_hours / max_remaining_hours) * 100, 1)
            if max_remaining_hours > 0 and remaining_hours > 0
            else 0
        )

    report_summary = dict(report_summary or {})
    report_summary.update(
        {
            "priority_subjects_with_gaps": priority_subjects_with_gaps,
            "priority_urgent_subjects": priority_urgent_subjects,
            "priority_uncovered_sections": priority_uncovered_sections,
            "priority_gap_hours": priority_gap_hours,
            "largest_priority_gap_subject_name": (
                largest_priority_row.get("subject_name", "")
                if largest_priority_row
                else ""
            ),
            "largest_priority_gap_hours": (
                int(largest_priority_row.get("remaining_hours", 0))
                if largest_priority_row
                else 0
            ),
            "largest_priority_gap_sections": (
                int(largest_priority_row.get("open_gap_sections_count", 0))
                if largest_priority_row
                else 0
            ),
            "largest_priority_gap_blocks": (
                int(
                    largest_priority_row.get(
                        "teacher_requirement_blocks",
                        largest_priority_row.get("additional_teachers_needed", 0),
                    )
                )
                if largest_priority_row
                else 0
            ),
        }
    )
    report_summary.update(
        _build_hiring_coverage_recommendation(decorated_rows)
    )
    report_summary = _enrich_report_summary_hiring_metrics(report_summary)
    return decorated_rows, report_summary


def _apply_excel_header_style(sheet, header_row: int, total_columns: int):
    header_fill = PatternFill(start_color="0A4EA3", end_color="0A4EA3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index in range(1, total_columns + 1):
        cell = sheet.cell(row=header_row, column=column_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _subject_fill_for_key(subject_key: str):
    if not subject_key:
        return None
    color_code = to_excel_hex(resolve_subject_color(subject_key))
    return PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")


def _build_report_allocation_filename(branch_name: str, academic_year_name: str) -> str:
    def _sanitize(text_value: str) -> str:
        normalized = re.sub(r"[^A-Za-z0-9]+", "-", str(text_value or "").strip())
        return normalized.strip("-").lower() or "scope"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_branch = _sanitize(branch_name)
    safe_year = _sanitize(academic_year_name)
    return f"teacher_allocation_plan_{safe_branch}_{safe_year}_{timestamp}.xlsx"


def _build_report_allocation_xlsx_bytes(
    branch_name: str,
    academic_year_name: str,
    subjects,
    planning_sections,
    reporting_context,
    allocation_data=None,
) -> bytes:
    allocation_data = allocation_data or _build_report_class_allocation_data(
        subjects=subjects,
        planning_sections=planning_sections,
        reporting_context=reporting_context,
    )

    class_rows = allocation_data["class_rows"]
    teacher_matrix_rows = allocation_data["teacher_matrix_rows"]
    assignment_rows = allocation_data["assignment_rows"]
    unassigned_rows = allocation_data["unassigned_rows"]
    subject_section_map = allocation_data.get("subject_section_map", {})

    report_summary = reporting_context.get("summary", {})
    report_subject_rows = [
        {
            **row,
            **subject_section_map.get(row["subject_key"], {}),
        }
        for row in reporting_context.get("subject_rows", [])
    ]
    report_subject_rows, report_summary = _decorate_staffing_report_rows(
        report_subject_rows,
        report_summary,
    )
    report_teacher_rows = reporting_context.get("teacher_rows", [])

    workbook = Workbook()
    matrix_sheet = workbook.active
    matrix_sheet.title = "Teacher_Class_Matrix"

    matrix_headers = [
        "Teacher ID",
        "Teacher Name",
        "Assigned Hours",
        "Remaining Capacity",
        "Assigned Subject",
        "Support Subject",
    ] + [class_row["class_label"] for class_row in class_rows]
    matrix_sheet.append(matrix_headers)
    _apply_excel_header_style(
        sheet=matrix_sheet,
        header_row=1,
        total_columns=len(matrix_headers),
    )
    matrix_sheet.freeze_panes = "A2"
    matrix_sheet.auto_filter.ref = f"A1:{get_column_letter(len(matrix_headers))}1"

    full_load_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    under_load_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    neutral_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    class_column_start = 7
    class_order = [class_row["class_key"] for class_row in class_rows]

    for row_data in teacher_matrix_rows:
        matrix_row = [
            row_data["teacher_id"],
            row_data["teacher_name"],
            row_data["expected_allocated_hours"],
            row_data["remaining_capacity_hours"],
            row_data["primary_subject_label"],
            row_data["support_subject_label"],
        ]
        matrix_row.extend(
            row_data["class_cells"].get(class_key, "")
            for class_key in class_order
        )
        matrix_sheet.append(matrix_row)
        excel_row_index = matrix_sheet.max_row

        expected_hours_cell = matrix_sheet.cell(row=excel_row_index, column=3)
        remaining_capacity_cell = matrix_sheet.cell(row=excel_row_index, column=4)
        if int(row_data["expected_allocated_hours"]) >= int(
            row_data.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
        ):
            expected_hours_cell.fill = full_load_fill
            remaining_capacity_cell.fill = full_load_fill
        else:
            expected_hours_cell.fill = under_load_fill
            remaining_capacity_cell.fill = under_load_fill

        for class_offset, class_key in enumerate(class_order):
            column_index = class_column_start + class_offset
            class_cell = matrix_sheet.cell(row=excel_row_index, column=column_index)
            if not class_cell.value:
                continue
            fill_subject_key = row_data["class_fill_subject_keys"].get(class_key, "")
            class_fill = _subject_fill_for_key(fill_subject_key)
            if class_fill:
                class_cell.fill = class_fill
            class_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

    matrix_sheet.column_dimensions["A"].width = 14
    matrix_sheet.column_dimensions["B"].width = 24
    matrix_sheet.column_dimensions["C"].width = 18
    matrix_sheet.column_dimensions["D"].width = 18
    matrix_sheet.column_dimensions["E"].width = 24
    matrix_sheet.column_dimensions["F"].width = 24
    for column_index in range(class_column_start, len(matrix_headers) + 1):
        matrix_sheet.column_dimensions[get_column_letter(column_index)].width = 18
    for row_index in range(2, matrix_sheet.max_row + 1):
        for col_index in range(1, 7):
            matrix_sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                horizontal="left" if col_index in {2, 5, 6} else "center",
                vertical="top",
                wrap_text=True,
            )
        if matrix_sheet.max_column >= 7:
            matrix_sheet.row_dimensions[row_index].height = 38

    details_sheet = workbook.create_sheet("Teacher_Details")
    detail_headers = [
        "Teacher ID",
        "Teacher Name",
        "Class",
        "Class Status",
        "Subject Code",
        "Subject Name",
        "Allocated Hours",
        "Coverage Type",
    ]
    details_sheet.append(detail_headers)
    _apply_excel_header_style(
        sheet=details_sheet,
        header_row=1,
        total_columns=len(detail_headers),
    )
    details_sheet.freeze_panes = "A2"
    details_sheet.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}1"

    support_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    homeroom_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    new_class_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    for item in assignment_rows:
        details_sheet.append(
            [
                item["teacher_id"],
                item["teacher_name"],
                item["class_label"],
                item["class_status"],
                item["subject_code"],
                item["subject_name"],
                item["allocated_hours"],
                item["coverage_type"],
            ]
        )
        row_index = details_sheet.max_row
        if item["coverage_type"] == "Support":
            details_sheet.cell(row=row_index, column=8).fill = support_fill
        elif item["coverage_type"] == "Homeroom":
            details_sheet.cell(row=row_index, column=8).fill = homeroom_fill
        if item["class_status"] == "New":
            details_sheet.cell(row=row_index, column=4).fill = new_class_fill
        details_sheet.cell(row=row_index, column=7).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    details_sheet.column_dimensions["A"].width = 14
    details_sheet.column_dimensions["B"].width = 24
    details_sheet.column_dimensions["C"].width = 14
    details_sheet.column_dimensions["D"].width = 12
    details_sheet.column_dimensions["E"].width = 14
    details_sheet.column_dimensions["F"].width = 26
    details_sheet.column_dimensions["G"].width = 15
    details_sheet.column_dimensions["H"].width = 14

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet["A1"] = "Teacher Allocation Planning Summary"
    summary_sheet["A1"].font = Font(bold=True, size=14, color="0A4EA3")
    summary_sheet["A2"] = f"Branch: {branch_name}"
    summary_sheet["A3"] = f"Academic Year: {academic_year_name}"
    summary_sheet["A4"] = f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    summary_metrics = [
        ("Existing Teachers", report_summary.get("total_existing_teachers", 0)),
        ("Total Required Hours", report_summary.get("total_required_hours", 0)),
        ("Covered by Assigned Teachers", report_summary.get("total_allocated_hours", 0)),
        ("Uncovered Hours", report_summary.get("total_remaining_hours", 0)),
        (
            "24h Staffing Blocks",
            report_summary.get("total_staffing_requirement_blocks", 0),
        ),
        (
            "Homeroom Default Coverage",
            report_summary.get("homeroom_default_coverage_hours", 0),
        ),
        (
            "Priority Subject Gap Hours",
            report_summary.get("priority_gap_hours", 0),
        ),
        (
            "Priority Subjects With Gaps",
            report_summary.get("priority_subjects_with_gaps", 0),
        ),
        (
            "Priority Uncovered Sections",
            report_summary.get("priority_uncovered_sections", 0),
        ),
        ("Coverage %", f"{report_summary.get('coverage_percentage', 0)}%"),
    ]
    summary_sheet.append([])
    summary_sheet.append(["Metric", "Value"])
    metric_header_row = summary_sheet.max_row
    _apply_excel_header_style(summary_sheet, metric_header_row, 2)
    for metric_label, metric_value in summary_metrics:
        summary_sheet.append([metric_label, metric_value])

    subject_table_row = summary_sheet.max_row + 2
    summary_sheet.cell(row=subject_table_row, column=1, value="Subject Summary")
    summary_sheet.cell(row=subject_table_row, column=1).font = Font(
        bold=True,
        color="0A4EA3",
    )
    subject_headers = [
        "Subject",
        "Current Hours",
        "New Hours",
        "Total Required",
        "Covered",
        "Uncovered",
        "24h Staffing Blocks",
        "Coverage Note",
    ]
    summary_sheet.append(subject_headers)
    _apply_excel_header_style(summary_sheet, subject_table_row + 1, len(subject_headers))
    for row in report_subject_rows:
        summary_sheet.append(
            [
                row["subject_name"],
                row["required_current_hours"],
                row["required_new_hours"],
                row["required_hours"],
                row["allocated_hours"],
                row["remaining_hours"],
                row.get("teacher_requirement_blocks", row["additional_teachers_needed"]),
                row.get("staffing_note", row.get("additional_teachers_note", "")),
            ]
        )
        row_index = summary_sheet.max_row
        uncovered_cell = summary_sheet.cell(row=row_index, column=6)
        staffing_blocks_cell = summary_sheet.cell(row=row_index, column=7)
        if int(row["remaining_hours"]) > 0:
            uncovered_cell.fill = under_load_fill
        else:
            uncovered_cell.fill = full_load_fill
        if int(row.get("teacher_requirement_blocks", row["additional_teachers_needed"])) > 0:
            staffing_blocks_cell.fill = under_load_fill
        else:
            staffing_blocks_cell.fill = full_load_fill

    teacher_table_row = summary_sheet.max_row + 2
    summary_sheet.cell(row=teacher_table_row, column=1, value="Teacher Load Summary")
    summary_sheet.cell(row=teacher_table_row, column=1).font = Font(
        bold=True,
        color="0A4EA3",
    )
    teacher_headers = [
        "Teacher ID",
        "Teacher Name",
        "Expected Hours",
        "Homeroom Allocated",
        "Primary Allocated",
        "Support Allocated",
        "Remaining Capacity",
    ]
    summary_sheet.append(teacher_headers)
    _apply_excel_header_style(summary_sheet, teacher_table_row + 1, len(teacher_headers))
    for row in report_teacher_rows:
        summary_sheet.append(
            [
                row["teacher_id"],
                row["teacher_name"],
                row["expected_allocated_hours"],
                row.get("homeroom_allocated_hours", 0),
                row["primary_allocated_hours"],
                row["support_allocated_hours"],
                row["remaining_capacity_hours"],
            ]
        )
        row_index = summary_sheet.max_row
        expected_cell = summary_sheet.cell(row=row_index, column=3)
        remaining_cell = summary_sheet.cell(row=row_index, column=7)
        if int(row["expected_allocated_hours"]) >= int(
            row.get("capacity_hours", REPORT_STANDARD_MAX_HOURS)
        ):
            expected_cell.fill = full_load_fill
            remaining_cell.fill = full_load_fill
        else:
            expected_cell.fill = under_load_fill
            remaining_cell.fill = under_load_fill

    unassigned_table_row = summary_sheet.max_row + 2
    summary_sheet.cell(
        row=unassigned_table_row,
        column=1,
        value="Unassigned Class Demand",
    )
    summary_sheet.cell(row=unassigned_table_row, column=1).font = Font(
        bold=True,
        color="0A4EA3",
    )
    unassigned_headers = [
        "Class",
        "Class Status",
        "Subject Code",
        "Subject Name",
        "Unassigned Hours",
    ]
    summary_sheet.append(unassigned_headers)
    _apply_excel_header_style(summary_sheet, unassigned_table_row + 1, len(unassigned_headers))
    if unassigned_rows:
        for row in unassigned_rows:
            summary_sheet.append(
                [
                    row["class_label"],
                    row["class_status"],
                    row["subject_code"],
                    row["subject_name"],
                    row["remaining_hours"],
                ]
            )
            row_index = summary_sheet.max_row
            summary_sheet.cell(row=row_index, column=5).fill = under_load_fill
    else:
        summary_sheet.append(
            [
                "All classes are fully covered by existing allocations.",
                "",
                "",
                "",
                0,
            ]
        )
        row_index = summary_sheet.max_row
        summary_sheet.cell(row=row_index, column=1).fill = full_load_fill
        summary_sheet.cell(row=row_index, column=5).fill = full_load_fill

    for column_key, width in {
        "A": 34,
        "B": 22,
        "C": 20,
        "D": 20,
        "E": 18,
        "F": 18,
        "G": 20,
        "H": 42,
    }.items():
        summary_sheet.column_dimensions[column_key].width = width

    for row_index in range(1, summary_sheet.max_row + 1):
        for col_index in range(1, 9):
            cell = summary_sheet.cell(row=row_index, column=col_index)
            if cell.value is None:
                continue
            if row_index <= 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if col_index in {2, 3, 4, 5, 6, 7}:
                    if not getattr(cell.fill, "fill_type", None):
                        cell.fill = neutral_fill

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


REPORT_EXPORT_SECTION_CHOICES = {"full", "summary", "hiring", "subjects", "teachers"}


def _normalize_report_export_section(section: str = "full") -> str:
    normalized = str(section or "full").strip().lower()
    return normalized if normalized in REPORT_EXPORT_SECTION_CHOICES else "full"


def _build_report_export_filename(
    branch_name: str,
    academic_year_name: str,
    extension: str,
    section: str = "full",
) -> str:
    def _sanitize(text_value: str) -> str:
        normalized = re.sub(r"[^A-Za-z0-9]+", "-", str(text_value or "").strip())
        return normalized.strip("-").lower() or "scope"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_branch = _sanitize(branch_name)
    safe_year = _sanitize(academic_year_name)
    safe_section = _sanitize(_normalize_report_export_section(section))
    safe_extension = str(extension or "xlsx").strip().lower().lstrip(".") or "xlsx"
    return f"tis_analytical_report_{safe_branch}_{safe_year}_{safe_section}_{timestamp}.{safe_extension}"


def _load_saved_or_auto_hiring_plan(
    db: Session,
    user,
    branch_id: int,
    academic_year_id: int,
    auto_payload: dict,
) -> dict:
    plan_payload = _normalize_hiring_plan_payload(auto_payload or {})
    source = "System suggested plan"
    updated_at = None

    draft = db.query(models.HiringPlanDraft).filter(
        models.HiringPlanDraft.branch_id == int(branch_id),
        models.HiringPlanDraft.academic_year_id == int(academic_year_id),
        models.HiringPlanDraft.user_id == int(user.id),
    ).first() if branch_id and academic_year_id and getattr(user, "id", None) else None

    if draft:
        try:
            saved_payload = json.loads(str(draft.plan_json or "{}"))
        except json.JSONDecodeError:
            saved_payload = {}
        try:
            saved_version = int(saved_payload.get("pool_logic_version", 0) or 0)
        except (TypeError, ValueError):
            saved_version = 0

        if saved_version >= HIRING_PLAN_POOL_LOGIC_VERSION:
            plan_payload = _normalize_hiring_plan_payload(saved_payload)
            plan_payload["locked"] = True
            source = "Saved hiring plan"
            updated_at = draft.updated_at
        else:
            source = "System suggested plan (saved plan was outdated)"

    return {
        "plan": plan_payload,
        "source": source,
        "updated_at": updated_at,
        "warnings": _collect_hiring_plan_warnings(plan_payload),
    }


def _summarize_hiring_plan_for_export(plan_payload: dict, report_summary: dict) -> dict:
    total_uncovered_hours = int(report_summary.get("total_uncovered_hours", 0) or 0)
    profile_rows = []
    item_rows = []
    unassigned_rows = []

    for profile in plan_payload.get("profiles", []) or []:
        items = list(profile.get("items", []) or [])
        total_hours = sum(int(item.get("hours", 0) or 0) for item in items)
        block_size = max(
            1,
            int(profile.get("block_size_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS),
        )
        max_hours = max(
            block_size,
            int(profile.get("max_hours", block_size) or block_size),
        )
        full_blocks = total_hours // block_size
        remainder_hours = total_hours % block_size
        coverage_pct = round((total_hours / max_hours) * 100) if max_hours else 0
        profile_name = str(profile.get("name", "Proposed Pool") or "Proposed Pool")

        profile_row = {
            "id": str(profile.get("id", "")),
            "name": profile_name,
            "group_key": str(profile.get("group_key", "") or ""),
            "accent_color": str(profile.get("accent_color", "#0A4EA3") or "#0A4EA3"),
            "assignment_note": str(profile.get("assignment_note", "") or ""),
            "total_hours": total_hours,
            "max_hours": max_hours,
            "block_size_hours": block_size,
            "full_blocks": full_blocks,
            "remainder_hours": remainder_hours,
            "remaining_capacity_hours": max(max_hours - total_hours, 0),
            "coverage_pct": min(max(coverage_pct, 0), 100),
            "subject_count": len(items),
            "subjects_label": ", ".join(
                f"{item.get('subject_name', 'Subject')} ({int(item.get('hours', 0) or 0)}h)"
                for item in items
            ),
            "is_manual": bool(profile.get("is_manual", False)),
        }
        profile_rows.append(profile_row)

        for item in items:
            hours = int(item.get("hours", 0) or 0)
            if hours <= 0:
                continue
            item_rows.append(
                {
                    "pool_name": profile_name,
                    "pool_group": profile_row["group_key"],
                    "pool_accent_color": profile_row["accent_color"],
                    "subject_name": str(item.get("subject_name", "Subject") or "Subject"),
                    "subject_code": str(item.get("subject_code", "") or ""),
                    "family": str(item.get("family", "") or ""),
                    "subject_color": str(item.get("subject_color", profile_row["accent_color"]) or profile_row["accent_color"]),
                    "hours": hours,
                }
            )

    for item in plan_payload.get("unassigned_items", []) or []:
        hours = int(item.get("hours", 0) or 0)
        if hours <= 0:
            continue
        unassigned_rows.append(
            {
                "subject_name": str(item.get("subject_name", "Subject") or "Subject"),
                "subject_code": str(item.get("subject_code", "") or ""),
                "family": str(item.get("family", "") or ""),
                "subject_color": str(item.get("subject_color", "#B91C1C") or "#B91C1C"),
                "hours": hours,
            }
        )

    planned_hours = sum(row["total_hours"] for row in profile_rows)
    planned_full_blocks = sum(row["full_blocks"] for row in profile_rows)
    unassigned_hours = sum(row["hours"] for row in unassigned_rows)
    planned_coverage_pct = (
        round((planned_hours / total_uncovered_hours) * 100)
        if total_uncovered_hours > 0
        else 100
    )

    return {
        "profiles": profile_rows,
        "items": item_rows,
        "unassigned_items": unassigned_rows,
        "summary": {
            "total_uncovered_hours": total_uncovered_hours,
            "planned_pool_hours": planned_hours,
            "planned_full_blocks": planned_full_blocks,
            "unassigned_hours": unassigned_hours,
            "planned_coverage_pct": max(0, min(100, planned_coverage_pct)),
            "global_full_hires": int(report_summary.get("total_new_teachers_required", 0) or 0),
            "global_remainder_hours": int(report_summary.get("remaining_uncovered_hours_after_hires", 0) or 0),
        },
    }


def _format_hiring_block_label(full_blocks: int, remainder_hours: int) -> str:
    full_blocks = int(full_blocks or 0)
    remainder_hours = int(remainder_hours or 0)
    label = f"{full_blocks} full hire{'s' if full_blocks != 1 else ''}"
    if remainder_hours > 0:
        label = f"{label} + {remainder_hours}h remainder"
    return label


def _build_subject_pool_distribution_rows(report_subject_rows: list[dict]) -> dict:
    pool_definitions = [
        "english_pool",
        "general_science_pool",
        "arabic_pool",
        "math_pool",
        "physical_education",
        "homeroom_pool",
    ]
    pool_sort_order = {
        group_key: index
        for index, group_key in enumerate(pool_definitions)
    }
    fallback_pool_key = "specialist_pool"
    pool_map: dict[str, dict[str, Any]] = {}

    for row in report_subject_rows or []:
        required_hours = int(row.get("required_hours", 0) or 0)
        allocated_hours = int(row.get("allocated_hours", 0) or 0)
        remaining_hours = int(row.get("remaining_hours", 0) or 0)
        if required_hours <= 0 and allocated_hours <= 0 and remaining_hours <= 0:
            continue

        family = _detect_hiring_subject_family(row)
        raw_group_key = HIRING_COMPATIBILITY_GROUPS.get(family, f"single_{family}")
        group_key = _normalize_hiring_pool_group_key(raw_group_key, family)
        if not group_key or group_key.startswith("single_"):
            group_key = fallback_pool_key

        group_label = HIRING_GROUP_LABELS.get(group_key)
        if not group_label:
            group_label = "Specialist / Other Pool"

        accent_color = HIRING_POOL_ACCENT_COLORS.get(
            group_key,
            str(row.get("subject_color", "#64748B") or "#64748B"),
        )
        pool = pool_map.setdefault(
            group_key,
            {
                "group_key": group_key,
                "group_label": group_label,
                "accent_color": accent_color,
                "subject_count": 0,
                "family_labels": [],
                "family_keys": [],
                "grade_labels": set(),
                "total_required_hours": 0,
                "total_current_hours": 0,
                "total_new_hours": 0,
                "allocated_hours": 0,
                "remaining_hours": 0,
                "subjects": [],
            },
        )

        family_label = HIRING_FAMILY_LABELS.get(family, family.replace("_", " ").title())
        if family not in pool["family_keys"]:
            pool["family_keys"].append(family)
            pool["family_labels"].append(family_label)

        grade_labels = [
            str(item)
            for item in row.get("grades", []) or []
            if str(item).strip()
        ]
        pool["grade_labels"].update(grade_labels)
        pool["subject_count"] += int(row.get("effective_subject_count", 1) or 1)
        pool["total_required_hours"] += required_hours
        pool["total_current_hours"] += int(row.get("required_current_hours", 0) or 0)
        pool["total_new_hours"] += int(row.get("required_new_hours", 0) or 0)
        pool["allocated_hours"] += allocated_hours
        pool["remaining_hours"] += remaining_hours

        subject_full_hires = remaining_hours // REPORT_STANDARD_MAX_HOURS
        subject_remainder_hours = remaining_hours % REPORT_STANDARD_MAX_HOURS
        subject_original_teacher_need_count = (
            math.ceil(required_hours / REPORT_STANDARD_MAX_HOURS)
            if required_hours > 0
            else 0
        )
        pool["subjects"].append(
            {
                "subject_key": str(row.get("subject_key", "") or ""),
                "subject_name": str(row.get("subject_name", "Subject") or "Subject"),
                "subject_code": str(row.get("subject_code", "") or ""),
                "subject_color": str(row.get("subject_color", accent_color) or accent_color),
                "family": family,
                "family_label": family_label,
                "grades": grade_labels,
                "grades_label": ", ".join(
                    "KG" if grade == "KG" else f"G{grade}"
                    for grade in grade_labels
                ) or "-",
                "required_hours": required_hours,
                "required_current_hours": int(row.get("required_current_hours", 0) or 0),
                "required_new_hours": int(row.get("required_new_hours", 0) or 0),
                "allocated_hours": allocated_hours,
                "remaining_hours": remaining_hours,
                "coverage_percentage": int(row.get("coverage_percentage", 0) or 0),
                "original_teacher_need_count": subject_original_teacher_need_count,
                "assigned_teacher_count": int(row.get("assigned_teacher_count", 0) or 0),
                "assigned_teacher_labels": list(row.get("assigned_teacher_labels", []) or []),
                "hire_full_count": subject_full_hires,
                "hire_remainder_hours": subject_remainder_hours,
                "hire_label": _format_hiring_block_label(
                    subject_full_hires,
                    subject_remainder_hours,
                ),
            }
        )

    pool_rows = []
    for pool in pool_map.values():
        group_key = pool["group_key"]
        family_order = HIRING_POOL_FAMILY_ORDER.get(group_key, [])
        family_order_map = {
            family: index
            for index, family in enumerate(family_order)
        }
        pool["subjects"].sort(
            key=lambda item: (
                family_order_map.get(item.get("family", ""), 999),
                -int(item.get("remaining_hours", 0) or 0),
                str(item.get("subject_name", "") or "").lower(),
            )
        )
        required_hours = int(pool["total_required_hours"] or 0)
        allocated_hours = int(pool["allocated_hours"] or 0)
        remaining_hours = int(pool["remaining_hours"] or 0)
        coverage_percentage = (
            round((allocated_hours / required_hours) * 100)
            if required_hours > 0
            else 0
        )
        full_hire_count = remaining_hours // REPORT_STANDARD_MAX_HOURS
        hire_remainder_hours = remaining_hours % REPORT_STANDARD_MAX_HOURS
        required_full_loads = required_hours // REPORT_STANDARD_MAX_HOURS
        required_remainder_hours = required_hours % REPORT_STANDARD_MAX_HOURS
        original_teacher_need_count = (
            math.ceil(required_hours / REPORT_STANDARD_MAX_HOURS)
            if required_hours > 0
            else 0
        )
        covered_width_pct = (
            round((allocated_hours / required_hours) * 100, 1)
            if required_hours > 0
            else 0
        )
        open_width_pct = (
            round((remaining_hours / required_hours) * 100, 1)
            if required_hours > 0
            else 0
        )
        hiring_coverage_items = [
            {
                "subject_name": item["subject_name"],
                "subject_code": item["subject_code"],
                "subject_key": item["subject_key"],
                "subject_color": item["subject_color"],
                "family": item["family"],
                "hours": int(item.get("remaining_hours", 0) or 0),
            }
            for item in pool["subjects"]
            if int(item.get("remaining_hours", 0) or 0) > 0
        ]
        if remaining_hours > 0:
            recommendation_note = _build_hiring_pool_reason(
                group_key,
                hiring_coverage_items,
                remaining_hours,
                full_hire_count,
                hire_remainder_hours,
            )
        else:
            recommendation_note = (
                f"{pool['group_label']} is fully covered by the current assignments; "
                "no new full hire is suggested from uncovered hours."
            )

        pool_rows.append(
            {
                **pool,
                "grade_labels": sorted(pool["grade_labels"], key=_grade_sort_key),
                "family_label": ", ".join(pool["family_labels"]) or "Specialist",
                "coverage_percentage": min(max(coverage_percentage, 0), 100),
                "full_hire_count": full_hire_count,
                "hire_remainder_hours": hire_remainder_hours,
                "minimum_teacher_blocks": (
                    math.ceil(remaining_hours / REPORT_STANDARD_MAX_HOURS)
                    if remaining_hours > 0
                    else 0
                ),
                "hire_label": _format_hiring_block_label(
                    full_hire_count,
                    hire_remainder_hours,
                ),
                "required_full_loads": required_full_loads,
                "required_remainder_hours": required_remainder_hours,
                "original_teacher_need_count": original_teacher_need_count,
                "covered_width_pct": min(max(covered_width_pct, 0), 100),
                "open_width_pct": min(max(open_width_pct, 0), 100),
                "required_load_label": _format_hiring_block_label(
                    required_full_loads,
                    required_remainder_hours,
                ),
                "recommendation_note": recommendation_note,
            }
        )

    pool_rows.sort(
        key=lambda pool: (
            pool_sort_order.get(pool.get("group_key", ""), 999),
            -int(pool.get("remaining_hours", 0) or 0),
            str(pool.get("group_label", "") or "").lower(),
        )
    )

    total_required_hours = sum(int(pool.get("total_required_hours", 0) or 0) for pool in pool_rows)
    total_allocated_hours = sum(int(pool.get("allocated_hours", 0) or 0) for pool in pool_rows)
    total_remaining_hours = sum(int(pool.get("remaining_hours", 0) or 0) for pool in pool_rows)
    full_hire_count = total_remaining_hours // REPORT_STANDARD_MAX_HOURS
    hire_remainder_hours = total_remaining_hours % REPORT_STANDARD_MAX_HOURS
    original_teacher_need_count = sum(
        int(pool.get("original_teacher_need_count", 0) or 0)
        for pool in pool_rows
    )

    return {
        "rows": pool_rows,
        "summary": {
            "pool_count": len(pool_rows),
            "subject_count": sum(int(pool.get("subject_count", 0) or 0) for pool in pool_rows),
            "total_required_hours": total_required_hours,
            "total_allocated_hours": total_allocated_hours,
            "total_remaining_hours": total_remaining_hours,
            "original_teacher_need_count": original_teacher_need_count,
            "full_hire_count": full_hire_count,
            "hire_remainder_hours": hire_remainder_hours,
            "hire_label": _format_hiring_block_label(
                full_hire_count,
                hire_remainder_hours,
            ),
        },
    }


def _build_current_report_package(db: Session, user) -> dict:
    scoped_branch_id = getattr(user, "scope_branch_id", user.branch_id)
    scoped_academic_year_id = getattr(
        user,
        "scope_academic_year_id",
        user.academic_year_id,
    )

    branch = db.query(models.Branch).filter(
        models.Branch.id == scoped_branch_id
    ).first()
    academic_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.id == scoped_academic_year_id
    ).first()

    subjects_query = db.query(models.Subject).filter(
        models.Subject.branch_id == scoped_branch_id,
        models.Subject.academic_year_id == scoped_academic_year_id,
    )
    teachers_query = db.query(models.Teacher).filter(
        models.Teacher.branch_id == scoped_branch_id,
        models.Teacher.academic_year_id == scoped_academic_year_id,
    )
    planning_sections_query = db.query(models.PlanningSection).filter(
        models.PlanningSection.branch_id == scoped_branch_id,
        models.PlanningSection.academic_year_id == scoped_academic_year_id,
    )

    subjects_dashboard_rows = subjects_query.order_by(
        models.Subject.grade.asc(),
        models.Subject.subject_code.asc(),
    ).all()
    for subject in subjects_dashboard_rows:
        bundle_subject_labels = list(
            get_homeroom_bundle_subject_labels(
                subject_code=subject.subject_code or "",
                subject_name=subject.subject_name or "",
                weekly_hours=subject.weekly_hours,
                grade_label=_normalize_grade_label(subject.grade),
            )
        )
        setattr(
            subject,
            "effective_subject_count",
            get_effective_subject_count(
                subject_code=subject.subject_code or "",
                subject_name=subject.subject_name or "",
                weekly_hours=subject.weekly_hours,
                grade_label=_normalize_grade_label(subject.grade),
            ),
        )
        setattr(subject, "homeroom_bundle_subject_labels", bundle_subject_labels)

    subject_count = sum(
        int(getattr(subject, "effective_subject_count", 1) or 1)
        for subject in subjects_dashboard_rows
    )
    teacher_count = teachers_query.count()
    planning_sections = planning_sections_query.all()
    planning_total_sections = len(planning_sections)
    planning_current_sections_count = sum(
        1
        for section in planning_sections
        if str(section.class_status or "").strip().lower() == "current"
    )
    planning_new_sections_count = sum(
        1
        for section in planning_sections
        if str(section.class_status or "").strip().lower() == "new"
    )
    teachers_for_reporting = teachers_query.order_by(models.Teacher.id.asc()).all()
    subject_hours_by_grade = {}
    for subject in subjects_dashboard_rows:
        grade_label = _normalize_grade_label(getattr(subject, "grade", None))
        if not grade_label:
            continue
        subject_hours_by_grade[grade_label] = (
            subject_hours_by_grade.get(grade_label, 0)
            + int(subject.weekly_hours or 0)
        )
    planning_total_allocated_hours = sum(
        subject_hours_by_grade.get(_normalize_grade_label(section.grade_level), 0)
        for section in planning_sections
    )

    planning_section_ids = [
        section.id
        for section in planning_sections
        if getattr(section, "id", None)
    ]
    section_assignments = []
    if planning_section_ids:
        section_assignments = db.query(models.TeacherSectionAssignment).filter(
            models.TeacherSectionAssignment.planning_section_id.in_(planning_section_ids)
        ).all()

    reporting_context = _build_reporting_context_from_section_assignments(
        db=db,
        subjects=subjects_dashboard_rows,
        planning_sections=planning_sections,
        teachers=teachers_for_reporting,
        section_assignments=section_assignments,
    )
    allocation_data = _build_report_class_allocation_data_from_section_assignments(
        db=db,
        subjects=subjects_dashboard_rows,
        planning_sections=planning_sections,
        teachers=teachers_for_reporting,
        reporting_context=reporting_context,
        section_assignments=section_assignments,
    )

    subject_section_map = allocation_data.get("subject_section_map", {})
    report_subject_rows = [
        {
            **row,
            **subject_section_map.get(row["subject_key"], {}),
        }
        for row in reporting_context.get("subject_rows", [])
    ]

    underloaded_teacher_map = {
        row.get("teacher_pk"): row
        for row in allocation_data.get("underloaded_teacher_rows", [])
    }
    report_teacher_rows = []
    for row in reporting_context.get("teacher_rows", []):
        underloaded_row = underloaded_teacher_map.get(row.get("teacher_pk"), {})
        report_teacher_rows.append(
            {
                **row,
                "is_underloaded": int(row.get("remaining_capacity_hours", 0)) > 0,
                "recommended_absorption_hours": int(
                    underloaded_row.get("recommended_absorption_hours", 0)
                ),
                "recommended_assignment_labels": list(
                    underloaded_row.get("recommended_assignment_labels", [])
                ),
                "projected_allocated_hours": int(
                    underloaded_row.get(
                        "projected_allocated_hours",
                        row.get("expected_allocated_hours", 0),
                    )
                ),
                "projected_remaining_capacity_hours": int(
                    underloaded_row.get(
                        "projected_remaining_capacity_hours",
                        row.get("remaining_capacity_hours", 0),
                    )
                ),
            }
        )

    report_subject_rows, report_summary = _decorate_staffing_report_rows(
        report_subject_rows,
        reporting_context.get("summary", {}),
    )
    report_gap_rows = [
        row
        for row in report_subject_rows
        if int(row.get("remaining_hours", 0) or 0) > 0
    ]
    report_summary["subjects_with_gaps"] = len(report_gap_rows)
    report_summary["underloaded_teachers"] = sum(
        1 for row in report_teacher_rows if row.get("is_underloaded")
    )
    report_summary["underloaded_teachers_with_recommendations"] = sum(
        1
        for row in report_teacher_rows
        if int(row.get("recommended_absorption_hours", 0) or 0) > 0
    )
    report_summary["recommended_internal_absorption_hours"] = sum(
        int(row.get("recommended_absorption_hours", 0) or 0)
        for row in report_teacher_rows
    )

    report_visuals = _build_dashboard_report_visuals(
        report_summary=report_summary,
        report_subject_rows=report_subject_rows,
        report_grade_rows=reporting_context.get("grade_rows", []),
        planning_current_sections_count=planning_current_sections_count,
        planning_new_sections_count=planning_new_sections_count,
    )
    report_subject_card_rows = sorted(
        report_subject_rows,
        key=lambda row: (
            -int(row.get("coverage_percentage", 0) or 0),
            int(row.get("remaining_hours", 0) or 0),
            str(row.get("subject_name", "") or "").lower(),
        ),
    )
    hiring_plan_editor_auto_payload = _build_hiring_plan_editor_payload(
        report_summary=report_summary,
        report_subject_rows=report_subject_rows,
    )
    hiring_plan_state = _load_saved_or_auto_hiring_plan(
        db=db,
        user=user,
        branch_id=scoped_branch_id,
        academic_year_id=scoped_academic_year_id,
        auto_payload=hiring_plan_editor_auto_payload,
    )
    hiring_plan_export = _summarize_hiring_plan_for_export(
        hiring_plan_state["plan"],
        report_summary,
    )
    subject_pool_distribution = _build_subject_pool_distribution_rows(
        report_subject_rows
    )

    branch_name = branch.name if branch else "Not assigned"
    academic_year_name = academic_year.year_name if academic_year else "Not assigned"

    return {
        "branch_name": branch_name,
        "academic_year_name": academic_year_name,
        "scoped_branch_id": scoped_branch_id,
        "scoped_academic_year_id": scoped_academic_year_id,
        "subject_count": subject_count,
        "teacher_count": teacher_count,
        "planning_total_sections": planning_total_sections,
        "planning_current_sections_count": planning_current_sections_count,
        "planning_new_sections_count": planning_new_sections_count,
        "planning_total_allocated_hours": planning_total_allocated_hours,
        "subjects_dashboard_rows": subjects_dashboard_rows,
        "planning_sections": planning_sections,
        "section_assignments": section_assignments,
        "reporting_context": reporting_context,
        "allocation_data": allocation_data,
        "report_summary": report_summary,
        "report_subject_count": sum(
            int(row.get("effective_subject_count", 1) or 1)
            for row in report_subject_rows
        ),
        "report_subject_rows": report_subject_rows,
        "report_subject_card_rows": report_subject_card_rows,
        "report_gap_rows": report_gap_rows,
        "report_teacher_rows": report_teacher_rows,
        "report_underloaded_teacher_rows": allocation_data.get(
            "underloaded_teacher_rows",
            [],
        ),
        "report_grade_rows": reporting_context.get("grade_rows", []),
        "report_visuals": report_visuals,
        "hiring_plan_editor_auto_payload": hiring_plan_editor_auto_payload,
        "hiring_plan_export": hiring_plan_export,
        "subject_pool_distribution_rows": subject_pool_distribution["rows"],
        "subject_pool_distribution_summary": subject_pool_distribution["summary"],
        "hiring_plan_source": hiring_plan_state["source"],
        "hiring_plan_updated_at": hiring_plan_state["updated_at"],
        "hiring_plan_warnings": hiring_plan_state["warnings"],
    }


def _excel_color(value: str, fallback: str = "0A4EA3") -> str:
    cleaned = str(value or "").strip().lstrip("#")
    if re.fullmatch(r"[0-9A-Fa-f]{6}", cleaned):
        return cleaned.upper()
    return fallback


def _make_fill(value: str, fallback: str = "F8FAFC") -> PatternFill:
    color = _excel_color(value, fallback)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _style_report_sheet(sheet, tab_color: str = "0A4EA3"):
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = _excel_color(tab_color)


def _apply_report_title(sheet, title: str, subtitle: str, last_column: int = 8):
    last_letter = get_column_letter(max(1, last_column))
    sheet.merge_cells(f"A1:{last_letter}1")
    sheet.merge_cells(f"A2:{last_letter}2")
    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=18, color="0A4EA3")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(size=10, color="536782")
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 26
    sheet.row_dimensions[2].height = 32


def _add_report_table(sheet, start_row: int, end_row: int, end_column: int, name: str):
    if end_row <= start_row:
        return
    safe_name = re.sub(r"[^A-Za-z0-9_]", "_", name).strip("_") or "ReportTable"
    safe_name = f"{safe_name[:220]}_{start_row}"
    table_ref = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    table = Table(displayName=safe_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _format_report_header_row(sheet, row_index: int, total_columns: int):
    _apply_excel_header_style(sheet, row_index, total_columns)
    for col_index in range(1, total_columns + 1):
        cell = sheet.cell(row=row_index, column=col_index)
        cell.border = Border(bottom=Side(style="thin", color="B8C6DA"))


def _write_kpi_cards(sheet, start_row: int, cards: list[tuple[str, Any, str]], columns: int = 4) -> int:
    row = start_row
    card_width = 2
    for index, (label, value, note) in enumerate(cards):
        col = 1 + (index % columns) * card_width
        if index and index % columns == 0:
            row += 4
        sheet.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + card_width - 1,
        )
        sheet.merge_cells(
            start_row=row + 1,
            start_column=col,
            end_row=row + 1,
            end_column=col + card_width - 1,
        )
        sheet.merge_cells(
            start_row=row + 2,
            start_column=col,
            end_row=row + 2,
            end_column=col + card_width - 1,
        )
        label_cell = sheet.cell(row=row, column=col, value=label)
        value_cell = sheet.cell(row=row + 1, column=col, value=value)
        note_cell = sheet.cell(row=row + 2, column=col, value=note)
        for r in range(row, row + 3):
            for c in range(col, col + card_width):
                sheet.cell(row=r, column=c).fill = _make_fill("EDF4FF")
                sheet.cell(row=r, column=c).border = Border(
                    left=Side(style="thin", color="D8E2F0"),
                    right=Side(style="thin", color="D8E2F0"),
                    top=Side(style="thin", color="D8E2F0"),
                    bottom=Side(style="thin", color="D8E2F0"),
                )
        label_cell.font = Font(bold=True, size=9, color="536782")
        value_cell.font = Font(bold=True, size=20, color="0A4EA3")
        note_cell.font = Font(size=9, color="536782")
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 4


def _build_summary_overview_sheet(workbook: Workbook, report_package: dict):
    summary = report_package["report_summary"]
    sheet = workbook.active
    sheet.title = "Summary Overview"
    _style_report_sheet(sheet, "0A4EA3")
    _apply_report_title(
        sheet,
        "TIS Analytical Staffing Report",
        (
            f"Branch: {report_package['branch_name']} | Academic Year: "
            f"{report_package['academic_year_name']} | Generated: "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ),
        last_column=8,
    )
    for col, width in {
        "A": 24,
        "B": 16,
        "C": 18,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "J": 18,
        "K": 16,
    }.items():
        sheet.column_dimensions[col].width = width

    next_row = _write_kpi_cards(
        sheet,
        4,
        [
            ("Total Demand", f"{summary.get('total_required_hours', 0)}h", "All planned section-subject hours"),
            ("Covered Hours", f"{summary.get('total_allocated_hours', 0)}h", "Matched current teacher allocations"),
            ("Uncovered Hours", f"{summary.get('total_uncovered_hours', 0)}h", "Remaining weekly hours"),
            ("Whole New Hires", summary.get("total_new_teachers_required", 0), "24h full teacher blocks"),
            ("Coverage", f"{summary.get('coverage_percentage', 0)}%", "Covered vs total demand"),
            ("Remaining After Hires", f"{summary.get('remaining_uncovered_hours_after_hires', 0)}h", "Remainder after whole hires"),
            ("Underloaded Teachers", summary.get("underloaded_teachers", 0), "Teachers with spare capacity"),
            ("Priority Gap Hours", f"{summary.get('priority_gap_hours', 0)}h", "Priority uncovered demand"),
        ],
    )

    metric_start = next_row + 1
    metrics = [
        ("Existing Teachers", summary.get("total_existing_teachers", 0)),
        ("Total Existing Capacity", f"{summary.get('total_existing_capacity_hours', 0)}h"),
        ("Unused Existing Capacity", f"{summary.get('unused_existing_capacity_hours', 0)}h"),
        ("Teachers Utilized", summary.get("teachers_utilized", 0)),
        ("Full Load Teachers", summary.get("teachers_full_load", 0)),
        ("Subjects With Gaps", summary.get("subjects_with_gaps", 0)),
        ("Priority Subjects With Gaps", summary.get("priority_subjects_with_gaps", 0)),
        ("Homeroom Default Coverage", f"{summary.get('homeroom_default_coverage_hours', 0)}h"),
        ("Hiring Plan Source", report_package.get("hiring_plan_source", "System suggested plan")),
    ]
    sheet.cell(row=metric_start, column=1, value="KPI Details").font = Font(
        bold=True,
        size=12,
        color="0A4EA3",
    )
    sheet.append(["Metric", "Value"])
    header_row = sheet.max_row
    _format_report_header_row(sheet, header_row, 2)
    for label, value in metrics:
        sheet.append([label, value])
    _add_report_table(sheet, header_row, sheet.max_row, 2, "SummaryKpis")

    chart_data_row = 4
    sheet["J2"] = "Coverage Chart"
    sheet["J2"].font = Font(bold=True, color="0A4EA3")
    sheet["J3"] = "Metric"
    sheet["K3"] = "Hours"
    chart_values = [
        ("Required", summary.get("total_required_hours", 0)),
        ("Covered", summary.get("total_allocated_hours", 0)),
        ("Uncovered", summary.get("total_uncovered_hours", 0)),
    ]
    for offset, (label, value) in enumerate(chart_values):
        sheet.cell(row=chart_data_row + offset, column=10, value=label)
        sheet.cell(row=chart_data_row + offset, column=11, value=value)

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 10
    bar_chart.title = "Demand vs Coverage"
    bar_chart.y_axis.title = "Metric"
    bar_chart.x_axis.title = "Hours"
    bar_chart.add_data(
        Reference(sheet, min_col=11, min_row=3, max_row=chart_data_row + len(chart_values) - 1),
        titles_from_data=True,
    )
    bar_chart.set_categories(
        Reference(sheet, min_col=10, min_row=chart_data_row, max_row=chart_data_row + len(chart_values) - 1)
    )
    bar_chart.height = 7
    bar_chart.width = 12
    sheet.add_chart(bar_chart, "J8")

    sheet["J20"] = "Coverage Mix"
    sheet["J20"].font = Font(bold=True, color="0A4EA3")
    sheet["J21"] = "Status"
    sheet["K21"] = "Hours"
    sheet["J22"] = "Covered"
    sheet["K22"] = summary.get("total_allocated_hours", 0)
    sheet["J23"] = "Uncovered"
    sheet["K23"] = summary.get("total_uncovered_hours", 0)
    pie_chart = PieChart()
    pie_chart.title = "Covered vs Uncovered"
    pie_chart.add_data(Reference(sheet, min_col=11, min_row=21, max_row=23), titles_from_data=True)
    pie_chart.set_categories(Reference(sheet, min_col=10, min_row=22, max_row=23))
    pie_chart.height = 7
    pie_chart.width = 9
    sheet.add_chart(pie_chart, "J25")


def _build_hiring_plan_sheet(workbook: Workbook, report_package: dict):
    plan = report_package["hiring_plan_export"]
    plan_summary = plan["summary"]
    sheet = workbook.create_sheet("Recommended Hiring Plan")
    _style_report_sheet(sheet, "C79A14")
    _apply_report_title(
        sheet,
        "Recommended Hiring Plan",
        (
            f"{report_package.get('hiring_plan_source', 'System suggested plan')} | "
            f"{plan_summary.get('global_full_hires', 0)} full hires | "
            f"{plan_summary.get('global_remainder_hours', 0)}h remainder"
        ),
        last_column=9,
    )
    for col, width in {
        "A": 24,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 16,
        "F": 16,
        "G": 18,
        "H": 46,
        "I": 16,
        "K": 24,
        "L": 14,
    }.items():
        sheet.column_dimensions[col].width = width

    next_row = _write_kpi_cards(
        sheet,
        4,
        [
            ("Pool Hours", f"{plan_summary.get('planned_pool_hours', 0)}h", "Uncovered hours assigned to pools"),
            ("Full Blocks In Pools", plan_summary.get("planned_full_blocks", 0), "Per-pool 24h blocks"),
            ("Unassigned Hours", f"{plan_summary.get('unassigned_hours', 0)}h", "Not placed in a pool"),
            ("Pool Coverage", f"{plan_summary.get('planned_coverage_pct', 0)}%", "Pool hours vs total uncovered"),
        ],
    )

    pool_header_row = next_row + 1
    sheet.cell(row=pool_header_row, column=1, value="Pools").font = Font(
        bold=True,
        size=12,
        color="0A4EA3",
    )
    headers = [
        "Pool",
        "Group",
        "Hours",
        "Max Hours",
        "Full 24h Blocks",
        "Remainder",
        "Progress %",
        "Subjects Inside Pool",
        "Note",
    ]
    sheet.append(headers)
    table_header_row = sheet.max_row
    _format_report_header_row(sheet, table_header_row, len(headers))
    for row in plan["profiles"]:
        sheet.append(
            [
                row["name"],
                row["group_key"],
                row["total_hours"],
                row["max_hours"],
                row["full_blocks"],
                row["remainder_hours"],
                row["coverage_pct"],
                row["subjects_label"],
                row["assignment_note"],
            ]
        )
        row_index = sheet.max_row
        for col_index in range(1, len(headers) + 1):
            sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                horizontal="left" if col_index in {1, 2, 8, 9} else "center",
                vertical="top",
                wrap_text=True,
            )
        sheet.cell(row=row_index, column=1).fill = _make_fill(row["accent_color"], "EDF4FF")
    if plan["profiles"]:
        progress_range = f"G{table_header_row + 1}:G{sheet.max_row}"
        sheet.conditional_formatting.add(
            progress_range,
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="0A4EA3"),
        )
        _add_report_table(sheet, table_header_row, sheet.max_row, len(headers), "HiringPools")

    item_header_row = sheet.max_row + 3
    sheet.cell(row=item_header_row, column=1, value="Subjects Inside Pools").font = Font(
        bold=True,
        size=12,
        color="0A4EA3",
    )
    item_headers = ["Pool", "Subject", "Code", "Family", "Hours"]
    sheet.append(item_headers)
    item_table_header = sheet.max_row
    _format_report_header_row(sheet, item_table_header, len(item_headers))
    for item in plan["items"]:
        sheet.append(
            [
                item["pool_name"],
                item["subject_name"],
                item["subject_code"],
                item["family"],
                item["hours"],
            ]
        )
        row_index = sheet.max_row
        sheet.cell(row=row_index, column=1).fill = _make_fill(item["pool_accent_color"], "EDF4FF")
        sheet.cell(row=row_index, column=2).fill = _make_fill(item["subject_color"], "EDF4FF")
    if plan["items"]:
        _add_report_table(sheet, item_table_header, sheet.max_row, len(item_headers), "HiringPoolSubjects")

    if plan["unassigned_items"]:
        unassigned_header_row = sheet.max_row + 3
        sheet.cell(row=unassigned_header_row, column=1, value="Unassigned Remaining Hours").font = Font(
            bold=True,
            size=12,
            color="B91C1C",
        )
        sheet.append(["Subject", "Code", "Family", "Hours"])
        unassigned_table_header = sheet.max_row
        _format_report_header_row(sheet, unassigned_table_header, 4)
        for item in plan["unassigned_items"]:
            sheet.append([item["subject_name"], item["subject_code"], item["family"], item["hours"]])
            sheet.cell(row=sheet.max_row, column=4).fill = _make_fill("FEE2E2")
        _add_report_table(sheet, unassigned_table_header, sheet.max_row, 4, "HiringUnassigned")

    if plan["profiles"]:
        chart_start_row = 4
        sheet["K2"] = "Pool Hours Chart"
        sheet["K2"].font = Font(bold=True, color="0A4EA3")
        sheet["K3"] = "Pool"
        sheet["L3"] = "Hours"
        for offset, row in enumerate(plan["profiles"]):
            sheet.cell(row=chart_start_row + offset, column=11, value=row["name"])
            sheet.cell(row=chart_start_row + offset, column=12, value=row["total_hours"])
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Hours by Pool"
        chart.add_data(
            Reference(sheet, min_col=12, min_row=3, max_row=chart_start_row + len(plan["profiles"]) - 1),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=11, min_row=chart_start_row, max_row=chart_start_row + len(plan["profiles"]) - 1)
        )
        chart.height = 7
        chart.width = 12
        sheet.add_chart(chart, "K8")


def _build_subjects_breakdown_sheet(workbook: Workbook, report_package: dict):
    rows = report_package["report_subject_rows"]
    sheet = workbook.create_sheet("Subjects Breakdown")
    _style_report_sheet(sheet, "0F7F7A")
    _apply_report_title(
        sheet,
        "Subjects Breakdown",
        "Demand, coverage, uncovered hours, section coverage, and staffing block requirements.",
        last_column=11,
    )
    widths = {
        "A": 26,
        "B": 14,
        "C": 18,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 18,
        "K": 44,
        "M": 24,
        "N": 14,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    headers = [
        "Subject",
        "Code",
        "Grades",
        "Current Hours",
        "New Hours",
        "Total Demand",
        "Covered",
        "Uncovered",
        "Coverage %",
        "24h Blocks",
        "Status / Note",
    ]
    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col_index, value=header)
    _format_report_header_row(sheet, header_row, len(headers))
    for row in rows:
        sheet.append(
            [
                row.get("subject_name", ""),
                row.get("subject_code", ""),
                ", ".join(str(item) for item in row.get("grades", [])),
                row.get("required_current_hours", 0),
                row.get("required_new_hours", 0),
                row.get("required_hours", 0),
                row.get("allocated_hours", 0),
                row.get("remaining_hours", 0),
                row.get("coverage_percentage", 0),
                row.get("teacher_requirement_blocks", row.get("additional_teachers_needed", 0)),
                row.get("staffing_status_label", "") or row.get("staffing_note", ""),
            ]
        )
        row_index = sheet.max_row
        subject_color = row.get("subject_color", "#EDF4FF")
        sheet.cell(row=row_index, column=1).fill = _make_fill(subject_color, "EDF4FF")
        sheet.cell(row=row_index, column=8).fill = (
            _make_fill("FEE2E2") if int(row.get("remaining_hours", 0) or 0) > 0 else _make_fill("DCFCE7")
        )
        sheet.cell(row=row_index, column=9).number_format = '0"%"'
        for col_index in range(1, len(headers) + 1):
            sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                horizontal="left" if col_index in {1, 2, 3, 11} else "center",
                vertical="top",
                wrap_text=True,
            )
    if rows:
        sheet.conditional_formatting.add(
            f"I{header_row + 1}:I{sheet.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="0F7F7A"),
        )
        _add_report_table(sheet, header_row, sheet.max_row, len(headers), "SubjectBreakdown")

    gap_rows = [
        row for row in rows if int(row.get("remaining_hours", 0) or 0) > 0
    ][:12]
    if gap_rows:
        chart_row = 4
        sheet["M2"] = "Top Uncovered Subjects"
        sheet["M2"].font = Font(bold=True, color="0A4EA3")
        sheet["M3"] = "Subject"
        sheet["N3"] = "Hours"
        for offset, row in enumerate(gap_rows):
            sheet.cell(row=chart_row + offset, column=13, value=row.get("subject_name", ""))
            sheet.cell(row=chart_row + offset, column=14, value=row.get("remaining_hours", 0))
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Uncovered Hours"
        chart.add_data(
            Reference(sheet, min_col=14, min_row=3, max_row=chart_row + len(gap_rows) - 1),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=13, min_row=chart_row, max_row=chart_row + len(gap_rows) - 1)
        )
        chart.height = 8
        chart.width = 12
        sheet.add_chart(chart, "M18")


def _build_teachers_overview_sheet(workbook: Workbook, report_package: dict):
    rows = report_package["report_teacher_rows"]
    sheet = workbook.create_sheet("Teachers Overview")
    _style_report_sheet(sheet, "7C3AED")
    _apply_report_title(
        sheet,
        "Teachers Overview",
        "Teacher load, assigned subjects, capacity, allocation, other-path hours, and remaining capacity.",
        last_column=11,
    )
    for col, width in {
        "A": 14,
        "B": 26,
        "C": 26,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 46,
        "M": 24,
        "N": 14,
    }.items():
        sheet.column_dimensions[col].width = width

    headers = [
        "Teacher ID",
        "Teacher Name",
        "Degree / Major",
        "Allocated Hours",
        "Capacity",
        "Load %",
        "Remaining Capacity",
        "Homeroom",
        "Primary",
        "Support",
        "Assigned Subjects",
    ]
    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col_index, value=header)
    _format_report_header_row(sheet, header_row, len(headers))
    for row in rows:
        capacity_hours = int(row.get("capacity_hours", REPORT_STANDARD_MAX_HOURS) or REPORT_STANDARD_MAX_HOURS)
        allocated_hours = int(row.get("expected_allocated_hours", 0) or 0)
        load_pct = round((allocated_hours / capacity_hours) * 100) if capacity_hours else 0
        sheet.append(
            [
                row.get("teacher_id", ""),
                row.get("teacher_name", ""),
                row.get("degree_major", ""),
                allocated_hours,
                capacity_hours,
                load_pct,
                row.get("remaining_capacity_hours", 0),
                row.get("homeroom_allocated_hours", 0),
                row.get("primary_allocated_hours", 0),
                row.get("support_allocated_hours", 0),
                ", ".join(row.get("allocation_labels", []) or row.get("subject_labels", [])),
            ]
        )
        row_index = sheet.max_row
        sheet.cell(row=row_index, column=6).number_format = '0"%"'
        sheet.cell(row=row_index, column=7).fill = (
            _make_fill("DCFCE7") if int(row.get("remaining_capacity_hours", 0) or 0) <= 0 else _make_fill("FEF3C7")
        )
        for col_index in range(1, len(headers) + 1):
            sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                horizontal="left" if col_index in {2, 3, 11} else "center",
                vertical="top",
                wrap_text=True,
            )
    if rows:
        sheet.conditional_formatting.add(
            f"F{header_row + 1}:F{sheet.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="7C3AED"),
        )
        _add_report_table(sheet, header_row, sheet.max_row, len(headers), "TeachersOverview")

    top_load_rows = rows[:12]
    if top_load_rows:
        chart_row = 4
        sheet["M2"] = "Teacher Load Chart"
        sheet["M2"].font = Font(bold=True, color="0A4EA3")
        sheet["M3"] = "Teacher"
        sheet["N3"] = "Hours"
        for offset, row in enumerate(top_load_rows):
            sheet.cell(row=chart_row + offset, column=13, value=row.get("teacher_name", ""))
            sheet.cell(row=chart_row + offset, column=14, value=row.get("expected_allocated_hours", 0))
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Allocated Hours by Teacher"
        chart.add_data(
            Reference(sheet, min_col=14, min_row=3, max_row=chart_row + len(top_load_rows) - 1),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=13, min_row=chart_row, max_row=chart_row + len(top_load_rows) - 1)
        )
        chart.height = 8
        chart.width = 12
        sheet.add_chart(chart, "M18")


def _build_class_matrix_sheet(workbook: Workbook, report_package: dict):
    allocation_data = report_package["allocation_data"]
    class_rows = allocation_data.get("class_rows", [])
    teacher_matrix_rows = allocation_data.get("teacher_matrix_rows", [])
    sheet = workbook.create_sheet("Class Allocation Matrix")
    _style_report_sheet(sheet, "1D4ED8")
    _apply_report_title(
        sheet,
        "Class Allocation Matrix",
        "Teacher-to-class coverage generated from current section assignments and homeroom-default allocation.",
        last_column=max(6 + len(class_rows), 8),
    )

    headers = [
        "Teacher ID",
        "Teacher Name",
        "Assigned Hours",
        "Remaining Capacity",
        "Assigned Subject",
        "Support Subject",
    ] + [row["class_label"] for row in class_rows]
    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col_index, value=header)
    _format_report_header_row(sheet, header_row, len(headers))
    class_order = [row["class_key"] for row in class_rows]
    for row_data in teacher_matrix_rows:
        row_values = [
            row_data.get("teacher_id", ""),
            row_data.get("teacher_name", ""),
            row_data.get("expected_allocated_hours", 0),
            row_data.get("remaining_capacity_hours", 0),
            row_data.get("primary_subject_label", ""),
            row_data.get("support_subject_label", ""),
        ] + [
            row_data.get("class_cells", {}).get(class_key, "")
            for class_key in class_order
        ]
        sheet.append(row_values)
        row_index = sheet.max_row
        for col_index in range(1, len(headers) + 1):
            sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                horizontal="left" if col_index in {2, 5, 6} or col_index >= 7 else "center",
                vertical="top",
                wrap_text=True,
            )
        sheet.cell(row=row_index, column=4).fill = (
            _make_fill("DCFCE7") if int(row_data.get("remaining_capacity_hours", 0) or 0) <= 0 else _make_fill("FEF3C7")
        )
        for offset, class_key in enumerate(class_order, start=7):
            if not sheet.cell(row=row_index, column=offset).value:
                continue
            fill_key = row_data.get("class_fill_subject_keys", {}).get(class_key, "")
            sheet.cell(row=row_index, column=offset).fill = _subject_fill_for_key(fill_key) or _make_fill("EDF4FF")

    for col_index in range(1, len(headers) + 1):
        width = 24 if col_index in {2, 5, 6} else 16
        if col_index >= 7:
            width = 18
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    if teacher_matrix_rows:
        _add_report_table(sheet, header_row, sheet.max_row, len(headers), "ClassAllocationMatrix")
    sheet.freeze_panes = "A5"


def _build_class_demand_sheet(workbook: Workbook, report_package: dict):
    allocation_data = report_package["allocation_data"]
    sheet = workbook.create_sheet("Class Demand Details")
    _style_report_sheet(sheet, "B91C1C")
    _apply_report_title(
        sheet,
        "Class Demand Details",
        "Uncovered class-level demand and detailed teacher allocation rows.",
        last_column=8,
    )
    headers = ["Class", "Class Status", "Subject Code", "Subject Name", "Uncovered Hours"]
    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col_index, value=header)
    _format_report_header_row(sheet, header_row, len(headers))
    unassigned_rows = allocation_data.get("unassigned_rows", [])
    if unassigned_rows:
        for row in unassigned_rows:
            sheet.append(
                [
                    row.get("class_label", ""),
                    row.get("class_status", ""),
                    row.get("subject_code", ""),
                    row.get("subject_name", ""),
                    row.get("remaining_hours", 0),
                ]
            )
            sheet.cell(row=sheet.max_row, column=5).fill = _make_fill("FEE2E2")
    else:
        sheet.append(["All classes are fully covered.", "", "", "", 0])
        sheet.cell(row=sheet.max_row, column=1).fill = _make_fill("DCFCE7")
    _add_report_table(sheet, header_row, sheet.max_row, len(headers), "ClassDemandGaps")

    detail_start = sheet.max_row + 3
    sheet.cell(row=detail_start, column=1, value="Teacher Allocation Rows").font = Font(
        bold=True,
        size=12,
        color="0A4EA3",
    )
    detail_headers = [
        "Teacher ID",
        "Teacher Name",
        "Class",
        "Class Status",
        "Subject Code",
        "Subject Name",
        "Allocated Hours",
        "Coverage Type",
    ]
    for col_index, header in enumerate(detail_headers, start=1):
        sheet.cell(row=detail_start + 1, column=col_index, value=header)
    _format_report_header_row(sheet, detail_start + 1, len(detail_headers))
    for item in allocation_data.get("assignment_rows", []):
        sheet.append(
            [
                item.get("teacher_id", ""),
                item.get("teacher_name", ""),
                item.get("class_label", ""),
                item.get("class_status", ""),
                item.get("subject_code", ""),
                item.get("subject_name", ""),
                item.get("allocated_hours", 0),
                item.get("coverage_type", ""),
            ]
        )
    if allocation_data.get("assignment_rows"):
        _add_report_table(sheet, detail_start + 1, sheet.max_row, len(detail_headers), "TeacherAllocationRows")

    for col, width in {
        "A": 18,
        "B": 24,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 28,
        "G": 16,
        "H": 16,
    }.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_professional_report_xlsx_bytes(report_package: dict, section: str = "full") -> bytes:
    section = _normalize_report_export_section(section)
    workbook = Workbook()
    _build_summary_overview_sheet(workbook, report_package)

    if section in {"full", "hiring"}:
        _build_hiring_plan_sheet(workbook, report_package)
    if section in {"full", "subjects"}:
        _build_subjects_breakdown_sheet(workbook, report_package)
    if section in {"full", "teachers"}:
        _build_teachers_overview_sheet(workbook, report_package)
    if section == "full":
        _build_class_matrix_sheet(workbook, report_package)
        _build_class_demand_sheet(workbook, report_package)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or "A4"
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and cell.alignment is None:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_escape_text(value: Any) -> str:
    text = str(value if value is not None else "")
    text = text.replace("\u2022", "-").replace("\u00d7", "x")
    text = text.encode("latin-1", "replace").decode("latin-1")
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_rgb(color_value: str, fallback: str = "#0A4EA3") -> tuple[float, float, float]:
    cleaned = str(color_value or fallback).strip().lstrip("#")
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", cleaned):
        cleaned = str(fallback).strip().lstrip("#")
    return (
        int(cleaned[0:2], 16) / 255,
        int(cleaned[2:4], 16) / 255,
        int(cleaned[4:6], 16) / 255,
    )


class _SimplePdfReport:
    width = 595.28
    height = 841.89
    margin = 42

    def __init__(self, title: str, subtitle: str):
        self.title = title
        self.subtitle = subtitle
        self.pages: list[list[str]] = []
        self.y = self.height - self.margin
        self.add_page()

    def _current(self) -> list[str]:
        return self.pages[-1]

    def _text_command(self, x: float, y: float, value: Any, size: float = 9, color: str = "#11243F", bold: bool = False) -> str:
        r, g, b = _pdf_rgb(color, "#11243F")
        font = "/F2" if bold else "/F1"
        return (
            f"{r:.4f} {g:.4f} {b:.4f} rg\n"
            f"BT {font} {size:.2f} Tf {x:.2f} {y:.2f} Td ({_pdf_escape_text(value)}) Tj ET\n"
        )

    def add_page(self):
        self.pages.append([])
        self.y = self.height - self.margin
        self.text(self.margin, self.y, self.title, size=15, color="#0A4EA3", bold=True)
        self.y -= 16
        self.text(self.margin, self.y, self.subtitle, size=8.5, color="#536782")
        self.y -= 22
        self.line(self.margin, self.y, self.width - self.margin, self.y, "#CAD9EA")
        self.y -= 18

    def ensure_space(self, height: float):
        if self.y - height < self.margin + 24:
            self.add_page()

    def text(self, x: float, y: float, value: Any, size: float = 9, color: str = "#11243F", bold: bool = False):
        self._current().append(self._text_command(x, y, value, size=size, color=color, bold=bold))

    def rect(self, x: float, y: float, width: float, height: float, color: str):
        r, g, b = _pdf_rgb(color)
        self._current().append(
            f"{r:.4f} {g:.4f} {b:.4f} rg\n{x:.2f} {y:.2f} {width:.2f} {height:.2f} re f\n"
        )

    def line(self, x1: float, y1: float, x2: float, y2: float, color: str = "#CAD9EA", width: float = 0.8):
        r, g, b = _pdf_rgb(color, "#CAD9EA")
        self._current().append(
            f"{r:.4f} {g:.4f} {b:.4f} RG\n{width:.2f} w\n{x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S\n"
        )

    def h2(self, value: str):
        self.ensure_space(34)
        self.text(self.margin, self.y, value, size=13, color="#0A4EA3", bold=True)
        self.y -= 18

    def paragraph(self, value: str, size: float = 8.5, color: str = "#536782", width: float = 500):
        words = str(value or "").split()
        line = ""
        max_chars = max(20, int(width / (size * 0.47)))
        for word in words:
            candidate = f"{line} {word}".strip()
            if len(candidate) > max_chars and line:
                self.ensure_space(11)
                self.text(self.margin, self.y, line, size=size, color=color)
                self.y -= 11
                line = word
            else:
                line = candidate
        if line:
            self.ensure_space(11)
            self.text(self.margin, self.y, line, size=size, color=color)
            self.y -= 13

    def kpi_grid(self, cards: list[tuple[str, str, str]]):
        card_width = (self.width - (2 * self.margin) - 18) / 4
        card_height = 58
        for index, (label, value, note) in enumerate(cards):
            if index and index % 4 == 0:
                self.y -= card_height + 10
            if index % 4 == 0:
                self.ensure_space(card_height + 10)
                row_y = self.y - card_height
            x = self.margin + (index % 4) * (card_width + 6)
            self.rect(x, row_y, card_width, card_height, "#EDF4FF")
            self.text(x + 8, row_y + 40, label, size=7.3, color="#536782", bold=True)
            self.text(x + 8, row_y + 22, value, size=16, color="#0A4EA3", bold=True)
            self.text(x + 8, row_y + 9, note[:32], size=6.5, color="#536782")
        self.y -= card_height + 16

    def progress_bar(self, label: str, value: int, maximum: int, color: str = "#0A4EA3"):
        maximum = max(int(maximum or 0), 1)
        value = max(0, int(value or 0))
        pct = max(0, min(1, value / maximum))
        self.ensure_space(28)
        self.text(self.margin, self.y, f"{label}: {value} / {maximum}", size=8, color="#11243F", bold=True)
        self.y -= 11
        bar_width = self.width - (2 * self.margin)
        self.rect(self.margin, self.y - 8, bar_width, 8, "#D9E3F1")
        self.rect(self.margin, self.y - 8, bar_width * pct, 8, color)
        self.y -= 18

    def bar_chart(self, title: str, items: list[tuple[str, int, str]], max_items: int = 8):
        chart_items = items[:max_items]
        if not chart_items:
            return
        self.h2(title)
        max_value = max(value for _, value, _ in chart_items) or 1
        bar_width = 310
        for label, value, color in chart_items:
            self.ensure_space(24)
            self.text(self.margin, self.y, label[:38], size=8, color="#11243F")
            self.text(self.margin + 390, self.y, f"{value}h", size=8, color="#11243F", bold=True)
            self.rect(self.margin + 150, self.y - 2, bar_width, 8, "#E5EDF7")
            self.rect(self.margin + 150, self.y - 2, bar_width * (value / max_value), 8, color)
            self.y -= 18
        self.y -= 8

    def table(self, title: str, headers: list[str], rows: list[list[Any]], widths: list[float], max_rows: int | None = None):
        if title:
            self.h2(title)
        rows = rows if max_rows is None else rows[:max_rows]
        row_height = 22
        total_width = sum(widths)
        self.ensure_space(row_height * 2)
        y = self.y - row_height
        self.rect(self.margin, y, total_width, row_height, "#0A4EA3")
        x = self.margin
        for header, width in zip(headers, widths):
            self.text(x + 4, y + 8, header, size=6.8, color="#FFFFFF", bold=True)
            x += width
        self.y = y
        for row in rows:
            self.ensure_space(row_height)
            y = self.y - row_height
            fill = "#F8FAFC" if (len(self._current()) % 2 == 0) else "#FFFFFF"
            self.rect(self.margin, y, total_width, row_height, fill)
            x = self.margin
            for value, width in zip(row, widths):
                text = str(value if value is not None else "")
                max_chars = max(8, int(width / 4.2))
                self.text(x + 4, y + 8, text[:max_chars], size=6.6, color="#11243F")
                x += width
            self.line(self.margin, y, self.margin + total_width, y, "#E6EDF6", 0.4)
            self.y = y
        self.y -= 14

    def build(self) -> bytes:
        for index, page in enumerate(self.pages, start=1):
            page.append(
                self._text_command(
                    self.margin,
                    24,
                    f"Generated by TIS | Page {index}",
                    size=7,
                    color="#536782",
                )
            )

        content_streams = [
            "".join(page).encode("latin-1", "replace")
            for page in self.pages
        ]
        kids = []
        objects: list[bytes] = []
        objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
        objects.append(b"")
        objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
        objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
        next_object_number = 5
        page_objects = []
        for content in content_streams:
            page_number = next_object_number
            content_number = next_object_number + 1
            kids.append(f"{page_number} 0 R")
            page_objects.append(
                (
                    page_number,
                    (
                        f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {self.width:.2f} {self.height:.2f}] "
                        f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
                        f"/Contents {content_number} 0 R >>"
                    ).encode("latin-1"),
                )
            )
            page_objects.append(
                (
                    content_number,
                    b"<< /Length " + str(len(content)).encode("ascii") + b" >>\nstream\n" + content + b"\nendstream",
                )
            )
            next_object_number += 2

        objects[1] = (
            f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(kids)} >>"
        ).encode("latin-1")
        for _, obj in page_objects:
            objects.append(obj)

        pdf = b"%PDF-1.4\n"
        offsets = [0]
        for object_index, obj in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf += f"{object_index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n"
        xref_offset = len(pdf)
        pdf += f"xref\n0 {len(objects) + 1}\n".encode("ascii")
        pdf += b"0000000000 65535 f \n"
        for offset in offsets[1:]:
            pdf += f"{offset:010d} 00000 n \n".encode("ascii")
        pdf += (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
        ).encode("ascii")
        return pdf


def _build_professional_report_pdf_bytes(report_package: dict, section: str = "full") -> bytes:
    section = _normalize_report_export_section(section)
    summary = report_package["report_summary"]
    title = "TIS Analytical Staffing Report"
    subtitle = (
        f"{report_package['branch_name']} | {report_package['academic_year_name']} | "
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    pdf = _SimplePdfReport(title, subtitle)
    pdf.h2("Summary Overview")
    pdf.paragraph(
        "This export uses the same current reporting calculations as the dashboard: planned section demand, "
        "teacher-section allocation, homeroom default coverage, uncovered hours, and the latest subject-pool hiring logic."
    )
    pdf.kpi_grid(
        [
            ("Total Demand", f"{summary.get('total_required_hours', 0)}h", "Section-subject hours"),
            ("Covered", f"{summary.get('total_allocated_hours', 0)}h", "Assigned coverage"),
            ("Uncovered", f"{summary.get('total_uncovered_hours', 0)}h", "Remaining hours"),
            ("Full Hires", str(summary.get("total_new_teachers_required", 0)), "24h teacher blocks"),
            ("Coverage", f"{summary.get('coverage_percentage', 0)}%", "Demand covered"),
            ("Remainder", f"{summary.get('remaining_uncovered_hours_after_hires', 0)}h", "After whole hires"),
            ("Underloaded", str(summary.get("underloaded_teachers", 0)), "Teachers with capacity"),
            ("Priority Gaps", f"{summary.get('priority_gap_hours', 0)}h", "Priority subjects"),
        ]
    )
    pdf.progress_bar(
        "Coverage",
        int(summary.get("total_allocated_hours", 0) or 0),
        int(summary.get("total_required_hours", 0) or 0),
        "#0A4EA3",
    )
    gap_items = [
        (
            row.get("subject_name", "Subject"),
            int(row.get("remaining_hours", 0) or 0),
            row.get("subject_color", "#B91C1C"),
        )
        for row in report_package.get("report_gap_rows", [])
    ]
    pdf.bar_chart("Top Uncovered Subjects", gap_items, max_items=8)

    if section in {"full", "hiring"}:
        plan = report_package["hiring_plan_export"]
        plan_summary = plan["summary"]
        pdf.h2("Recommended Hiring Plan")
        pdf.paragraph(
            f"{report_package.get('hiring_plan_source', 'System suggested plan')}. "
            f"Pools cover {plan_summary.get('planned_pool_hours', 0)}h out of "
            f"{plan_summary.get('total_uncovered_hours', 0)}h uncovered demand."
        )
        pdf.bar_chart(
            "Hiring Pool Hours",
            [
                (row["name"], int(row["total_hours"]), row["accent_color"])
                for row in plan["profiles"]
            ],
            max_items=10,
        )
        pdf.table(
            "Pool Summary",
            ["Pool", "Hours", "Full", "Rem", "Subjects"],
            [
                [
                    row["name"],
                    row["total_hours"],
                    row["full_blocks"],
                    row["remainder_hours"],
                    row["subject_count"],
                ]
                for row in plan["profiles"]
            ],
            [210, 55, 45, 45, 65],
        )

    if section in {"full", "subjects"}:
        pdf.table(
            "Subjects Breakdown",
            ["Subject", "Demand", "Covered", "Uncov.", "Cov %", "Blocks"],
            [
                [
                    row.get("subject_name", ""),
                    row.get("required_hours", 0),
                    row.get("allocated_hours", 0),
                    row.get("remaining_hours", 0),
                    f"{row.get('coverage_percentage', 0)}%",
                    row.get("teacher_requirement_blocks", 0),
                ]
                for row in report_package.get("report_subject_rows", [])
            ],
            [170, 55, 55, 55, 50, 50],
            max_rows=None if section == "subjects" else 28,
        )

    if section in {"full", "teachers"}:
        pdf.table(
            "Teachers Overview",
            ["Teacher", "ID", "Load", "Cap", "Remain", "Subjects"],
            [
                [
                    row.get("teacher_name", ""),
                    row.get("teacher_id", ""),
                    row.get("expected_allocated_hours", 0),
                    row.get("capacity_hours", 0),
                    row.get("remaining_capacity_hours", 0),
                    ", ".join(row.get("allocation_labels", []) or row.get("subject_labels", [])),
                ]
                for row in report_package.get("report_teacher_rows", [])
            ],
            [145, 55, 45, 45, 55, 175],
            max_rows=None if section == "teachers" else 24,
        )

    return pdf.build()


@app.middleware("http")
async def inactivity_timeout_middleware(request: Request, call_next):
    path = request.url.path or ""
    if any(
        path.startswith(prefix) for prefix in IDLE_TIMEOUT_EXEMPT_PREFIXES
    ):
        return await call_next(request)

    session_user_id = auth.get_session_user_id(request)
    if not session_user_id:
        return await call_next(request)

    db = SessionLocal()
    try:
        user = None
        current_user = None
        if session_user_id:
            user = db.query(models.User).filter(
                models.User.user_id == session_user_id
            ).first()
            if not user:
                response = RedirectResponse(url="/", status_code=302)
                return _clear_auth_session_cookies(response)

            if not auth.is_user_active(user):
                response = RedirectResponse(url="/?inactive=1", status_code=302)
                return _clear_auth_session_cookies(response)

            if path not in IDLE_TIMEOUT_EXEMPT_PATHS and not auth.is_platform_user(user):
                now_ts = int(time.time())
                last_activity_raw = str(request.cookies.get(IDLE_TIMEOUT_COOKIE_KEY) or "").strip()
                if last_activity_raw:
                    try:
                        last_activity_ts = int(last_activity_raw)
                    except ValueError:
                        last_activity_ts = 0

                    if last_activity_ts > 0 and (now_ts - last_activity_ts) > IDLE_TIMEOUT_SECONDS:
                        timeout_response = RedirectResponse(url="/?timeout=1", status_code=302)
                        return _clear_auth_session_cookies(timeout_response)
            current_user = auth.get_current_user(request, db)

        permission_response = authorization.enforce_route_permission(
            request,
            db,
            current_user=current_user,
        )
        if permission_response is not None:
            return permission_response

        response = await call_next(request)
        if session_user_id and user and path not in IDLE_TIMEOUT_EXEMPT_PATHS and not auth.is_platform_user(user):
            response.set_cookie(
                key=IDLE_TIMEOUT_COOKIE_KEY,
                value=str(int(time.time())),
                **auth.secure_cookie_kwargs(request),
            )
        return response
    finally:
        db.close()


@app.middleware("http")
async def audit_logging_middleware(request: Request, call_next):
    started_at = time.perf_counter()
    status_code = 500
    error_name = ""

    try:
        response = await call_next(request)
        status_code = response.status_code
        return response
    except Exception as exc:
        error_name = exc.__class__.__name__
        raise
    finally:
        duration_ms = round((time.perf_counter() - started_at) * 1000, 2)
        _write_request_audit_log(
            request=request,
            status_code=status_code,
            duration_ms=duration_ms,
            error_name=error_name,
        )


def _build_login_context(
    db: Session,
    username: str = "",
    error: Optional[str] = None,
):
    active_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.is_active == True
    ).first()

    return {
        "username": username,
        "active_year_name": active_year.year_name if active_year else "Not configured",
        "error": error,
    }


def _render_login_page(
    request: Request,
    db: Session,
    username: str = "",
    error: Optional[str] = None,
    status_code: int = 200,
):
    context = _build_login_context(
        db=db,
        username=username,
        error=error,
    )
    context["request"] = request
    return templates.TemplateResponse(
        request,
        "index.html",
        context,
        status_code=status_code,
    )


def _render_login_entrypoint(
    request: Request,
    db: Session,
):
    current_user = auth.get_current_user(request, db)
    if current_user:
        return RedirectResponse(
            url=("/platform" if auth.is_platform_user(current_user) else "/dashboard?info=already-logged-in"),
            status_code=302
        )

    timeout_notice = str(request.query_params.get("timeout", "")).strip().lower()
    timeout_error = IDLE_TIMEOUT_LOGIN_MESSAGE if timeout_notice in {"1", "true", "yes"} else None
    inactive_notice = str(request.query_params.get("inactive", "")).strip().lower()
    inactive_error = (
        auth.INACTIVE_ACCOUNT_MESSAGE
        if inactive_notice in {"1", "true", "yes"}
        else None
    )

    return _render_login_page(
        request=request,
        db=db,
        error=inactive_error or timeout_error,
    )

# ---------------------------------------
# Include Routers
# ---------------------------------------
app.include_router(subjects.router)
app.include_router(users.router)
app.include_router(teachers.router)
app.include_router(planning.router)
app.include_router(timetable.router)
app.include_router(academic_calendar.router)
app.include_router(observations.router)

# ---------------------------------------
# ROOT (public landing on tisplatform.com, app login elsewhere)
# ---------------------------------------
@app.get("/", response_class=HTMLResponse)
def read_root(
    request: Request,
    db: Session = Depends(get_db)
):
    if _is_public_landing_host(request):
        return templates.TemplateResponse(
            request,
            "landing.html",
            {"request": request},
        )

    return _render_login_entrypoint(request=request, db=db)


def _clean_demo_field(value: Any, max_length: int = 1200) -> str:
    cleaned_value = str(value or "").replace("\r", " ").strip()
    cleaned_value = re.sub(r"[ \t]+", " ", cleaned_value)
    return cleaned_value[:max_length]


def _clean_email_header(value: Any, fallback: str) -> str:
    cleaned_value = re.sub(r"[\r\n]+", " ", str(value or "").strip())
    return cleaned_value[:160] or fallback


def _build_demo_email_body(demo_fields: dict[str, str], request: Request) -> str:
    submitted_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    source_host = _request_hostname(request) or "unknown"
    source_ip = getattr(getattr(request, "client", None), "host", "unknown")

    lines = [
        "New TIS landing page demo request",
        "",
        f"Submitted At: {submitted_at}",
        f"Source Host: {source_host}",
        f"Source IP: {source_ip}",
        "",
    ]
    for field_key, field_label in DEMO_REQUEST_FIELDS:
        lines.append(f"{field_label}: {demo_fields.get(field_key, '') or '-'}")
    return "\n".join(lines)


def _create_demo_request_record(db: Session, demo_fields: dict[str, str], request: Request):
    demo_request = models.DemoRequest(
        school_name=demo_fields.get("school_name", ""),
        full_name=demo_fields.get("full_name", ""),
        email=demo_fields.get("email", ""),
        phone=demo_fields.get("phone", ""),
        country=demo_fields.get("country", ""),
        school_type=demo_fields.get("school_type", ""),
        number_of_teachers=demo_fields.get("teachers", ""),
        number_of_students=demo_fields.get("students", ""),
        number_of_branches=demo_fields.get("branches", ""),
        interested_plan=demo_fields.get("interested_plan", ""),
        message=demo_fields.get("message", ""),
        status=DEMO_REQUEST_STATUSES[0],
        source_host=_request_hostname(request),
        source_ip=str(getattr(getattr(request, "client", None), "host", "") or ""),
    )
    db.add(demo_request)
    db.commit()
    db.refresh(demo_request)
    return demo_request


def _send_smtp_message(message: EmailMessage) -> None:
    smtp_host = str(os.getenv("TIS_SMTP_HOST") or "").strip()
    if not smtp_host:
        raise RuntimeError("TIS_SMTP_HOST is not configured")

    smtp_port = _get_positive_int_env("TIS_SMTP_PORT", 587)
    smtp_timeout = _get_positive_int_env("TIS_SMTP_TIMEOUT_SECONDS", 12)
    smtp_user = str(os.getenv("TIS_SMTP_USER") or "").strip()
    smtp_password = str(os.getenv("TIS_SMTP_PASSWORD") or "").strip()
    use_ssl = _get_bool_env("TIS_SMTP_SSL", False)
    use_tls = _get_bool_env("TIS_SMTP_TLS", not use_ssl)

    context = ssl.create_default_context()
    if use_ssl:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=smtp_timeout, context=context) as server:
            if smtp_user and smtp_password:
                server.login(smtp_user, smtp_password)
            server.send_message(message)
        return

    with smtplib.SMTP(smtp_host, smtp_port, timeout=smtp_timeout) as server:
        server.ehlo()
        if use_tls:
            server.starttls(context=context)
            server.ehlo()
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        server.send_message(message)


def _send_demo_request_email(demo_fields: dict[str, str], request: Request) -> None:
    smtp_user = str(os.getenv("TIS_SMTP_USER") or "").strip()
    smtp_from = str(os.getenv("TIS_SMTP_FROM") or smtp_user or DEMO_REQUEST_EMAIL_TO).strip()

    school_name = _clean_email_header(demo_fields.get("school_name"), "New school")
    contact_name = _clean_email_header(demo_fields.get("full_name"), "Landing visitor")

    message = EmailMessage()
    message["Subject"] = f"TIS demo request - {school_name}"
    message["From"] = smtp_from
    message["To"] = DEMO_REQUEST_EMAIL_TO
    reply_to = _clean_email_header(demo_fields.get("email"), "")
    if reply_to:
        message["Reply-To"] = reply_to
    message.set_content(_build_demo_email_body(demo_fields, request))
    message.add_header("X-TIS-Demo-Contact", contact_name)
    _send_smtp_message(message)


@app.post("/request-demo")
async def submit_demo_request(
    request: Request,
    db: Session = Depends(get_db),
):
    try:
        form_data = await request.form()
    except Exception:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "message": "Please review the form and try again."},
        )

    demo_fields = {
        field_key: _clean_demo_field(form_data.get(field_key), 2400 if field_key == "message" else 500)
        for field_key, _field_label in DEMO_REQUEST_FIELDS
    }

    missing_fields = [
        field_label
        for field_key, field_label in DEMO_REQUEST_FIELDS
        if field_key in {"school_name", "full_name", "email", "country", "interested_plan"}
        and not demo_fields.get(field_key)
    ]
    if missing_fields:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "message": "Please complete: " + ", ".join(missing_fields) + ".",
            },
        )

    try:
        demo_request = _create_demo_request_record(db, demo_fields, request)
    except Exception as exc:
        db.rollback()
        logging.error(
            "TIS landing demo request save failed host=%s school=%s email=%s: %s",
            _request_hostname(request),
            demo_fields.get("school_name"),
            demo_fields.get("email"),
            exc,
            exc_info=True,
        )
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "message": (
                    "We could not save the request right now. "
                    "Please email info@tisplatform.com directly."
                ),
            },
        )

    try:
        _send_demo_request_email(demo_fields, request)
    except Exception as exc:
        logging.warning(
            "TIS landing demo request saved id=%s but email failed host=%s school=%s email=%s: %s",
            demo_request.id,
            _request_hostname(request),
            demo_fields.get("school_name"),
            demo_fields.get("email"),
            exc,
            exc_info=True,
        )

    logging.info(
        "TIS landing demo request saved id=%s host=%s school=%s email=%s plan=%s",
        demo_request.id,
        _request_hostname(request),
        demo_fields.get("school_name"),
        demo_fields.get("email"),
        demo_fields.get("interested_plan"),
    )
    return JSONResponse(
        content={
            "ok": True,
            "message": (
                "Thank you. Your demo request has been received by the TIS team. "
                "We will contact you shortly."
            ),
        }
    )


@app.get("/login", response_class=HTMLResponse)
def login_page(
    request: Request,
    db: Session = Depends(get_db)
):
    return _render_login_entrypoint(request=request, db=db)


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return FileResponse(
        FAVICON_IMAGE_PATH,
        media_type="image/png",
        headers=FAVICON_CACHE_HEADERS,
    )


@app.get(
    "/organization-assets/{school_group_id}/{asset_path:path}",
    name="organization_asset",
    include_in_schema=False,
)
def organization_asset(
    school_group_id: int,
    asset_path: str,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return PlainTextResponse("Not found", status_code=404)
    requesting_group_id = _get_user_school_group_id(db, current_user)
    can_manage_all = _can_manage_all_school_scopes(db, current_user)
    if not branding_storage.can_access_organization_assets(
        requesting_group_id,
        school_group_id,
        can_manage_all=can_manage_all,
    ):
        return PlainTextResponse("Not found", status_code=404)
    try:
        asset = branding_storage.resolve_organization_asset_path(
            school_group_id,
            asset_path,
            require_file=True,
        )
    except (branding_storage.BrandingStorageError, FileNotFoundError):
        return PlainTextResponse("Not found", status_code=404)
    return FileResponse(
        asset,
        headers={
            "Cache-Control": "private, max-age=300",
            "X-Content-Type-Options": "nosniff",
        },
    )

# ---------------------------------------
# LOGIN
# ---------------------------------------

DEVELOPER_USER_ID = os.getenv("TIS_DEVELOPER_USER_ID", "2623252018")
NOTIFICATION_STATUS_NEW = "New"
NOTIFICATION_STATUS_SEEN = "Seen"
NOTIFICATION_STATUS_RESOLVED = "Resolved"
NOTIFICATION_TYPE_FORGOT_PASSWORD = "Forgot Password"
NOTIFICATION_TYPE_MESSAGE = "Message"
NOTIFICATION_SCOPE_USER = "User"
NOTIFICATION_SCOPE_ALL = "All"
MESSAGE_PAGE_SIZE = 50
NOTIFICATION_AUTO_ARCHIVE_DAYS = 90


def _notification_logger():
    return logging.getLogger("uvicorn.error")


def _is_notification_diagnostic_mode_enabled() -> bool:
    raw_value = str(os.getenv("TIS_NOTIFICATION_DIAGNOSTIC", "") or "").strip().lower()
    if raw_value in {"1", "true", "yes", "on"}:
        return True
    if raw_value in {"0", "false", "no", "off"}:
        return False

    # Default to enabled for local/development diagnostics.
    env_value = str(
        os.getenv("TIS_ENV")
        or os.getenv("ENV")
        or os.getenv("FASTAPI_ENV")
        or ""
    ).strip().lower()
    return env_value in {"", "dev", "development", "local", "debug", "test", "testing"}


def _collect_system_notification_schema_snapshot() -> dict:
    snapshot = {
        "table_name": "system_notifications",
        "table_exists": False,
        "expected_columns": [
            col.name for col in models.SystemNotification.__table__.columns
        ],
        "actual_columns": [],
        "missing_columns": [],
        "expected_indexes": sorted(
            index.name for index in models.SystemNotification.__table__.indexes
        ),
        "actual_indexes": [],
        "missing_indexes": [],
    }

    try:
        inspector = inspect(engine)
        table_names = set(inspector.get_table_names())
        if "system_notifications" not in table_names:
            snapshot["missing_columns"] = list(snapshot["expected_columns"])
            snapshot["missing_indexes"] = list(snapshot["expected_indexes"])
            return snapshot

        snapshot["table_exists"] = True
        snapshot["actual_columns"] = [
            col["name"] for col in inspector.get_columns("system_notifications")
        ]
        snapshot["actual_indexes"] = sorted(
            idx.get("name") for idx in inspector.get_indexes("system_notifications")
        )
        snapshot["missing_columns"] = [
            col for col in snapshot["expected_columns"] if col not in snapshot["actual_columns"]
        ]
        snapshot["missing_indexes"] = [
            idx for idx in snapshot["expected_indexes"] if idx not in snapshot["actual_indexes"]
        ]
    except Exception as exc:
        snapshot["schema_introspection_error"] = f"{exc.__class__.__name__}: {exc}"

    return snapshot


def _ensure_system_notifications_indexes() -> None:
    # Notification indexes are managed by db_migrations.
    return


def _get_notification_total_count_safe(db: Session, user_id: Optional[str]) -> Optional[int]:
    if not user_id:
        return None
    try:
        return db.query(func.count(models.SystemNotification.id)).filter(
            models.SystemNotification.recipient_user_id == str(user_id)
        ).scalar()
    except Exception:
        return None


def _render_notification_error_fallback(
    request: Request,
    *,
    route_name: str,
    error_message: str,
    diagnostic_payload: Optional[dict] = None,
):
    diagnostic_enabled = _is_notification_diagnostic_mode_enabled()
    heading = "Notification Center Error"
    safe_error = html.escape(str(error_message or "Unknown error"))
    safe_path = html.escape(str(getattr(request, "url", "") or ""))

    body_html = (
        "<p style='margin:0 0 12px 0;'>The notification page could not be rendered.</p>"
        "<p style='margin:0 0 16px 0;'>"
        "Please check server logs for full diagnostics and refresh after fixing the issue."
        "</p>"
    )
    if diagnostic_enabled:
        payload_block = ""
        if diagnostic_payload:
            payload_block = (
                "<h2 style='font-size:16px;margin:20px 0 8px 0;'>Captured diagnostics</h2>"
                f"<pre style='background:#f4f4f4;border:1px solid #ddd;padding:12px;overflow:auto;'>{html.escape(json.dumps(diagnostic_payload, indent=2, default=str))}</pre>"
            )
        body_html += (
            "<h2 style='font-size:16px;margin:0 0 8px 0;'>Exception message</h2>"
            f"<pre style='background:#fff7f7;border:1px solid #f5c2c7;padding:12px;overflow:auto;'>{safe_error}</pre>"
            f"<p style='margin:12px 0 0 0;color:#555;'>Route: {safe_path}</p>"
            f"<p style='margin:4px 0 0 0;color:#555;'>Handler: {html.escape(route_name)}</p>"
            f"{payload_block}"
        )

    html_content = (
        "<!doctype html>"
        "<html><head><meta charset='utf-8'><title>Notification Error</title></head>"
        "<body style='font-family:Segoe UI,Arial,sans-serif;background:#fafafa;padding:24px;'>"
        "<main style='max-width:860px;margin:0 auto;background:#fff;border:1px solid #e5e5e5;border-radius:8px;padding:24px;'>"
        f"<h1 style='margin-top:0;'>{heading}</h1>"
        f"{body_html}"
        "<p style='margin-top:20px;'><a href='/notifications'>Back to Notification Center</a></p>"
        "</main></body></html>"
    )
    return HTMLResponse(content=html_content, status_code=500)


DEFAULT_DISPLAY_TIMEZONE = "Asia/Riyadh"


def _resolve_display_timezone(timezone_name: str):
    raw_timezone = str(timezone_name or "").strip() or DEFAULT_DISPLAY_TIMEZONE
    try:
        return raw_timezone, ZoneInfo(raw_timezone)
    except ZoneInfoNotFoundError:
        if raw_timezone == "Asia/Riyadh":
            return raw_timezone, timezone(timedelta(hours=3), name="KSA")
        return DEFAULT_DISPLAY_TIMEZONE, timezone(timedelta(hours=3), name="KSA")


def _get_request_timezone(request: Request | None) -> str:
    raw_timezone = ""
    if request is not None:
        raw_timezone = str(request.cookies.get("tis_timezone", "") or "").strip()
    timezone_name, _ = _resolve_display_timezone(raw_timezone)
    return timezone_name


def _timezone_label(timezone_name: str) -> str:
    if timezone_name == "Asia/Riyadh":
        return "KSA"
    return timezone_name


def _format_notification_timestamp(
    value,
    fallback: str = "Unknown",
    timezone_name: str = DEFAULT_DISPLAY_TIMEZONE,
) -> str:
    if not value:
        return fallback

    parsed_value = value
    if not isinstance(parsed_value, datetime):
        cleaned = str(value or "").strip()
        if not cleaned:
            return fallback
        try:
            parsed_value = datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
        except ValueError:
            return cleaned

    if parsed_value.tzinfo is None:
        parsed_value = parsed_value.replace(tzinfo=timezone.utc)
    target_timezone_name, target_timezone = _resolve_display_timezone(timezone_name)
    localized_value = parsed_value.astimezone(target_timezone)
    return f"{localized_value.strftime('%d %b %Y %H:%M')} {_timezone_label(target_timezone_name)}"


def _notification_timestamp_iso(value) -> str:
    if not value:
        return ""
    parsed_value = value
    if not isinstance(parsed_value, datetime):
        cleaned = str(value or "").strip()
        if not cleaned:
            return ""
        try:
            parsed_value = datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
        except ValueError:
            return ""
    if parsed_value.tzinfo is None:
        parsed_value = parsed_value.replace(tzinfo=timezone.utc)
    return parsed_value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def _build_user_display(user_id: str, user=None) -> str:
    if not user:
        return f"User ID {user_id} (not found in system)"

    full_name = f"{getattr(user, 'first_name', '') or ''} {getattr(user, 'last_name', '') or ''}".strip()
    return f"{full_name} (User ID {user_id})" if full_name else f"User ID {user_id}"


def _create_system_notification(
    db: Session,
    *,
    recipient_user_id: str,
    requesting_user_id: str,
    request_type: str,
    title: str,
    message: str = "",
    details: str = "",
    recipient_scope: str = NOTIFICATION_SCOPE_USER,
):
    _ensure_system_notifications_table_columns()
    recipient_user = db.query(models.User).filter(
        models.User.user_id == str(recipient_user_id).strip()
    ).first()
    requesting_user = None
    if requesting_user_id:
        requesting_user = db.query(models.User).filter(
            models.User.user_id == str(requesting_user_id).strip()
        ).first()
    notification_school_group_id = auth.get_notification_school_group_id(
        db,
        recipient_user=recipient_user,
        current_user=requesting_user,
    )
    notification = models.SystemNotification(
        school_group_id=notification_school_group_id,
        branch_id=getattr(recipient_user, "branch_id", None),
        academic_year_id=getattr(recipient_user, "academic_year_id", None),
        recipient_user_id=str(recipient_user_id).strip(),
        requesting_user_id=str(requesting_user_id or "").strip(),
        request_type=str(request_type).strip(),
        title=str(title).strip(),
        message=str(message or "").strip(),
        details=str(details or "").strip(),
        status=NOTIFICATION_STATUS_NEW,
        recipient_scope=str(recipient_scope).strip(),
        created_at=datetime.now(timezone.utc).replace(tzinfo=None),
    )
    db.add(notification)
    db.commit()
    db.refresh(notification)
    logging.info(
        "TIS notification: created id=%s type=%s recipient_user_id=%s requesting_user_id=%s scope=%s",
        notification.id,
        notification.request_type,
        notification.recipient_user_id,
        notification.requesting_user_id,
        notification.recipient_scope,
    )
    return notification


def _auto_archive_old_notifications(db: Session, user_id: str) -> int:
    _ensure_system_notifications_table_columns()
    cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=NOTIFICATION_AUTO_ARCHIVE_DAYS)
    updated_count = 0
    recipient_rows = db.query(models.SystemNotification).filter(
        models.SystemNotification.recipient_user_id == user_id,
        models.SystemNotification.recipient_archived_at.is_(None),
        models.SystemNotification.created_at < cutoff,
    ).all()
    for notification in recipient_rows:
        notification.recipient_archived_at = cutoff
        notification.recipient_archived_by_user_id = "system"
        updated_count += 1

    requester_rows = db.query(models.SystemNotification).filter(
        models.SystemNotification.requesting_user_id == user_id,
        models.SystemNotification.requester_archived_at.is_(None),
        models.SystemNotification.created_at < cutoff,
    ).all()
    for notification in requester_rows:
        notification.requester_archived_at = cutoff
        notification.requester_archived_by_user_id = "system"
        updated_count += 1
    if updated_count:
        db.commit()
    return updated_count


def _base_notification_query_for_box(db: Session, user_id: str, box: str):
    normalized_box = str(box or "inbox").strip().lower()
    if normalized_box == "sent":
        return db.query(models.SystemNotification).filter(
            models.SystemNotification.requesting_user_id == user_id,
            models.SystemNotification.requester_archived_at.is_(None),
        )
    if normalized_box == "archive":
        return db.query(models.SystemNotification).filter(
            (
                (models.SystemNotification.recipient_user_id == user_id)
                & (models.SystemNotification.recipient_archived_at.isnot(None))
            )
            | (
                (models.SystemNotification.requesting_user_id == user_id)
                & (models.SystemNotification.requester_archived_at.isnot(None))
            )
        )
    return db.query(models.SystemNotification).filter(
        models.SystemNotification.recipient_user_id == user_id,
        models.SystemNotification.recipient_archived_at.is_(None),
    )


def _get_notification_counts(db: Session, user_id: str) -> dict:
    _ensure_system_notifications_table_columns()
    counts = {
        NOTIFICATION_STATUS_NEW: 0,
        NOTIFICATION_STATUS_SEEN: 0,
        NOTIFICATION_STATUS_RESOLVED: 0,
    }
    rows = db.query(
        models.SystemNotification.status,
        func.count(models.SystemNotification.id),
    ).filter(
        models.SystemNotification.recipient_user_id == user_id,
        models.SystemNotification.recipient_archived_at.is_(None),
    ).group_by(
        models.SystemNotification.status
    ).all()
    for status, count in rows:
        counts[str(status or "")] = count
    counts["All"] = sum(
        counts.get(status, 0)
        for status in (
            NOTIFICATION_STATUS_NEW,
            NOTIFICATION_STATUS_SEEN,
            NOTIFICATION_STATUS_RESOLVED,
        )
    )
    counts["Sent"] = db.query(func.count(models.SystemNotification.id)).filter(
        models.SystemNotification.requesting_user_id == user_id,
        models.SystemNotification.requester_archived_at.is_(None),
    ).scalar() or 0
    counts["Archived"] = db.query(func.count(models.SystemNotification.id)).filter(
        (
            (models.SystemNotification.recipient_user_id == user_id)
            & (models.SystemNotification.recipient_archived_at.isnot(None))
        )
        | (
            (models.SystemNotification.requesting_user_id == user_id)
            & (models.SystemNotification.requester_archived_at.isnot(None))
        )
    ).scalar() or 0
    return counts


def _notification_group_key(notification) -> str:
    details = str(getattr(notification, "details", "") or "").strip()
    observation_match = re.match(r"^observation:(\d+):", details)
    if observation_match:
        return f"observation:{observation_match.group(1)}"
    return f"single:{getattr(notification, 'id', '')}"


def _notification_group_matches(notification, group_key: str) -> bool:
    normalized_key = str(group_key or "").strip()
    if normalized_key.startswith("single:"):
        return normalized_key == f"single:{getattr(notification, 'id', '')}"
    if normalized_key.startswith("observation:"):
        return str(getattr(notification, "details", "") or "").strip().startswith(f"{normalized_key}:")
    return False


def _user_notification_display_map(db: Session, messages: list) -> dict:
    user_ids = {
        str(value or "").strip()
        for notification in messages
        for value in (
            getattr(notification, "recipient_user_id", ""),
            getattr(notification, "requesting_user_id", ""),
            getattr(notification, "resolved_by_user_id", ""),
        )
        if str(value or "").strip()
    }
    if not user_ids:
        return {}
    users = db.query(models.User).filter(models.User.user_id.in_(user_ids)).all()
    display_map = {}
    for user in users:
        name = " ".join(
            part
            for part in [
                str(getattr(user, "first_name", "") or "").strip(),
                str(getattr(user, "last_name", "") or "").strip(),
            ]
            if part
        ).strip()
        display_map[user.user_id] = name or str(user.username or user.user_id)
    return display_map


def _build_notification_groups(messages: list) -> list[dict]:
    grouped: dict[str, dict] = {}
    for notification in messages:
        group_key = _notification_group_key(notification)
        group = grouped.get(group_key)
        if not group:
            group = {
                "key": group_key,
                "latest": notification,
                "title": notification.title,
                "request_type": notification.request_type,
                "items": [],
                "count": 0,
                "new_count": 0,
                "seen_count": 0,
                "done_count": 0,
                "status_label": "Seen",
                "status_class": "",
                "is_grouped": group_key.startswith("observation:"),
            }
            if group["is_grouped"]:
                group["title"] = "Observation updates"
            grouped[group_key] = group
        group["items"].append(notification)
        group["count"] += 1
        if notification.status == NOTIFICATION_STATUS_NEW:
            group["new_count"] += 1
        elif notification.status == NOTIFICATION_STATUS_RESOLVED:
            group["done_count"] += 1
        else:
            group["seen_count"] += 1

    for group in grouped.values():
        if group["new_count"]:
            group["status_label"] = "New"
            group["status_class"] = "is-new"
        elif group["done_count"] == group["count"]:
            group["status_label"] = "Done"
            group["status_class"] = "is-resolved"
        else:
            group["status_label"] = "Seen"
            group["status_class"] = ""
    return list(grouped.values())


def _get_user_notification_or_redirect(
    db: Session,
    current_user,
    notification_id: int,
):
    _ensure_system_notifications_table_columns()
    notification = db.query(models.SystemNotification).filter(
        models.SystemNotification.id == notification_id,
        (
            (models.SystemNotification.recipient_user_id == current_user.user_id)
            | (models.SystemNotification.requesting_user_id == current_user.user_id)
        ),
    ).first()
    if not notification:
        return None, RedirectResponse(
            url="/notifications?notice=Message%20not%20found.",
            status_code=302,
        )
    return notification, None


@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    username = username.strip()
    request.state.audit_actor_user_id = username or "anonymous"
    request.state.audit_actor_username = username
    request.state.audit_actor_role = "Unauthenticated"
    request.state.audit_actor_branch_id = None
    user = auth.authenticate_user(db, username, password)

    if not user:
        return _render_login_page(
            request=request,
            db=db,
            username=username,
            error="Invalid User ID, email, or password.",
            status_code=401
        )

    if not auth.is_user_active(user):
        response = _render_login_page(
            request=request,
            db=db,
            username=username,
            error=auth.INACTIVE_ACCOUNT_MESSAGE,
            status_code=403
        )
        return _clear_auth_session_cookies(response)

    can_all_branch_scope = auth.can_access_all_branches(user, db)
    active_branches = auth.get_accessible_branch_query(db, user).order_by(
        models.Branch.name.asc()
    ).all()
    active_branch_map = {
        branch.id: branch for branch in active_branches
    }
    assigned_branch = (
        None
        if auth.is_platform_user(user)
        else active_branch_map.get(user.branch_id)
    )

    if not assigned_branch and not can_all_branch_scope:
        return _render_login_page(
            request=request,
            db=db,
            username=username,
            error="Your assigned branch is inactive or not configured.",
            status_code=400
        )

    if (
        not assigned_branch
        and can_all_branch_scope
        and not active_branches
        and not auth.is_platform_user(user)
    ):
        return _render_login_page(
            request=request,
            db=db,
            username=username,
            error="No active branch is available in the system.",
            status_code=400
        )

    branch_scope_id = None if auth.is_platform_user(user) else (
        assigned_branch.id
        if assigned_branch
        else active_branches[0].id
        if active_branches
        else None
    )
    branch_scope = active_branch_map.get(branch_scope_id)
    scope_school_group_id = getattr(branch_scope, "school_group_id", None) if branch_scope else None

    active_year = auth.get_active_academic_year_for_school_group(db, scope_school_group_id)
    if not active_year and auth.is_platform_user(user):
        active_year = auth.get_latest_academic_year_for_school_group(db, scope_school_group_id)
    if not active_year and not auth.is_platform_user(user):
        return _render_login_page(
            request=request,
            db=db,
            username=username,
            error="No active academic year set by administrator.",
            status_code=400
        )

    if (
        not auth.is_platform_user(user)
        and scope_school_group_id
        and getattr(user, "school_group_id", None) != scope_school_group_id
    ):
        user.school_group_id = scope_school_group_id

    user.last_login_at = datetime.utcnow()
    db.commit()

    response = RedirectResponse(
        url="/platform" if auth.is_platform_user(user) else "/dashboard",
        status_code=302,
    )
    request.state.audit_actor_user_id = user.user_id
    request.state.audit_actor_username = user.username or ""
    request.state.audit_actor_role = (
        auth.normalize_platform_role(getattr(user, "platform_role", ""))
        or auth.normalize_role(user.role)
    )
    request.state.audit_actor_branch_id = branch_scope_id
    auth.set_auth_session_cookie(response, user, request)
    if auth.is_platform_user(user):
        response.delete_cookie("school_group_id")
        response.delete_cookie("branch_id")
        response.delete_cookie("academic_year_id")
    elif branch_scope_id:
        auth.set_scope_cookie(response, "branch_id", branch_scope_id, request)
    if active_year and not auth.is_platform_user(user):
        auth.set_scope_cookie(response, "academic_year_id", active_year.id, request)
    auth.set_scope_cookie(response, IDLE_TIMEOUT_COOKIE_KEY, int(time.time()), request)
    return response


@app.get("/platform", response_class=HTMLResponse)
def platform_console(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.is_platform_user(current_user):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("system_owner.full_access",),
            page_key="dashboard",
            message="The Platform Console is available only to platform identities.",
        )

    school_groups = db.query(models.SchoolGroup).order_by(models.SchoolGroup.name.asc()).all()
    branches = db.query(models.Branch).all()
    tenant_user_count_by_branch = {
        branch_id: count
        for branch_id, count in db.query(
            models.User.branch_id,
            func.count(models.User.id),
        ).filter(
            models.User.user_type == auth.USER_TYPE_TENANT,
            models.User.branch_id.isnot(None),
        ).group_by(models.User.branch_id).all()
    }
    teacher_count_by_branch = {
        branch_id: count
        for branch_id, count in db.query(
            models.Teacher.branch_id,
            func.count(models.Teacher.id),
        ).filter(
            models.Teacher.branch_id.isnot(None),
        ).group_by(models.Teacher.branch_id).all()
    }
    branches.sort(
        key=lambda branch: (
            branch.school_group_id or 0,
            not bool(branch.status),
            -(
                int(tenant_user_count_by_branch.get(branch.id, 0) or 0)
                + int(teacher_count_by_branch.get(branch.id, 0) or 0)
            ),
            str(branch.name or "").casefold(),
        )
    )
    branch_count_by_group = {
        group_id: count
        for group_id, count in db.query(
            models.Branch.school_group_id,
            func.count(models.Branch.id),
        ).group_by(models.Branch.school_group_id).all()
    }
    tenant_user_count_by_group = {
        group_id: count
        for group_id, count in db.query(
            models.User.school_group_id,
            func.count(models.User.id),
        ).filter(
            models.User.user_type == auth.USER_TYPE_TENANT,
        ).group_by(models.User.school_group_id).all()
    }
    platform_owners = db.query(models.User).filter(
        models.User.user_type == auth.USER_TYPE_PLATFORM,
        models.User.platform_role == auth.PLATFORM_ROLE_OWNER,
    ).order_by(models.User.id.asc()).all()
    platform_developers = db.query(models.User).filter(
        models.User.user_type == auth.USER_TYPE_PLATFORM,
        models.User.platform_role == auth.PLATFORM_ROLE_DEVELOPER,
    ).order_by(models.User.id.asc()).all()
    developer_permission_keys = {
        row.id: (
            {
                permission.permission_key
                for permission in db.query(models.PlatformUserPermission).filter(
                    models.PlatformUserPermission.platform_user_id == row.id,
                    models.PlatformUserPermission.is_allowed == True,
                ).all()
            }
            if bool(getattr(row, "platform_permissions_initialized", False))
            else set(permission_registry.PLATFORM_DEVELOPER_DEFAULT_PERMISSION_KEYS)
        )
        for row in platform_developers
    }
    developer_permission_groups = []
    for group in permission_registry.PERMISSION_GROUPS:
        permissions = [
            {"key": key, "label": label}
            for key, label in group["permissions"]
            if key in permission_registry.DEVELOPER_ASSIGNABLE_PERMISSION_KEYS
        ]
        if permissions:
            developer_permission_groups.append(
                {"key": group["key"], "label": group["label"], "permissions": permissions}
            )

    return templates.TemplateResponse(
        request,
        "platform_console.html",
        {
            "request": request,
            "school_groups": school_groups,
            "branches": branches,
            "branch_count_by_group": branch_count_by_group,
            "tenant_user_count_by_group": tenant_user_count_by_group,
            "selected_school_group_id": getattr(current_user, "scope_school_group_id", None),
            "platform_owners": platform_owners,
            "platform_developers": platform_developers,
            "developer_permission_keys": developer_permission_keys,
            "developer_permission_groups": developer_permission_groups,
            "developer_default_permission_keys": permission_registry.PLATFORM_DEVELOPER_DEFAULT_PERMISSION_KEYS,
            "can_manage_ownership": auth.is_platform_owner(current_user),
            "can_transfer_ownership": auth.is_primary_platform_owner(current_user),
            "owner_account": current_user,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="platform",
                title="Platform Console",
                eyebrow="Global Access",
                intro="Select any organization and branch, then enter its tenant workspace without changing your platform identity.",
                icon="shield",
            ),
        },
    )


def _get_platform_owner_access(request: Request, db: Session, *, primary: bool = False):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    allowed = (
        auth.is_primary_platform_owner(current_user)
        if primary
        else auth.is_platform_owner(current_user)
    )
    if allowed:
        return current_user, None
    return (
        None,
        authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=(
                "system_owner.transfer_ownership"
                if primary
                else "system_owner.manage_ownership"
            ,),
            page_key="platform",
            message="Only the Platform Owner can perform this action.",
        ),
    )


def _platform_account_storage_group_id(db: Session) -> int | None:
    columns = inspect(db.get_bind()).get_columns("users")
    school_group_column = next(
        (column for column in columns if column.get("name") == "school_group_id"),
        None,
    )
    if not school_group_column or bool(school_group_column.get("nullable", True)):
        return None
    return db.query(models.SchoolGroup.id).order_by(models.SchoolGroup.id.asc()).scalar()


def _validate_platform_account_fields(
    db: Session,
    *,
    user_id: str,
    username: str,
    email: str,
    password: str,
) -> list[str]:
    errors = []
    if not re.fullmatch(r"\d{1,10}", user_id):
        errors.append("Platform user ID must be numeric and up to 10 digits.")
    if not re.fullmatch(r"[A-Za-z0-9_.-]{3,50}", username):
        errors.append("Username must be 3-50 letters, numbers, dots, dashes, or underscores.")
    if not auth.is_valid_email(email):
        errors.append("Enter a valid email address.")
    if len(password) < 12:
        errors.append("Platform account passwords must be at least 12 characters.")
    duplicate_conditions = [
        models.User.user_id == user_id,
        models.User.username == username,
    ]
    duplicate = db.query(models.User).filter(or_(*duplicate_conditions)).first()
    if duplicate:
        errors.append("User ID or username already exists.")
    email_error = auth.get_email_registration_error(db, email)
    if email_error:
        errors.append(email_error)
    return errors


def _platform_console_error(message: str):
    return RedirectResponse(
        url=f"/platform?notice={quote_plus(message)}",
        status_code=302,
    )


def _build_platform_owner_verification_url(user, request: Request) -> str:
    token = auth.create_email_verification_token(user)
    return str(
        request.url_for("verify_platform_owner_email").include_query_params(token=token)
    )


def _email_public_base_url(request: Request) -> str:
    configured_url = str(os.getenv("TIS_PUBLIC_BASE_URL") or "").strip()
    return (configured_url or str(request.base_url)).rstrip("/")


def _email_public_asset_url(request: Request, static_path: str) -> str:
    configured_url = str(os.getenv("TIS_PUBLIC_BASE_URL") or "").strip()
    if configured_url:
        return f"{configured_url.rstrip('/')}/static/{quote(static_path, safe='/')}"
    return str(request.url_for("static", path=static_path))


def _email_tis_logo_url(request: Request) -> str:
    return _email_public_asset_url(
        request,
        branding_storage.tis_logo_relative_path(theme="light", compact=True),
    )


def _local_email_verification_log_path() -> str:
    configured_path = str(
        os.getenv("TIS_LOCAL_EMAIL_VERIFICATION_LOG")
        or os.path.join("logs", "email_verification.log")
    ).strip()
    if not os.path.isabs(configured_path):
        configured_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), configured_path)
    return os.path.normpath(os.path.abspath(configured_path))


def _ensure_local_email_verification_log_file() -> str:
    log_path = _local_email_verification_log_path()
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    with open(log_path, "a", encoding="utf-8"):
        pass
    return log_path


def _write_local_email_verification_link(user, verification_url: str) -> str:
    log_path = _ensure_local_email_verification_log_file()
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(
            "\n".join(
                (
                    f"timestamp_utc={datetime.now(timezone.utc).isoformat()}",
                    "purpose=owner_email_verification",
                    f"user_id={str(getattr(user, 'user_id', '') or '')}",
                    f"email={str(getattr(user, 'email', '') or '')}",
                    f"verification_url={verification_url}",
                    "---",
                    "",
                )
            )
        )
        log_file.flush()
        os.fsync(log_file.fileno())
    return log_path


if not auth.is_production_environment() and not email_service.is_resend_configured():
    try:
        _ensure_local_email_verification_log_file()
    except OSError:
        logging.exception("Could not initialize the local email verification log file.")


def _send_platform_owner_verification_email(user, request: Request) -> None:
    verification_url = _build_platform_owner_verification_url(user, request)
    recipient = _clean_email_header(getattr(user, "email", ""), "")
    if not recipient:
        raise ValueError("The Owner account does not have a valid email address.")
    email_content = email_templates.build_email_verification_email(
        verification_url=verification_url,
        logo_url=_email_tis_logo_url(request),
    )
    email_service.send_email(
        to=recipient,
        subject=email_content.subject,
        text=email_content.text,
        html=email_content.html,
    )


@app.post("/platform/account/email")
def update_platform_owner_email(
    request: Request,
    email: str = Form(...),
    current_password: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db)
    if denied:
        return denied

    email = str(email or "").strip()
    email_normalized = auth.normalize_email(email)
    if not auth.verify_password(current_password, current_user.password):
        return _platform_console_error("Current password is incorrect. Email was not changed.")
    if not auth.is_valid_email(email) or not email_normalized:
        return _platform_console_error("Enter a valid email address.")

    email_error = auth.get_email_registration_error(
        db,
        email,
        exclude_user_pk=current_user.id,
    )
    if email_error:
        return _platform_console_error(email_error)

    current_normalized = (
        getattr(current_user, "email_normalized", None)
        or auth.normalize_email(getattr(current_user, "email", None))
    )
    identity_changed = current_normalized != email_normalized
    current_user.email = email
    current_user.email_normalized = email_normalized
    if identity_changed:
        current_user.email_verified_at = None

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return _platform_console_error(auth.EMAIL_ALREADY_REGISTERED_MESSAGE)

    notice = (
        "Owner email updated. Verification is required."
        if identity_changed
        else "Owner email saved."
    )
    return _redirect_with_notice("/platform", notice)


@app.post("/platform/account/request-email-verification")
def request_platform_owner_email_verification(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db)
    if denied:
        return denied
    if not auth.is_valid_email(getattr(current_user, "email", None)):
        return _platform_console_error("Save a valid email address before requesting verification.")
    if getattr(current_user, "email_verified_at", None):
        return _redirect_with_notice("/platform", "Owner email is already verified.")
    if not email_service.is_resend_configured():
        if auth.is_production_environment():
            return _platform_console_error("Email service is not configured.")
        try:
            verification_url = _build_platform_owner_verification_url(current_user, request)
            log_path = _write_local_email_verification_link(current_user, verification_url)
        except Exception:
            logging.exception(
                "Local Owner email verification link creation failed for user_id=%s",
                current_user.user_id,
            )
            return _platform_console_error("Verification link could not be created.")
        return _redirect_with_notice(
            "/platform?verification_local=1",
            "Email service is not configured. Verification link is available in local logs: "
            f"{log_path}"
        )

    try:
        _send_platform_owner_verification_email(current_user, request)
    except email_service.EmailDeliveryError as exc:
        logging.error(
            "Owner verification email delivery failed user_id=%s provider=resend error=%s",
            current_user.user_id,
            exc,
        )
        return _platform_console_error(
            "Verification email could not be sent. Please try again or contact support."
        )
    return _redirect_with_notice(
        "/platform?verification_sent=1",
        "Verification email has been sent.",
    )


@app.get("/platform/account/verify-email")
def verify_platform_owner_email(
    token: str = Query(...),
    db: Session = Depends(get_db),
):
    max_age_seconds = _get_positive_int_env("TIS_EMAIL_VERIFICATION_MAX_AGE_SECONDS", 3600)
    payload = auth.decode_email_verification_token(token, max_age_seconds=max_age_seconds)
    if not payload:
        return PlainTextResponse("This email verification link is invalid or expired.", status_code=400)

    user = db.query(models.User).filter(
        models.User.user_id == str(payload.get("user_id") or "").strip(),
        models.User.user_type == auth.USER_TYPE_PLATFORM,
        models.User.platform_role == auth.PLATFORM_ROLE_OWNER,
    ).first()
    token_email = auth.normalize_email(payload.get("email"))
    current_email = (
        getattr(user, "email_normalized", None)
        or auth.normalize_email(getattr(user, "email", None))
        if user
        else None
    )
    if not user or not token_email or token_email != current_email:
        return PlainTextResponse(
            "This email verification link no longer matches the Owner account.",
            status_code=400,
        )
    if not getattr(user, "email_verified_at", None):
        user.email_verified_at = datetime.utcnow()
        db.commit()
    return _redirect_with_notice("/platform", "Owner email verified successfully.")


@app.post("/platform/account/password")
def update_platform_owner_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db)
    if denied:
        return denied

    current_password = str(current_password or "")
    new_password = str(new_password or "")
    confirm_password = str(confirm_password or "")
    if not auth.verify_password(current_password, current_user.password):
        return _platform_console_error("Current password is incorrect. Password was not changed.")
    if len(new_password) < 12:
        return _platform_console_error("New password must be at least 12 characters.")
    if new_password != confirm_password:
        return _platform_console_error("New password confirmation does not match.")
    if auth.verify_password(new_password, current_user.password):
        return _platform_console_error("New password must be different from the current password.")

    current_user.password = get_password_hash(new_password)
    db.commit()
    return _redirect_with_notice("/platform", "Owner password changed successfully.")


@app.post("/platform/developers")
def create_platform_developer(
    request: Request,
    user_id: str = Form(..., alias="developer_user_id"),
    username: str = Form(...),
    email: str = Form(...),
    first_name: str = Form(...),
    last_name: str = Form(...),
    password: str = Form(...),
    permission_keys: list[str] = Form([]),
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db)
    if denied:
        return denied
    user_id = str(user_id or "").strip()
    username = str(username or "").strip().lower()
    email = str(email or "").strip()
    email_normalized = auth.normalize_email(email)
    password = str(password or "")
    errors = _validate_platform_account_fields(
        db,
        user_id=user_id,
        username=username,
        email=email,
        password=password,
    )
    if errors:
        return _platform_console_error(" ".join(errors))
    allowed_permissions = set(permission_keys) & set(
        permission_registry.DEVELOPER_ASSIGNABLE_PERMISSION_KEYS
    )
    developer = models.User(
        user_id=user_id,
        username=username,
        email=email,
        email_normalized=email_normalized,
        first_name=str(first_name or "").strip(),
        last_name=str(last_name or "").strip(),
        password=get_password_hash(password),
        user_type=auth.USER_TYPE_PLATFORM,
        platform_role=auth.PLATFORM_ROLE_DEVELOPER,
        platform_owner_kind=None,
        platform_permissions_initialized=True,
        access_scope=auth.ACCESS_SCOPE_GLOBAL,
        role=None,
        position=None,
        school_group_id=_platform_account_storage_group_id(db),
        branch_id=None,
        academic_year_id=None,
        is_active=True,
    )
    db.add(developer)
    db.flush()
    now = datetime.utcnow()
    for permission_key in sorted(allowed_permissions):
        db.add(
            models.PlatformUserPermission(
                platform_user_id=developer.id,
                permission_key=permission_key,
                is_allowed=True,
                updated_by_user_id=current_user.user_id,
                created_at=now,
                updated_at=now,
            )
        )
    db.commit()
    return _redirect_with_notice("/platform", "Platform Developer created.")


@app.post("/platform/developers/{developer_id}/permissions")
def update_platform_developer_permissions(
    developer_id: int,
    request: Request,
    permission_keys: list[str] = Form([]),
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db)
    if denied:
        return denied
    developer = db.query(models.User).filter(
        models.User.id == developer_id,
        models.User.user_type == auth.USER_TYPE_PLATFORM,
        models.User.platform_role == auth.PLATFORM_ROLE_DEVELOPER,
    ).first()
    if not developer:
        return _platform_console_error("Developer account not found.")
    allowed_permissions = set(permission_keys) & set(
        permission_registry.DEVELOPER_ASSIGNABLE_PERMISSION_KEYS
    )
    db.query(models.PlatformUserPermission).filter(
        models.PlatformUserPermission.platform_user_id == developer.id
    ).delete(synchronize_session=False)
    now = datetime.utcnow()
    for permission_key in sorted(allowed_permissions):
        db.add(
            models.PlatformUserPermission(
                platform_user_id=developer.id,
                permission_key=permission_key,
                is_allowed=True,
                updated_by_user_id=current_user.user_id,
                created_at=now,
                updated_at=now,
            )
        )
    developer.platform_permissions_initialized = True
    db.commit()
    return _redirect_with_notice("/platform", "Developer permissions updated.")


@app.post("/platform/owners")
def create_platform_co_owner(
    request: Request,
    user_id: str = Form(..., alias="co_owner_user_id"),
    username: str = Form(...),
    email: str = Form(...),
    first_name: str = Form(...),
    last_name: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db)
    if denied:
        return denied
    user_id = str(user_id or "").strip()
    username = str(username or "").strip().lower()
    email = str(email or "").strip()
    email_normalized = auth.normalize_email(email)
    password = str(password or "")
    errors = _validate_platform_account_fields(
        db,
        user_id=user_id,
        username=username,
        email=email,
        password=password,
    )
    if errors:
        return _platform_console_error(" ".join(errors))
    co_owner = models.User(
        user_id=user_id,
        username=username,
        email=email,
        email_normalized=email_normalized,
        first_name=str(first_name or "").strip(),
        last_name=str(last_name or "").strip(),
        password=get_password_hash(password),
        user_type=auth.USER_TYPE_PLATFORM,
        platform_role=auth.PLATFORM_ROLE_OWNER,
        platform_owner_kind=auth.PLATFORM_OWNER_CO_OWNER,
        platform_permissions_initialized=False,
        access_scope=auth.ACCESS_SCOPE_GLOBAL,
        role=None,
        position=None,
        school_group_id=_platform_account_storage_group_id(db),
        branch_id=None,
        academic_year_id=None,
        is_active=True,
    )
    db.add(co_owner)
    db.commit()
    return _redirect_with_notice("/platform", "Co-Owner created.")


@app.post("/platform/ownership/transfer")
def transfer_platform_ownership(
    request: Request,
    target_user_id: str = Form(...),
    current_password: str = Form(...),
    confirmation: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user, denied = _get_platform_owner_access(request, db, primary=True)
    if denied:
        return denied
    if str(confirmation or "").strip() != "TRANSFER OWNERSHIP":
        return _platform_console_error("Type TRANSFER OWNERSHIP exactly to confirm.")
    if not auth.verify_password(str(current_password or ""), current_user.password):
        return _platform_console_error("Current owner password is incorrect.")
    target = db.query(models.User).filter(
        models.User.user_id == str(target_user_id or "").strip(),
        models.User.user_type == auth.USER_TYPE_PLATFORM,
        models.User.is_active == True,
    ).first()
    if not target or target.id == current_user.id:
        return _platform_console_error("Select another active platform user.")
    db.query(models.User).filter(
        models.User.user_type == auth.USER_TYPE_PLATFORM,
        models.User.platform_owner_kind == auth.PLATFORM_OWNER_PRIMARY,
    ).update(
        {models.User.platform_owner_kind: auth.PLATFORM_OWNER_CO_OWNER},
        synchronize_session=False,
    )
    target.platform_role = auth.PLATFORM_ROLE_OWNER
    target.platform_owner_kind = auth.PLATFORM_OWNER_PRIMARY
    target.platform_permissions_initialized = False
    db.query(models.PlatformUserPermission).filter(
        models.PlatformUserPermission.platform_user_id == target.id
    ).delete(synchronize_session=False)
    db.commit()
    return _redirect_with_notice("/platform", "Primary ownership transferred.")


# ---------------------------------------
# LOGOUT
# ---------------------------------------

# ---------------------------------------
# FORGOT PASSWORD
# ---------------------------------------
def _password_reset_email_recipient(db: Session) -> str:
    platform_recipient = db.query(models.User).filter(
        models.User.user_id == DEVELOPER_USER_ID,
        models.User.user_type == auth.USER_TYPE_PLATFORM,
    ).first()
    account_email = str(getattr(platform_recipient, "email", "") or "").strip()
    if auth.is_valid_email(account_email):
        return account_email
    fallback_email = str(os.getenv("EMAIL_REPLY_TO") or "").strip()
    return fallback_email if auth.is_valid_email(fallback_email) else ""


def _send_password_reset_request_email(
    db: Session,
    *,
    request: Request,
    user_id: str,
    user_display: str,
) -> None:
    recipient = _password_reset_email_recipient(db)
    if not recipient:
        raise email_service.EmailDeliveryError(
            "No valid platform recipient email is configured for password reset requests."
        )
    email_content = email_templates.build_password_reset_request_email(
        requester_display=user_display,
        user_id=user_id,
        platform_url=f"{_email_public_base_url(request)}/notifications",
        logo_url=_email_tis_logo_url(request),
    )
    email_service.send_email(
        to=recipient,
        subject=email_content.subject,
        text=email_content.text,
        html=email_content.html,
    )


@app.post("/forgot-password")
async def forgot_password(
    request: Request,
    db: Session = Depends(get_db),
):
    logging.info(
        "TIS notification: forgot-password request received route=/forgot-password method=%s ip=%s",
        request.method,
        getattr(getattr(request, "client", None), "host", "unknown"),
    )

    payload = None
    try:
        payload = await request.json()
    except Exception:
        payload = None

    if payload is None or not isinstance(payload, dict):
        try:
            form_data = await request.form()
            payload = dict(form_data)
        except Exception:
            payload = None

    if payload is None or not isinstance(payload, dict):
        logging.error("TIS notification: forgot-password payload could not be parsed")
        return JSONResponse(
            status_code=400,
            content={"ok": False, "message": "Invalid request format."},
        )

    user_id = str(payload.get("user_id") or "").strip()
    if not user_id:
        logging.warning("TIS notification: forgot-password request missing user_id")
        return JSONResponse(
            status_code=400,
            content={"ok": False, "message": "Please enter your User ID."},
        )

    logging.info("TIS notification: forgot-password processing user_id=%s", user_id)

    try:
        user = db.query(models.User).filter(
            models.User.user_id == user_id
        ).first()
    except Exception as exc:
        logging.error(
            "TIS notification: forgot-password user lookup failed for user_id=%s: %s",
            user_id,
            exc,
            exc_info=True,
        )
        user = None

    user_display = _build_user_display(user_id, user)
    logging.info(
        "TIS notification: forgot-password user lookup complete user_id=%s found=%s recipient_user_id=%s",
        user_id,
        bool(user),
        DEVELOPER_USER_ID,
    )

    try:
        notification = _create_system_notification(
            db,
            recipient_user_id=DEVELOPER_USER_ID,
            requesting_user_id=user_id,
            request_type=NOTIFICATION_TYPE_FORGOT_PASSWORD,
            title="Forgot Password Request",
            message=f"{user_display} requested a password reset.",
            details=(
                "Password reset requested from the login page. "
                "Review the user account and reset the password manually."
            ),
        )
    except Exception as exc:
        db.rollback()
        logging.error(
            "TIS notification: failed to create forgot-password notification for user_id=%s recipient_user_id=%s: %s",
            user_id,
            DEVELOPER_USER_ID,
            exc,
            exc_info=True,
        )
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "message": "The request could not be saved. Please contact the system administrator.",
            },
        )

    logging.info(
        "TIS notification: forgot-password notification saved id=%s recipient_user_id=%s requesting_user_id=%s status=%s",
        notification.id,
        notification.recipient_user_id,
        notification.requesting_user_id,
        notification.status,
    )

    if email_service.is_resend_configured():
        try:
            _send_password_reset_request_email(
                db,
                request=request,
                user_id=user_id,
                user_display=user_display,
            )
        except email_service.EmailDeliveryError as exc:
            logging.error(
                "Password reset email delivery failed user_id=%s provider=resend error=%s",
                user_id,
                exc,
            )
            return JSONResponse(
                status_code=502,
                content={
                    "ok": False,
                    "message": (
                        "Your request was saved inside TIS, but the email notification "
                        "could not be delivered. Please contact the system administrator."
                    ),
                },
            )
    elif auth.is_production_environment():
        return JSONResponse(
            status_code=503,
            content={
                "ok": False,
                "message": (
                    "Your request was saved inside TIS, but email service is not configured. "
                    "Please contact the system administrator."
                ),
            },
        )

    return JSONResponse(
        content={
            "ok": True,
            "message": (
                "Your request has been sent to the system administrator inside TIS. "
                "They will review it and reset your password manually."
            ),
        }
    )


@app.get("/notifications")
def notification_center(
    request: Request,
    status: str = Query(""),
    box: str = Query("inbox"),
    db: Session = Depends(get_db),
):
    current_user = None
    messages = []
    counts = {}
    selected_status = ""
    selected_box = "inbox"
    template_context = None
    try:
        _ensure_system_notifications_table_columns()
        current_user = auth.get_current_user(request, db)
        if not current_user:
            _notification_logger().info(
                "TIS notification center opened without authenticated user path=%s",
                request.url.path,
            )
            return RedirectResponse(url="/", status_code=302)
        archived_count = _auto_archive_old_notifications(db, current_user.user_id)
        if archived_count:
            _notification_logger().info(
                "TIS notification: auto-archived old messages user_id=%s count=%s",
                current_user.user_id,
                archived_count,
            )

        selected_box = str(box or "inbox").strip().lower()
        if selected_box not in {"inbox", "sent", "archive"}:
            selected_box = "inbox"

        allowed_statuses = {
            NOTIFICATION_STATUS_NEW,
            NOTIFICATION_STATUS_SEEN,
            NOTIFICATION_STATUS_RESOLVED,
        }
        selected_status = str(status or "").strip()
        if selected_status not in allowed_statuses:
            selected_status = ""

        query = _base_notification_query_for_box(db, current_user.user_id, selected_box)
        if selected_status:
            query = query.filter(models.SystemNotification.status == selected_status)

        messages = query.order_by(
            models.SystemNotification.created_at.desc(),
            models.SystemNotification.id.desc(),
        ).limit(MESSAGE_PAGE_SIZE).all()
        user_display_map = _user_notification_display_map(db, messages)
        notification_groups = _build_notification_groups(messages)
        counts = _get_notification_counts(db, current_user.user_id)
        _notification_logger().info(
            (
                "TIS notification center opened user_id=%s selected_status=%s "
                "message_count=%s counts=%s recipient_filter=%s"
            ),
            current_user.user_id,
            selected_status or "All",
            len(messages),
            counts,
            current_user.user_id,
        )

        can_compose = auth.has_any_permission(
            db,
            current_user,
            "notifications.send_direct",
            "notifications.broadcast",
        )
        all_users = []
        if can_compose:
            all_users = auth.get_notification_recipient_query(db, current_user).order_by(
                models.User.first_name.asc(),
                models.User.last_name.asc(),
            ).all()

        user_timezone = _get_request_timezone(request)
        template_context = {
            "request": request,
            "current_user": current_user,
            "messages": messages,
            "notification_groups": notification_groups,
            "user_display_map": user_display_map,
            "selected_status": selected_status,
            "selected_box": selected_box,
            "notification_counts": counts,
            "user_timezone": user_timezone,
            "format_notification_timestamp": (
                lambda value, fallback="Unknown": _format_notification_timestamp(
                    value,
                    fallback,
                    user_timezone,
                )
            ),
            "notification_timestamp_iso": _notification_timestamp_iso,
            "status_options": [
                NOTIFICATION_STATUS_NEW,
                NOTIFICATION_STATUS_SEEN,
                NOTIFICATION_STATUS_RESOLVED,
            ],
            "can_compose": can_compose,
            "all_users": all_users,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="notifications",
            ),
        }
        return templates.TemplateResponse(
            request,
            "notifications.html",
            template_context,
        )
    except Exception as exc:
        user_id = getattr(current_user, "user_id", None)
        safe_total_count = _get_notification_total_count_safe(db, user_id)
        schema_snapshot = _collect_system_notification_schema_snapshot()
        context_keys = sorted(template_context.keys()) if isinstance(template_context, dict) else []
        diagnostic_payload = {
            "route": request.url.path,
            "handler": "notification_center",
            "user_id": user_id,
            "selected_status": selected_status or "All",
            "selected_box": selected_box,
            "messages_count": len(messages),
            "notification_total_count": safe_total_count,
            "template_context_keys": context_keys,
            "schema": schema_snapshot,
        }
        _notification_logger().error(
            (
                "TIS notification diagnostic failure route=%s handler=%s user_id=%s "
                "messages_count=%s notification_total_count=%s schema_table=%s "
                "schema_actual_columns=%s schema_missing_columns=%s template_context_keys=%s "
                "exception=%s"
            ),
            request.url.path,
            "notification_center",
            user_id,
            len(messages),
            safe_total_count,
            schema_snapshot.get("table_name"),
            schema_snapshot.get("actual_columns"),
            schema_snapshot.get("missing_columns"),
            context_keys,
            f"{exc.__class__.__name__}: {exc}",
            exc_info=True,
        )
        return _render_notification_error_fallback(
            request,
            route_name="notification_center",
            error_message=f"{exc.__class__.__name__}: {exc}",
            diagnostic_payload=diagnostic_payload,
        )


@app.get("/notifications/compose")
def compose_message_form(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    can_compose = auth.has_any_permission(
        db,
        current_user,
        "notifications.send_direct",
        "notifications.broadcast",
    )
    if not can_compose:
        return RedirectResponse(
            url="/notifications?notice=You%20do%20not%20have%20permission%20to%20send%20messages.",
            status_code=302,
        )

    all_users = auth.get_notification_recipient_query(db, current_user).order_by(
        models.User.first_name.asc(),
        models.User.last_name.asc(),
    ).all()

    return templates.TemplateResponse(
        request,
        "compose_message.html",
        {
            "request": request,
            "current_user": current_user,
            "all_users": all_users,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="notifications",
                eyebrow="Notification Center",
                title="Compose Message",
            ),
        },
    )


@app.post("/notifications/mark-all-read")
def mark_all_notifications_read(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    _ensure_system_notifications_table_columns()
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    updated_count = db.query(models.SystemNotification).filter(
        models.SystemNotification.recipient_user_id == current_user.user_id,
        models.SystemNotification.status == NOTIFICATION_STATUS_NEW,
        models.SystemNotification.recipient_archived_at.is_(None),
    ).update(
        {
            models.SystemNotification.status: NOTIFICATION_STATUS_SEEN,
            models.SystemNotification.seen_at: now,
        },
        synchronize_session=False,
    )
    db.commit()
    _notification_logger().info(
        "TIS notification: marked all read user_id=%s count=%s",
        current_user.user_id,
        updated_count,
    )
    return RedirectResponse(
        url=f"/notifications?notice={updated_count}%20message(s)%20marked%20as%20read.",
        status_code=302,
    )


@app.post("/notifications/group-action")
def update_notification_group(
    request: Request,
    group_key: str = Form(...),
    action: str = Form(...),
    box: str = Form("inbox"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    normalized_action = str(action or "").strip().lower()
    if normalized_action not in {"read", "done", "archive"}:
        return RedirectResponse(url="/notifications?notice=Unknown%20notification%20action.", status_code=302)
    normalized_box = str(box or "inbox").strip().lower()
    if normalized_box not in {"inbox", "sent", "archive"}:
        normalized_box = "inbox"
    if normalized_action in {"read", "done"} and normalized_box != "inbox":
        return RedirectResponse(
            url=f"/notifications?box={normalized_box}&notice=Only%20the%20recipient%20can%20mark%20a%20task%20read%20or%20done.",
            status_code=302,
        )

    _ensure_system_notifications_table_columns()
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    notifications = _base_notification_query_for_box(db, current_user.user_id, normalized_box).all()
    updated_count = 0
    for notification in notifications:
        if not _notification_group_matches(notification, group_key):
            continue
        if normalized_action == "read" and notification.status == NOTIFICATION_STATUS_NEW:
            notification.status = NOTIFICATION_STATUS_SEEN
            notification.seen_at = now
            updated_count += 1
        elif normalized_action == "done" and notification.status != NOTIFICATION_STATUS_RESOLVED:
            notification.status = NOTIFICATION_STATUS_RESOLVED
            if not notification.seen_at:
                notification.seen_at = now
            notification.resolved_at = now
            notification.resolved_by_user_id = current_user.user_id
            updated_count += 1
        elif normalized_action == "archive":
            if normalized_box == "sent":
                notification.requester_archived_at = now
                notification.requester_archived_by_user_id = current_user.user_id
            else:
                notification.recipient_archived_at = now
                notification.recipient_archived_by_user_id = current_user.user_id
            updated_count += 1
    db.commit()
    _notification_logger().info(
        "TIS notification: group action user_id=%s group_key=%s action=%s count=%s",
        current_user.user_id,
        group_key,
        normalized_action,
        updated_count,
    )
    if normalized_action == "read":
        notice = "marked%20as%20read"
    elif normalized_action == "done":
        notice = "marked%20done"
    else:
        notice = "archived"
    return RedirectResponse(
        url=f"/notifications?box={normalized_box}&notice={updated_count}%20message(s)%20{notice}.",
        status_code=302,
    )


@app.post("/notifications/compose")
def send_message(
    request: Request,
    title: str = Form(...),
    recipient: str = Form(...),
    message: str = Form(...),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    title = str(title or "").strip()
    recipient = str(recipient or "").strip()
    message = str(message or "").strip()

    if not title or not recipient or not message:
        return RedirectResponse(
            url="/notifications/compose?notice=Title%2C%20recipient%2C%20and%20message%20are%20required.",
            status_code=302,
        )

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    allowed_recipient_query = auth.get_notification_recipient_query(db, current_user)
    required_notification_permission = (
        "notifications.broadcast"
        if recipient == "ALL"
        else "notifications.send_direct"
    )
    if not auth.has_permission(db, current_user, required_notification_permission):
        return RedirectResponse(
            url="/notifications?notice=You%20do%20not%20have%20permission%20to%20send%20messages.",
            status_code=302,
        )

    if recipient == "ALL":
        active_users = allowed_recipient_query.all()
        count = 0
        for user in active_users:
            n = models.SystemNotification(
                school_group_id=auth.get_notification_school_group_id(
                    db,
                    recipient_user=user,
                    current_user=current_user,
                ),
                branch_id=getattr(user, "branch_id", None),
                academic_year_id=getattr(user, "academic_year_id", None),
                recipient_user_id=user.user_id,
                requesting_user_id=current_user.user_id,
                request_type=NOTIFICATION_TYPE_MESSAGE,
                title=title,
                message=message,
                details="",
                status=NOTIFICATION_STATUS_NEW,
                recipient_scope=NOTIFICATION_SCOPE_ALL,
                created_at=now,
            )
            db.add(n)
            count += 1
        db.commit()
        logging.info(
            "TIS message: broadcast sent by user_id=%s to %d users title=%s",
            current_user.user_id,
            count,
            title,
        )
    else:
        target_user = allowed_recipient_query.filter(
            models.User.user_id == recipient,
        ).first()
        if not target_user:
            return RedirectResponse(
                url="/notifications/compose?notice=Selected%20user%20not%20found.",
                status_code=302,
            )
        n = models.SystemNotification(
            school_group_id=auth.get_notification_school_group_id(
                db,
                recipient_user=target_user,
                current_user=current_user,
            ),
            branch_id=getattr(target_user, "branch_id", None),
            academic_year_id=getattr(target_user, "academic_year_id", None),
            recipient_user_id=target_user.user_id,
            requesting_user_id=current_user.user_id,
            request_type=NOTIFICATION_TYPE_MESSAGE,
            title=title,
            message=message,
            details="",
            status=NOTIFICATION_STATUS_NEW,
            recipient_scope=NOTIFICATION_SCOPE_USER,
            created_at=now,
        )
        db.add(n)
        db.commit()
        logging.info(
            "TIS message: sent by user_id=%s to user_id=%s title=%s",
            current_user.user_id,
            target_user.user_id,
            title,
        )

    return RedirectResponse(
        url="/notifications?notice=Message%20sent%20successfully.",
        status_code=302,
    )


@app.get("/notifications/{notification_id}")
def notification_detail(
    notification_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = None
    notification = None
    template_context = None
    try:
        _ensure_system_notifications_table_columns()
        current_user = auth.get_current_user(request, db)
        if not current_user:
            _notification_logger().info(
                "TIS notification detail opened without authenticated user notification_id=%s",
                notification_id,
            )
            return RedirectResponse(url="/", status_code=302)

        notification, redirect_response = _get_user_notification_or_redirect(
            db,
            current_user,
            notification_id,
        )
        if redirect_response:
            _notification_logger().info(
                "TIS notification detail not found user_id=%s notification_id=%s",
                current_user.user_id,
                notification_id,
            )
            return redirect_response

        is_recipient = notification.recipient_user_id == current_user.user_id
        # Auto-mark as seen only for the assigned recipient.
        if is_recipient and notification.status == NOTIFICATION_STATUS_NEW:
            notification.status = NOTIFICATION_STATUS_SEEN
            notification.seen_at = datetime.now(timezone.utc).replace(tzinfo=None)
            db.commit()
            _notification_logger().info(
                "TIS notification detail auto-marked seen user_id=%s notification_id=%s",
                current_user.user_id,
                notification.id,
            )

        resolved_by_user = None
        if notification.resolved_by_user_id:
            resolved_by_user = db.query(models.User).filter(
                models.User.user_id == notification.resolved_by_user_id
            ).first()

        sender_user = None
        if notification.requesting_user_id:
            sender_user = db.query(models.User).filter(
                models.User.user_id == notification.requesting_user_id
            ).first()
        user_timezone = _get_request_timezone(request)

        _notification_logger().info(
            "TIS notification detail opened user_id=%s notification_id=%s status=%s request_type=%s",
            current_user.user_id,
            notification.id,
            notification.status,
            notification.request_type,
        )

        template_context = {
            "request": request,
            "current_user": current_user,
            "notification": notification,
            "resolved_by_user": resolved_by_user,
            "sender_user": sender_user,
            "is_recipient": is_recipient,
            "can_mark_done": bool(is_recipient and notification.status != NOTIFICATION_STATUS_RESOLVED),
            "can_archive_notification": bool(
                (is_recipient and not notification.recipient_archived_at)
                or (
                    notification.requesting_user_id == current_user.user_id
                    and not notification.requester_archived_at
                )
            ),
            "user_timezone": user_timezone,
            "format_notification_timestamp": (
                lambda value, fallback="Unknown": _format_notification_timestamp(
                    value,
                    fallback,
                    user_timezone,
                )
            ),
            "notification_timestamp_iso": _notification_timestamp_iso,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="notifications",
            ),
        }
        return templates.TemplateResponse(
            request,
            "notification_detail.html",
            template_context,
        )
    except Exception as exc:
        user_id = getattr(current_user, "user_id", None)
        safe_total_count = _get_notification_total_count_safe(db, user_id)
        schema_snapshot = _collect_system_notification_schema_snapshot()
        context_keys = sorted(template_context.keys()) if isinstance(template_context, dict) else []
        diagnostic_payload = {
            "route": request.url.path,
            "handler": "notification_detail",
            "notification_id": notification_id,
            "loaded_notification_id": getattr(notification, "id", None),
            "user_id": user_id,
            "notification_total_count": safe_total_count,
            "template_context_keys": context_keys,
            "schema": schema_snapshot,
        }
        _notification_logger().error(
            (
                "TIS notification diagnostic failure route=%s handler=%s notification_id=%s "
                "loaded_notification_id=%s user_id=%s notification_total_count=%s "
                "schema_table=%s schema_actual_columns=%s schema_missing_columns=%s "
                "template_context_keys=%s exception=%s"
            ),
            request.url.path,
            "notification_detail",
            notification_id,
            getattr(notification, "id", None),
            user_id,
            safe_total_count,
            schema_snapshot.get("table_name"),
            schema_snapshot.get("actual_columns"),
            schema_snapshot.get("missing_columns"),
            context_keys,
            f"{exc.__class__.__name__}: {exc}",
            exc_info=True,
        )
        return _render_notification_error_fallback(
            request,
            route_name="notification_detail",
            error_message=f"{exc.__class__.__name__}: {exc}",
            diagnostic_payload=diagnostic_payload,
        )


@app.post("/notifications/{notification_id}/resolved")
def mark_notification_resolved(
    notification_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    notification, redirect_response = _get_user_notification_or_redirect(
        db,
        current_user,
        notification_id,
    )
    if redirect_response:
        return redirect_response
    if notification.recipient_user_id != current_user.user_id:
        return RedirectResponse(
            url=f"/notifications/{notification.id}?notice=Only%20the%20assigned%20recipient%20can%20mark%20this%20task%20done.",
            status_code=302,
        )

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    notification.status = NOTIFICATION_STATUS_RESOLVED
    if not notification.seen_at:
        notification.seen_at = now
    notification.resolved_at = now
    notification.resolved_by_user_id = current_user.user_id
    db.commit()
    logging.info(
        "TIS notification: marked done id=%s user_id=%s",
        notification.id,
        current_user.user_id,
    )

    return RedirectResponse(
        url=f"/notifications/{notification.id}?notice=Message%20marked%20done.",
        status_code=302,
    )


@app.post("/notifications/{notification_id}/archive")
def archive_notification(
    notification_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    notification, redirect_response = _get_user_notification_or_redirect(
        db,
        current_user,
        notification_id,
    )
    if redirect_response:
        return redirect_response

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    archived = False
    if notification.recipient_user_id == current_user.user_id and not notification.recipient_archived_at:
        notification.recipient_archived_at = now
        notification.recipient_archived_by_user_id = current_user.user_id
        archived = True
    if notification.requesting_user_id == current_user.user_id and not notification.requester_archived_at:
        notification.requester_archived_at = now
        notification.requester_archived_by_user_id = current_user.user_id
        archived = True
    if archived:
        db.commit()
    return RedirectResponse(
        url="/notifications?notice=Message%20archived.",
        status_code=302,
    )


@app.get("/logout")
def logout():
    response = RedirectResponse(url="/", status_code=302)
    return _clear_auth_session_cookies(response)


@app.post("/profile/photo")
async def upload_profile_photo(
    request: Request,
    profile_photo: UploadFile = File(...),
    return_to: str = Form("/dashboard"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    safe_return_to = _safe_redirect_path(return_to)
    if not profile_photo or not profile_photo.filename:
        return _redirect_with_notice(
            safe_return_to,
            "Choose an image file before saving your profile photo.",
        )

    file_bytes = await profile_photo.read()
    if not file_bytes:
        return _redirect_with_notice(
            safe_return_to,
            "The selected file was empty. Choose another image.",
        )
    if len(file_bytes) > PROFILE_PHOTO_MAX_BYTES:
        return _redirect_with_notice(
            safe_return_to,
            "Profile photo must be 3 MB or smaller.",
        )

    detected_extension = _detect_profile_photo_extension(file_bytes)
    if not detected_extension:
        return _redirect_with_notice(
            safe_return_to,
            "Use a PNG, JPG, GIF, or WEBP image for the profile photo.",
        )

    _ensure_profile_photo_upload_dir()
    file_name = f"user_{current_user.id}_{int(time.time() * 1000)}{detected_extension}"
    relative_path = f"{PROFILE_PHOTO_RELATIVE_DIR}/{file_name}"
    absolute_path = os.path.join(PROFILE_PHOTO_UPLOAD_DIR, file_name)
    profile_media_type = _profile_photo_media_type_from_extension(detected_extension)

    try:
        with open(absolute_path, "wb") as output_file:
            output_file.write(file_bytes)
    except OSError:
        # Database-backed storage is the primary persistence mechanism.
        pass

    previous_profile_image_path = str(
        getattr(current_user, "profile_image_path", "") or ""
    ).strip()
    current_user.profile_image_path = relative_path
    current_user.profile_image_content_type = profile_media_type
    current_user.profile_image_data = file_bytes
    db.commit()
    _delete_profile_photo_file(previous_profile_image_path)

    return _redirect_with_notice(
        safe_return_to,
        "Profile photo updated successfully.",
    )


@app.get("/profile/photo/current")
def get_current_profile_photo(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return PlainTextResponse("Unauthorized", status_code=401)

    profile_image_data = getattr(current_user, "profile_image_data", None)
    if profile_image_data:
        payload = bytes(profile_image_data)
        media_type = str(
            getattr(current_user, "profile_image_content_type", "") or ""
        ).strip() or "application/octet-stream"
        return StreamingResponse(
            io.BytesIO(payload),
            media_type=media_type,
            headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
        )

    normalized_relative_path = _normalize_profile_photo_relative_path(
        getattr(current_user, "profile_image_path", "") or ""
    )
    if normalized_relative_path.startswith(f"{PROFILE_PHOTO_RELATIVE_DIR}/"):
        absolute_path = os.path.abspath(
            os.path.join("static", *normalized_relative_path.split("/"))
        )
        upload_root = os.path.abspath(PROFILE_PHOTO_UPLOAD_DIR)
        if absolute_path.startswith(upload_root) and os.path.exists(absolute_path):
            return FileResponse(
                absolute_path,
                headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
            )

    return PlainTextResponse("Profile photo not found", status_code=404)


# ---------------------------------------
# DEVELOPER: DOWNLOAD AUDIT LOG
# ---------------------------------------
@app.get("/admin/audit-log")
def download_audit_log(
    request: Request,
    format: str = Query(default="xlsx"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    if not auth.has_permission(db, current_user, "configuration.export_audit_log"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("configuration.export_audit_log",),
            page_key="system-configuration",
        )

    audit_log_path = get_audit_log_path()
    if not audit_log_path.exists():
        return PlainTextResponse(
            "Audit log file has not been created yet.",
            status_code=404
        )

    download_format = str(format).strip().lower()
    if download_format not in {"xlsx", "csv", "raw"}:
        return PlainTextResponse(
            "Unsupported format. Use ?format=xlsx or ?format=csv or ?format=raw.",
            status_code=400
        )

    if download_format == "xlsx":
        try:
            payload = build_audit_xlsx_bytes(audit_log_path)
        except OSError:
            return PlainTextResponse(
                "Audit log file is temporarily unavailable. Please retry in a moment.",
                status_code=503
            )

        response = StreamingResponse(
            iter([payload]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Content-Disposition"] = (
            f"attachment; filename={get_audit_xlsx_filename()}"
        )
        return response

    if download_format == "csv":
        response = StreamingResponse(
            iter_audit_csv_bytes(audit_log_path),
            media_type="text/csv",
        )
        response.headers["Content-Disposition"] = (
            f"attachment; filename={get_audit_csv_filename()}"
        )
        return response

    try:
        file_handle = open(audit_log_path, "rb")
    except OSError:
        return PlainTextResponse(
            "Audit log file is temporarily unavailable. Please retry in a moment.",
            status_code=503
        )

    def _iter_audit_file():
        with file_handle:
            while True:
                chunk = file_handle.read(64 * 1024)
                if not chunk:
                    break
                yield chunk

    response = StreamingResponse(
        _iter_audit_file(),
        media_type="text/plain",
    )
    response.headers["Content-Disposition"] = (
        f"attachment; filename={audit_log_path.name}"
    )
    return response


# ---------------------------------------
# PLATFORM: DEMO REQUESTS
# ---------------------------------------
def _get_demo_requests_access(request: Request, db: Session, *permission_keys: str):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    required_permissions = permission_keys or ("demo_requests.view",)
    if not any(auth.has_permission(db, current_user, permission_key) for permission_key in required_permissions):
        return (
            None,
            authorization.build_access_denied_response(
                request,
                db,
                current_user=current_user,
                permission_keys=required_permissions,
                page_key="demo-requests",
            ),
        )
    return current_user, None


def _normalize_demo_request_status(status: str) -> str:
    cleaned_status = str(status or "").strip()
    for option in DEMO_REQUEST_STATUSES:
        if cleaned_status.lower() == option.lower():
            return option
    return DEMO_REQUEST_STATUSES[0]


def _demo_request_filters(request: Request) -> dict:
    return {
        "school_name": str(request.query_params.get("school_name", "") or "").strip(),
        "email": str(request.query_params.get("email", "") or "").strip(),
        "interested_plan": str(request.query_params.get("interested_plan", "") or "").strip(),
        "status": str(request.query_params.get("status", "") or "").strip(),
    }


def _apply_demo_request_filters(query, filters: dict):
    school_name = filters.get("school_name", "")
    email = filters.get("email", "")
    interested_plan = filters.get("interested_plan", "")
    status = filters.get("status", "")

    if school_name:
        query = query.filter(models.DemoRequest.school_name.ilike(f"%{school_name}%"))
    if email:
        query = query.filter(models.DemoRequest.email.ilike(f"%{email}%"))
    if interested_plan:
        query = query.filter(models.DemoRequest.interested_plan == interested_plan)
    if status:
        query = query.filter(models.DemoRequest.status == _normalize_demo_request_status(status))
    return query


def _demo_request_filter_query_string(filters: dict) -> str:
    return urlencode(
        {
            key: value
            for key, value in filters.items()
            if str(value or "").strip()
        }
    )


def _demo_request_status_counts(db: Session) -> dict[str, int]:
    counts = {status: 0 for status in DEMO_REQUEST_STATUSES}
    rows = (
        db.query(models.DemoRequest.status, func.count(models.DemoRequest.id))
        .group_by(models.DemoRequest.status)
        .all()
    )
    for status, count in rows:
        counts[_normalize_demo_request_status(status)] = int(count or 0)
    return counts


def _demo_request_export_filename() -> str:
    return f"demo_requests_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"


def _write_demo_requests_csv(rows) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Submitted At",
            "School Name",
            "Full Name",
            "Email",
            "Phone",
            "Country",
            "School Type",
            "Number of Teachers",
            "Number of Students",
            "Number of Branches",
            "Interested Plan",
            "Status",
            "Message",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                _format_notification_timestamp(row.submitted_at, ""),
                row.school_name or "",
                row.full_name or "",
                row.email or "",
                row.phone or "",
                row.country or "",
                row.school_type or "",
                row.number_of_teachers or "",
                row.number_of_students or "",
                row.number_of_branches or "",
                row.interested_plan or "",
                row.status or "",
                row.message or "",
            ]
        )
    return output.getvalue().encode("utf-8-sig")


@app.get("/demo-requests", response_class=HTMLResponse)
def list_demo_requests(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_demo_requests_access(request, db, "demo_requests.view")
    if redirect_response:
        return redirect_response

    filters = _demo_request_filters(request)
    query = db.query(models.DemoRequest)
    filtered_query = _apply_demo_request_filters(query, filters)
    demo_requests = (
        filtered_query
        .order_by(models.DemoRequest.submitted_at.desc(), models.DemoRequest.id.desc())
        .all()
    )
    filter_query_string = _demo_request_filter_query_string(filters)
    export_url = "/demo-requests/export"
    if filter_query_string:
        export_url = f"{export_url}?{filter_query_string}"
    return_to = str(request.url.path)
    if request.url.query:
        return_to = f"{return_to}?{request.url.query}"

    return templates.TemplateResponse(
        request,
        "demo_requests.html",
        {
            "request": request,
            "demo_requests": demo_requests,
            "filters": filters,
            "status_options": DEMO_REQUEST_STATUSES,
            "status_counts": _demo_request_status_counts(db),
            "export_url": export_url,
            "return_to": return_to,
            "format_timestamp": _format_notification_timestamp,
            "timestamp_iso": _notification_timestamp_iso,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="demo-requests",
            ),
        },
    )


@app.get("/demo-requests/export")
def export_demo_requests(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_demo_requests_access(request, db, "demo_requests.export")
    if redirect_response:
        return redirect_response

    filters = _demo_request_filters(request)
    rows = (
        _apply_demo_request_filters(db.query(models.DemoRequest), filters)
        .order_by(models.DemoRequest.submitted_at.desc(), models.DemoRequest.id.desc())
        .all()
    )
    payload = _write_demo_requests_csv(rows)
    response = StreamingResponse(
        iter([payload]),
        media_type="text/csv",
    )
    response.headers["Content-Disposition"] = (
        f'attachment; filename="{_demo_request_export_filename()}"'
    )
    return response


@app.get("/demo-requests/{demo_request_id}", response_class=HTMLResponse)
def view_demo_request(
    demo_request_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_demo_requests_access(request, db, "demo_requests.view")
    if redirect_response:
        return redirect_response

    demo_request = db.query(models.DemoRequest).filter(
        models.DemoRequest.id == demo_request_id
    ).first()
    if not demo_request:
        return RedirectResponse(url="/demo-requests?notice=Demo%20request%20not%20found.", status_code=302)
    if not getattr(demo_request, "seen_at", None):
        demo_request.seen_at = datetime.now(timezone.utc).replace(tzinfo=None)
        demo_request.seen_by_user_id = getattr(current_user, "user_id", "")
        demo_request.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
        db.commit()
        db.refresh(demo_request)

    return templates.TemplateResponse(
        request,
        "demo_request_detail.html",
        {
            "request": request,
            "demo_request": demo_request,
            "status_options": DEMO_REQUEST_STATUSES,
            "return_to": f"/demo-requests/{demo_request.id}",
            "format_timestamp": _format_notification_timestamp,
            "timestamp_iso": _notification_timestamp_iso,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="demo-requests",
                title="Demo Request Detail",
                intro="Review the submitted marketing lead and update its follow-up status.",
            ),
        },
    )


@app.post("/demo-requests/{demo_request_id}/status")
def update_demo_request_status(
    demo_request_id: int,
    request: Request,
    status: str = Form(...),
    return_to: str = Form("/demo-requests"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_demo_requests_access(request, db, "demo_requests.update_status")
    if redirect_response:
        return redirect_response

    demo_request = db.query(models.DemoRequest).filter(
        models.DemoRequest.id == demo_request_id
    ).first()
    if not demo_request:
        return _redirect_with_notice("/demo-requests", "Demo request not found.")

    demo_request.status = _normalize_demo_request_status(status)
    demo_request.status_updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    demo_request.status_updated_by_user_id = getattr(current_user, "user_id", "")
    demo_request.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.commit()

    safe_return_to = str(return_to or "/demo-requests").strip()
    if not safe_return_to.startswith("/demo-requests"):
        safe_return_to = "/demo-requests"
    return _redirect_with_notice(safe_return_to, "Demo request status updated.")


@app.post("/demo-requests/{demo_request_id}/delete")
def delete_demo_request(
    demo_request_id: int,
    request: Request,
    return_to: str = Form("/demo-requests"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_demo_requests_access(request, db, "demo_requests.update_status")
    if redirect_response:
        return redirect_response

    demo_request = db.query(models.DemoRequest).filter(
        models.DemoRequest.id == demo_request_id
    ).first()
    if not demo_request:
        return _redirect_with_notice("/demo-requests", "Demo request not found.")

    db.delete(demo_request)
    db.commit()

    safe_return_to = str(return_to or "/demo-requests").strip()
    if not safe_return_to.startswith("/demo-requests"):
        safe_return_to = "/demo-requests"
    if safe_return_to == f"/demo-requests/{demo_request_id}":
        safe_return_to = "/demo-requests"
    return _redirect_with_notice(safe_return_to, "Demo request deleted.")


@app.post("/demo-requests/bulk-status")
def bulk_update_demo_request_statuses(
    request: Request,
    payload: dict,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_demo_requests_access(
        request,
        db,
        "demo_requests.update_status",
    )
    if redirect_response:
        status_code = 401 if not current_user else 403
        return _bulk_json_error(
            [
                _bulk_item_error(
                    item_id=None,
                    label="Demo Requests",
                    field="permission",
                    message=(
                        "Authentication required."
                        if not current_user
                        else "You do not have permission to update demo request statuses."
                    ),
                )
            ],
            status_code=status_code,
        )

    items, errors = _bulk_payload_items(payload)
    if errors:
        return _bulk_json_error(errors, status_code=400)

    parsed_items: list[tuple[int, dict[str, object]]] = []
    seen_ids: set[int] = set()
    for item in items:
        raw_id = item.get("id")
        try:
            demo_request_id = int(raw_id)
        except (TypeError, ValueError):
            errors.append(
                _bulk_item_error(
                    item_id=raw_id,
                    label="Demo Request",
                    field="id",
                    message="A valid demo request ID is required.",
                )
            )
            continue
        if demo_request_id in seen_ids:
            errors.append(
                _bulk_item_error(
                    item_id=demo_request_id,
                    label=f"Demo Request {demo_request_id}",
                    field="id",
                    message="This demo request was submitted more than once.",
                )
            )
            continue
        seen_ids.add(demo_request_id)
        parsed_items.append((demo_request_id, item))

    rows = db.query(models.DemoRequest).filter(
        models.DemoRequest.id.in_([item_id for item_id, _item in parsed_items])
    ).all() if parsed_items else []
    row_map = {row.id: row for row in rows}
    update_plan: list[tuple[object, str]] = []

    for demo_request_id, item in parsed_items:
        row = row_map.get(demo_request_id)
        label = str(getattr(row, "school_name", "") or f"Demo Request {demo_request_id}").strip()
        if not row:
            errors.append(
                _bulk_item_error(
                    item_id=demo_request_id,
                    label=label,
                    field="id",
                    message="Demo request was not found.",
                )
            )
            continue
        submitted_status = str(item.get("status") or "").strip()
        normalized_status = next(
            (
                option
                for option in DEMO_REQUEST_STATUSES
                if option.casefold() == submitted_status.casefold()
            ),
            None,
        )
        if not normalized_status:
            errors.append(
                _bulk_item_error(
                    item_id=demo_request_id,
                    label=label,
                    field="status",
                    message="Select a valid demo request status.",
                )
            )
            continue
        update_plan.append((row, normalized_status))

    if errors:
        return _bulk_json_error(errors)

    now = datetime.now(timezone.utc).replace(tzinfo=None)
    for row, normalized_status in update_plan:
        row.status = normalized_status
        row.status_updated_at = now
        row.status_updated_by_user_id = getattr(current_user, "user_id", "")
        row.updated_at = now
    db.commit()
    return JSONResponse(
        {
            "ok": True,
            "saved_count": len(update_plan),
            "message": f"Saved {len(update_plan)} modified demo request status"
            + ("." if len(update_plan) == 1 else "es."),
        }
    )


# ---------------------------------------
# DEVELOPER: SYSTEM CONFIGURATION
# ---------------------------------------
def _get_configuration_access(request: Request, db: Session):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    if not auth.has_any_permission(
        db,
        current_user,
        "configuration.view",
        "schools.view",
        "branches.view",
        "academic_years.view",
        "branding.view",
        "configuration.manage_permissions",
        "configuration.manage_degrees",
        "configuration.manage_specializations",
        "timetable.manage_settings",
        "timetable.manage_blocks",
        "calendar.manage_event_types",
    ):
        return (
            None,
            authorization.build_access_denied_response(
                request,
                db,
                current_user=current_user,
                permission_keys=(
                    "configuration.view",
                    "schools.view",
                    "branches.view",
                    "academic_years.view",
                    "branding.view",
                ),
                page_key="system-configuration",
            ),
        )
    return current_user, None


def _get_school_branding_access(request: Request, db: Session):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    if not auth.has_any_permission(
        db,
        current_user,
        "branding.view",
        "branding.manage_school_logos",
        "branding.manage_branch_logos",
    ):
        return (
            None,
            authorization.build_access_denied_response(
                request,
                db,
                current_user=current_user,
                permission_keys=(
                    "branding.view",
                    "branding.manage_school_logos",
                    "branding.manage_branch_logos",
                ),
                page_key="school-branding",
            ),
        )
    return current_user, None


def _get_role_permissions_access(request: Request, db: Session):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "configuration.manage_permissions"):
        return (
            None,
            authorization.build_access_denied_response(
                request,
                db,
                current_user=current_user,
                permission_keys=("configuration.manage_permissions",),
                page_key="system-configuration",
            ),
        )
    return current_user, None


def _get_design_control_access(request: Request, db: Session):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    if not auth.is_platform_user(current_user) or not auth.has_permission(db, current_user, "design_control.manage"):
        return (
            None,
            authorization.build_access_denied_response(
                request,
                db,
                current_user=current_user,
                permission_keys=("design_control.manage",),
                page_key="system-configuration",
            ),
        )
    return current_user, None


def _render_configuration_template(
    *,
    request: Request,
    db: Session,
    current_user,
    template_name: str,
    active_module_key: str,
    title: str,
    intro: str,
    extra_context: dict | None = None,
):
    context = _build_configuration_context(request, db, current_user)
    context["configuration_modules"] = _get_configuration_modules(
        active_module_key,
        db,
        current_user,
    )
    context.update(extra_context or {})
    return templates.TemplateResponse(
        request,
        template_name,
        {
            "request": request,
            "user": current_user,
            "saudi_regions": SAUDI_REGIONS,
            **context,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="system-configuration",
                title=title,
                intro=intro,
            ),
        },
    )


def _build_timetable_settings_module_context(
    request: Request,
    db: Session,
    current_user,
):
    scoped_branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    scoped_academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    timetable_settings = get_timetable_settings_payload(
        db,
        scoped_branch_id,
        scoped_academic_year_id,
    )
    return {
        "timetable_settings": timetable_settings,
        "working_day_options": list(WORKING_DAY_OPTIONS),
        "block_type_options": list(BLOCK_TYPE_OPTIONS),
        "all_day_key": ALL_DAY_KEY,
        "timetable_settings_notice": str(
            request.query_params.get("notice", "") or ""
        ).strip(),
    }


def _logo_source_label(source: str) -> str:
    if source == "branch":
        return "Branch override"
    if source == "school_group":
        return "School group logo"
    return "Not configured"


def _get_user_school_group_id(db: Session, current_user) -> int | None:
    return getattr(current_user, "scope_school_group_id", None) or auth.get_user_school_group_id(
        db,
        current_user,
    )


def _can_manage_all_school_scopes(db: Session, current_user) -> bool:
    return auth.has_any_permission(
        db,
        current_user,
        "schools.manage_all_schools",
        "system_owner.switch_all_schools",
    )


def _can_manage_global_role_permissions(db: Session, current_user) -> bool:
    return auth.has_permission(
        db,
        current_user,
        "system_owner.manage_global_role_permissions",
    )


def _build_school_logo_module_context(
    request: Request,
    db: Session,
    current_user,
    *,
    administrator_mode: bool = False,
):
    can_manage_all = _can_manage_all_school_scopes(db, current_user)
    user_school_group_id = _get_user_school_group_id(db, current_user)

    school_groups_query = db.query(models.SchoolGroup).order_by(
        models.SchoolGroup.status.desc(),
        models.SchoolGroup.name.asc(),
    )
    if not can_manage_all:
        if not user_school_group_id:
            school_groups = []
        else:
            school_groups = school_groups_query.filter(
                models.SchoolGroup.id == user_school_group_id
            ).all()
    else:
        school_groups = school_groups_query.all()

    requested_group_id = request.query_params.get("school_group_id")
    selected_group = None
    if requested_group_id and can_manage_all:
        try:
            requested_group_id_int = int(requested_group_id)
        except ValueError:
            requested_group_id_int = None
        if (
            auth.is_platform_user(current_user)
            and requested_group_id_int != user_school_group_id
        ):
            requested_group_id_int = None
        selected_group = next(
            (group for group in school_groups if group.id == requested_group_id_int),
            None,
        )
    if not selected_group and user_school_group_id:
        selected_group = next(
            (group for group in school_groups if group.id == user_school_group_id),
            None,
        )
    if not selected_group and school_groups:
        selected_group = school_groups[0]

    branch_rows = []
    if selected_group:
        branch_query = db.query(models.Branch).filter(
            models.Branch.school_group_id == selected_group.id
        ).order_by(models.Branch.status.desc(), models.Branch.name.asc())
        if not can_manage_all:
            branch_query = branch_query.filter(models.Branch.id == current_user.branch_id)
        for branch in branch_query.all():
            branch_rows.append(
                {
                    "branch": branch,
                    "logos": [
                        {
                            **logo,
                            "source_label": _logo_source_label(logo.get("source")),
                        }
                        for logo in get_school_logo_slots(
                            request,
                            db,
                            branch.id,
                            include_empty=True,
                            include_all_slots=True,
                        )
                    ],
                }
            )

    group_logos = []
    if selected_group:
        group_logos = [
            {
                **logo,
                "source_label": _logo_source_label(logo.get("source")),
            }
            for logo in get_school_logo_slots(
                request,
                db,
                None,
                school_group_id=selected_group.id,
                include_empty=True,
                include_all_slots=True,
            )
        ]

    return {
        "logo_school_groups": school_groups,
        "selected_school_group": selected_group,
        "group_logos": group_logos,
        "logo_branch_rows": branch_rows,
        "logo_slots": list(DEFAULT_SCHOOL_LOGO_SLOTS),
        "can_manage_all_logo_scopes": can_manage_all,
        "administrator_branding_mode": administrator_mode,
        "logo_return_to": (
            "/school-branding"
            if administrator_mode
            else f"/system-configuration/logos?school_group_id={selected_group.id}"
            if selected_group
            else "/system-configuration/logos"
        ),
        "school_logo_notice": str(request.query_params.get("notice", "") or "").strip(),
        "school_logo_error": str(request.query_params.get("error", "") or "").strip(),
    }


def _ensure_timetable_setting_scope_row(
    db: Session,
    branch_id: int,
    academic_year_id: int,
):
    timetable_setting_row = get_timetable_setting_row(
        db,
        branch_id,
        academic_year_id,
    )
    if timetable_setting_row:
        return timetable_setting_row

    default_settings = get_timetable_settings_payload(
        db,
        branch_id,
        academic_year_id,
    )
    timetable_setting_row = models.TimetableSetting(
        branch_id=branch_id,
        academic_year_id=academic_year_id,
        working_days_csv=",".join(default_settings["working_day_keys"]),
        periods_per_day=default_settings["periods_per_day"],
        period_duration_minutes=default_settings["period_duration_minutes"],
        school_start_time=default_settings["school_start_time"],
        school_end_time=default_settings["school_end_time"],
    )
    db.add(timetable_setting_row)
    db.commit()
    db.refresh(timetable_setting_row)
    return timetable_setting_row


@app.get("/system-configuration")
def system_configuration(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_hub.html",
        active_module_key="overview",
        title="Configuration Hub",
        intro="Open each configuration module from a clean landing page instead of managing everything on one screen.",
    )


def _location_api_user_or_response(request: Request, db: Session):
    current_user = auth.get_current_user(request, db)
    if current_user:
        return current_user, None
    return None, JSONResponse(
        {"detail": "Authentication required."},
        status_code=401,
    )


@app.get("/api/locations/countries")
def location_countries(
    request: Request,
    db: Session = Depends(get_db),
):
    _current_user, error_response = _location_api_user_or_response(request, db)
    if error_response:
        return error_response
    return JSONResponse(
        {"items": location_service.list_countries()},
        headers={"Cache-Control": "private, max-age=86400"},
    )


@app.get("/api/locations/regions")
def location_regions(
    request: Request,
    country_code: str = Query(..., min_length=2, max_length=2),
    db: Session = Depends(get_db),
):
    _current_user, error_response = _location_api_user_or_response(request, db)
    if error_response:
        return error_response
    try:
        items = location_service.list_regions(country_code)
    except location_service.LocationValidationError as exc:
        return JSONResponse({"detail": str(exc)}, status_code=400)
    return JSONResponse(
        {"items": items},
        headers={"Cache-Control": "private, max-age=86400"},
    )


@app.get("/api/locations/cities")
def location_cities(
    request: Request,
    country_code: str = Query(..., min_length=2, max_length=2),
    region_id: int = Query(..., gt=0),
    db: Session = Depends(get_db),
):
    _current_user, error_response = _location_api_user_or_response(request, db)
    if error_response:
        return error_response
    try:
        items = location_service.list_cities(country_code, region_id)
    except location_service.LocationValidationError as exc:
        return JSONResponse({"detail": str(exc)}, status_code=400)
    return JSONResponse(
        {"items": items},
        headers={"Cache-Control": "private, max-age=86400"},
    )


def _build_school_management_context(request: Request, db: Session, current_user):
    context = _build_school_logo_module_context(request, db, current_user)
    selected_school_group = context.get("selected_school_group")
    school_group_id = getattr(selected_school_group, "id", None)
    school_academic_year_rows = []
    school_branch_rows = []
    if school_group_id:
        school_academic_year_rows = _build_academic_year_configuration_rows(
            db,
            school_group_id,
        )
        school_branch_rows = [
            row
            for row in _build_branch_configuration_rows(db)
            if row.get("school_group_id") == school_group_id
        ]
    return {
        **context,
        "school_academic_year_rows": school_academic_year_rows,
        "school_branch_rows": school_branch_rows,
        "school_branch_region_options": sorted(
            {str(row.get("region") or "").strip() for row in school_branch_rows}
            - {""},
            key=str.casefold,
        ),
        "school_delete_summary": _build_school_delete_summary(db, school_group_id),
    }


def _build_role_permissions_context(request: Request, db: Session, current_user):
    can_manage_all = _can_manage_global_role_permissions(db, current_user)
    user_school_group_id = _get_user_school_group_id(db, current_user)
    school_groups = db.query(models.SchoolGroup).order_by(
        models.SchoolGroup.status.desc(),
        models.SchoolGroup.name.asc(),
    ).all()

    requested_role = request.query_params.get("role", auth.ROLE_ADMINISTRATOR)
    selected_role = permission_registry.normalize_managed_role(requested_role) or auth.ROLE_ADMINISTRATOR
    requested_scope = str(request.query_params.get("scope", "") or "").strip().lower()
    requested_group_id = request.query_params.get("school_group_id")
    selected_school_group = None
    selected_school_group_id = None
    if can_manage_all and requested_scope == "global":
        selected_school_group_id = None
    else:
        if can_manage_all and requested_group_id:
            try:
                selected_school_group_id = int(requested_group_id)
            except ValueError:
                selected_school_group_id = None
        else:
            selected_school_group_id = user_school_group_id
        if selected_school_group_id:
            selected_school_group = next(
                (school for school in school_groups if school.id == selected_school_group_id),
                None,
            )
        if not selected_school_group and school_groups and not can_manage_all:
            selected_school_group = school_groups[0]
            selected_school_group_id = selected_school_group.id

    editable_roles = list(permission_registry.MANAGED_ROLES)
    permission_payload = _build_role_permission_payload(
        db,
        selected_role,
        selected_school_group_id,
    )

    return {
        "permission_groups": permission_registry.PERMISSION_GROUPS,
        "managed_roles": permission_registry.MANAGED_ROLES,
        "editable_roles": editable_roles,
        "selected_permission_role": selected_role,
        "selected_permission_school_group": selected_school_group,
        "selected_permission_school_group_id": selected_school_group_id,
        "selected_permission_scope": "global" if selected_school_group_id is None else "school",
        "role_permission_payload": permission_payload,
        "role_permission_summary_map": _build_role_permission_summary_map(
            db,
            selected_school_group_id,
        ),
        "role_permission_school_groups": school_groups,
        "can_manage_all_role_permissions": can_manage_all,
        "can_edit_selected_role_permissions": (
            auth.has_permission(db, current_user, "configuration.manage_permissions")
        ),
        "notice_message": str(request.query_params.get("notice", "") or "").strip(),
    }


@app.get("/system-configuration/role-permissions")
def system_configuration_role_permissions(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_role_permissions_access(request, db)
    if redirect_response:
        return redirect_response

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_role_permissions.html",
        active_module_key="role-permissions",
        title="Role Permissions",
        intro="Assign detailed TIS permissions to each role package before giving that role to users.",
        extra_context=_build_role_permissions_context(request, db, current_user),
    )


@app.post("/system-configuration/role-permissions")
def update_role_permissions(
    request: Request,
    role: str = Form(...),
    scope_type: str = Form("school"),
    school_group_id: int | None = Form(None),
    permission_keys: list[str] = Form([]),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_role_permissions_access(request, db)
    if redirect_response:
        return redirect_response

    can_manage_all = _can_manage_global_role_permissions(db, current_user)
    selected_role = permission_registry.normalize_managed_role(role)
    if not selected_role:
        return _redirect_with_error(
            "/system-configuration/role-permissions",
            "Select a valid role.",
        )
    target_school_group_id = None
    if can_manage_all and str(scope_type or "").strip().lower() == "global":
        target_school_group_id = None
    else:
        target_school_group_id = school_group_id or _get_user_school_group_id(db, current_user)
        if not target_school_group_id:
            return _redirect_with_error(
                "/system-configuration/role-permissions",
                "No school context was found for these permissions.",
            )
        if not can_manage_all and target_school_group_id != _get_user_school_group_id(db, current_user):
            return RedirectResponse(url="/dashboard", status_code=302)

    allowed_keys = {
        key for key in permission_keys
        if key in permission_registry.PERMISSION_LABELS
    }
    _set_role_permission_rows(
        db,
        role=selected_role,
        allowed_keys=allowed_keys,
        school_group_id=target_school_group_id,
        updated_by_user_id=getattr(current_user, "user_id", None),
    )
    db.commit()

    if target_school_group_id is None:
        return_to = f"/system-configuration/role-permissions?role={quote_plus(selected_role)}&scope=global"
    else:
        return_to = (
            f"/system-configuration/role-permissions?role={quote_plus(selected_role)}"
            f"&school_group_id={target_school_group_id}"
        )
    return _redirect_with_notice(return_to, f"{selected_role} permissions updated.")


@app.get("/api/design-studio/config")
def get_design_studio_config(
    request: Request,
    page_key: str = Query("dashboard"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_design_control_access(request, db)
    if redirect_response:
        return redirect_response

    normalized_page_key = str(page_key or "dashboard").strip() or "dashboard"
    rows = db.query(models.VisualDesignSetting).filter(
        models.VisualDesignSetting.page_key.in_(("global", normalized_page_key)),
        models.VisualDesignSetting.is_active == True,
    ).all()
    return JSONResponse(build_visual_design_config(normalized_page_key, rows_to_visual_settings(rows)))


@app.post("/api/design-studio/component-settings")
async def update_design_studio_component_settings(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_design_control_access(request, db)
    if redirect_response:
        return redirect_response

    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "Invalid design settings payload."}, status_code=400)

    component_key = str(payload.get("component_key", "") or "").strip()
    component = VISUAL_COMPONENT_MAP.get(component_key)
    is_custom_component = is_custom_component_key(component_key)
    if not component and not is_custom_component:
        return JSONResponse({"ok": False, "error": "Unknown design component."}, status_code=400)

    component_type = str(payload.get("component_type", "") or getattr(component, "component_type", "") or "element").strip()
    try:
        normalized_settings = normalize_visual_payload(
            component_key,
            payload.get("settings") or {},
            component_type,
        )
    except ValueError as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

    page_key = getattr(component, "page_key", "") or str(payload.get("page_key", "") or "").strip() or "global"
    db.query(models.VisualDesignSetting).filter(
        models.VisualDesignSetting.page_key == page_key,
        models.VisualDesignSetting.component_key == component_key,
    ).delete(synchronize_session=False)

    now = datetime.utcnow()
    for setting_key, setting_value in normalized_settings.items():
        db.add(
            models.VisualDesignSetting(
                page_key=page_key,
                component_key=component_key,
                component_type=component_type,
                setting_key=setting_key,
                setting_value=setting_value,
                scope_type="global",
                updated_by_user_id=getattr(current_user, "user_id", None),
                created_at=now,
                updated_at=now,
            )
        )
    db.commit()
    return JSONResponse({"ok": True, "component_key": component_key, "settings": normalized_settings})


@app.post("/api/design-studio/reset")
async def reset_design_studio_settings(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_design_control_access(request, db)
    if redirect_response:
        return redirect_response

    try:
        payload = await request.json()
    except Exception:
        payload = {}

    component_key = str((payload or {}).get("component_key", "") or "").strip()
    page_key = str((payload or {}).get("page_key", "") or "").strip()
    query = db.query(models.VisualDesignSetting)
    if component_key:
        component = VISUAL_COMPONENT_MAP.get(component_key)
        if not component and not is_custom_component_key(component_key):
            return JSONResponse({"ok": False, "error": "Unknown design component."}, status_code=400)
        component_page_key = getattr(component, "page_key", "") or page_key
        query = query.filter(
            models.VisualDesignSetting.component_key == component_key,
        )
        if component_page_key:
            query = query.filter(models.VisualDesignSetting.page_key == component_page_key)
    elif page_key:
        query = query.filter(models.VisualDesignSetting.page_key == page_key)
    else:
        return JSONResponse({"ok": False, "error": "Select a component or page to reset."}, status_code=400)

    deleted_count = query.delete(synchronize_session=False)
    db.commit()
    return JSONResponse({"ok": True, "deleted_count": deleted_count})


@app.post("/api/design-studio/reset-hidden")
async def reset_hidden_design_studio_settings(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_design_control_access(request, db)
    if redirect_response:
        return redirect_response

    try:
        payload = await request.json()
    except Exception:
        payload = {}

    page_key = str((payload or {}).get("page_key", "") or "").strip()
    query = db.query(models.VisualDesignSetting).filter(
        models.VisualDesignSetting.setting_key == "visibility",
        models.VisualDesignSetting.setting_value == "hidden",
    )
    if page_key:
        query = query.filter(models.VisualDesignSetting.page_key == page_key)
    deleted_count = query.delete(synchronize_session=False)
    db.commit()
    return JSONResponse({"ok": True, "deleted_count": deleted_count})


@app.post("/api/design-studio/reset-all")
async def reset_all_design_studio_settings(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_design_control_access(request, db)
    if redirect_response:
        return redirect_response

    visual_count = db.query(models.VisualDesignSetting).delete(synchronize_session=False)
    token_count = db.query(models.SystemDesignSetting).delete(synchronize_session=False)
    db.commit()
    return JSONResponse(
        {
            "ok": True,
            "visual_deleted_count": visual_count,
            "theme_deleted_count": token_count,
        }
    )


@app.get("/system-configuration/schools")
def system_configuration_schools(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_schools.html",
        active_module_key="school-management",
        title="School Management",
        intro="Manage school organizations, branches, academic years, branding, access, and subscription-ready setup.",
        extra_context=_build_school_management_context(request, db, current_user),
    )


@app.post("/system-configuration/schools")
def create_school_group(
    request: Request,
    name: str = Form(...),
    country_code: str = Form(""),
    region_id: str = Form(""),
    region_manual: str = Form(""),
    city_id: str = Form(""),
    city_manual: str = Form(""),
    district_name: str = Form(""),
    neighborhood_name: str = Form(""),
    return_to: str = Form("/system-configuration/schools"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    cleaned_name = " ".join(str(name or "").split())
    if not cleaned_name:
        return _redirect_with_error(return_to, "School/organization name is required.")

    existing = db.query(models.SchoolGroup).filter(
        func.lower(models.SchoolGroup.name) == cleaned_name.lower()
    ).first()
    if existing:
        return _redirect_with_error(return_to, "A school/organization with that name already exists.")

    try:
        resolved_location = _resolve_submitted_location(
            country_code=country_code,
            region_id=region_id,
            region_manual=region_manual,
            city_id=city_id,
            city_manual=city_manual,
        )
        cleaned_district = _normalize_location_detail(district_name, "District")
        cleaned_neighborhood = _normalize_location_detail(
            neighborhood_name,
            "Neighborhood",
        )
    except location_service.LocationValidationError as exc:
        return _redirect_with_error(return_to, str(exc))

    school_group = models.SchoolGroup(
        name=cleaned_name,
        status=True,
        country_code=resolved_location.country_code,
        country_name=resolved_location.country_name,
        region_name=resolved_location.region_name,
        city_name=resolved_location.city_name,
        district_name=cleaned_district,
        neighborhood_name=cleaned_neighborhood,
    )
    db.add(school_group)
    db.flush()
    try:
        branding_storage.ensure_organization_logo_dir(school_group.id)
    except OSError:
        db.rollback()
        return _redirect_with_error(
            return_to,
            "Unable to initialize organization branding storage.",
        )
    db.add(
        models.Branch(
            school_group_id=school_group.id,
            name=cleaned_name,
            location=resolved_location.region_name,
            country_code=resolved_location.country_code,
            country_name=resolved_location.country_name,
            region_name=resolved_location.region_name,
            city_name=resolved_location.city_name,
            district_name=cleaned_district,
            neighborhood_name=cleaned_neighborhood,
            status=True,
        )
    )
    db.commit()
    return _redirect_with_notice(
        f"/system-configuration/schools?school_group_id={school_group.id}",
        "School created with a default branch.",
    )


@app.post("/system-configuration/schools/{school_group_id}")
def update_school_group(
    school_group_id: int,
    request: Request,
    name: str = Form(...),
    status: str = Form("active"),
    country_code: str = Form(""),
    region_id: str = Form(""),
    region_manual: str = Form(""),
    city_id: str = Form(""),
    city_manual: str = Form(""),
    district_name: str = Form(""),
    neighborhood_name: str = Form(""),
    return_to: str = Form("/system-configuration/schools"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    school_group = db.query(models.SchoolGroup).filter(
        models.SchoolGroup.id == school_group_id
    ).first()
    if not school_group:
        return _redirect_with_error(return_to, "School/organization record not found.")

    cleaned_name = " ".join(str(name or "").split())
    if not cleaned_name:
        return _redirect_with_error(return_to, "School/organization name is required.")

    duplicate = db.query(models.SchoolGroup).filter(
        func.lower(models.SchoolGroup.name) == cleaned_name.lower(),
        models.SchoolGroup.id != school_group_id,
    ).first()
    if duplicate:
        return _redirect_with_error(return_to, "Another school/organization already uses that name.")

    try:
        resolved_location = _resolve_submitted_location(
            country_code=country_code,
            region_id=region_id,
            region_manual=region_manual,
            city_id=city_id,
            city_manual=city_manual,
            required=False,
        )
        cleaned_district = _normalize_location_detail(district_name, "District")
        cleaned_neighborhood = _normalize_location_detail(
            neighborhood_name,
            "Neighborhood",
        )
    except location_service.LocationValidationError as exc:
        return _redirect_with_error(return_to, str(exc))

    school_group.name = cleaned_name
    school_group.status = str(status or "").strip().lower() != "inactive"
    if resolved_location:
        school_group.country_code = resolved_location.country_code
        school_group.country_name = resolved_location.country_name
        school_group.region_name = resolved_location.region_name
        school_group.city_name = resolved_location.city_name
    school_group.district_name = cleaned_district
    school_group.neighborhood_name = cleaned_neighborhood
    school_group.updated_at = datetime.utcnow()
    db.commit()
    return _redirect_with_notice(
        f"/system-configuration/schools?school_group_id={school_group.id}",
        "School information updated.",
    )


@app.post("/system-configuration/schools/{school_group_id}/delete")
def delete_school_group(
    school_group_id: int,
    request: Request,
    return_to: str = Form("/system-configuration/schools"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    safe_return_to = _safe_redirect_path(return_to)
    school_group = db.query(models.SchoolGroup).filter(
        models.SchoolGroup.id == school_group_id
    ).first()
    if not school_group:
        return _redirect_with_error(safe_return_to, "School/organization record not found.")

    delete_summary = _build_school_delete_summary(db, school_group_id)
    if not delete_summary["can_delete"]:
        reason = str(delete_summary.get("reason") or "the school still has linked data")
        return _redirect_with_error(
            safe_return_to,
            f"This school cannot be deleted because {reason}. Deactivate it instead.",
        )

    db.query(models.Branch).filter(
        models.Branch.school_group_id == school_group_id
    ).delete(synchronize_session=False)
    db.delete(school_group)
    db.commit()
    _ensure_default_school_group(db)

    return _redirect_with_notice(
        "/system-configuration/schools",
        "School deleted successfully.",
    )


@app.get("/system-configuration/branches")
def system_configuration_branches(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response
    return RedirectResponse(url="/system-configuration/schools", status_code=302)

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_branches.html",
        active_module_key="branches",
        title="Branch Management",
        intro="Manage branch records in a compact operational table.",
    )


@app.get("/system-configuration/logos")
def system_configuration_logos(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_logos.html",
        active_module_key="school-management",
        title="School Logos",
        intro="Manage school-group logos and branch logo overrides used across the system.",
        extra_context=_build_school_logo_module_context(request, db, current_user),
    )


@app.get("/school-branding")
def school_branding(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_school_branding_access(request, db)
    if redirect_response:
        return redirect_response

    return templates.TemplateResponse(
        request,
        "system_configuration_logos.html",
        {
            "request": request,
            "user": current_user,
            "saudi_regions": SAUDI_REGIONS,
            **_build_configuration_context(request, db, current_user),
            **_build_school_logo_module_context(
                request,
                db,
                current_user,
                administrator_mode=not _can_manage_all_school_scopes(db, current_user),
            ),
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="school-branding",
                title="School Branding",
                intro="Manage the logo references for your school subscription.",
            ),
        },
    )


@app.post("/system-configuration/logos")
async def save_system_configuration_logo(
    request: Request,
    scope_type: str = Form("branch"),
    school_group_id: int | None = Form(None),
    branch_id: int | None = Form(None),
    slot_key: str = Form(...),
    label: str = Form(""),
    return_to: str = Form("/system-configuration/logos"),
    logo_file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_school_branding_access(request, db)
    if redirect_response:
        return redirect_response

    can_manage_all = _can_manage_all_school_scopes(db, current_user)
    normalized_scope_type = str(scope_type or "branch").strip().lower()
    if normalized_scope_type not in {"school_group", "branch"}:
        return _redirect_with_error(return_to, "Select a valid logo scope.")

    slot = next(
        (item for item in DEFAULT_SCHOOL_LOGO_SLOTS if item["slot_key"] == slot_key),
        None,
    )
    if not slot:
        return _redirect_with_error(return_to, "Select a valid logo slot.")

    user_school_group_id = _get_user_school_group_id(db, current_user)
    branch = None
    school_group = None
    if normalized_scope_type == "school_group":
        if not school_group_id:
            return _redirect_with_error(return_to, "Select a valid school group.")
        school_group = db.query(models.SchoolGroup).filter(
            models.SchoolGroup.id == school_group_id
        ).first()
        if not school_group:
            return _redirect_with_error(return_to, "Select a valid school group.")
        if not can_manage_all and school_group.id != user_school_group_id:
            return RedirectResponse(url="/dashboard", status_code=302)
    else:
        if not branch_id:
            return _redirect_with_error(return_to, "Select a valid branch.")
        branch = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
        if not branch:
            return _redirect_with_error(return_to, "Select a valid branch.")
        if not can_manage_all and (
            branch.id
            != getattr(current_user, "scope_branch_id", current_user.branch_id)
            or branch.school_group_id != user_school_group_id
        ):
            return RedirectResponse(url="/dashboard", status_code=302)

    normalized_label = str(label or "").strip() or slot["label"]
    if normalized_scope_type == "school_group":
        existing_logo = db.query(models.SchoolGroupLogo).filter(
            models.SchoolGroupLogo.school_group_id == school_group.id,
            models.SchoolGroupLogo.slot_key == slot_key,
        ).first()
    else:
        existing_logo = db.query(models.BranchLogo).filter(
            models.BranchLogo.branch_id == branch.id,
            models.BranchLogo.slot_key == slot_key,
        ).first()
    target_school_group_id = (
        school_group.id
        if normalized_scope_type == "school_group"
        else getattr(branch, "school_group_id", None)
    )
    if not target_school_group_id:
        return _redirect_with_error(
            return_to,
            "The selected branding scope has no organization owner.",
        )

    file_bytes = b""
    if logo_file and logo_file.filename:
        file_bytes = await logo_file.read()

    if not file_bytes:
        if existing_logo:
            existing_logo.label = normalized_label
            existing_logo.updated_by_user_id = getattr(current_user, "user_id", None)
            existing_logo.updated_at = datetime.utcnow()
            db.commit()
            return _redirect_with_notice(return_to, "Logo label updated.")
        return _redirect_with_error(return_to, "Choose a logo file to upload.")

    try:
        upload_info = branding_storage.validate_logo_upload(
            file_bytes,
            logo_file.filename,
            slot_key=slot_key,
        )
        relative_path = branding_storage.write_logo_file(
            file_bytes,
            school_group_id=target_school_group_id,
            branch_id=branch.id if normalized_scope_type == "branch" else None,
            slot_key=slot_key,
            extension=upload_info.extension,
        )
    except (branding_storage.BrandingStorageError, OSError) as exc:
        return _redirect_with_error(return_to, str(exc))

    old_path = ""
    if not existing_logo:
        logo_payload = {
            "slot_key": slot_key,
            "label": normalized_label,
            "image_path": relative_path,
            "content_type": upload_info.content_type,
            "sort_order": slot["sort_order"],
            "updated_by_user_id": getattr(current_user, "user_id", None),
        }
        if normalized_scope_type == "school_group":
            existing_logo = models.SchoolGroupLogo(
                school_group_id=school_group.id,
                **logo_payload,
            )
        else:
            existing_logo = models.BranchLogo(
                branch_id=branch.id,
                **logo_payload,
            )
        db.add(existing_logo)
    else:
        old_path = existing_logo.image_path
        existing_logo.label = normalized_label
        existing_logo.image_path = relative_path
        existing_logo.content_type = upload_info.content_type
        existing_logo.sort_order = slot["sort_order"]
        existing_logo.updated_by_user_id = getattr(current_user, "user_id", None)
        existing_logo.updated_at = datetime.utcnow()

    db.commit()
    if old_path and old_path != relative_path:
        branding_storage.delete_owned_logo_file(
            old_path,
            school_group_id=target_school_group_id,
            branch_id=branch.id if normalized_scope_type == "branch" else None,
        )
    return _redirect_with_notice(return_to, "Logo updated.")


@app.post("/system-configuration/logos/reset")
def reset_system_configuration_logo(
    request: Request,
    scope_type: str = Form("branch"),
    school_group_id: int | None = Form(None),
    branch_id: int | None = Form(None),
    slot_key: str = Form(...),
    return_to: str = Form("/system-configuration/logos"),
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_school_branding_access(request, db)
    if redirect_response:
        return redirect_response

    can_manage_all = _can_manage_all_school_scopes(db, current_user)
    normalized_scope_type = str(scope_type or "branch").strip().lower()
    if normalized_scope_type not in {"school_group", "branch"}:
        return _redirect_with_error(return_to, "Select a valid logo scope.")
    user_school_group_id = _get_user_school_group_id(db, current_user)
    existing_logo = None
    target_school_group_id = None
    target_branch_id = None
    if normalized_scope_type == "school_group":
        if not school_group_id:
            return _redirect_with_error(return_to, "Select a valid school group.")
        school_group = db.query(models.SchoolGroup).filter(
            models.SchoolGroup.id == school_group_id
        ).first()
        if not school_group:
            return _redirect_with_error(return_to, "Select a valid school group.")
        if not can_manage_all and school_group_id != user_school_group_id:
            return RedirectResponse(url="/dashboard", status_code=302)
        target_school_group_id = school_group.id
        existing_logo = db.query(models.SchoolGroupLogo).filter(
            models.SchoolGroupLogo.school_group_id == school_group_id,
            models.SchoolGroupLogo.slot_key == slot_key,
        ).first()
    else:
        if not branch_id:
            return _redirect_with_error(return_to, "Select a valid branch.")
        branch = db.query(models.Branch).filter(
            models.Branch.id == branch_id
        ).first()
        if not branch or not getattr(branch, "school_group_id", None):
            return _redirect_with_error(return_to, "Select a valid branch.")
        if not can_manage_all and (
            branch_id
            != getattr(current_user, "scope_branch_id", current_user.branch_id)
            or branch.school_group_id != user_school_group_id
        ):
            return RedirectResponse(url="/dashboard", status_code=302)
        target_school_group_id = branch.school_group_id
        target_branch_id = branch.id
        existing_logo = db.query(models.BranchLogo).filter(
            models.BranchLogo.branch_id == branch_id,
            models.BranchLogo.slot_key == slot_key,
        ).first()
    if existing_logo:
        old_path = existing_logo.image_path
        db.delete(existing_logo)
        db.commit()
        branding_storage.delete_owned_logo_file(
            old_path,
            school_group_id=target_school_group_id,
            branch_id=target_branch_id,
        )
    return _redirect_with_notice(return_to, "Logo reset.")


@app.get("/system-configuration/degrees")
def system_configuration_degrees(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response
    context = _build_configuration_context(request, db, current_user)

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_qualifications.html",
        active_module_key="degrees",
        title="Degrees Management",
        intro="Manage academic degree options in a direct editable list.",
        extra_context={
            "module_title": "Degrees",
            "module_description": "These options appear in the teacher qualification form.",
            "module_icon": "copy",
            "module_kind": QUALIFICATION_KIND_DEGREE,
            "module_rows": context["degree_rows"],
            "create_label": "Add Degree",
            "name_label": "Degree Name",
            "empty_message": "No degrees are configured yet.",
        },
    )


@app.get("/system-configuration/specializations")
def system_configuration_specializations(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response
    context = _build_configuration_context(request, db, current_user)

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_qualifications.html",
        active_module_key="specializations",
        title="Majors / Specializations",
        intro="Manage majors and teaching specializations in a direct editable list.",
        extra_context={
            "module_title": "Majors / Teaching Specializations",
            "module_description": "These options are used directly by the teacher qualification form.",
            "module_icon": "subjects",
            "module_kind": QUALIFICATION_KIND_SPECIALIZATION,
            "module_rows": context["specialization_rows"],
            "create_label": "Add Specialization",
            "name_label": "Specialization Name",
            "empty_message": "No specializations are configured yet.",
        },
    )


@app.get("/system-configuration/academic-years")
def system_configuration_academic_years(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response
    return RedirectResponse(url="/system-configuration/schools", status_code=302)

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_years.html",
        active_module_key="academic-years",
        title="Academic Year Management",
        intro="Open and switch academic years from a dedicated configuration module.",
    )


@app.get("/system-configuration/timetable-settings")
def system_configuration_timetable_settings(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_configuration_access(request, db)
    if redirect_response:
        return redirect_response

    return _render_configuration_template(
        request=request,
        db=db,
        current_user=current_user,
        template_name="system_configuration_timetable.html",
        active_module_key="timetable-settings",
        title="Timetable Settings",
        intro="Define the school week, period structure, and non-teaching timetable blocks for the active branch and academic year.",
        extra_context=_build_timetable_settings_module_context(
            request,
            db,
            current_user,
        ),
    )


@app.post("/system-configuration/timetable-settings")
def save_timetable_settings(
    request: Request,
    working_days: list[str] = Form([]),
    periods_per_day: str = Form("8"),
    period_duration_minutes: str = Form("45"),
    school_start_time: str = Form("07:00"),
    school_end_time: str = Form("13:00"),
    return_to: str = Form("/system-configuration/timetable-settings"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "timetable.manage_settings"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("timetable.manage_settings",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    normalized_settings = normalize_timetable_settings_values(
        working_days,
        periods_per_day,
        period_duration_minutes,
        school_start_time,
        school_end_time,
    )
    if normalized_settings["errors"]:
        return _redirect_with_error(
            safe_return_to,
            " ".join(normalized_settings["errors"]),
        )

    existing_settings = get_timetable_settings_payload(
        db,
        branch_id,
        academic_year_id,
    )
    updated_time_slots = build_time_slots(
        normalized_settings["periods_per_day"],
        normalized_settings["period_duration_minutes"],
        normalized_settings["school_start_time"],
    )
    invalid_existing_blocks = []
    for block in existing_settings.get("blocks", []):
        if (
            block["day_key"] != ALL_DAY_KEY
            and block["day_key"] not in normalized_settings["working_day_keys"]
        ):
            invalid_existing_blocks.append(
                f"{block['label']} uses {block['day_label']}, which is no longer part of the working week."
            )
        normalized_existing_block = normalize_non_teaching_block_values(
            block_type=block.get("block_type"),
            label=block.get("label"),
            day_key=block.get("day_key"),
            start_time=block.get("start_time"),
            end_time=block.get("end_time"),
            start_period=block.get("start_period"),
            end_period=block.get("end_period"),
            periods_per_day=normalized_settings["periods_per_day"],
            working_day_keys=normalized_settings["working_day_keys"],
            time_slots=updated_time_slots,
        )
        if normalized_existing_block.get("errors"):
            invalid_existing_blocks.append(
                f"{block['label']} no longer fits the updated timetable period structure."
            )

    if invalid_existing_blocks:
        return _redirect_with_error(
            safe_return_to,
            "Update or delete the affected non-teaching blocks first: "
            + " ".join(invalid_existing_blocks),
        )

    timetable_setting_row = get_timetable_setting_row(
        db,
        branch_id,
        academic_year_id,
    )
    if not timetable_setting_row:
        timetable_setting_row = models.TimetableSetting(
            branch_id=branch_id,
            academic_year_id=academic_year_id,
        )
        db.add(timetable_setting_row)

    timetable_setting_row.working_days_csv = ",".join(
        normalized_settings["working_day_keys"]
    )
    timetable_setting_row.periods_per_day = normalized_settings["periods_per_day"]
    timetable_setting_row.period_duration_minutes = normalized_settings["period_duration_minutes"]
    timetable_setting_row.school_start_time = normalized_settings["school_start_time"]
    timetable_setting_row.school_end_time = normalized_settings["school_end_time"]
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        "Timetable settings saved for the current branch and academic year.",
    )


@app.post("/system-configuration/timetable-settings/recalculate")
def recalculate_timetable_settings_structure(
    request: Request,
    return_to: str = Form("/system-configuration/timetable-settings"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "timetable.manage_settings"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("timetable.manage_settings",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    timetable_setting_row = _ensure_timetable_setting_scope_row(
        db,
        branch_id,
        academic_year_id,
    )
    refreshed_settings = get_timetable_settings_payload(
        db,
        branch_id,
        academic_year_id,
    )
    timetable_setting_row.school_end_time = refreshed_settings.get("school_end_time") or timetable_setting_row.school_end_time
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        "Timetable timeline recalculated from the latest period duration and non-teaching blocks.",
    )


@app.post("/system-configuration/timetable-settings/blocks")
def create_timetable_block(
    request: Request,
    block_type: str = Form(...),
    label: str = Form(...),
    day_key: str = Form(ALL_DAY_KEY),
    start_time: str = Form(...),
    end_time: str = Form(...),
    return_to: str = Form("/system-configuration/timetable-settings"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "timetable.manage_blocks"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("timetable.manage_blocks",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    timetable_setting_row = _ensure_timetable_setting_scope_row(
        db,
        branch_id,
        academic_year_id,
    )
    timetable_settings = get_timetable_settings_payload(
        db,
        branch_id,
        academic_year_id,
    )
    normalized_block = normalize_non_teaching_block_values(
        block_type=block_type,
        label=label,
        day_key=day_key,
        start_time=start_time,
        end_time=end_time,
        start_period=None,
        end_period=None,
        periods_per_day=timetable_settings["periods_per_day"],
        working_day_keys=timetable_settings["working_day_keys"],
        time_slots=timetable_settings["time_slots"],
    )
    block_errors = list(normalized_block["errors"])
    block_errors.extend(
        validate_non_teaching_block_overlap(
            timetable_settings.get("blocks", []),
            normalized_block,
        )
    )
    if block_errors:
        return _redirect_with_error(
            safe_return_to,
            " ".join(block_errors),
        )

    db.add(
        models.TimetableNonTeachingBlock(
            timetable_setting_id=timetable_setting_row.id,
            block_type=normalized_block["block_type"],
            label=normalized_block["label"],
            day_key=normalized_block["day_key"],
            start_time=normalized_block["start_time"],
            end_time=normalized_block["end_time"],
            start_period=normalized_block["start_period"],
            end_period=normalized_block["end_period"],
        )
    )
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        f"{normalized_block['label']} added to timetable settings.",
    )


@app.post("/system-configuration/timetable-settings/blocks/{block_id}")
def update_timetable_block(
    block_id: int,
    request: Request,
    block_type: str = Form(...),
    label: str = Form(...),
    day_key: str = Form(ALL_DAY_KEY),
    start_time: str = Form(...),
    end_time: str = Form(...),
    return_to: str = Form("/system-configuration/timetable-settings"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "timetable.manage_blocks"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("timetable.manage_blocks",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    timetable_setting_row = _ensure_timetable_setting_scope_row(
        db,
        branch_id,
        academic_year_id,
    )
    block_row = db.query(models.TimetableNonTeachingBlock).filter(
        models.TimetableNonTeachingBlock.id == block_id,
        models.TimetableNonTeachingBlock.timetable_setting_id == timetable_setting_row.id,
    ).first()
    if not block_row:
        return _redirect_with_error(
            safe_return_to,
            "Timetable block record not found for the active branch/year scope.",
        )

    timetable_settings = get_timetable_settings_payload(
        db,
        branch_id,
        academic_year_id,
    )
    normalized_block = normalize_non_teaching_block_values(
        block_type=block_type,
        label=label,
        day_key=day_key,
        start_time=start_time,
        end_time=end_time,
        start_period=None,
        end_period=None,
        periods_per_day=timetable_settings["periods_per_day"],
        working_day_keys=timetable_settings["working_day_keys"],
        time_slots=timetable_settings["time_slots"],
    )
    block_errors = list(normalized_block["errors"])
    block_errors.extend(
        validate_non_teaching_block_overlap(
            timetable_settings.get("blocks", []),
            normalized_block,
            ignore_block_id=block_id,
        )
    )
    if block_errors:
        return _redirect_with_error(
            safe_return_to,
            " ".join(block_errors),
        )

    block_row.block_type = normalized_block["block_type"]
    block_row.label = normalized_block["label"]
    block_row.day_key = normalized_block["day_key"]
    block_row.start_time = normalized_block["start_time"]
    block_row.end_time = normalized_block["end_time"]
    block_row.start_period = normalized_block["start_period"]
    block_row.end_period = normalized_block["end_period"]
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        f"{normalized_block['label']} updated successfully.",
    )


@app.post("/system-configuration/timetable-settings/blocks/{block_id}/delete")
def delete_timetable_block(
    block_id: int,
    request: Request,
    return_to: str = Form("/system-configuration/timetable-settings"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "timetable.manage_blocks"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("timetable.manage_blocks",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    timetable_setting_row = _ensure_timetable_setting_scope_row(
        db,
        branch_id,
        academic_year_id,
    )
    block_row = db.query(models.TimetableNonTeachingBlock).filter(
        models.TimetableNonTeachingBlock.id == block_id,
        models.TimetableNonTeachingBlock.timetable_setting_id == timetable_setting_row.id,
    ).first()
    if not block_row:
        return _redirect_with_error(
            safe_return_to,
            "Timetable block record not found for the active branch/year scope.",
        )

    block_label = str(block_row.label or "Timetable block").strip() or "Timetable block"
    db.delete(block_row)
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        f"{block_label} deleted successfully.",
    )


# ---------------------------------------
# DEVELOPER: CREATE BRANCH
# ---------------------------------------
@app.post("/system-configuration/branches")
def create_branch(
    request: Request,
    name: str = Form(...),
    region: str = Form(""),
    country_code: str = Form(""),
    region_id: str = Form(""),
    region_manual: str = Form(""),
    city_id: str = Form(""),
    city_manual: str = Form(""),
    district_name: str = Form(""),
    neighborhood_name: str = Form(""),
    school_group_id: int | None = Form(None),
    return_to: str = Form("/system-configuration/branches"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "branches.create"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("branches.create",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    cleaned_name = " ".join(str(name or "").split())

    if not cleaned_name:
        return _redirect_with_error(
            safe_return_to,
            "Branch name is required.",
        )

    try:
        resolved_location = _resolve_submitted_location(
            country_code=country_code,
            region_id=region_id,
            region_manual=region_manual,
            city_id=city_id,
            city_manual=city_manual,
            legacy_region=region,
        )
        cleaned_district = _normalize_location_detail(district_name, "District")
        cleaned_neighborhood = _normalize_location_detail(
            neighborhood_name,
            "Neighborhood",
        )
    except location_service.LocationValidationError as exc:
        return _redirect_with_error(safe_return_to, str(exc))

    existing_branch = db.query(models.Branch).filter(
        func.lower(models.Branch.name) == cleaned_name.lower()
    ).first()
    if existing_branch:
        return _redirect_with_error(
            safe_return_to,
            "A branch with that name already exists.",
        )

    school_group = None
    if school_group_id:
        school_group = db.query(models.SchoolGroup).filter(
            models.SchoolGroup.id == school_group_id
        ).first()
    if not school_group:
        school_group = db.query(models.SchoolGroup).filter(
            models.SchoolGroup.name == DEFAULT_SCHOOL_GROUP_NAME
        ).first()
    if not school_group:
        school_group = _ensure_default_school_group(db)

    db.add(
        models.Branch(
            name=cleaned_name,
            location=resolved_location.region_name,
            country_code=resolved_location.country_code,
            country_name=resolved_location.country_name,
            region_name=resolved_location.region_name,
            city_name=resolved_location.city_name,
            district_name=cleaned_district,
            neighborhood_name=cleaned_neighborhood,
            status=True,
            school_group_id=getattr(school_group, "id", None),
        )
    )
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        "Branch created successfully.",
    )


# ---------------------------------------
# DEVELOPER: UPDATE BRANCH
# ---------------------------------------
def _branch_record_is_in_user_scope(db: Session, current_user, branch_row) -> bool:
    if auth.is_platform_user(current_user):
        return True
    access_scope = auth.get_access_scope(current_user)
    if access_scope == auth.ACCESS_SCOPE_ORGANIZATION:
        return getattr(branch_row, "school_group_id", None) == auth.get_user_school_group_id(
            db,
            current_user,
        )
    scoped_branch_id = getattr(current_user, "scope_branch_id", None) or getattr(
        current_user,
        "branch_id",
        None,
    )
    return getattr(branch_row, "id", None) == scoped_branch_id


@app.post("/system-configuration/branches/bulk-update")
def bulk_update_branches(
    request: Request,
    payload: dict,
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return _bulk_json_error(
            [_bulk_item_error(item_id=None, label="Session", field="auth", message="Authentication required.")],
            status_code=401,
        )

    can_edit = auth.has_permission(db, current_user, "branches.edit")
    can_change_status = auth.has_permission(
        db,
        current_user,
        "branches.activate_deactivate",
    )
    if not can_edit and not can_change_status:
        return _bulk_json_error(
            [
                _bulk_item_error(
                    item_id=None,
                    label="Branches",
                    field="permission",
                    message="You do not have permission to update branches.",
                )
            ],
            status_code=403,
        )

    items, errors = _bulk_payload_items(payload)
    if errors:
        return _bulk_json_error(errors, status_code=400)

    parsed_items: list[tuple[int, dict[str, object]]] = []
    seen_ids: set[int] = set()
    for item in items:
        raw_id = item.get("id")
        try:
            branch_id = int(raw_id)
        except (TypeError, ValueError):
            errors.append(
                _bulk_item_error(
                    item_id=raw_id,
                    label="Branch",
                    field="id",
                    message="A valid branch ID is required.",
                )
            )
            continue
        if branch_id in seen_ids:
            errors.append(
                _bulk_item_error(
                    item_id=branch_id,
                    label=f"Branch {branch_id}",
                    field="id",
                    message="This branch was submitted more than once.",
                )
            )
            continue
        seen_ids.add(branch_id)
        parsed_items.append((branch_id, item))

    branch_rows = db.query(models.Branch).filter(
        models.Branch.id.in_([branch_id for branch_id, _item in parsed_items])
    ).all() if parsed_items else []
    branch_map = {branch.id: branch for branch in branch_rows}
    update_plan: list[dict[str, object]] = []

    for branch_id, item in parsed_items:
        branch_row = branch_map.get(branch_id)
        label = str(getattr(branch_row, "name", "") or f"Branch {branch_id}").strip()
        if not branch_row or not _branch_record_is_in_user_scope(db, current_user, branch_row):
            errors.append(
                _bulk_item_error(
                    item_id=branch_id,
                    label=label,
                    field="id",
                    message="Branch was not found or is outside your accessible scope.",
                )
            )
            continue

        submitted_changed_fields = item.get("_changed_fields")
        changed_fields = (
            {str(field) for field in submitted_changed_fields}
            if isinstance(submitted_changed_fields, list)
            else {str(field) for field in item.keys() if field not in {"id", "_changed_fields"}}
        )
        name_requested = "name" in changed_fields
        status_requested = "status" in changed_fields
        location_core_field_names = {
            "region",
            "country_code",
            "region_id",
            "region_manual",
            "city_id",
            "city_manual",
        }
        address_detail_field_names = {"district_name", "neighborhood_name"}
        location_core_requested = bool(changed_fields & location_core_field_names)
        address_detail_requested = bool(changed_fields & address_detail_field_names)
        location_requested = location_core_requested or address_detail_requested

        submitted_name = item.get("name") if name_requested else branch_row.name
        cleaned_name = " ".join(str(submitted_name or "").split())
        if name_requested and not can_edit:
            errors.append(
                _bulk_item_error(
                    item_id=branch_id,
                    label=label,
                    field="name",
                    message="You do not have permission to edit branch details.",
                )
            )
        elif not cleaned_name:
            errors.append(
                _bulk_item_error(
                    item_id=branch_id,
                    label=label,
                    field="name",
                    message="Branch name is required.",
                )
            )

        next_status = bool(branch_row.status)
        if status_requested:
            normalized_status = str(item.get("status") or "").strip().lower()
            if normalized_status not in {"active", "inactive"}:
                errors.append(
                    _bulk_item_error(
                        item_id=branch_id,
                        label=label,
                        field="status",
                        message="Select a valid branch status.",
                    )
                )
            else:
                next_status = normalized_status == "active"
                if next_status != bool(branch_row.status) and not can_change_status:
                    errors.append(
                        _bulk_item_error(
                            item_id=branch_id,
                            label=label,
                            field="status",
                            message="You do not have permission to activate or deactivate branches.",
                        )
                    )

        resolved_location = None
        cleaned_district = str(getattr(branch_row, "district_name", "") or "").strip()
        cleaned_neighborhood = str(
            getattr(branch_row, "neighborhood_name", "") or ""
        ).strip()
        if location_requested:
            if not can_edit:
                errors.append(
                    _bulk_item_error(
                        item_id=branch_id,
                        label=label,
                        field="location",
                        message="You do not have permission to edit branch details.",
                    )
                )
            elif location_core_requested:
                try:
                    resolved_location = _resolve_submitted_location(
                        country_code=str(item.get("country_code") or ""),
                        region_id=str(item.get("region_id") or ""),
                        region_manual=str(item.get("region_manual") or ""),
                        city_id=str(item.get("city_id") or ""),
                        city_manual=str(item.get("city_manual") or ""),
                        legacy_region=str(item.get("region") or ""),
                        required=False,
                    )
                except location_service.LocationValidationError as exc:
                    errors.append(
                        _bulk_item_error(
                            item_id=branch_id,
                            label=label,
                            field="location",
                            message=str(exc),
                        )
                    )
            if can_edit and address_detail_requested:
                try:
                    if "district_name" in changed_fields:
                        cleaned_district = _normalize_location_detail(
                            item.get("district_name"),
                            "District",
                        )
                    if "neighborhood_name" in changed_fields:
                        cleaned_neighborhood = _normalize_location_detail(
                            item.get("neighborhood_name"),
                            "Neighborhood",
                        )
                except location_service.LocationValidationError as exc:
                    errors.append(
                        _bulk_item_error(
                            item_id=branch_id,
                            label=label,
                            field="address_detail",
                            message=str(exc),
                        )
                    )

        update_plan.append(
            {
                "row": branch_row,
                "id": branch_id,
                "label": label,
                "name": cleaned_name,
                "name_changed": name_requested and cleaned_name != str(branch_row.name or "").strip(),
                "status": next_status,
                "status_changed": status_requested and next_status != bool(branch_row.status),
                "location": resolved_location,
                "district": cleaned_district,
                "district_changed": (
                    "district_name" in changed_fields
                    and cleaned_district
                    != str(getattr(branch_row, "district_name", "") or "").strip()
                ),
                "neighborhood": cleaned_neighborhood,
                "neighborhood_changed": (
                    "neighborhood_name" in changed_fields
                    and cleaned_neighborhood
                    != str(getattr(branch_row, "neighborhood_name", "") or "").strip()
                ),
            }
        )

    final_names = {
        branch.id: str(branch.name or "").strip()
        for branch in db.query(models.Branch).all()
    }
    for planned in update_plan:
        final_names[int(planned["id"])] = str(planned["name"])
    name_counts: dict[str, int] = {}
    for final_name in final_names.values():
        normalized_name = final_name.casefold()
        if normalized_name:
            name_counts[normalized_name] = name_counts.get(normalized_name, 0) + 1
    for planned in update_plan:
        normalized_name = str(planned["name"]).casefold()
        if planned["name_changed"] and name_counts.get(normalized_name, 0) > 1:
            errors.append(
                _bulk_item_error(
                    item_id=planned["id"],
                    label=str(planned["label"]),
                    field="name",
                    message="Another branch already uses this name.",
                )
            )

    status_changed_group_ids = {
        getattr(planned["row"], "school_group_id", None)
        for planned in update_plan
        if planned["status_changed"]
    }
    if status_changed_group_ids:
        final_statuses = {
            branch.id: bool(branch.status)
            for branch in db.query(models.Branch).filter(
                models.Branch.school_group_id.in_(status_changed_group_ids)
            ).all()
        }
        for planned in update_plan:
            final_statuses[int(planned["id"])] = bool(planned["status"])
        for group_id in status_changed_group_ids:
            group_branch_ids = {
                branch.id
                for branch in db.query(models.Branch.id).filter(
                    models.Branch.school_group_id == group_id
                ).all()
            }
            if group_branch_ids and not any(final_statuses.get(branch_id, False) for branch_id in group_branch_ids):
                for planned in update_plan:
                    if (
                        getattr(planned["row"], "school_group_id", None) == group_id
                        and planned["status_changed"]
                        and not planned["status"]
                    ):
                        errors.append(
                            _bulk_item_error(
                                item_id=planned["id"],
                                label=str(planned["label"]),
                                field="status",
                                message="At least one active branch must remain in the organization.",
                            )
                        )

    if errors:
        return _bulk_json_error(errors)

    for planned in update_plan:
        branch_row = planned["row"]
        if planned["name_changed"]:
            branch_row.name = planned["name"]
        if planned["status_changed"]:
            branch_row.status = planned["status"]
        resolved_location = planned["location"]
        if resolved_location:
            branch_row.location = resolved_location.region_name
            branch_row.country_code = resolved_location.country_code
            branch_row.country_name = resolved_location.country_name
            branch_row.region_name = resolved_location.region_name
            branch_row.city_name = resolved_location.city_name
        if planned["district_changed"]:
            branch_row.district_name = planned["district"]
        if planned["neighborhood_changed"]:
            branch_row.neighborhood_name = planned["neighborhood"]

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return _bulk_json_error(
            [
                _bulk_item_error(
                    item_id=None,
                    label="Branches",
                    field="database",
                    message="Branch changes conflict with existing records.",
                )
            ],
            status_code=409,
        )

    return JSONResponse(
        {
            "ok": True,
            "saved_count": len(update_plan),
            "message": f"Saved {len(update_plan)} modified branch"
            + ("." if len(update_plan) == 1 else "es."),
        }
    )


@app.post("/system-configuration/branches/{branch_id}")
def update_branch(
    branch_id: int,
    request: Request,
    name: str = Form(...),
    region: str = Form(""),
    country_code: str = Form(""),
    region_id: str = Form(""),
    region_manual: str = Form(""),
    city_id: str = Form(""),
    city_manual: str = Form(""),
    district_name: str = Form(""),
    neighborhood_name: str = Form(""),
    status: str = Form("active"),
    return_to: str = Form("/system-configuration/branches"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_any_permission(db, current_user, "branches.edit", "branches.activate_deactivate"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("branches.edit", "branches.activate_deactivate"),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    branch_row = db.query(models.Branch).filter(
        models.Branch.id == branch_id
    ).first()
    if not branch_row:
        return _redirect_with_error(
            safe_return_to,
            "Branch record not found.",
        )

    safe_return_to, _, _ = _safe_redirect_path(return_to).partition("#")

    cleaned_name = " ".join(str(name or "").split())
    normalized_status = str(status or "").strip().lower()
    next_status = normalized_status != "inactive"

    if not cleaned_name:
        return _redirect_with_error(
            safe_return_to,
            "Branch name is required.",
        )

    try:
        resolved_location = _resolve_submitted_location(
            country_code=country_code,
            region_id=region_id,
            region_manual=region_manual,
            city_id=city_id,
            city_manual=city_manual,
            legacy_region=region,
            required=False,
        )
        cleaned_district = _normalize_location_detail(district_name, "District")
        cleaned_neighborhood = _normalize_location_detail(
            neighborhood_name,
            "Neighborhood",
        )
    except location_service.LocationValidationError as exc:
        return _redirect_with_error(safe_return_to, str(exc))

    duplicate_branch = db.query(models.Branch).filter(
        func.lower(models.Branch.name) == cleaned_name.lower(),
        models.Branch.id != branch_id,
    ).first()
    if duplicate_branch:
        return _redirect_with_error(
            safe_return_to,
            "Another branch already uses that name.",
        )

    active_branch_count = db.query(models.Branch).filter(
        models.Branch.status == True
    ).count()
    if (
        branch_row.status
        and not next_status
        and active_branch_count <= 1
    ):
        return _redirect_with_error(
            safe_return_to,
            "At least one active branch must remain available.",
        )

    branch_row.name = cleaned_name
    if resolved_location:
        branch_row.location = resolved_location.region_name
        branch_row.country_code = resolved_location.country_code
        branch_row.country_name = resolved_location.country_name
        branch_row.region_name = resolved_location.region_name
        branch_row.city_name = resolved_location.city_name
    branch_row.district_name = cleaned_district
    branch_row.neighborhood_name = cleaned_neighborhood
    branch_row.status = next_status
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        "Branch updated successfully.",
    )


# ---------------------------------------
# DEVELOPER: CREATE QUALIFICATION OPTION
# ---------------------------------------
@app.post("/system-configuration/qualifications")
def create_qualification_option(
    request: Request,
    label: str = Form(...),
    kind: str = Form(...),
    return_to: str = Form("/system-configuration"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "branches.delete"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("branches.delete",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    normalized_kind = _normalize_qualification_kind(kind)
    cleaned_label = _normalize_qualification_label(label)

    if not cleaned_label:
        return _redirect_with_error(
            safe_return_to,
            "Qualification label is required.",
        )

    ensure_qualification_options_seeded(db)
    duplicate_label = db.query(models.QualificationOption).filter(
        func.lower(models.QualificationOption.label) == cleaned_label.lower(),
        models.QualificationOption.kind == normalized_kind,
    ).first()
    if duplicate_label:
        return _redirect_with_error(
            safe_return_to,
            f"{cleaned_label} already exists in this qualification section.",
        )

    base_key = build_qualification_key(cleaned_label)
    if not base_key:
        return _redirect_with_error(
            safe_return_to,
            "Unable to create a qualification key from the provided label.",
        )

    candidate_key = base_key
    suffix = 2
    while db.query(models.QualificationOption).filter(
        models.QualificationOption.qualification_key == candidate_key
    ).first():
        candidate_key = f"{base_key}_{suffix}"
        suffix += 1

    db.add(
        models.QualificationOption(
            qualification_key=candidate_key,
            label=cleaned_label,
            kind=normalized_kind,
            alignment_keys=",".join(
                get_subject_alignment_group_keys(cleaned_label)
                if normalized_kind == QUALIFICATION_KIND_SPECIALIZATION
                else []
            ),
            legacy_aliases=cleaned_label.lower(),
            sort_order=0,
        )
    )
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        f"{cleaned_label} added to configuration.",
    )


# ---------------------------------------
# DEVELOPER: UPDATE QUALIFICATION OPTION
# ---------------------------------------
@app.post("/system-configuration/qualifications/{qualification_key}")
def update_qualification_option(
    qualification_key: str,
    request: Request,
    label: str = Form(...),
    return_to: str = Form("/system-configuration"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "academic_years.activate"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("academic_years.activate",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    ensure_qualification_options_seeded(db)
    option_row = db.query(models.QualificationOption).filter(
        models.QualificationOption.qualification_key == qualification_key
    ).first()
    if not option_row:
        return _redirect_with_error(
            safe_return_to,
            "Qualification record not found.",
        )

    cleaned_label = _normalize_qualification_label(label)

    if not cleaned_label:
        return _redirect_with_error(
            safe_return_to,
            "Qualification label is required.",
        )

    duplicate_label = db.query(models.QualificationOption).filter(
        func.lower(models.QualificationOption.label) == cleaned_label.lower(),
        models.QualificationOption.kind == option_row.kind,
        models.QualificationOption.qualification_key != option_row.qualification_key,
    ).first()
    if duplicate_label:
        return _redirect_with_error(
            safe_return_to,
            f"Another {option_row.kind} already uses the label {cleaned_label}.",
        )

    option_row.label = cleaned_label
    if option_row.kind == QUALIFICATION_KIND_SPECIALIZATION:
        option_row.alignment_keys = ",".join(
            get_subject_alignment_group_keys(cleaned_label)
        )
    normalized_aliases = set(
        filter(
            None,
            [
                alias.strip()
                for alias in str(option_row.legacy_aliases or "").split(",")
            ],
        )
    )
    normalized_aliases.add(build_qualification_key(cleaned_label).replace("_", " "))
    option_row.legacy_aliases = ",".join(sorted(normalized_aliases))
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        f"{cleaned_label} updated successfully.",
    )


# ---------------------------------------
# DEVELOPER: DELETE QUALIFICATION OPTION
# ---------------------------------------
@app.post("/system-configuration/qualifications/{qualification_key}/delete")
def delete_qualification_option(
    qualification_key: str,
    request: Request,
    return_to: str = Form("/system-configuration"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "academic_years.create"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("academic_years.create",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    ensure_qualification_options_seeded(db)
    option_row = db.query(models.QualificationOption).filter(
        models.QualificationOption.qualification_key == qualification_key
    ).first()
    if not option_row:
        return _redirect_with_error(
            safe_return_to,
            "Qualification record not found.",
        )

    usage_count = db.query(models.TeacherQualificationSelection).filter(
        models.TeacherQualificationSelection.qualification_key == qualification_key
    ).count()
    if usage_count > 0:
        return _redirect_with_error(
            safe_return_to,
            f"{option_row.label} is already used by {usage_count} teacher selection"
            + ("" if usage_count == 1 else "s")
            + " and cannot be deleted yet.",
        )

    db.delete(option_row)
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        f"{option_row.label} deleted successfully.",
    )


# ---------------------------------------
# DEVELOPER: DELETE BRANCH
# ---------------------------------------
@app.post("/system-configuration/branches/{branch_id}/delete")
def delete_branch(
    branch_id: int,
    request: Request,
    return_to: str = Form("/system-configuration/branches"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user or not auth.can_manage_system_settings(current_user):
        return RedirectResponse(url="/", status_code=302)

    safe_return_to = _safe_redirect_path(return_to)
    branch_row = db.query(models.Branch).filter(
        models.Branch.id == branch_id
    ).first()
    if not branch_row:
        return _redirect_with_error(
            safe_return_to,
            "Branch record not found.",
        )

    usage_counts = _branch_usage_counts(db, branch_id)
    linked_records_count = sum(int(value or 0) for value in usage_counts.values())
    if linked_records_count > 0:
        return _redirect_with_error(
            safe_return_to,
            "This branch is already linked to system records. Update or deactivate it instead of deleting it.",
        )

    active_branch_count = db.query(models.Branch).filter(
        models.Branch.status == True
    ).count()
    if branch_row.status and active_branch_count <= 1:
        return _redirect_with_error(
            safe_return_to,
            "At least one active branch must remain available.",
        )

    db.delete(branch_row)
    db.commit()

    return _redirect_with_notice(
        safe_return_to,
        "Branch deleted successfully.",
    )


# ---------------------------------------
# ADMIN: DELETE EMPTY ACADEMIC YEAR
# ---------------------------------------
@app.post("/system-configuration/academic-years/{academic_year_id}/delete")
def delete_academic_year(
    academic_year_id: int,
    request: Request,
    return_to: str = Form("/system-configuration/schools"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.has_permission(db, current_user, "academic_years.delete"):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("academic_years.delete",),
            page_key="system-configuration",
        )

    safe_return_to = _safe_redirect_path(return_to)
    academic_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.id == academic_year_id
    ).first()
    if not academic_year:
        return _redirect_with_error(safe_return_to, "Academic year record not found.")

    if (
        not _can_manage_all_school_scopes(db, current_user)
        and academic_year.school_group_id != _get_user_school_group_id(db, current_user)
    ):
        return RedirectResponse(url="/dashboard", status_code=302)

    usage_counts = _academic_year_usage_counts(db, academic_year.id)
    linked_records_count = sum(int(count or 0) for count in usage_counts.values())
    if linked_records_count:
        usage_summary = ", ".join(
            f"{usage_counts[key]} {label}"
            for key, label, _model in ACADEMIC_YEAR_USAGE_MODELS
            if usage_counts[key]
        )
        return _redirect_with_error(
            safe_return_to,
            "This academic year cannot be deleted because it contains "
            f"{usage_summary}.",
        )

    db.delete(academic_year)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return _redirect_with_error(
            safe_return_to,
            "This academic year cannot be deleted because linked system data still exists.",
        )

    response = _redirect_with_notice(safe_return_to, "Academic year deleted successfully.")
    if str(request.cookies.get("academic_year_id") or "") == str(academic_year_id):
        response.delete_cookie("academic_year_id")
    return response


# ---------------------------------------
# ADMIN: SET CURRENT YEAR
# ---------------------------------------
@app.post("/admin/current-year")
def set_current_year(
    request: Request,
    academic_year_id: int = Form(...),
    return_to: str = Form("/dashboard"),
    db: Session = Depends(get_db)
):
    current_user = auth.get_current_user(request, db)

    if not current_user or not auth.can_manage_system_settings(current_user):
        return RedirectResponse(url="/", status_code=302)

    target_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.id == academic_year_id
    ).first()

    if not target_year:
        return RedirectResponse(
            url=_safe_redirect_path(return_to),
            status_code=302,
        )

    target_group_id = getattr(target_year, "school_group_id", None)
    deactivate_query = db.query(models.AcademicYear)
    if target_group_id:
        deactivate_query = deactivate_query.filter(models.AcademicYear.school_group_id == target_group_id)
    deactivate_query.update({models.AcademicYear.is_active: False}, synchronize_session=False)
    target_year.is_active = True
    db.commit()

    response = _redirect_with_notice(
        return_to,
        "Current academic year updated successfully.",
    )
    auth.set_scope_cookie(response, "academic_year_id", target_year.id, request)
    return response


# ---------------------------------------
# DEVELOPER: OPEN NEW ACADEMIC YEAR
# ---------------------------------------
@app.post("/developer/open-academic-year")
def open_new_academic_year(
    request: Request,
    year_name: str = Form(...),
    school_group_id: int | None = Form(None),
    return_to: str = Form("/dashboard"),
    db: Session = Depends(get_db)
):
    current_user = auth.get_current_user(request, db)
    if not current_user or not auth.can_manage_system_settings(current_user):
        return RedirectResponse(url="/", status_code=302)

    cleaned_year_name = year_name.strip()
    if not ACADEMIC_YEAR_NAME_PATTERN.match(cleaned_year_name):
        return _redirect_with_error(
            return_to,
            "Academic year names must use the YYYY-YYYY format.",
        )

    selected_group_id = school_group_id or _get_user_school_group_id(db, current_user)
    existing_query = db.query(models.AcademicYear).filter(
        models.AcademicYear.year_name == cleaned_year_name
    )
    if selected_group_id:
        existing_query = existing_query.filter(models.AcademicYear.school_group_id == selected_group_id)
    existing_year = existing_query.first()
    if existing_year:
        target_year = existing_year
        deactivate_query = db.query(models.AcademicYear)
        if selected_group_id:
            deactivate_query = deactivate_query.filter(models.AcademicYear.school_group_id == selected_group_id)
        deactivate_query.update({models.AcademicYear.is_active: False}, synchronize_session=False)
        target_year.is_active = True
        db.commit()
    else:
        deactivate_query = db.query(models.AcademicYear)
        if selected_group_id:
            deactivate_query = deactivate_query.filter(models.AcademicYear.school_group_id == selected_group_id)
        deactivate_query.update({models.AcademicYear.is_active: False}, synchronize_session=False)
        target_year = models.AcademicYear(
            school_group_id=selected_group_id,
            year_name=cleaned_year_name,
            is_active=True
        )
        db.add(target_year)
        db.commit()
        db.refresh(target_year)

    notice_message = (
        "Academic year reactivated and set as current."
        if existing_year
        else "Academic year opened successfully."
    )
    response = _redirect_with_notice(
        return_to,
        notice_message,
    )
    auth.set_scope_cookie(response, "academic_year_id", target_year.id, request)
    return response


# ---------------------------------------
# SCOPE: SET CURRENT ACADEMIC YEAR
# ---------------------------------------
@app.post("/scope/academic-year")
def set_scope_academic_year(
    request: Request,
    academic_year_id: int = Form(...),
    return_to: str = Form("/dashboard"),
    db: Session = Depends(get_db)
):
    current_user = auth.get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    if not auth.can_access_all_branches(current_user, db):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("branches.view", "schools.manage_all_schools"),
            page_key="dashboard",
        )

    target_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.id == academic_year_id
    ).first()
    if not target_year:
        return RedirectResponse(
            url=_safe_redirect_path(return_to),
            status_code=302,
        )
    scoped_branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    if not auth.validate_branch_year_scope(
        db,
        branch_id=scoped_branch_id,
        academic_year_id=target_year.id,
        current_user=current_user,
    ):
        return RedirectResponse(
            url=_safe_redirect_path(return_to),
            status_code=302,
        )

    response = RedirectResponse(
        url=_safe_redirect_path(return_to),
        status_code=302,
    )
    auth.set_scope_cookie(response, "academic_year_id", target_year.id, request)
    return response


# ---------------------------------------
# SCOPE: SET CURRENT ORGANIZATION
# ---------------------------------------
@app.post("/scope/organization")
def set_scope_organization(
    request: Request,
    school_group_id: int = Form(...),
    return_to: str = Form("/platform"),
    db: Session = Depends(get_db),
):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/", status_code=302)
    if not auth.is_platform_user(current_user):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("system_owner.switch_all_schools",),
            page_key="platform",
            message="Only platform users can select an organization context.",
        )

    target_group = db.query(models.SchoolGroup).filter(
        models.SchoolGroup.id == school_group_id
    ).first()
    if not target_group:
        return RedirectResponse(url=_safe_redirect_path(return_to), status_code=302)

    response = RedirectResponse(url=_safe_redirect_path(return_to), status_code=302)
    auth.set_scope_cookie(response, "school_group_id", target_group.id, request)
    response.delete_cookie("branch_id")
    response.delete_cookie("academic_year_id")
    return response


# ---------------------------------------
# SCOPE: SET CURRENT BRANCH
# ---------------------------------------
@app.post("/scope/branch")
def set_scope_branch(
    request: Request,
    branch_id: int = Form(...),
    return_to: str = Form("/dashboard"),
    db: Session = Depends(get_db)
):
    current_user = auth.get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/", status_code=302)

    if not auth.can_access_all_branches(current_user, db):
        return authorization.build_access_denied_response(
            request,
            db,
            current_user=current_user,
            permission_keys=("schools.manage_all_schools", "dashboard.view_all_schools", "system_owner.switch_all_schools"),
            page_key="dashboard",
        )

    target_branch = auth.get_accessible_branch_query(db, current_user).filter(
        models.Branch.id == branch_id
    ).first()
    if not target_branch:
        return RedirectResponse(
            url=_safe_redirect_path(return_to),
            status_code=302,
        )
    target_group_id = getattr(target_branch, "school_group_id", None)
    current_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id,
    )
    target_year = auth.get_academic_year_for_school_group(db, current_year_id, target_group_id)
    if not target_year:
        target_year = auth.get_active_academic_year_for_school_group(db, target_group_id)
    if not target_year and auth.is_platform_user(current_user):
        target_year = auth.get_latest_academic_year_for_school_group(db, target_group_id)
    if not target_year:
        return RedirectResponse(
            url=_safe_redirect_path(return_to),
            status_code=302,
        )

    response = RedirectResponse(
        url=_safe_redirect_path(return_to),
        status_code=302,
    )
    auth.set_scope_cookie(response, "branch_id", target_branch.id, request)
    auth.set_scope_cookie(response, "school_group_id", target_group_id, request)
    auth.set_scope_cookie(response, "academic_year_id", target_year.id, request)
    return response


# ---------------------------------------
# DASHBOARD
# ---------------------------------------
@app.get("/dashboard")
def dashboard(
    request: Request,
    db: Session = Depends(get_db)
):
    user = auth.get_current_user(request, db)

    if not user:
        return RedirectResponse(url="/")

    report_package = _build_current_report_package(db, user)
    branch_name = report_package["branch_name"]
    academic_year_name = report_package["academic_year_name"]
    subject_count = report_package["subject_count"]
    teacher_count = report_package["teacher_count"]
    planning_total_sections = report_package["planning_total_sections"]
    planning_current_sections_count = report_package["planning_current_sections_count"]
    planning_new_sections_count = report_package["planning_new_sections_count"]
    planning_total_allocated_hours = report_package["planning_total_allocated_hours"]
    subjects_dashboard_rows = report_package["subjects_dashboard_rows"]
    report_summary = report_package["report_summary"]
    report_subject_count = report_package["report_subject_count"]
    report_subject_rows = report_package["report_subject_rows"]
    report_subject_card_rows = report_package["report_subject_card_rows"]
    report_gap_rows = report_package["report_gap_rows"]
    report_teacher_rows = report_package["report_teacher_rows"]
    report_visuals = report_package["report_visuals"]
    subject_pool_distribution_rows = report_package["subject_pool_distribution_rows"]
    subject_pool_distribution_summary = report_package["subject_pool_distribution_summary"]
    hiring_plan_editor_auto_payload = report_package["hiring_plan_editor_auto_payload"]
    allocation_data = report_package["allocation_data"]
    reporting_context = report_package["reporting_context"]
    scoped_academic_year_id = report_package["scoped_academic_year_id"]
    scoped_branch_id = report_package["scoped_branch_id"]
    all_years = db.query(models.AcademicYear).order_by(
        models.AcademicYear.year_name.desc()
    ).all()
    year_map = {
        year.id: year.year_name for year in all_years
    }
    available_scope_branches = db.query(models.Branch).filter(
        models.Branch.status == True
    ).order_by(models.Branch.name.asc()).all()
    can_manage_system_settings = auth.can_manage_system_settings(user)
    info_message = ""
    if request.query_params.get("info") == "already-logged-in":
        info_message = "You are already logged in."
    active_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.is_active == True
    ).first()

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "request": request,
            "user": user,
            "branch_name": branch_name,
            "academic_year_name": academic_year_name,
            "subject_count": subject_count,
            "teacher_count": teacher_count,
            "planning_total_sections": planning_total_sections,
            "planning_current_sections_count": planning_current_sections_count,
            "planning_new_sections_count": planning_new_sections_count,
            "planning_total_allocated_hours": planning_total_allocated_hours,
            "subjects_dashboard_rows": subjects_dashboard_rows,
            "report_summary": report_summary,
            "report_subject_count": report_subject_count,
            "report_subject_rows": report_subject_rows,
            "report_subject_card_rows": report_subject_card_rows,
            "report_gap_rows": report_gap_rows,
            "report_teacher_rows": report_teacher_rows,
            "report_underloaded_teacher_rows": allocation_data.get(
                "underloaded_teacher_rows",
                [],
            ),
            "report_grade_rows": reporting_context["grade_rows"],
            "report_visuals": report_visuals,
            "subject_pool_distribution_rows": subject_pool_distribution_rows,
            "subject_pool_distribution_summary": subject_pool_distribution_summary,
            "hiring_plan_editor_auto_payload": hiring_plan_editor_auto_payload,
            "all_years": all_years,
            "year_map": year_map,
            "can_manage_system_settings": can_manage_system_settings,
            "info_message": info_message,
            "scoped_academic_year_id": scoped_academic_year_id,
            "available_scope_branches": available_scope_branches,
            "scoped_branch_id": scoped_branch_id,
            "active_year_id": active_year.id if active_year else None,
            **build_shell_context(
                request,
                db,
                user,
                page_key="dashboard",
                notice=info_message,
            ),
        }
    )


def _build_hiring_plan_auto_payload_for_scope(
    db: Session,
    scoped_branch_id: int,
    scoped_academic_year_id: int,
) -> dict:
    subjects_dashboard_rows = db.query(models.Subject).filter(
        models.Subject.branch_id == scoped_branch_id,
        models.Subject.academic_year_id == scoped_academic_year_id,
    ).order_by(
        models.Subject.grade.asc(),
        models.Subject.subject_code.asc(),
    ).all()
    for subject in subjects_dashboard_rows:
        bundle_subject_labels = list(
            get_homeroom_bundle_subject_labels(
                subject_code=subject.subject_code or "",
                subject_name=subject.subject_name or "",
                weekly_hours=subject.weekly_hours,
                grade_label=_normalize_grade_label(subject.grade),
            )
        )
        setattr(
            subject,
            "effective_subject_count",
            get_effective_subject_count(
                subject_code=subject.subject_code or "",
                subject_name=subject.subject_name or "",
                weekly_hours=subject.weekly_hours,
                grade_label=_normalize_grade_label(subject.grade),
            ),
        )
        setattr(subject, "homeroom_bundle_subject_labels", bundle_subject_labels)

    planning_sections = db.query(models.PlanningSection).filter(
        models.PlanningSection.branch_id == scoped_branch_id,
        models.PlanningSection.academic_year_id == scoped_academic_year_id,
    ).order_by(
        models.PlanningSection.grade_level.asc(),
        models.PlanningSection.section_name.asc(),
        models.PlanningSection.id.asc(),
    ).all()
    teachers_for_reporting = db.query(models.Teacher).filter(
        models.Teacher.branch_id == scoped_branch_id,
        models.Teacher.academic_year_id == scoped_academic_year_id,
    ).order_by(models.Teacher.id.asc()).all()

    planning_section_ids = [
        section.id
        for section in planning_sections
        if getattr(section, "id", None)
    ]
    section_assignments = []
    if planning_section_ids:
        section_assignments = db.query(models.TeacherSectionAssignment).filter(
            models.TeacherSectionAssignment.planning_section_id.in_(planning_section_ids)
        ).all()

    reporting_context = _build_reporting_context_from_section_assignments(
        db=db,
        subjects=subjects_dashboard_rows,
        planning_sections=planning_sections,
        teachers=teachers_for_reporting,
        section_assignments=section_assignments,
    )
    allocation_data = _build_report_class_allocation_data_from_section_assignments(
        db=db,
        subjects=subjects_dashboard_rows,
        planning_sections=planning_sections,
        teachers=teachers_for_reporting,
        reporting_context=reporting_context,
        section_assignments=section_assignments,
    )
    subject_section_map = allocation_data.get("subject_section_map", {})
    report_subject_rows = [
        {
            **row,
            **subject_section_map.get(row["subject_key"], {}),
        }
        for row in reporting_context["subject_rows"]
    ]
    report_subject_rows, report_summary = _decorate_staffing_report_rows(
        report_subject_rows,
        reporting_context["summary"],
    )
    return _build_hiring_plan_editor_payload(
        report_summary=report_summary,
        report_subject_rows=report_subject_rows,
    )


@app.get("/dashboard/api/hiring-plan")
def load_dashboard_hiring_plan(
    request: Request,
    db: Session = Depends(get_db),
):
    user = auth.get_current_user(request, db)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Authentication required."})

    scoped_branch_id = getattr(user, "scope_branch_id", user.branch_id)
    scoped_academic_year_id = getattr(user, "scope_academic_year_id", user.academic_year_id)
    if not scoped_branch_id or not scoped_academic_year_id:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Scope is not configured."})

    draft = db.query(models.HiringPlanDraft).filter(
        models.HiringPlanDraft.branch_id == int(scoped_branch_id),
        models.HiringPlanDraft.academic_year_id == int(scoped_academic_year_id),
        models.HiringPlanDraft.user_id == int(user.id),
    ).first()
    if not draft:
        return {"ok": True, "source": "none", "plan": None}

    try:
        plan_payload = json.loads(str(draft.plan_json or "{}"))
    except json.JSONDecodeError:
        plan_payload = {}
    try:
        saved_pool_logic_version = int(plan_payload.get("pool_logic_version", 0) or 0)
    except (TypeError, ValueError):
        saved_pool_logic_version = 0
    if saved_pool_logic_version < HIRING_PLAN_POOL_LOGIC_VERSION:
        return {
            "ok": True,
            "source": "outdated",
            "plan": None,
            "updated_at": draft.updated_at.isoformat() if getattr(draft, "updated_at", None) else None,
        }
    normalized_plan = _normalize_hiring_plan_payload(plan_payload)
    normalized_plan["locked"] = True
    return {
        "ok": True,
        "source": "saved",
        "plan": normalized_plan,
        "updated_at": draft.updated_at.isoformat() if getattr(draft, "updated_at", None) else None,
    }


@app.get("/dashboard/api/hiring-plan/effective")
def load_dashboard_effective_hiring_plan(
    request: Request,
    db: Session = Depends(get_db),
):
    user = auth.get_current_user(request, db)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Authentication required."})

    scoped_branch_id = getattr(user, "scope_branch_id", user.branch_id)
    scoped_academic_year_id = getattr(user, "scope_academic_year_id", user.academic_year_id)
    if not scoped_branch_id or not scoped_academic_year_id:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Scope is not configured."})

    draft = db.query(models.HiringPlanDraft).filter(
        models.HiringPlanDraft.branch_id == int(scoped_branch_id),
        models.HiringPlanDraft.academic_year_id == int(scoped_academic_year_id),
        models.HiringPlanDraft.user_id == int(user.id),
    ).first()
    if draft:
        try:
            plan_payload = json.loads(str(draft.plan_json or "{}"))
        except json.JSONDecodeError:
            plan_payload = {}
        try:
            saved_pool_logic_version = int(plan_payload.get("pool_logic_version", 0) or 0)
        except (TypeError, ValueError):
            saved_pool_logic_version = 0
        if saved_pool_logic_version >= HIRING_PLAN_POOL_LOGIC_VERSION:
            normalized_plan = _normalize_hiring_plan_payload(plan_payload)
            normalized_plan["locked"] = True
            return {
                "ok": True,
                "source": "saved",
                "plan": normalized_plan,
                "updated_at": draft.updated_at.isoformat() if getattr(draft, "updated_at", None) else None,
            }

    auto_plan = _build_hiring_plan_auto_payload_for_scope(
        db=db,
        scoped_branch_id=int(scoped_branch_id),
        scoped_academic_year_id=int(scoped_academic_year_id),
    )
    auto_plan["locked"] = True
    return {
        "ok": True,
        "source": "auto",
        "plan": auto_plan,
    }


@app.post("/dashboard/api/hiring-plan/save")
async def save_dashboard_hiring_plan(
    request: Request,
    db: Session = Depends(get_db),
):
    user = auth.get_current_user(request, db)
    if not user:
        return JSONResponse(status_code=401, content={"ok": False, "error": "Authentication required."})

    scoped_branch_id = getattr(user, "scope_branch_id", user.branch_id)
    scoped_academic_year_id = getattr(user, "scope_academic_year_id", user.academic_year_id)
    if not scoped_branch_id or not scoped_academic_year_id:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Scope is not configured."})

    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Invalid JSON payload."})

    normalized_plan = _normalize_hiring_plan_payload(body.get("plan", {}))
    warnings = _collect_hiring_plan_warnings(normalized_plan)
    now_utc = datetime.utcnow()

    draft = db.query(models.HiringPlanDraft).filter(
        models.HiringPlanDraft.branch_id == int(scoped_branch_id),
        models.HiringPlanDraft.academic_year_id == int(scoped_academic_year_id),
        models.HiringPlanDraft.user_id == int(user.id),
    ).first()

    if not draft:
        draft = models.HiringPlanDraft(
            branch_id=int(scoped_branch_id),
            academic_year_id=int(scoped_academic_year_id),
            user_id=int(user.id),
            plan_json=json.dumps(normalized_plan, ensure_ascii=False),
            updated_at=now_utc,
        )
        db.add(draft)
    else:
        draft.plan_json = json.dumps(normalized_plan, ensure_ascii=False)
        draft.updated_at = now_utc

    db.commit()

    return {
        "ok": True,
        "warnings": warnings,
        "updated_at": now_utc.isoformat(),
    }


# ---------------------------------------
# REPORT EXPORT
# ---------------------------------------
@app.get("/reports/allocation-plan.xlsx")
def download_report_allocation_plan(
    request: Request,
    section: str = Query("full"),
    db: Session = Depends(get_db),
):
    user = auth.get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/")

    normalized_section = _normalize_report_export_section(section)
    report_package = _build_current_report_package(db, user)
    payload = _build_professional_report_xlsx_bytes(
        report_package,
        section=normalized_section,
    )
    file_name = _build_report_export_filename(
        report_package["branch_name"],
        report_package["academic_year_name"],
        "xlsx",
        normalized_section,
    )

    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@app.get("/reports/allocation-plan.pdf")
def download_report_allocation_plan_pdf(
    request: Request,
    section: str = Query("full"),
    db: Session = Depends(get_db),
):
    user = auth.get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/")

    normalized_section = _normalize_report_export_section(section)
    report_package = _build_current_report_package(db, user)
    payload = _build_professional_report_pdf_bytes(
        report_package,
        section=normalized_section,
    )
    file_name = _build_report_export_filename(
        report_package["branch_name"],
        report_package["academic_year_name"],
        "pdf",
        normalized_section,
    )

    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


# ---------------------------------------
# Startup Schema Compatibility
# ---------------------------------------
def _ensure_users_table_columns():
    inspector = inspect(engine)
    if "users" not in inspector.get_table_names():
        return

    existing_columns = {
        col["name"] for col in inspector.get_columns("users")
    }

    if "is_active" not in existing_columns:
        return

    with engine.begin() as connection:
        connection.execute(
            text("UPDATE users SET is_active = TRUE WHERE is_active IS NULL")
        )


def _ensure_teachers_table_columns():
    inspector = inspect(engine)
    if "teachers" not in inspector.get_table_names():
        return

    existing_columns = {
        col["name"] for col in inspector.get_columns("teachers")
    }

    with engine.begin() as connection:
        for column_name, default_value in (
            ("extra_hours_allowed", "FALSE"),
            ("extra_hours_count", "0"),
            ("teaches_national_section", "FALSE"),
            ("national_section_hours", "0"),
            ("is_new_teacher", "FALSE"),
        ):
            if column_name not in existing_columns:
                continue
            connection.execute(
                text(
                    f"UPDATE teachers SET {column_name} = {default_value} "
                    f"WHERE {column_name} IS NULL"
                )
            )


def _ensure_teacher_subject_allocation_columns():
    inspector = inspect(engine)
    if "teacher_subject_allocations" not in inspector.get_table_names():
        return

    existing_columns = {
        col["name"] for col in inspector.get_columns("teacher_subject_allocations")
    }

    if "compatibility_override" not in existing_columns:
        return

    with engine.begin() as connection:
        connection.execute(
            text(
                "UPDATE teacher_subject_allocations "
                "SET compatibility_override = FALSE "
                "WHERE compatibility_override IS NULL"
            )
        )


def _ensure_timetable_non_teaching_block_columns():
    # Timetable block schema is managed by db_migrations.
    return


def _ensure_system_notifications_table_columns():
    inspector = inspect(engine)
    if "system_notifications" not in inspector.get_table_names():
        return

    existing_columns = {
        col["name"] for col in inspector.get_columns("system_notifications")
    }

    with engine.begin() as connection:
        for column_name, default_value in (
            ("recipient_scope", "User"),
            ("status", "New"),
            ("request_type", "Message"),
            ("title", "System Notification"),
        ):
            if column_name not in existing_columns:
                continue
            connection.execute(
                text(
                    "UPDATE system_notifications "
                    f"SET {column_name} = :default_value "
                    f"WHERE {column_name} IS NULL OR {column_name} = ''"
                ),
                {"default_value": default_value},
            )

        if "created_at" in existing_columns:
            created_at_missing_predicate = (
                "created_at IS NULL"
                if engine.dialect.name == "postgresql"
                else "created_at IS NULL OR created_at = '' OR datetime(created_at) IS NULL"
            )
            connection.execute(
                text(
                    "UPDATE system_notifications "
                    "SET created_at = CURRENT_TIMESTAMP "
                    f"WHERE {created_at_missing_predicate}"
                )
            )

        if engine.dialect.name != "postgresql":
            for column_name in (
                "seen_at",
                "resolved_at",
                "recipient_archived_at",
                "requester_archived_at",
            ):
                if column_name not in existing_columns:
                    continue
                connection.execute(
                    text(
                        "UPDATE system_notifications "
                        f"SET {column_name} = NULL "
                        f"WHERE {column_name} = '' "
                        f"OR ({column_name} IS NOT NULL AND datetime({column_name}) IS NULL)"
                    )
                )

    _ensure_system_notifications_indexes()


def _log_notification_schema_compatibility(source: str):
    snapshot = _collect_system_notification_schema_snapshot()
    _notification_logger().info(
        (
            "TIS notification schema compatibility source=%s table_exists=%s "
            "expected_columns=%s actual_columns=%s missing_columns=%s "
            "expected_indexes=%s actual_indexes=%s missing_indexes=%s"
        ),
        source,
        snapshot.get("table_exists"),
        snapshot.get("expected_columns"),
        snapshot.get("actual_columns"),
        snapshot.get("missing_columns"),
        snapshot.get("expected_indexes"),
        snapshot.get("actual_indexes"),
        snapshot.get("missing_indexes"),
    )


def _is_scope_teacher_unique_definition(columns) -> bool:
    return tuple(columns or []) == ("branch_id", "academic_year_id", "teacher_id")


def _is_global_teacher_unique_definition(columns) -> bool:
    return tuple(columns or []) == ("teacher_id",)


def _ensure_teacher_scope_schema_sqlite():
    return


def _ensure_teacher_scope_schema_non_sqlite(
    teacher_unique_constraints,
    teacher_indexes,
):
    return


def _ensure_teacher_scope_schema():
    inspector = inspect(engine)
    if "teachers" not in inspector.get_table_names():
        return

    teacher_unique_constraints = inspector.get_unique_constraints("teachers")
    teacher_indexes = inspector.get_indexes("teachers")

    has_scoped_unique = any(
        _is_scope_teacher_unique_definition(
            unique_constraint.get("column_names") or []
        )
        for unique_constraint in teacher_unique_constraints
    ) or any(
        teacher_index.get("unique")
        and _is_scope_teacher_unique_definition(
            teacher_index.get("column_names") or []
        )
        for teacher_index in teacher_indexes
    )

    has_global_teacher_unique = any(
        _is_global_teacher_unique_definition(
            unique_constraint.get("column_names") or []
        )
        for unique_constraint in teacher_unique_constraints
    ) or any(
        teacher_index.get("unique")
        and _is_global_teacher_unique_definition(
            teacher_index.get("column_names") or []
        )
        for teacher_index in teacher_indexes
    )

    if has_scoped_unique and not has_global_teacher_unique:
        return

    logging.getLogger("uvicorn.error").warning(
        "Teacher scoped uniqueness is not fully migrated; run pending database migrations."
    )


def _is_subject_code_foreign_key(foreign_key) -> bool:
    constrained_columns = set(foreign_key.get("constrained_columns") or [])
    referred_columns = set(foreign_key.get("referred_columns") or [])
    return (
        foreign_key.get("referred_table") == "subjects"
        and "subject_code" in constrained_columns
        and "subject_code" in referred_columns
    )


def _ensure_subject_scope_schema_sqlite(
    rebuild_teachers: bool,
    rebuild_allocations: bool,
    reset_subject_indexes: bool,
):
    return


def _ensure_subject_scope_schema_non_sqlite(
    rebuild_teachers: bool,
    rebuild_allocations: bool,
    reset_subject_indexes: bool,
    has_subject_code_index: bool,
    has_scope_unique_index: bool,
):
    return


def _ensure_subject_scope_schema():
    inspector = inspect(engine)
    table_names = set(inspector.get_table_names())
    required_tables = {"subjects", "teachers", "teacher_subject_allocations"}
    if not required_tables.issubset(table_names):
        return

    subject_indexes = {
        index["name"]: index
        for index in inspector.get_indexes("subjects")
    }
    teacher_subject_fks = [
        foreign_key
        for foreign_key in inspector.get_foreign_keys("teachers")
        if _is_subject_code_foreign_key(foreign_key)
    ]
    allocation_subject_fks = [
        foreign_key
        for foreign_key in inspector.get_foreign_keys("teacher_subject_allocations")
        if _is_subject_code_foreign_key(foreign_key)
    ]

    subject_code_index = subject_indexes.get("ix_subjects_subject_code")
    has_subject_code_index = subject_code_index is not None
    has_scope_unique_index = "uq_subjects_scope_code" in subject_indexes
    reset_subject_indexes = bool(subject_code_index and subject_code_index.get("unique"))
    rebuild_teachers = bool(teacher_subject_fks)
    rebuild_allocations = bool(allocation_subject_fks)

    if not any(
        [
            rebuild_teachers,
            rebuild_allocations,
            reset_subject_indexes,
            not has_scope_unique_index,
        ]
    ):
        return

    logging.getLogger("uvicorn.error").warning(
        "Subject scoped uniqueness is not fully migrated; run pending database migrations."
    )


def _seed_teacher_subject_allocations():
    inspector = inspect(engine)
    if (
        "teachers" not in inspector.get_table_names()
        or "teacher_subject_allocations" not in inspector.get_table_names()
    ):
        return

    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO teacher_subject_allocations (teacher_id, subject_code)
                SELECT t.id, t.subject_code
                FROM teachers t
                WHERE t.subject_code IS NOT NULL
                  AND t.subject_code <> ''
                  AND NOT EXISTS (
                      SELECT 1
                      FROM teacher_subject_allocations a
                      WHERE a.teacher_id = t.id
                        AND a.subject_code = t.subject_code
                  )
                """
            )
        )


def _normalize_branch_name(value: str) -> str:
    return " ".join(str(value or "").split()).strip().lower()


def _get_branch_data_ids(db: Session):
    data_branch_ids = set()
    for model in (User, models.Subject, models.Teacher, models.PlanningSection):
        rows = (
            db.query(model.branch_id)
            .filter(model.branch_id.isnot(None))
            .distinct()
            .all()
        )
        for row in rows:
            if row[0] is not None:
                data_branch_ids.add(row[0])
    return data_branch_ids


def _reassign_branch_scope_data(
    db: Session,
    from_branch_id: int,
    to_branch_id: int,
) -> bool:
    if from_branch_id == to_branch_id:
        return False

    changed = False
    for model in (User, models.Subject, models.Teacher, models.PlanningSection):
        moved_count = (
            db.query(model)
            .filter(model.branch_id == from_branch_id)
            .update({model.branch_id: to_branch_id}, synchronize_session=False)
        )
        if moved_count:
            changed = True
    return changed


def _ensure_gender_branches(db: Session):
    required_base_branch_names = [
        "Hamadania",
        "Manar",
        "Obhor",
        "Alshaati",
        "Fayha",
        "Najran",
        "Zahra",
        "Khamis Msheit",
        "Abha",
        "Rawda",
    ]

    existing_branches = db.query(Branch).order_by(Branch.id.asc()).all()
    branches_by_name = {}
    for branch_row in existing_branches:
        key = _normalize_branch_name(branch_row.name)
        if key:
            branches_by_name.setdefault(key, []).append(branch_row)

    data_branch_ids = _get_branch_data_ids(db)
    branch_changes = False
    default_branch = None

    for base_name in required_base_branch_names:
        boys_name = f"{base_name}-Boys"
        girls_name = f"{base_name}-Girls"

        base_key = _normalize_branch_name(base_name)
        boys_key = _normalize_branch_name(boys_name)
        girls_key = _normalize_branch_name(girls_name)

        legacy_rows = list(branches_by_name.get(base_key, []))
        boys_rows = branches_by_name.get(boys_key, [])
        girls_rows = branches_by_name.get(girls_key, [])

        boys_row = boys_rows[0] if boys_rows else None
        girls_row = girls_rows[0] if girls_rows else None

        if not boys_row:
            rename_candidate = None
            if legacy_rows:
                rename_candidate = next(
                    (row for row in legacy_rows if row.id in data_branch_ids),
                    legacy_rows[0],
                )
                legacy_rows.remove(rename_candidate)

            if rename_candidate:
                if rename_candidate.name != boys_name:
                    rename_candidate.name = boys_name
                    branch_changes = True
                boys_row = rename_candidate
                branches_by_name.setdefault(boys_key, []).append(boys_row)
            else:
                boys_row = Branch(
                    name=boys_name,
                    location="Makkah",
                    country_code="SA",
                    country_name="Saudi Arabia",
                    region_name="Makkah",
                    status=True,
                )
                db.add(boys_row)
                db.flush()
                branch_changes = True
                branches_by_name.setdefault(boys_key, []).append(boys_row)

        if not girls_row:
            rename_candidate = legacy_rows.pop(0) if legacy_rows else None
            if rename_candidate:
                if rename_candidate.id in data_branch_ids:
                    if _reassign_branch_scope_data(db, rename_candidate.id, boys_row.id):
                        branch_changes = True
                    data_branch_ids.discard(rename_candidate.id)
                    data_branch_ids.add(boys_row.id)

                if rename_candidate.name != girls_name:
                    rename_candidate.name = girls_name
                    branch_changes = True
                girls_row = rename_candidate
                branches_by_name.setdefault(girls_key, []).append(girls_row)
            else:
                girls_row = Branch(
                    name=girls_name,
                    location="Makkah",
                    country_code="SA",
                    country_name="Saudi Arabia",
                    region_name="Makkah",
                    status=True,
                )
                db.add(girls_row)
                db.flush()
                branch_changes = True
                branches_by_name.setdefault(girls_key, []).append(girls_row)

        for legacy_row in legacy_rows:
            if legacy_row.id in data_branch_ids:
                if _reassign_branch_scope_data(db, legacy_row.id, boys_row.id):
                    branch_changes = True
                data_branch_ids.discard(legacy_row.id)
                data_branch_ids.add(boys_row.id)

            retired_name = f"{base_name}-Legacy-{legacy_row.id}"
            if legacy_row.name != retired_name:
                legacy_row.name = retired_name
                branch_changes = True
            if legacy_row.status:
                legacy_row.status = False
                branch_changes = True

        if base_name == "Hamadania":
            default_branch = boys_row

    if branch_changes:
        db.commit()

    return default_branch


# ---------------------------------------
# Startup Initialization
# ---------------------------------------
@app.on_event("startup")
def setup_initial_data():

    auth.validate_security_configuration()
    _ensure_school_group_schema()
    _ensure_users_table_columns()
    _ensure_teachers_table_columns()
    _ensure_teacher_scope_schema()
    _ensure_subject_scope_schema()
    _ensure_subject_color_schema()
    _ensure_role_permission_schema()
    _ensure_teacher_subject_allocation_columns()
    _ensure_timetable_non_teaching_block_columns()
    _ensure_system_notifications_table_columns()
    _log_notification_schema_compatibility("startup")
    _seed_teacher_subject_allocations()
    _ensure_profile_photo_upload_dir()
    observations.ensure_observation_schema()
    db = SessionLocal()
    observations.ensure_observation_seed_data(db)
    _seed_global_role_permissions(db)
    _backfill_subject_colors(db)
    _migrate_profile_photos_to_database(db)
    admin_user_id = os.getenv("ADMIN_USER_ID", "2623252018")
    admin_username = os.getenv("ADMIN_USERNAME", "developer")
    admin_password = os.getenv("ADMIN_PASSWORD")

    default_branch = _ensure_gender_branches(db)
    _ensure_default_school_group(db)
    _ensure_organization_branding_storage(db)

    if not default_branch:
        default_branch = db.query(Branch).filter(
            Branch.name == "Hamadania-Boys"
        ).first()
    if not default_branch:
        default_branch = db.query(Branch).filter(
            Branch.status == True
        ).order_by(Branch.id.asc()).first()
    default_school_group_id = getattr(default_branch, "school_group_id", None)
    user_school_group_nullable = next(
        (
            bool(column.get("nullable", True))
            for column in inspect(engine).get_columns("users")
            if column.get("name") == "school_group_id"
        ),
        True,
    )
    platform_storage_group_id = None if user_school_group_nullable else default_school_group_id

    legacy_position_map = {
        "Education Excelency": "Education Excellence",
        "Principle": "Principal",
        "Priciple": "Principal",
    }
    legacy_position_users = db.query(User).filter(
        User.position.in_(legacy_position_map.keys())
    ).all()
    if legacy_position_users:
        for user_row in legacy_position_users:
            normalized_position = legacy_position_map.get(user_row.position)
            if normalized_position:
                user_row.position = normalized_position
        db.commit()

    # Create Academic Year if not exists
    academic_year = db.query(AcademicYear).filter(
        AcademicYear.year_name == "2025-2026"
    ).first()

    if not academic_year:
        academic_year = AcademicYear(
            school_group_id=default_school_group_id,
            year_name="2025-2026",
            is_active=True
        )
        db.add(academic_year)
        db.commit()
        db.refresh(academic_year)
    else:
        active_year = db.query(AcademicYear).filter(
            AcademicYear.is_active == True
        ).first()
        if not active_year:
            academic_year.is_active = True
            if not getattr(academic_year, "school_group_id", None):
                academic_year.school_group_id = default_school_group_id
            db.commit()

    # Create Admin User if not exists
    existing_user = db.query(User).filter(
        User.user_id == admin_user_id
    ).first()

    if not existing_user:
        bootstrap_password = admin_password or secrets.token_urlsafe(24)
        if not admin_password:
            logging.warning(
                "ADMIN_PASSWORD is not set; created bootstrap developer account with a random password. "
                "Set ADMIN_PASSWORD explicitly before production deployment."
            )
        admin_user = User(
            user_id=admin_user_id,
            username=admin_username,
            first_name="mohamad",
            last_name="El Ghoche",
            position=None,
            password=get_password_hash(bootstrap_password),
            role=None,
            user_type=auth.USER_TYPE_PLATFORM,
            platform_role=auth.PLATFORM_ROLE_OWNER,
            platform_owner_kind=auth.PLATFORM_OWNER_PRIMARY,
            access_scope=auth.ACCESS_SCOPE_GLOBAL,
            school_group_id=platform_storage_group_id,
            branch_id=None,
            academic_year_id=None,
            is_active=True
        )
        db.add(admin_user)
        db.commit()
    else:
        updated = False

        if admin_password and not auth.verify_password(admin_password, existing_user.password):
            existing_user.password = get_password_hash(admin_password)
            updated = True

        if not existing_user.username:
            existing_user.username = admin_username
            updated = True

        if existing_user.position is not None:
            existing_user.position = None
            updated = True

        if existing_user.role is not None:
            existing_user.role = None
            updated = True

        if auth.normalize_user_type(getattr(existing_user, "user_type", "")) != auth.USER_TYPE_PLATFORM:
            existing_user.user_type = auth.USER_TYPE_PLATFORM
            updated = True

        if auth.normalize_platform_role(getattr(existing_user, "platform_role", "")) != auth.PLATFORM_ROLE_OWNER:
            existing_user.platform_role = auth.PLATFORM_ROLE_OWNER
            updated = True

        if auth.normalize_platform_owner_kind(
            getattr(existing_user, "platform_owner_kind", "")
        ) != auth.PLATFORM_OWNER_PRIMARY:
            existing_user.platform_owner_kind = auth.PLATFORM_OWNER_PRIMARY
            updated = True

        if auth.normalize_access_scope(getattr(existing_user, "access_scope", "")) != auth.ACCESS_SCOPE_GLOBAL:
            existing_user.access_scope = auth.ACCESS_SCOPE_GLOBAL
            updated = True

        if existing_user.branch_id is not None:
            existing_user.branch_id = None
            updated = True

        if getattr(existing_user, "school_group_id", None) != platform_storage_group_id:
            existing_user.school_group_id = platform_storage_group_id
            updated = True

        if existing_user.academic_year_id is not None:
            existing_user.academic_year_id = None
            updated = True

        if not existing_user.is_active:
            existing_user.is_active = True
            updated = True

        if updated:
            db.commit()

    db.close()
