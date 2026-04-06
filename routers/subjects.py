from collections import Counter
from datetime import datetime
from io import BytesIO
import re

from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from fastapi.templating import Jinja2Templates
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import auth
import models
from dependencies import get_db
from auth import get_current_user
from ui_shell import build_shell_context
from year_copy import get_copy_year_choices, get_academic_year

router = APIRouter(prefix="/subjects", tags=["Subjects"])
templates = Jinja2Templates(directory="templates")
SUBJECT_CODE_PATTERN = re.compile(r"^[A-Z]{3}\d{3}$")
SUBJECT_NAME_PATTERN = re.compile(r"^[A-Za-z][A-Za-z\s'\-]*$")
EXPORT_GRADE_ORDER = list(range(0, 13))
EXPORT_GRADE_COLORS = {
    0: "FF0F766E",   # KG
    1: "FF1D4ED8",
    2: "FF2563EB",
    3: "FF3B82F6",
    4: "FF0891B2",
    5: "FF0EA5E9",
    6: "FF0D9488",
    7: "FF059669",
    8: "FF65A30D",
    9: "FFCA8A04",
    10: "FFD97706",
    11: "FFEA580C",
    12: "FFDC2626",
}


def _get_scope_ids(current_user):
    branch_id = getattr(current_user, "scope_branch_id", current_user.branch_id)
    academic_year_id = getattr(
        current_user,
        "scope_academic_year_id",
        current_user.academic_year_id
    )
    return branch_id, academic_year_id


def _normalize_spaces(value: str) -> str:
    return " ".join(value.split())


def _normalize_subject_code(value) -> str:
    if value is None:
        return ""
    return _normalize_spaces(str(value).strip()).upper().replace(" ", "")


def _normalize_subject_name(value) -> str:
    if value is None:
        return ""
    cleaned = _normalize_spaces(str(value).strip())
    if not cleaned:
        return ""
    return " ".join(part.capitalize() for part in cleaned.split(" "))


def _parse_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None

    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else None


def _grade_label(grade: int) -> str:
    if grade == 0:
        return "KG"
    return f"Grade {grade}"


def _grade_sheet_name(grade: int) -> str:
    if grade == 0:
        return "KG"
    return f"Grade {grade}"


def _apply_sheet_header_style(sheet, fill_color: str):
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(
        start_color=fill_color,
        end_color=fill_color,
        fill_type="solid",
    )
    thin_border = Border(
        left=Side(style="thin", color="D6DEEA"),
        right=Side(style="thin", color="D6DEEA"),
        top=Side(style="thin", color="D6DEEA"),
        bottom=Side(style="thin", color="D6DEEA"),
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    return thin_border


def _apply_alternating_rows(sheet, start_row: int, end_row: int, border: Border):
    stripe_fill = PatternFill(
        start_color="FFF6FAFF",
        end_color="FFF6FAFF",
        fill_type="solid",
    )
    for row_index in range(start_row, end_row + 1):
        for cell in sheet[row_index]:
            if row_index % 2 == 0:
                cell.fill = stripe_fill
            cell.border = border
            if cell.column in {3, 4}:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")


def _validate_subject_payload(subject_code, subject_name, weekly_hours, grade):
    errors = []
    if not subject_code:
        errors.append("Subject code is required.")
    elif not SUBJECT_CODE_PATTERN.match(subject_code):
        errors.append("Subject code must follow format AAA999 (3 letters + 3 digits).")

    if not subject_name:
        errors.append("Subject name is required.")
    elif not SUBJECT_NAME_PATTERN.match(subject_name):
        errors.append("Subject name should contain letters only and start with an uppercase letter.")

    if weekly_hours is None or weekly_hours <= 0:
        errors.append("Weekly hours must be a positive whole number.")

    if grade is None or grade < 0:
        errors.append("Grade must be KG (0) or a whole number from 1 to 12.")
    elif grade > 12:
        errors.append("Grade must be KG (0) or a whole number from 1 to 12.")

    return errors


def _get_existing_subject_codes_in_scope(db: Session, subject_codes, branch_id: int, academic_year_id: int):
    normalized_codes = sorted({code for code in subject_codes if code})
    if not normalized_codes:
        return []

    existing_codes = db.query(models.Subject.subject_code).filter(
        models.Subject.subject_code.in_(normalized_codes),
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id,
    ).all()
    return sorted({code for (code,) in existing_codes if code})


def _has_subject_assignment_in_scope(
    db: Session,
    subject_codes,
    branch_id: int,
    academic_year_id: int,
):
    normalized_codes = [code for code in subject_codes if code]
    if not normalized_codes:
        return False

    has_teacher_reference = db.query(models.Teacher).filter(
        models.Teacher.subject_code.in_(normalized_codes),
        models.Teacher.branch_id == branch_id,
        models.Teacher.academic_year_id == academic_year_id,
    ).first()
    if has_teacher_reference:
        return True

    has_allocation_reference = (
        db.query(models.TeacherSubjectAllocation)
        .join(models.Teacher, models.Teacher.id == models.TeacherSubjectAllocation.teacher_id)
        .filter(
            models.TeacherSubjectAllocation.subject_code.in_(normalized_codes),
            models.Teacher.branch_id == branch_id,
            models.Teacher.academic_year_id == academic_year_id,
        )
        .first()
    )
    if has_allocation_reference is not None:
        return True

    has_section_assignment_reference = (
        db.query(models.TeacherSectionAssignment)
        .join(
            models.PlanningSection,
            models.PlanningSection.id == models.TeacherSectionAssignment.planning_section_id,
        )
        .filter(
            models.TeacherSectionAssignment.subject_code.in_(normalized_codes),
            models.PlanningSection.branch_id == branch_id,
            models.PlanningSection.academic_year_id == academic_year_id,
        )
        .first()
    )
    return has_section_assignment_reference is not None


def _render_subjects_page(
    request: Request,
    db: Session,
    current_user,
    error: str = "",
    success: str = "",
    detail_errors=None
):
    branch_id, academic_year_id = _get_scope_ids(current_user)
    can_modify = auth.can_modify_data(current_user)
    can_edit = auth.can_edit_data(current_user)
    can_delete = auth.can_delete_data(current_user)
    subjects = db.query(models.Subject).filter(
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id
    ).order_by(models.Subject.id.desc()).all()
    copy_year_choices = get_copy_year_choices(db, academic_year_id)

    return templates.TemplateResponse(
        request,
        "subjects.html",
        {
            "request": request,
            "subjects": subjects,
            "user": current_user,
            "can_modify": can_modify,
            "can_edit": can_edit,
            "can_delete": can_delete,
            "error": error,
            "success": success,
            "detail_errors": detail_errors or [],
            "copy_year_choices": copy_year_choices,
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="subjects",
            ),
        }
    )


@router.get("/")
def subjects_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):

    if not current_user:
        return RedirectResponse(url="/")

    return _render_subjects_page(
        request=request,
        db=db,
        current_user=current_user
    )


@router.post("/copy-from-year")
def copy_subjects_from_year(
    request: Request,
    source_academic_year_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_modify_data(current_user):
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Your role has read-only access and cannot copy subjects.",
        )

    branch_id, target_academic_year_id = _get_scope_ids(current_user)
    if source_academic_year_id == target_academic_year_id:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Select a different academic year to copy subjects from.",
        )

    source_year = get_academic_year(db, source_academic_year_id)
    target_year = get_academic_year(db, target_academic_year_id)
    if not source_year or not target_year:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="The selected academic year was not found.",
        )

    source_subjects = (
        db.query(models.Subject)
        .filter(
            models.Subject.branch_id == branch_id,
            models.Subject.academic_year_id == source_academic_year_id,
        )
        .order_by(models.Subject.grade.asc(), models.Subject.subject_code.asc())
        .all()
    )
    if not source_subjects:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error=f"No subjects were found in {source_year.year_name} for the current branch.",
        )

    target_subject_codes = {
        subject_code
        for (subject_code,) in (
            db.query(models.Subject.subject_code)
            .filter(
                models.Subject.branch_id == branch_id,
                models.Subject.academic_year_id == target_academic_year_id,
            )
            .all()
        )
        if subject_code
    }

    copied_count = 0
    skipped_existing_count = 0
    for source_subject in source_subjects:
        if not source_subject.subject_code:
            continue
        if source_subject.subject_code in target_subject_codes:
            skipped_existing_count += 1
            continue

        db.add(
            models.Subject(
                subject_code=source_subject.subject_code,
                subject_name=source_subject.subject_name,
                weekly_hours=source_subject.weekly_hours,
                grade=source_subject.grade,
                branch_id=branch_id,
                academic_year_id=target_academic_year_id,
            )
        )
        target_subject_codes.add(source_subject.subject_code)
        copied_count += 1

    if copied_count:
        db.commit()

    success_parts = [
        f"Subjects copied from {source_year.year_name} to {target_year.year_name}: {copied_count} added.",
    ]
    if skipped_existing_count:
        success_parts.append(f"{skipped_existing_count} already existed and were skipped.")
    if not copied_count and not skipped_existing_count:
        success_parts.append("No subject rows were eligible to copy.")

    return _render_subjects_page(
        request=request,
        db=db,
        current_user=current_user,
        success=" ".join(success_parts),
    )


# --------------------------------------------------
# DOWNLOAD SUBJECT EXCEL TEMPLATE
# URL: GET /subjects/template
# --------------------------------------------------
@router.get("/template")
def download_subject_template(
    request: Request,
    db: Session = Depends(get_db)
):
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Subjects"

    headers = ["subject_code", "subject_name", "weekly_hours", "grade"]
    sheet.append(headers)
    sheet.append(["ENG101", "English", 4, 5])
    sheet.append(["MAT102", "Mathematics", 5, 6])

    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 12
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=subjects_template.xlsx"}
    )


# --------------------------------------------------
# EXPORT SUBJECTS TO EXCEL (BY GRADE SHEETS)
# URL: GET /subjects/export
# --------------------------------------------------
@router.get("/export")
def export_subjects_excel(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/")

    branch_id, academic_year_id = _get_scope_ids(current_user)
    subjects = db.query(models.Subject).filter(
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id,
    ).order_by(
        models.Subject.grade.asc(),
        models.Subject.subject_code.asc(),
    ).all()

    subjects_by_grade = {grade: [] for grade in EXPORT_GRADE_ORDER}
    for subject in subjects:
        if subject.grade is None:
            continue
        parsed_grade = int(subject.grade)
        if parsed_grade not in subjects_by_grade:
            continue
        subjects_by_grade[parsed_grade].append(subject)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_headers = [
        "Grade Level",
        "Subjects Count",
        "Total Weekly Hours",
        "Open Sheet",
    ]
    summary_sheet.append(summary_headers)
    summary_border = _apply_sheet_header_style(summary_sheet, "FF0F766E")

    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 16
    summary_sheet.column_dimensions["C"].width = 20
    summary_sheet.column_dimensions["D"].width = 14
    summary_sheet.freeze_panes = "A2"

    total_subjects_count = 0
    total_weekly_hours = 0
    summary_row_index = 2

    for grade in EXPORT_GRADE_ORDER:
        grade_label = _grade_label(grade)
        sheet_name = _grade_sheet_name(grade)
        grade_subjects = subjects_by_grade.get(grade, [])
        grade_subjects_count = len(grade_subjects)
        grade_hours_total = sum(
            int(subject.weekly_hours or 0)
            for subject in grade_subjects
        )

        total_subjects_count += grade_subjects_count
        total_weekly_hours += grade_hours_total

        grade_sheet = workbook.create_sheet(title=sheet_name)
        grade_sheet.append(
            ["Subject Code", "Subject Name", "Weekly Hours", "Grade Level"]
        )
        grade_border = _apply_sheet_header_style(
            grade_sheet,
            EXPORT_GRADE_COLORS.get(grade, "FF0F766E"),
        )
        grade_sheet.freeze_panes = "A2"
        grade_sheet.sheet_properties.tabColor = EXPORT_GRADE_COLORS.get(grade, "FF0F766E")

        if grade_subjects_count == 0:
            grade_sheet.append(["-", "No subjects found for this grade.", "-", grade_label])
            empty_cell = grade_sheet.cell(row=2, column=2)
            empty_cell.font = Font(italic=True, color="667085")
            for cell in grade_sheet[2]:
                cell.border = grade_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            for subject in grade_subjects:
                grade_sheet.append(
                    [
                        subject.subject_code or "",
                        subject.subject_name or "",
                        int(subject.weekly_hours or 0),
                        grade_label,
                    ]
                )
            _apply_alternating_rows(
                sheet=grade_sheet,
                start_row=2,
                end_row=grade_sheet.max_row,
                border=grade_border,
            )

        grade_sheet.column_dimensions["A"].width = 18
        grade_sheet.column_dimensions["B"].width = 34
        grade_sheet.column_dimensions["C"].width = 16
        grade_sheet.column_dimensions["D"].width = 14
        grade_sheet.auto_filter.ref = f"A1:D{grade_sheet.max_row}"

        summary_sheet.cell(row=summary_row_index, column=1, value=grade_label)
        summary_sheet.cell(row=summary_row_index, column=2, value=grade_subjects_count)
        summary_sheet.cell(row=summary_row_index, column=3, value=grade_hours_total)
        link_cell = summary_sheet.cell(row=summary_row_index, column=4, value="Open")
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color="2563EB", bold=True, underline="single")
        link_cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in range(1, 4):
            cell = summary_sheet.cell(row=summary_row_index, column=column)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = summary_border
        summary_sheet.cell(row=summary_row_index, column=4).border = summary_border
        summary_row_index += 1

    totals_row_index = summary_row_index
    summary_sheet.cell(row=totals_row_index, column=1, value="All Levels")
    summary_sheet.cell(row=totals_row_index, column=2, value=total_subjects_count)
    summary_sheet.cell(row=totals_row_index, column=3, value=total_weekly_hours)
    summary_sheet.cell(row=totals_row_index, column=4, value="-")

    totals_fill = PatternFill(
        start_color="FFE3F3EF",
        end_color="FFE3F3EF",
        fill_type="solid",
    )
    for cell in summary_sheet[totals_row_index]:
        cell.font = Font(bold=True, color="0B4F4A")
        cell.fill = totals_fill
        cell.border = summary_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_sheet.auto_filter.ref = f"A1:D{totals_row_index}"

    generated_stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=subjects_export_{generated_stamp}.xlsx"
            )
        },
    )


# --------------------------------------------------
# IMPORT SUBJECTS FROM EXCEL
# URL: POST /subjects/import
# --------------------------------------------------
@router.post("/import")
def import_subjects(
    request: Request,
    subject_file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    current_user = get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_modify_data(current_user):
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Your role has read-only access and cannot import subjects."
        )

    if not subject_file or not subject_file.filename:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Please choose an Excel file before importing."
        )

    if not subject_file.filename.lower().endswith(".xlsx"):
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Only .xlsx Excel files are supported."
        )

    try:
        workbook = load_workbook(subject_file.file, data_only=True)
    except Exception:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Unable to read the Excel file. Please use the template and try again."
        )

    sheet = workbook.active
    expected_headers = ["subject_code", "subject_name", "weekly_hours", "grade"]
    header_cells = [cell.value for cell in sheet[1]]
    actual_headers = []
    for value in header_cells[:4]:
        if value is None:
            actual_headers.append("")
        else:
            actual_headers.append(_normalize_spaces(str(value).strip().lower()))

    if actual_headers != expected_headers:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Invalid template format.",
            detail_errors=[
                "Expected first row headers: subject_code, subject_name, weekly_hours, grade."
            ]
        )

    branch_id, academic_year_id = _get_scope_ids(current_user)
    row_errors = []
    prepared_rows = []
    imported_codes = []

    for row_number, row_values in enumerate(
        sheet.iter_rows(min_row=2, max_col=4, values_only=True),
        start=2
    ):
        if not row_values:
            continue

        raw_code, raw_name, raw_weekly_hours, raw_grade = row_values
        if all(
            value is None or str(value).strip() == ""
            for value in [raw_code, raw_name, raw_weekly_hours, raw_grade]
        ):
            continue

        subject_code = _normalize_subject_code(raw_code)
        subject_name = _normalize_subject_name(raw_name)
        weekly_hours = _parse_int(raw_weekly_hours)
        grade = _parse_int(raw_grade)

        validation_errors = _validate_subject_payload(
            subject_code,
            subject_name,
            weekly_hours,
            grade
        )

        if validation_errors:
            row_errors.append(f"Row {row_number}: {' '.join(validation_errors)}")
            continue

        imported_codes.append(subject_code)
        prepared_rows.append(
            {
                "subject_code": subject_code,
                "subject_name": subject_name,
                "weekly_hours": weekly_hours,
                "grade": grade,
                "branch_id": branch_id,
                "academic_year_id": academic_year_id
            }
        )

    if not prepared_rows and not row_errors:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="No valid data rows found in the file."
        )

    duplicate_codes_in_file = [
        code for code, count in Counter(imported_codes).items() if count > 1
    ]
    if duplicate_codes_in_file:
        row_errors.append(
            "Duplicate subject codes inside Excel file: "
            + ", ".join(sorted(duplicate_codes_in_file))
        )

    if imported_codes:
        existing_code_set = _get_existing_subject_codes_in_scope(
            db,
            imported_codes,
            branch_id,
            academic_year_id,
        )
        if existing_code_set:
            row_errors.append(
                "These subject codes already exist in the current branch and academic year: "
                + ", ".join(existing_code_set)
            )

    if row_errors:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Import blocked. Please fix the file and try again.",
            detail_errors=row_errors
        )

    new_subjects = [
        models.Subject(**subject_data) for subject_data in prepared_rows
    ]

    try:
        db.add_all(new_subjects)
        db.commit()
    except IntegrityError:
        db.rollback()
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Import failed due to duplicate subject code conflict."
        )

    return _render_subjects_page(
        request=request,
        db=db,
        current_user=current_user,
        success=f"Import successful. {len(new_subjects)} subject(s) added."
    )


# --------------------------------------------------
# ADD SUBJECT
# URL: POST /subjects
# --------------------------------------------------
@router.post("/")
def add_subject(
    request: Request,
    subject_code: str = Form(...),
    subject_name: str = Form(...),
    weekly_hours: int = Form(...),
    grade: int = Form(...),
    db: Session = Depends(get_db),
):

    current_user = get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_modify_data(current_user):
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Your role has read-only access and cannot add subjects."
        )

    subject_code = _normalize_subject_code(subject_code)
    subject_name = _normalize_subject_name(subject_name)
    validation_errors = _validate_subject_payload(
        subject_code,
        subject_name,
        weekly_hours,
        grade
    )
    if validation_errors:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Unable to add subject. Please correct the data.",
            detail_errors=validation_errors
        )

    branch_id, academic_year_id = _get_scope_ids(current_user)

    existing_code = db.query(models.Subject).filter(
        models.Subject.subject_code == subject_code,
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id,
    ).first()
    if existing_code:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Subject code already exists in the current branch and academic year. Please use another code."
        )

    new_subject = models.Subject(
        subject_code=subject_code,
        subject_name=subject_name,
        weekly_hours=weekly_hours,
        grade=grade,
        branch_id=branch_id,
        academic_year_id=academic_year_id
    )

    try:
        db.add(new_subject)
        db.commit()
    except IntegrityError:
        db.rollback()

        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Duplicate subject code is not allowed."
        )

    return RedirectResponse(url="/subjects", status_code=302)


# --------------------------------------------------
# EDIT SUBJECT (GET)
# URL: /subjects/edit/{id}
# --------------------------------------------------
@router.get("/edit/{subject_id}")
def edit_subject_page(
    request: Request,
    subject_id: int,
    db: Session = Depends(get_db),
):

    current_user = get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_edit_data(current_user):
        return RedirectResponse(url="/subjects", status_code=302)

    branch_id, academic_year_id = _get_scope_ids(current_user)
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id
    ).first()

    if not subject:
        return RedirectResponse(url="/subjects")

    return templates.TemplateResponse(
        request,
        "edit_subject.html",
        {
            "request": request,
            "subject": subject,
            "user": current_user,
            "error": "",
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="subjects",
                title="Edit Subject",
                eyebrow="Academic Catalog",
                intro="Refine subject details while keeping branch and year context visible in the shared workspace.",
                icon="subjects",
            ),
        }
    )


# --------------------------------------------------
# UPDATE SUBJECT (POST)
# --------------------------------------------------
@router.post("/edit/{subject_id}")
def update_subject(
    request: Request,
    subject_id: int,
    subject_code: str = Form(...),
    subject_name: str = Form(...),
    weekly_hours: int = Form(...),
    grade: int = Form(...),
    db: Session = Depends(get_db),
):

    current_user = get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_edit_data(current_user):
        return RedirectResponse(url="/subjects", status_code=302)

    branch_id, academic_year_id = _get_scope_ids(current_user)
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id
    ).first()

    if not subject:
        return RedirectResponse(url="/subjects")

    subject_code = _normalize_subject_code(subject_code)
    subject_name = _normalize_subject_name(subject_name)
    validation_errors = _validate_subject_payload(
        subject_code,
        subject_name,
        weekly_hours,
        grade
    )
    if validation_errors:
        return templates.TemplateResponse(
            request,
            "edit_subject.html",
            {
                "request": request,
                "subject": subject,
                "user": current_user,
                "error": " ".join(validation_errors),
                **build_shell_context(
                    request,
                    db,
                    current_user,
                    page_key="subjects",
                    title="Edit Subject",
                    eyebrow="Academic Catalog",
                    intro="Refine subject details while keeping branch and year context visible in the shared workspace.",
                    icon="subjects",
                ),
            },
            status_code=400
        )

    existing_code = db.query(models.Subject).filter(
        models.Subject.subject_code == subject_code,
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id,
        models.Subject.id != subject.id
    ).first()
    if existing_code:
        return templates.TemplateResponse(
            request,
            "edit_subject.html",
            {
                "request": request,
                "subject": subject,
                "user": current_user,
                "error": "Subject code already exists in the current branch and academic year. Please use another code.",
                **build_shell_context(
                    request,
                    db,
                    current_user,
                    page_key="subjects",
                    title="Edit Subject",
                    eyebrow="Academic Catalog",
                    intro="Refine subject details while keeping branch and year context visible in the shared workspace.",
                    icon="subjects",
                ),
            },
            status_code=400
        )

    previous_subject_code = subject.subject_code
    subject.subject_code = subject_code
    subject.subject_name = subject_name
    subject.weekly_hours = weekly_hours
    subject.grade = grade

    try:
        if previous_subject_code and previous_subject_code != subject_code:
            teacher_ids_in_scope = [
                teacher_id
                for (teacher_id,) in db.query(models.Teacher.id).filter(
                    models.Teacher.branch_id == branch_id,
                    models.Teacher.academic_year_id == academic_year_id,
                ).all()
            ]
            if teacher_ids_in_scope:
                db.query(models.Teacher).filter(
                    models.Teacher.id.in_(teacher_ids_in_scope),
                    models.Teacher.subject_code == previous_subject_code,
                ).update(
                    {models.Teacher.subject_code: subject_code},
                    synchronize_session=False,
                )
                db.query(models.TeacherSubjectAllocation).filter(
                    models.TeacherSubjectAllocation.teacher_id.in_(teacher_ids_in_scope),
                    models.TeacherSubjectAllocation.subject_code == previous_subject_code,
                ).update(
                    {models.TeacherSubjectAllocation.subject_code: subject_code},
                    synchronize_session=False,
                )

            planning_section_ids_in_scope = [
                planning_section_id
                for (planning_section_id,) in db.query(models.PlanningSection.id).filter(
                    models.PlanningSection.branch_id == branch_id,
                    models.PlanningSection.academic_year_id == academic_year_id,
                ).all()
            ]
            if planning_section_ids_in_scope:
                db.query(models.TeacherSectionAssignment).filter(
                    models.TeacherSectionAssignment.planning_section_id.in_(planning_section_ids_in_scope),
                    models.TeacherSectionAssignment.subject_code == previous_subject_code,
                ).update(
                    {models.TeacherSectionAssignment.subject_code: subject_code},
                    synchronize_session=False,
                )
        db.commit()
    except IntegrityError:
        db.rollback()
        return RedirectResponse(url="/subjects", status_code=302)

    return RedirectResponse(url="/subjects", status_code=302)


# --------------------------------------------------
# DELETE SUBJECT
# --------------------------------------------------
@router.get("/delete/{subject_id}")
def delete_subject(
    request: Request,
    subject_id: int,
    db: Session = Depends(get_db),
):

    current_user = get_current_user(request, db)

    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_delete_data(current_user):
        return RedirectResponse(url="/subjects")

    branch_id, academic_year_id = _get_scope_ids(current_user)
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id
    ).first()

    if subject:
        if _has_subject_assignment_in_scope(
            db,
            [subject.subject_code],
            branch_id,
            academic_year_id,
        ):
            return _render_subjects_page(
                request=request,
                db=db,
                current_user=current_user,
                error="Cannot delete this subject because it is assigned to one or more teachers or planning sections.",
            )

        try:
            db.delete(subject)
            db.commit()
        except IntegrityError:
            db.rollback()
            return _render_subjects_page(
                request=request,
                db=db,
                current_user=current_user,
                error="Unable to delete subject due to related records.",
            )

    return RedirectResponse(url="/subjects", status_code=302)


@router.post("/delete-bulk")
def delete_subjects_bulk(
    request: Request,
    selected_subject_ids: list[int] = Form([]),
    db: Session = Depends(get_db),
):
    current_user = get_current_user(request, db)
    if not current_user:
        return RedirectResponse(url="/")

    if not auth.can_delete_data(current_user):
        return RedirectResponse(url="/subjects", status_code=302)

    unique_subject_ids = sorted({
        int(subject_id)
        for subject_id in selected_subject_ids
        if subject_id
    })
    if not unique_subject_ids:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Select at least one subject to delete.",
        )

    branch_id, academic_year_id = _get_scope_ids(current_user)
    subject_rows = db.query(models.Subject).filter(
        models.Subject.id.in_(unique_subject_ids),
        models.Subject.branch_id == branch_id,
        models.Subject.academic_year_id == academic_year_id,
    ).all()
    subject_map = {subject.id: subject for subject in subject_rows}
    missing_ids = [
        subject_id for subject_id in unique_subject_ids
        if subject_id not in subject_map
    ]
    if missing_ids:
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="One or more selected subjects were not found in your current scope.",
        )

    selected_subject_codes = [
        subject.subject_code
        for subject in subject_rows
        if subject.subject_code
    ]
    if _has_subject_assignment_in_scope(
        db,
        selected_subject_codes,
        branch_id,
        academic_year_id,
    ):
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="One or more selected subjects cannot be deleted because they are assigned to teachers or planning sections.",
        )

    try:
        db.query(models.Subject).filter(
            models.Subject.id.in_(unique_subject_ids),
            models.Subject.branch_id == branch_id,
            models.Subject.academic_year_id == academic_year_id,
        ).delete(synchronize_session=False)
        db.commit()
    except IntegrityError:
        db.rollback()
        return _render_subjects_page(
            request=request,
            db=db,
            current_user=current_user,
            error="Bulk delete failed due to related records.",
        )

    deleted_count = len(unique_subject_ids)
    success_message = (
        "Subject deleted successfully."
        if deleted_count == 1
        else f"{deleted_count} subjects deleted successfully."
    )

    return _render_subjects_page(
        request=request,
        db=db,
        current_user=current_user,
        success=success_message,
    )
