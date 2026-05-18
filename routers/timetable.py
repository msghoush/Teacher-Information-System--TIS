from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, Request
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi.templating import Jinja2Templates
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

import auth
import models
from dependencies import get_db
from timetable_logic import (
    build_timetable_workspace_payload,
    get_scope_ids,
    normalize_day_key,
)
from ui_shell import build_shell_context


router = APIRouter(prefix="/timetable", tags=["Timetable"])
templates = Jinja2Templates(directory="templates")
BASE_DIR = Path(__file__).resolve().parent.parent
EXPORT_LOGO_ASSETS = (
    {
        "path": BASE_DIR / "static" / "images" / "TIS_Logo_Adjusted.png",
        "anchor": "A1",
        "width": 108,
        "height": 54,
        "fallback": "Teacher Information System",
    },
    {
        "path": BASE_DIR / "static" / "images" / "andalus-logo.png",
        "anchor": "D1",
        "width": 112,
        "height": 54,
        "fallback": "Little Andalus International Schools",
    },
    {
        "path": BASE_DIR / "static" / "images" / "cognia-logo.png",
        "anchor": "G1",
        "width": 108,
        "height": 45,
        "fallback": "Cognia",
    },
    {
        "path": BASE_DIR / "static" / "images" / "andalus-logo-main.png",
        "anchor": "J1",
        "width": 132,
        "height": 52,
        "fallback": "Andalus International Schools",
    },
)
EXCEL_BRAND_BLUE = "0A4EA3"
EXCEL_TEXT = "17365D"
EXCEL_MUTED = "60728C"
EXCEL_BORDER = "D8E5F4"
EXCEL_SOFT = "F4F8FF"


def _parse_int(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _get_current_user_or_redirect(request: Request, db: Session):
    current_user = auth.get_current_user(request, db)
    if not current_user:
        return None, RedirectResponse(url="/", status_code=302)
    return current_user, None


def _find_section_by_id(workspace_payload: dict, section_id: int):
    return next(
        (
            section
            for section in workspace_payload.get("sections", [])
            if int(section.get("id") or 0) == int(section_id or 0)
        ),
        None,
    )


def _find_teacher_by_id(workspace_payload: dict, teacher_id: int):
    return next(
        (
            teacher
            for teacher in workspace_payload.get("teachers", [])
            if int(teacher.get("id") or 0) == int(teacher_id or 0)
        ),
        None,
    )


def _find_entry_by_section_slot(
    workspace_payload: dict,
    *,
    section_id: int,
    day_key: str,
    period_index: int,
):
    return next(
        (
            entry
            for entry in workspace_payload.get("entries", [])
            if int(entry.get("section_id") or 0) == int(section_id or 0)
            and str(entry.get("day_key") or "") == day_key
            and int(entry.get("period_index") or 0) == int(period_index or 0)
        ),
        None,
    )


def _find_section_option(
    workspace_payload: dict,
    *,
    section_id: int,
    subject_code: str,
):
    section_payload = _find_section_by_id(workspace_payload, section_id)
    if not section_payload:
        return None
    return next(
        (
            option
            for option in section_payload.get("options", [])
            if str(option.get("subject_code") or "").strip().upper() == subject_code
        ),
        None,
    )


def _find_teacher_conflict(
    workspace_payload: dict,
    *,
    teacher_id: int,
    day_key: str,
    period_index: int,
    ignore_entry_id: int | None = None,
):
    return next(
        (
            entry
            for entry in workspace_payload.get("entries", [])
            if int(entry.get("teacher_id") or 0) == int(teacher_id or 0)
            and str(entry.get("day_key") or "") == day_key
            and int(entry.get("period_index") or 0) == int(period_index or 0)
            and int(entry.get("id") or 0) != int(ignore_entry_id or 0)
        ),
        None,
    )


def _json_error(message: str, status_code: int = 400):
    return JSONResponse(
        status_code=status_code,
        content={
            "ok": False,
            "message": message,
        },
    )


def _json_success(workspace_payload: dict, *, message: str = ""):
    return JSONResponse(
        content={
            "ok": True,
            "message": message,
            "payload": workspace_payload,
        }
    )


def _sanitize_filename_part(value: Any, fallback: str = "scope") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", str(value or "").strip())
    return cleaned.strip("-").lower() or fallback


def _get_scope_labels(db: Session, branch_id: int, academic_year_id: int) -> tuple[str, str]:
    branch = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    academic_year = (
        db.query(models.AcademicYear)
        .filter(models.AcademicYear.id == academic_year_id)
        .first()
    )
    return (
        str(getattr(branch, "name", "") or f"Branch {branch_id}").strip(),
        str(getattr(academic_year, "year_name", "") or f"Year {academic_year_id}").strip(),
    )


def _build_timetable_export_filename(
    branch_name: str,
    academic_year_name: str,
    extension: str,
) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_branch = _sanitize_filename_part(branch_name, "branch")
    safe_year = _sanitize_filename_part(academic_year_name, "year")
    safe_extension = str(extension or "xlsx").strip().lower().lstrip(".") or "xlsx"
    return f"tis_timetable_{safe_branch}_{safe_year}_{timestamp}.{safe_extension}"


def _safe_excel_hex(value: Any, fallback: str = EXCEL_SOFT) -> str:
    cleaned = str(value or "").strip().lstrip("#").upper()
    if re.fullmatch(r"[0-9A-F]{6}", cleaned):
        return cleaned
    fallback_cleaned = str(fallback or EXCEL_SOFT).strip().lstrip("#").upper()
    return fallback_cleaned if re.fullmatch(r"[0-9A-F]{6}", fallback_cleaned) else EXCEL_SOFT


def _add_excel_logo_strip(sheet):
    sheet.row_dimensions[1].height = 44
    for asset in EXPORT_LOGO_ASSETS:
        anchor = asset["anchor"]
        try:
            image_path = asset["path"]
            if not image_path.exists():
                raise FileNotFoundError(str(image_path))
            logo = ExcelImage(str(image_path))
            logo.width = asset["width"]
            logo.height = asset["height"]
            sheet.add_image(logo, anchor)
        except Exception:
            fallback_cell = sheet[anchor]
            fallback_cell.value = asset["fallback"]
            fallback_cell.font = Font(bold=True, color=EXCEL_TEXT, size=9)
            fallback_cell.alignment = Alignment(horizontal="center", vertical="center")


def _merge_sheet_title(sheet, row: int, title: str, subtitle: str, total_columns: int):
    last_column = get_column_letter(max(total_columns, 1))
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
    title_cell = sheet.cell(row=row, column=1, value=title)
    title_cell.font = Font(bold=True, size=18, color=EXCEL_BRAND_BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=total_columns)
    subtitle_cell = sheet.cell(row=row + 1, column=1, value=subtitle)
    subtitle_cell.font = Font(size=10, color=EXCEL_MUTED)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[row].height = 24
    sheet.row_dimensions[row + 1].height = 18
    sheet.auto_filter.ref = f"A{row + 4}:{last_column}{row + 4}"


def _prepare_excel_sheet(
    sheet,
    *,
    title: str,
    subtitle: str,
    total_columns: int,
    include_logos: bool = False,
) -> int:
    sheet.sheet_view.showGridLines = False
    if include_logos:
        _add_excel_logo_strip(sheet)
        title_row = 4
    else:
        title_row = 1
    _merge_sheet_title(sheet, title_row, title, subtitle, total_columns)
    return title_row + 4


def _style_excel_header_row(sheet, row_index: int, total_columns: int):
    header_fill = PatternFill(start_color=EXCEL_BRAND_BLUE, end_color=EXCEL_BRAND_BLUE, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color=EXCEL_BORDER),
        right=Side(style="thin", color=EXCEL_BORDER),
        top=Side(style="thin", color=EXCEL_BORDER),
        bottom=Side(style="thin", color=EXCEL_BORDER),
    )
    for column_index in range(1, total_columns + 1):
        cell = sheet.cell(row=row_index, column=column_index)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def _style_excel_table_area(sheet, start_row: int, end_row: int, total_columns: int):
    thin_border = Border(
        left=Side(style="thin", color=EXCEL_BORDER),
        right=Side(style="thin", color=EXCEL_BORDER),
        top=Side(style="thin", color=EXCEL_BORDER),
        bottom=Side(style="thin", color=EXCEL_BORDER),
    )
    for row_index in range(start_row, end_row + 1):
        for column_index in range(1, total_columns + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_entry_lookup(workspace_payload: dict, entity_key: str) -> dict[tuple[int, str, int], dict]:
    lookup = {}
    for entry in workspace_payload.get("entries", []):
        entity_id = _parse_int(entry.get(entity_key))
        period_index = _parse_int(entry.get("period_index"))
        day_key = str(entry.get("day_key") or "").strip()
        if entity_id is None or period_index is None or not day_key:
            continue
        lookup[(entity_id, day_key, period_index)] = entry
    return lookup


def _build_blocked_slot_lookup(workspace_payload: dict) -> dict[tuple[str, int], dict]:
    lookup = {}
    for slot in workspace_payload.get("blocked_slots", []):
        period_index = _parse_int(slot.get("period_index"))
        day_key = str(slot.get("day_key") or "").strip()
        if period_index is None or not day_key:
            continue
        lookup[(day_key, period_index)] = slot
    return lookup


def _format_timetable_entry(entry: dict | None, view: str) -> str:
    if not entry:
        return ""
    subject_label = " - ".join(
        part
        for part in [
            str(entry.get("subject_code") or "").strip(),
            str(entry.get("subject_name") or "").strip(),
        ]
        if part
    )
    secondary = (
        str(entry.get("teacher_name") or "").strip()
        if view == "section"
        else str(entry.get("section_label") or "").strip()
    )
    status = str(entry.get("status") or "").strip()
    lines = [subject_label or "Scheduled Lesson"]
    if secondary:
        lines.append(secondary)
    if status == "stale":
        lines.append("Needs review")
    return "\n".join(lines)


def _write_overview_sheet(
    workbook: Workbook,
    workspace_payload: dict,
    branch_name: str,
    academic_year_name: str,
    generated_at: datetime,
):
    sheet = workbook.active
    sheet.title = "Overview"
    row = _prepare_excel_sheet(
        sheet,
        title="Weekly Timetable Export",
        subtitle=(
            f"{branch_name} | Academic Year {academic_year_name} | "
            f"Generated {generated_at.strftime('%Y-%m-%d %H:%M')}"
        ),
        total_columns=8,
        include_logos=True,
    )
    summary = workspace_payload.get("summary", {})
    metrics = [
        ("Sections", summary.get("section_count", 0), "Planning sections"),
        ("Teachers", summary.get("teacher_count", 0), "Active teachers"),
        ("Required", f"{summary.get('required_hours', 0)}h", "Total required periods"),
        ("Scheduled", f"{summary.get('scheduled_hours', 0)}h", "Placed periods"),
        ("Remaining", f"{summary.get('remaining_hours', 0)}h", "Still unplaced"),
        ("Teaching Slots", summary.get("teaching_slot_count", 0), "Configured grid slots"),
        ("Blocked Slots", summary.get("blocked_slot_count", 0), "Non-teaching slots"),
        ("Working Days", len(workspace_payload.get("days", [])), "Days in week"),
    ]
    for index, (label, value, note) in enumerate(metrics):
        column = 1 + (index % 4) * 2
        metric_row = row + (index // 4) * 3
        sheet.merge_cells(start_row=metric_row, start_column=column, end_row=metric_row, end_column=column + 1)
        sheet.merge_cells(start_row=metric_row + 1, start_column=column, end_row=metric_row + 1, end_column=column + 1)
        sheet.merge_cells(start_row=metric_row + 2, start_column=column, end_row=metric_row + 2, end_column=column + 1)
        sheet.cell(metric_row, column, label).font = Font(bold=True, color=EXCEL_MUTED, size=9)
        sheet.cell(metric_row + 1, column, value).font = Font(bold=True, color=EXCEL_BRAND_BLUE, size=16)
        sheet.cell(metric_row + 2, column, note).font = Font(color=EXCEL_MUTED, size=8)
        for offset in range(3):
            cell = sheet.cell(metric_row + offset, column)
            cell.fill = PatternFill(start_color=EXCEL_SOFT, end_color=EXCEL_SOFT, fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
    row += 8

    warnings = workspace_payload.get("warnings", [])
    if warnings:
        sheet.cell(row=row, column=1, value="Warnings").font = Font(bold=True, color="9F2D1F", size=12)
        row += 1
        for warning in warnings[:5]:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            warning_cell = sheet.cell(row=row, column=1, value=warning)
            warning_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
            warning_cell.font = Font(color="8A4F00")
            warning_cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    sheet.cell(row=row, column=1, value="Section Requirement Status").font = Font(
        bold=True,
        color=EXCEL_BRAND_BLUE,
        size=12,
    )
    row += 1
    headers = ["Section", "Grade", "Class Status", "Subjects", "Required", "Scheduled", "Remaining", "Teacher Gaps"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=column_index, value=header)
    _style_excel_header_row(sheet, row, len(headers))
    table_start = row
    for section in workspace_payload.get("sections", []):
        row += 1
        values = [
            section.get("section_label", ""),
            section.get("grade_label", ""),
            section.get("class_status", ""),
            section.get("subject_count", 0),
            section.get("total_required_hours", 0),
            section.get("scheduled_hours", 0),
            section.get("remaining_hours", 0),
            section.get("missing_teacher_subjects", 0),
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column_index, value=value)
        if int(section.get("remaining_hours") or 0) == 0:
            sheet.cell(row=row, column=7).fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        else:
            sheet.cell(row=row, column=7).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    _style_excel_table_area(sheet, table_start, row, len(headers))

    for column_key, width in {
        "A": 28,
        "B": 12,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 14,
    }.items():
        sheet.column_dimensions[column_key].width = width
    sheet.freeze_panes = "A8"


def _write_entity_timetable_sheet(
    workbook: Workbook,
    workspace_payload: dict,
    *,
    title: str,
    sheet_name: str,
    entity_kind: str,
):
    days = workspace_payload.get("days", [])
    time_slots = workspace_payload.get("time_slots", [])
    entity_key = "section_id" if entity_kind == "section" else "teacher_id"
    entities = (
        workspace_payload.get("sections", [])
        if entity_kind == "section"
        else workspace_payload.get("teachers", [])
    )
    entry_lookup = _build_entry_lookup(workspace_payload, entity_key)
    blocked_lookup = _build_blocked_slot_lookup(workspace_payload)
    total_columns = 2 + len(days)
    sheet = workbook.create_sheet(sheet_name)
    row = _prepare_excel_sheet(
        sheet,
        title=title,
        subtitle="Each cell shows the saved subject placement from the live timetable workspace.",
        total_columns=total_columns,
    )
    for column_index in range(1, total_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 16 if column_index <= 2 else 24

    if not entities:
        sheet.cell(row=row, column=1, value=f"No {sheet_name.lower()} are available.")
        return

    for entity in entities:
        entity_id = int(entity.get("id") or 0)
        label = entity.get("section_label") if entity_kind == "section" else entity.get("teacher_name")
        label = str(label or f"{entity_kind.title()} #{entity_id}").strip()
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
        title_cell = sheet.cell(row=row, column=1, value=label)
        title_cell.fill = PatternFill(start_color=EXCEL_TEXT, end_color=EXCEL_TEXT, fill_type="solid")
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[row].height = 22
        row += 1

        if entity_kind == "section":
            meta = (
                f"Class Status: {entity.get('class_status', '')} | "
                f"Scheduled {entity.get('scheduled_hours', 0)}/{entity.get('total_required_hours', 0)}h | "
                f"Remaining {entity.get('remaining_hours', 0)}h"
            )
        else:
            meta = (
                f"Teacher ID: {entity.get('teacher_id', '')} | "
                f"Scheduled {entity.get('scheduled_hours', 0)}/{entity.get('required_hours', 0)}h | "
                f"Capacity {entity.get('capacity_hours', 0)}h"
            )
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
        meta_cell = sheet.cell(row=row, column=1, value=meta)
        meta_cell.fill = PatternFill(start_color=EXCEL_SOFT, end_color=EXCEL_SOFT, fill_type="solid")
        meta_cell.font = Font(color=EXCEL_MUTED, size=9)
        row += 1

        headers = ["Period", "Time"] + [day.get("label", "") for day in days]
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row, column=column_index, value=header)
        _style_excel_header_row(sheet, row, total_columns)
        table_start = row
        row += 1

        for slot in time_slots:
            period_index = int(slot.get("period_index") or 0)
            sheet.cell(row=row, column=1, value=slot.get("short_label") or f"P{period_index}")
            sheet.cell(row=row, column=2, value=slot.get("time_range") or "")
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for day_offset, day in enumerate(days, start=3):
                day_key = str(day.get("key") or "")
                blocked_slot = blocked_lookup.get((day_key, period_index))
                cell = sheet.cell(row=row, column=day_offset)
                if blocked_slot:
                    cell.value = f"{blocked_slot.get('label', 'Blocked')}\n{blocked_slot.get('block_type_label', '')}"
                    cell.fill = PatternFill(start_color="F3F6F9", end_color="F3F6F9", fill_type="solid")
                    cell.font = Font(color="475569")
                else:
                    entry = entry_lookup.get((entity_id, day_key, period_index))
                    cell.value = _format_timetable_entry(entry, entity_kind)
                    if entry:
                        cell.fill = PatternFill(
                            start_color=_safe_excel_hex(entry.get("subject_color_soft"), "F4F8FF"),
                            end_color=_safe_excel_hex(entry.get("subject_color_soft"), "F4F8FF"),
                            fill_type="solid",
                        )
                        cell.font = Font(
                            color=_safe_excel_hex(entry.get("subject_color_text"), EXCEL_TEXT),
                            bold=True,
                            size=9,
                        )
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        cell.font = Font(color=EXCEL_MUTED, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.row_dimensions[row].height = 42
            row += 1
        _style_excel_table_area(sheet, table_start, row - 1, total_columns)
        row += 2
    sheet.freeze_panes = "A5"


def _write_subject_remaining_sheet(workbook: Workbook, workspace_payload: dict):
    sheet = workbook.create_sheet("Subject Remaining")
    headers = [
        "Section",
        "Grade",
        "Subject Code",
        "Subject Name",
        "Teacher",
        "Weekly Hours",
        "Scheduled",
        "Remaining",
        "Status",
    ]
    row = _prepare_excel_sheet(
        sheet,
        title="Subject Remaining Hours",
        subtitle="Per-section subject requirements after the latest timetable assignments.",
        total_columns=len(headers),
    )
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=column_index, value=header)
    _style_excel_header_row(sheet, row, len(headers))
    table_start = row
    for section in workspace_payload.get("sections", []):
        for option in section.get("options", []):
            row += 1
            remaining_hours = int(option.get("remaining_hours") or 0)
            status = (
                "Complete"
                if remaining_hours == 0
                else "Teacher missing"
                if not option.get("is_schedulable")
                else "Needs scheduling"
            )
            values = [
                section.get("section_label", ""),
                section.get("grade_label", ""),
                option.get("subject_code", ""),
                option.get("subject_name", ""),
                option.get("teacher_name", "") or "Not assigned",
                option.get("weekly_hours", 0),
                option.get("scheduled_count", 0),
                remaining_hours,
                status,
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row, column=column_index, value=value)
            fill_color = "DCFCE7" if status == "Complete" else "FEE2E2" if status == "Teacher missing" else "FFF4E6"
            sheet.cell(row=row, column=8).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            sheet.cell(row=row, column=9).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    _style_excel_table_area(sheet, table_start, row, len(headers))
    for column_index, width in enumerate([24, 10, 14, 30, 28, 14, 12, 12, 18], start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A{table_start}:{get_column_letter(len(headers))}{row}"


def _build_timetable_xlsx_bytes(
    workspace_payload: dict,
    branch_name: str,
    academic_year_name: str,
) -> bytes:
    generated_at = datetime.now()
    workbook = Workbook()
    _write_overview_sheet(workbook, workspace_payload, branch_name, academic_year_name, generated_at)
    _write_entity_timetable_sheet(
        workbook,
        workspace_payload,
        title="Timetable By Section",
        sheet_name="By Section",
        entity_kind="section",
    )
    _write_entity_timetable_sheet(
        workbook,
        workspace_payload,
        title="Timetable By Teacher",
        sheet_name="By Teacher",
        entity_kind="teacher",
    )
    _write_subject_remaining_sheet(workbook, workspace_payload)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _pdf_escape_text(value: Any) -> str:
    text = str(value if value is not None else "")
    text = text.replace("\u2022", "-").replace("\u00d7", "x")
    text = text.encode("latin-1", "replace").decode("latin-1")
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_rgb(color_value: Any, fallback: str = "#0A4EA3") -> tuple[float, float, float]:
    cleaned = str(color_value or fallback).strip().lstrip("#")
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", cleaned):
        cleaned = str(fallback).strip().lstrip("#")
    return (
        int(cleaned[0:2], 16) / 255,
        int(cleaned[2:4], 16) / 255,
        int(cleaned[4:6], 16) / 255,
    )


class _TimetablePdf:
    width = 841.89
    height = 595.28
    margin = 30

    def __init__(self, title: str, subtitle: str):
        self.title = title
        self.subtitle = subtitle
        self.pages: list[list[str]] = []
        self.y = self.height - self.margin
        self.add_page()

    def _current(self) -> list[str]:
        return self.pages[-1]

    def _text_command(
        self,
        x: float,
        y: float,
        value: Any,
        *,
        size: float = 8,
        color: str = "#17365D",
        bold: bool = False,
    ) -> str:
        r, g, b = _pdf_rgb(color, "#17365D")
        font = "/F2" if bold else "/F1"
        return (
            f"{r:.4f} {g:.4f} {b:.4f} rg\n"
            f"BT {font} {size:.2f} Tf {x:.2f} {y:.2f} Td ({_pdf_escape_text(value)}) Tj ET\n"
        )

    def add_page(self):
        self.pages.append([])
        self.y = self.height - self.margin
        self.text(self.margin, self.y, self.title, size=14, color="#0A4EA3", bold=True)
        self.text(self.width - self.margin - 170, self.y, "TIS | Andalus | Cognia", size=8, color="#60728C", bold=True)
        self.y -= 16
        self.text(self.margin, self.y, self.subtitle, size=8, color="#60728C")
        self.y -= 14
        self.line(self.margin, self.y, self.width - self.margin, self.y, "#CAD9EA")
        self.y -= 16

    def text(self, x: float, y: float, value: Any, *, size: float = 8, color: str = "#17365D", bold: bool = False):
        self._current().append(self._text_command(x, y, value, size=size, color=color, bold=bold))

    def rect(self, x: float, y: float, width: float, height: float, color: str):
        r, g, b = _pdf_rgb(color)
        self._current().append(
            f"{r:.4f} {g:.4f} {b:.4f} rg\n{x:.2f} {y:.2f} {width:.2f} {height:.2f} re f\n"
        )

    def line(self, x1: float, y1: float, x2: float, y2: float, color: str = "#CAD9EA", width: float = 0.7):
        r, g, b = _pdf_rgb(color, "#CAD9EA")
        self._current().append(
            f"{r:.4f} {g:.4f} {b:.4f} RG\n{width:.2f} w\n{x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S\n"
        )

    def paragraph(self, value: str, *, width: float = 720, size: float = 8, color: str = "#60728C"):
        words = str(value or "").split()
        line = ""
        max_chars = max(20, int(width / (size * 0.48)))
        for word in words:
            candidate = f"{line} {word}".strip()
            if len(candidate) > max_chars and line:
                self.text(self.margin, self.y, line, size=size, color=color)
                self.y -= 11
                line = word
            else:
                line = candidate
        if line:
            self.text(self.margin, self.y, line, size=size, color=color)
            self.y -= 14

    def build(self) -> bytes:
        for index, page in enumerate(self.pages, start=1):
            page.append(
                self._text_command(
                    self.margin,
                    18,
                    f"Generated by TIS | Page {index}",
                    size=7,
                    color="#60728C",
                )
            )

        content_streams = ["".join(page).encode("latin-1", "replace") for page in self.pages]
        kids = []
        objects: list[bytes] = []
        objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
        objects.append(b"")
        objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
        objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
        next_object_number = 5
        page_objects = []
        for content in content_streams:
            page_number = next_object_number
            content_number = next_object_number + 1
            kids.append(f"{page_number} 0 R")
            page_objects.append(
                (
                    page_number,
                    (
                        f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {self.width:.2f} {self.height:.2f}] "
                        f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
                        f"/Contents {content_number} 0 R >>"
                    ).encode("latin-1"),
                )
            )
            page_objects.append(
                (
                    content_number,
                    b"<< /Length "
                    + str(len(content)).encode("ascii")
                    + b" >>\nstream\n"
                    + content
                    + b"\nendstream",
                )
            )
            next_object_number += 2

        objects[1] = f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(kids)} >>".encode("latin-1")
        for _, obj in page_objects:
            objects.append(obj)

        pdf = b"%PDF-1.4\n"
        offsets = [0]
        for object_index, obj in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf += f"{object_index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n"
        xref_offset = len(pdf)
        pdf += f"xref\n0 {len(objects) + 1}\n".encode("ascii")
        pdf += b"0000000000 65535 f \n"
        for offset in offsets[1:]:
            pdf += f"{offset:010d} 00000 n \n".encode("ascii")
        pdf += (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
        ).encode("ascii")
        return pdf


def _pdf_truncate(value: Any, max_chars: int) -> str:
    text = " ".join(str(value or "").replace("\n", " ").split())
    if len(text) <= max_chars:
        return text
    return text[: max(max_chars - 1, 1)].rstrip() + "..."


def _draw_pdf_metric_cards(pdf: _TimetablePdf, workspace_payload: dict):
    summary = workspace_payload.get("summary", {})
    metrics = [
        ("Sections", summary.get("section_count", 0)),
        ("Teachers", summary.get("teacher_count", 0)),
        ("Required", f"{summary.get('required_hours', 0)}h"),
        ("Scheduled", f"{summary.get('scheduled_hours', 0)}h"),
        ("Remaining", f"{summary.get('remaining_hours', 0)}h"),
    ]
    card_width = 145
    card_height = 42
    y = pdf.y - card_height
    for index, (label, value) in enumerate(metrics):
        x = pdf.margin + index * (card_width + 8)
        pdf.rect(x, y, card_width, card_height, "#EDF4FF")
        pdf.text(x + 8, y + 25, label, size=7, color="#60728C", bold=True)
        pdf.text(x + 8, y + 9, value, size=14, color="#0A4EA3", bold=True)
    pdf.y = y - 22


def _draw_pdf_timetable_grid(
    pdf: _TimetablePdf,
    workspace_payload: dict,
    entity: dict,
    *,
    entity_kind: str,
):
    days = workspace_payload.get("days", [])
    time_slots = workspace_payload.get("time_slots", [])
    if not days or not time_slots:
        pdf.paragraph("No timetable days or periods are configured for this scope.")
        return

    entity_id = int(entity.get("id") or 0)
    entry_lookup = _build_entry_lookup(
        workspace_payload,
        "section_id" if entity_kind == "section" else "teacher_id",
    )
    blocked_lookup = _build_blocked_slot_lookup(workspace_payload)

    period_width = 76
    grid_width = pdf.width - (2 * pdf.margin)
    day_width = (grid_width - period_width) / len(days)
    header_height = 23
    row_height = min(43, max(27, (pdf.y - pdf.margin - 30 - header_height) / max(len(time_slots), 1)))
    y_top = pdf.y

    pdf.rect(pdf.margin, y_top - header_height, period_width, header_height, "#0A4EA3")
    pdf.text(pdf.margin + 8, y_top - 15, "Period", size=7, color="#FFFFFF", bold=True)
    x = pdf.margin + period_width
    for day in days:
        pdf.rect(x, y_top - header_height, day_width, header_height, "#0A4EA3")
        pdf.text(x + 5, y_top - 15, _pdf_truncate(day.get("label", ""), 14), size=7, color="#FFFFFF", bold=True)
        x += day_width

    y = y_top - header_height
    for slot in time_slots:
        period_index = int(slot.get("period_index") or 0)
        y -= row_height
        pdf.rect(pdf.margin, y, period_width, row_height, "#F8FBFF")
        pdf.text(pdf.margin + 5, y + row_height - 13, slot.get("short_label") or f"P{period_index}", size=7, color="#17365D", bold=True)
        pdf.text(pdf.margin + 5, y + 7, _pdf_truncate(slot.get("time_range", ""), 16), size=6, color="#60728C")
        x = pdf.margin + period_width
        for day in days:
            day_key = str(day.get("key") or "")
            blocked_slot = blocked_lookup.get((day_key, period_index))
            entry = entry_lookup.get((entity_id, day_key, period_index))
            if blocked_slot:
                fill_color = "#F3F6F9"
                top_line = blocked_slot.get("label", "Blocked")
                bottom_line = blocked_slot.get("block_type_label", "")
                text_color = "#475569"
            elif entry:
                fill_color = f"#{_safe_excel_hex(entry.get('subject_color_soft'), 'F4F8FF')}"
                top_line = f"{entry.get('subject_code', '')} {entry.get('subject_name', '')}".strip()
                bottom_line = entry.get("teacher_name") if entity_kind == "section" else entry.get("section_label")
                if entry.get("status") == "stale":
                    bottom_line = "Needs review"
                text_color = f"#{_safe_excel_hex(entry.get('subject_color_text'), EXCEL_TEXT)}"
            else:
                fill_color = "#FFFFFF"
                top_line = ""
                bottom_line = ""
                text_color = "#60728C"
            pdf.rect(x, y, day_width, row_height, fill_color)
            pdf.line(x, y, x + day_width, y, "#D8E5F4", 0.35)
            pdf.line(x, y, x, y + row_height, "#D8E5F4", 0.35)
            if top_line:
                pdf.text(x + 4, y + row_height - 13, _pdf_truncate(top_line, max(10, int(day_width / 4.5))), size=6.4, color=text_color, bold=True)
            if bottom_line:
                pdf.text(x + 4, y + 8, _pdf_truncate(bottom_line, max(10, int(day_width / 4.5))), size=5.8, color="#60728C")
            x += day_width
        pdf.line(pdf.margin, y, pdf.margin + grid_width, y, "#D8E5F4", 0.35)
    pdf.y = y - 12


def _build_timetable_pdf_bytes(
    workspace_payload: dict,
    branch_name: str,
    academic_year_name: str,
) -> bytes:
    generated_at = datetime.now()
    pdf = _TimetablePdf(
        "Weekly Timetable",
        f"{branch_name} | Academic Year {academic_year_name} | Generated {generated_at.strftime('%Y-%m-%d %H:%M')}",
    )
    pdf.paragraph(
        "Printable timetable export generated from the same live section, teacher, subject, and school-day data shown in the TIS timetable workspace."
    )
    _draw_pdf_metric_cards(pdf, workspace_payload)
    warnings = workspace_payload.get("warnings", [])
    if warnings:
        pdf.text(pdf.margin, pdf.y, "Warnings", size=11, color="#9F2D1F", bold=True)
        pdf.y -= 13
        for warning in warnings[:4]:
            pdf.paragraph(f"- {warning}", size=7.5, color="#8A4F00")

    for section in workspace_payload.get("sections", []):
        pdf.add_page()
        pdf.text(pdf.margin, pdf.y, str(section.get("section_label") or "Section"), size=12, color="#0A4EA3", bold=True)
        pdf.y -= 14
        pdf.paragraph(
            f"Class Status: {section.get('class_status', '')} | "
            f"Scheduled {section.get('scheduled_hours', 0)}/{section.get('total_required_hours', 0)}h | "
            f"Remaining {section.get('remaining_hours', 0)}h",
            size=7.5,
        )
        _draw_pdf_timetable_grid(pdf, workspace_payload, section, entity_kind="section")

    for teacher in workspace_payload.get("teachers", []):
        pdf.add_page()
        pdf.text(pdf.margin, pdf.y, str(teacher.get("teacher_name") or "Teacher"), size=12, color="#0A4EA3", bold=True)
        pdf.y -= 14
        pdf.paragraph(
            f"Teacher ID: {teacher.get('teacher_id', '')} | "
            f"Scheduled {teacher.get('scheduled_hours', 0)}/{teacher.get('required_hours', 0)}h | "
            f"Capacity {teacher.get('capacity_hours', 0)}h",
            size=7.5,
        )
        _draw_pdf_timetable_grid(pdf, workspace_payload, teacher, entity_kind="teacher")

    return pdf.build()


@router.get("/")
def timetable_page(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_current_user_or_redirect(request, db)
    if redirect_response:
        return redirect_response

    branch_id, academic_year_id = get_scope_ids(current_user)
    workspace_payload = build_timetable_workspace_payload(
        db,
        branch_id,
        academic_year_id,
    )

    return templates.TemplateResponse(
        request,
        "timetable.html",
        {
            "request": request,
            "timetable_payload": workspace_payload,
            "can_edit_timetable": auth.can_edit_data(current_user),
            "can_manage_system_settings": auth.can_manage_system_settings(current_user),
            **build_shell_context(
                request,
                db,
                current_user,
                page_key="timetable",
                title="Timetable",
                intro=(
                    "Build the weekly teaching schedule from your current branch planning, "
                    "teacher assignments, and timetable settings."
                ),
                icon="timetable",
            ),
        },
    )


@router.get("/export.xlsx")
def export_timetable_excel(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_current_user_or_redirect(request, db)
    if redirect_response:
        return redirect_response

    branch_id, academic_year_id = get_scope_ids(current_user)
    branch_name, academic_year_name = _get_scope_labels(db, branch_id, academic_year_id)
    workspace_payload = build_timetable_workspace_payload(
        db,
        branch_id,
        academic_year_id,
    )
    payload = _build_timetable_xlsx_bytes(
        workspace_payload,
        branch_name,
        academic_year_name,
    )
    file_name = _build_timetable_export_filename(
        branch_name,
        academic_year_name,
        "xlsx",
    )
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={file_name}"},
    )


@router.get("/export.pdf")
def export_timetable_pdf(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_current_user_or_redirect(request, db)
    if redirect_response:
        return redirect_response

    branch_id, academic_year_id = get_scope_ids(current_user)
    branch_name, academic_year_name = _get_scope_labels(db, branch_id, academic_year_id)
    workspace_payload = build_timetable_workspace_payload(
        db,
        branch_id,
        academic_year_id,
    )
    payload = _build_timetable_pdf_bytes(
        workspace_payload,
        branch_name,
        academic_year_name,
    )
    file_name = _build_timetable_export_filename(
        branch_name,
        academic_year_name,
        "pdf",
    )
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={file_name}"},
    )


@router.post("/api/assign")
async def assign_timetable_slot(
    request: Request,
    db: Session = Depends(get_db),
):
    current_user, redirect_response = _get_current_user_or_redirect(request, db)
    if redirect_response:
        return _json_error("Please sign in again to continue.", status_code=401)

    if not auth.can_edit_data(current_user):
        return _json_error(
            "Your role can view the timetable but cannot change timetable slots.",
            status_code=403,
        )

    try:
        payload = await request.json()
    except Exception:
        return _json_error("Unable to read the timetable request payload.")

    branch_id, academic_year_id = get_scope_ids(current_user)
    workspace_payload = build_timetable_workspace_payload(
        db,
        branch_id,
        academic_year_id,
    )
    settings_payload = workspace_payload.get("settings", {})

    section_id = _parse_int(payload.get("section_id"))
    period_index = _parse_int(payload.get("period_index"))
    day_key = normalize_day_key(payload.get("day_key"))
    subject_code = str(payload.get("subject_code") or "").strip().upper()

    if section_id is None:
        return _json_error("Select a valid section before assigning a timetable slot.")
    if not day_key or day_key not in {
        day_item.get("key")
        for day_item in workspace_payload.get("days", [])
    }:
        return _json_error("Selected timetable day is not valid for the current settings.")
    if (
        period_index is None
        or period_index <= 0
        or period_index > int(settings_payload.get("periods_per_day") or 0)
    ):
        return _json_error("Selected timetable period is outside the configured school day.")
    if any(
        str(block_slot.get("day_key") or "") == day_key
        and int(block_slot.get("period_index") or 0) == period_index
        for block_slot in workspace_payload.get("blocked_slots", [])
    ):
        return _json_error("That slot is blocked by a break, prayer time, or another non-teaching rule.")

    section_payload = _find_section_by_id(workspace_payload, section_id)
    if not section_payload:
        return _json_error("Selected section is not available in the active planning scope.")

    existing_entry_payload = _find_entry_by_section_slot(
        workspace_payload,
        section_id=section_id,
        day_key=day_key,
        period_index=period_index,
    )
    existing_entry_row = None
    if existing_entry_payload and existing_entry_payload.get("id"):
        existing_entry_row = db.query(models.TimetableEntry).filter(
            models.TimetableEntry.id == int(existing_entry_payload["id"]),
            models.TimetableEntry.branch_id == branch_id,
            models.TimetableEntry.academic_year_id == academic_year_id,
        ).first()

    if not subject_code:
        if not existing_entry_row:
            return _json_error("That slot is already empty.")
        db.delete(existing_entry_row)
        db.commit()
        refreshed_payload = build_timetable_workspace_payload(db, branch_id, academic_year_id)
        return _json_success(refreshed_payload, message="Timetable slot cleared.")

    option_payload = _find_section_option(
        workspace_payload,
        section_id=section_id,
        subject_code=subject_code,
    )
    if not option_payload:
        return _json_error("Selected subject is not part of this section timetable plan.")
    if not option_payload.get("is_schedulable") or not option_payload.get("teacher_id"):
        return _json_error(
            "This subject does not currently have an assigned teacher in planning, so it cannot be placed yet."
        )

    teacher_id = int(option_payload["teacher_id"])
    teacher_payload = _find_teacher_by_id(workspace_payload, teacher_id)
    if not teacher_payload:
        return _json_error("Assigned teacher is not available in the active branch and academic year.")

    other_entry_for_teacher = _find_teacher_conflict(
        workspace_payload,
        teacher_id=teacher_id,
        day_key=day_key,
        period_index=period_index,
        ignore_entry_id=existing_entry_payload.get("id") if existing_entry_payload else None,
    )
    if other_entry_for_teacher:
        return _json_error(
            f"{teacher_payload['teacher_name']} is already teaching "
            f"{other_entry_for_teacher.get('section_label', 'another section')} in that slot."
        )

    ignore_entry_id = (
        int(existing_entry_payload.get("id") or 0)
        if existing_entry_payload
        else 0
    )
    scheduled_count = sum(
        1
        for entry in workspace_payload.get("entries", [])
        if int(entry.get("section_id") or 0) == section_id
        and str(entry.get("subject_code") or "").strip().upper() == subject_code
        and str(entry.get("status") or "") == "scheduled"
        and int(entry.get("id") or 0) != ignore_entry_id
    )
    weekly_hours = int(option_payload.get("weekly_hours") or 0)
    if scheduled_count >= weekly_hours:
        return _json_error(
            f"{subject_code} already reached its required {weekly_hours} hour"
            + ("" if weekly_hours == 1 else "s")
            + f" for {section_payload['section_label']}."
        )

    try:
        if existing_entry_row is None:
            db.add(
                models.TimetableEntry(
                    branch_id=branch_id,
                    academic_year_id=academic_year_id,
                    planning_section_id=section_id,
                    subject_code=subject_code,
                    teacher_id=teacher_id,
                    day_key=day_key,
                    period_index=period_index,
                )
            )
        else:
            existing_entry_row.subject_code = subject_code
            existing_entry_row.teacher_id = teacher_id
            existing_entry_row.day_key = day_key
            existing_entry_row.period_index = period_index
        db.commit()
    except IntegrityError:
        db.rollback()
        return _json_error(
            "This slot could not be saved because it conflicts with another timetable entry. Refresh the timetable and try again."
        )

    refreshed_payload = build_timetable_workspace_payload(db, branch_id, academic_year_id)
    return _json_success(
        refreshed_payload,
        message=(
            f"{subject_code} assigned to {section_payload['section_label']} "
            f"with {teacher_payload['teacher_name']}."
        ),
    )
