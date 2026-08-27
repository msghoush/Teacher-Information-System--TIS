import pathlib
import io
import re

import pytest
from sqlalchemy import create_engine, inspect, text
from openpyxl import load_workbook
from starlette.requests import Request

import db_migrations
import main
import models
from routers import timetable as timetable_router
from timetable_logic import BLOCK_TYPE_OPTIONS, normalize_block_type
from timetable_slot_service import build_canonical_slot_projection
from test_timetable_versioning import db  # noqa: F401 - shared route-render fixture


def compose(*, days=("monday",), blocks=(), periods=8, duration=45, start="07:15"):
    return build_canonical_slot_projection(
        school_group_id=1, branch_id=10, academic_year_id=100,
        working_day_keys=list(days), periods_per_day=periods,
        period_duration_minutes=duration, school_start_time=start,
        blocks=list(blocks),
    )


def after(period, minutes, *, day="all", label="Break", block_type="break"):
    return {"block_type": block_type, "label": label, "day_key": day,
            "placement_mode": "after_period", "insert_after_period": period,
            "duration_minutes": minutes}


def test_base_and_after_period_composition_matches_product_example():
    base = compose()
    assert base["calculated_school_end_time"] == "13:15"
    result = compose(blocks=[after(3, 25, label="Morning Break")])
    assert result["valid"] is True
    assert result["calculated_school_end_time"] == "13:40"
    assert [(item["label"], item["start_time"], item["end_time"])
            for item in result["timelines"][0]["items"]][2:5] == [
        ("Period 3", "08:45", "09:30"),
        ("Morning Break", "09:30", "09:55"),
        ("Period 4", "09:55", "10:40"),
    ]
    assert [slot["period_index"] for slot in result["teaching_slots"]] == list(range(1, 9))


def test_multiple_and_day_specific_blocks_shift_only_applicable_days():
    result = compose(days=("monday", "tuesday"), blocks=[
        after(3, 20, label="Break"),
        after(6, 15, day="monday", label="Prayer", block_type="prayer"),
    ])
    monday, tuesday = result["timelines"]
    assert monday["end_time"] == "13:50" and monday["block_minutes"] == 35
    assert tuesday["end_time"] == "13:35" and tuesday["block_minutes"] == 20
    assert result["slot_map"][("monday", 7)]["start_time"] == "12:20"
    assert result["slot_map"][("tuesday", 7)]["start_time"] == "12:05"


def test_fixed_time_boundary_is_inserted_and_partial_overlap_is_rejected():
    boundary = compose(periods=2, blocks=[{"block_type": "break", "label": "Break", "day_key": "all",
        "placement_mode": "fixed_time", "start_time": "08:00", "end_time": "08:15"}])
    assert boundary["valid"] is True
    assert boundary["slot_map"][("monday", 2)]["start_time"] == "08:15"
    partial = compose(periods=2, blocks=[{"block_type": "break", "label": "Bad", "day_key": "all",
        "placement_mode": "fixed_time", "start_time": "07:30", "end_time": "07:40"}])
    assert partial["valid"] is False
    assert "inside a teaching period" in partial["issues"][0]["message"]


def test_multiple_fixed_blocks_conflict_instead_of_overwriting():
    result = compose(periods=2, blocks=[
        {"block_type": "break", "label": "First", "day_key": "all", "placement_mode": "fixed_time", "start_time": "08:00", "end_time": "08:20"},
        {"block_type": "prayer", "label": "Conflict", "day_key": "all", "placement_mode": "fixed_time", "start_time": "08:00", "end_time": "08:10"},
    ])
    assert result["valid"] is False


def test_controlled_block_types_and_unknown_rejection():
    keys = {item["key"] for item in BLOCK_TYPE_OPTIONS}
    assert {"break", "prayer", "non_teaching", "recess", "lunch", "assembly",
            "whole_school_event", "advisory", "intervention", "transition",
            "dismissal_preparation", "other"} <= keys
    assert all(normalize_block_type(key) == key for key in keys)
    assert normalize_block_type("unknown") == ""


def test_fingerprint_changes_and_templates_consume_composed_timeline():
    assert compose()["fingerprint"] != compose(blocks=[after(3, 25)])["fingerprint"]
    config = pathlib.Path("templates/system_configuration_timetable.html").read_text(encoding="utf-8")
    board = pathlib.Path("templates/timetable.html").read_text(encoding="utf-8")
    assert "Calculated School End" in config and 'day_timeline["items"]' in config
    assert "timeline_rows" in board and "daySlot.time_range" in board


def test_timetable_settings_and_main_routes_render_composed_timeline(db, monkeypatch):
    setting = db.query(models.TimetableSetting).filter_by(id=5000).one()
    setting.periods_per_day = 8
    setting.period_duration_minutes = 45
    setting.school_start_time = "07:15"
    setting.school_end_time = "13:40"
    db.add(models.TimetableNonTeachingBlock(
        timetable_setting_id=setting.id,
        block_type="break",
        label="Morning Break",
        day_key="all",
        start_time="",
        end_time="",
        start_period=3,
        end_period=3,
        placement_mode="after_period",
        insert_after_period=3,
        duration_minutes=25,
    ))
    db.commit()

    user = db.query(models.User).filter_by(user_id="U1").one()
    user.scope_school_group_id = 1
    user.scope_branch_id = 10
    user.scope_academic_year_id = 100
    monkeypatch.setattr(main, "_get_configuration_access", lambda request, session: (user, None))
    monkeypatch.setattr(main.auth, "has_permission", lambda *args, **kwargs: True)
    monkeypatch.setattr(main.auth, "has_any_permission", lambda *args, **kwargs: True)
    monkeypatch.setattr(main.auth, "is_platform_user", lambda *args, **kwargs: False)
    monkeypatch.setattr(timetable_router, "_get_current_user_or_redirect", lambda request, session: (user, None))

    def request(path):
        return Request({
            "type": "http",
            "http_version": "1.1",
            "method": "GET",
            "path": path,
            "raw_path": path.encode("utf-8"),
            "query_string": b"",
            "headers": [],
            "scheme": "http",
            "server": ("testserver", 80),
            "client": ("testclient", 50000),
            "root_path": "",
            "app": main.app,
        })

    settings_response = main.system_configuration_timetable_settings(
        request("/system-configuration/timetable-settings"), db
    )
    settings_html = settings_response.body.decode("utf-8")
    assert settings_response.status_code == 200
    assert "P1" in settings_html and "07:15–08:00" in settings_html
    assert "Morning Break" in settings_html and "09:30–09:55" in settings_html
    assert "P4" in settings_html and "09:55–10:40" in settings_html

    timetable_response = timetable_router.timetable_page(request("/timetable/"), db=db)
    assert timetable_response.status_code == 200
    assert "Timetable Grid" in timetable_response.body.decode("utf-8")


def test_stage35_migration_is_additive_idempotent_and_preserves_legacy_row(tmp_path):
    engine = create_engine(f"sqlite:///{tmp_path / 'stage35.db'}")
    with engine.begin() as connection:
        connection.execute(text("""CREATE TABLE timetable_non_teaching_blocks (
            id INTEGER PRIMARY KEY, timetable_setting_id INTEGER NOT NULL,
            block_type VARCHAR(32) NOT NULL, label VARCHAR(80) NOT NULL,
            day_key VARCHAR(16) NOT NULL, start_time VARCHAR(5), end_time VARCHAR(5),
            start_period INTEGER NOT NULL, end_period INTEGER NOT NULL)"""))
        connection.execute(text("""INSERT INTO timetable_non_teaching_blocks
            VALUES (1, 10, 'break', 'Legacy Break', 'all', '09:30', '09:50', 3, 3)"""))
    for _ in range(2):
        with engine.begin() as connection:
            db_migrations._smart_timetable_stage35_composed_timeline(engine, connection)
    columns = {column["name"] for column in inspect(engine).get_columns("timetable_non_teaching_blocks")}
    assert {"placement_mode", "insert_after_period", "duration_minutes"} <= columns
    with engine.connect() as connection:
        row = connection.execute(text("SELECT label, placement_mode, start_time, end_time FROM timetable_non_teaching_blocks")).one()
    assert tuple(row) == ("Legacy Break", "fixed_time", "09:30", "09:50")


def test_export_rows_keep_day_specific_composed_times():
    projection = compose(days=("monday", "tuesday"), periods=4,
        blocks=[after(2, 20, day="monday", label="Monday Break")])
    workspace = {
        "days": [{"key": "monday", "label": "Monday"}, {"key": "tuesday", "label": "Tuesday"}],
        "timeline_rows": projection["timeline_rows"], "time_slots": projection["periods"],
        "sections": [{"id": 1, "section_label": "Grade 1 - A", "class_status": "active",
                      "scheduled_hours": 0, "total_required_hours": 0, "remaining_hours": 0}],
        "teachers": [], "entries": [], "blocked_slots": [], "settings": {"blocks": []},
        "warnings": [], "readiness": {},
    }
    rows = timetable_router._build_export_board_rows(workspace)
    p3 = next(row for row in rows if row["row_type"] == "period" and row["slot"]["period_index"] == 3)
    assert p3["items_by_day"]["monday"]["start_time"] == "09:05"
    assert p3["items_by_day"]["tuesday"]["start_time"] == "08:45"
    output = timetable_router._build_timetable_xlsx_bytes(workspace, "Branch", "2026-2027")
    sheet = load_workbook(io.BytesIO(output))["By Section"]
    values = [str(cell.value or "") for row in sheet.iter_rows() for cell in row]
    assert any("09:05" in value for value in values)
    assert any("08:45" in value for value in values)
    assert any("Monday Break" in value for value in values)
    assert any(" - " in value for value in values)
    assert all("?" not in value for value in values)
    pdf = timetable_router._build_timetable_pdf_bytes(workspace, "Branch", "2026-2027", logo_assets=[])
    pdf_text = b"\n".join(re.findall(rb"\((.*?)\) Tj", pdf)).decode("latin-1")
    assert "Monday Break" in pdf_text
    assert "?" not in pdf_text
