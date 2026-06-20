import io
import json
import sys
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect, text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import auth
import main
import models
from teacher_qualifications import ensure_qualification_options_seeded
from database import SessionLocal, engine


def _assert(condition, message):
    if not condition:
        raise AssertionError(message)


def _session_for(user, branch, year):
    return {
        auth.SESSION_COOKIE_KEY: auth.create_session_token(user),
        "branch_id": str(branch.id),
        "academic_year_id": str(year.id),
    }


def _decode_response(response):
    try:
        return response.text
    except UnicodeDecodeError:
        return response.content.decode("utf-8", errors="ignore")


def _post(client, url, data=None, files=None):
    return client.post(url, data=data or {}, files=files, follow_redirects=False)


def _get(client, url):
    return client.get(url, follow_redirects=False)


def _seed(db):
    ensure_qualification_options_seeded(db)

    school_a = models.SchoolGroup(name="School A", status=True)
    school_b = models.SchoolGroup(name="School B", status=True)
    db.add_all([school_a, school_b])
    db.flush()

    branch_a = models.Branch(name="School A Main", school_group_id=school_a.id, status=True)
    branch_b = models.Branch(name="School B Main", school_group_id=school_b.id, status=True)
    db.add_all([branch_a, branch_b])
    db.flush()

    year_a = models.AcademicYear(school_group_id=school_a.id, year_name="2026-2027", is_active=True)
    year_b = models.AcademicYear(school_group_id=school_b.id, year_name="2026-2027", is_active=True)
    db.add_all([year_a, year_b])
    db.flush()

    dev_a = models.User(
        user_id="1000000001",
        username="school_a_dev",
        first_name="Ada",
        last_name="Admin",
        position="Principal",
        role=auth.ROLE_ADMINISTRATOR,
        user_type=auth.USER_TYPE_TENANT,
        access_scope=auth.ACCESS_SCOPE_ORGANIZATION,
        password=auth.get_password_hash("password123"),
        school_group_id=school_a.id,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        is_active=True,
    )
    admin_a = models.User(
        user_id="1000000002",
        username="school_a_admin",
        first_name="Ava",
        last_name="Admin",
        position="Principal",
        role=auth.ROLE_ADMINISTRATOR,
        password=auth.get_password_hash("password123"),
        school_group_id=school_a.id,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        is_active=True,
    )
    teacher_user_a = models.User(
        user_id="1000000003",
        username="school_a_teacher",
        first_name="Alice",
        last_name="Teacher",
        position="Teacher",
        role=auth.ROLE_USER,
        password=auth.get_password_hash("password123"),
        school_group_id=school_a.id,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        is_active=True,
    )
    admin_b = models.User(
        user_id="2000000002",
        username="school_b_admin",
        first_name="Bea",
        last_name="Admin",
        position="Principal",
        role=auth.ROLE_ADMINISTRATOR,
        password=auth.get_password_hash("password123"),
        school_group_id=school_b.id,
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
        is_active=True,
    )
    teacher_user_b = models.User(
        user_id="2000000003",
        username="school_b_teacher",
        first_name="Bella",
        last_name="Teacher",
        position="Teacher",
        role=auth.ROLE_USER,
        password=auth.get_password_hash("password123"),
        school_group_id=school_b.id,
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
        is_active=True,
    )
    db.add_all([dev_a, admin_a, teacher_user_a, admin_b, teacher_user_b])
    db.flush()

    subject_a = models.Subject(
        subject_code="AAA101",
        subject_name="Alpha Arithmetic",
        weekly_hours=5,
        grade=1,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
    )
    subject_b = models.Subject(
        subject_code="BBB101",
        subject_name="Beta Biology",
        weekly_hours=5,
        grade=1,
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
    )
    db.add_all([subject_a, subject_b])
    db.flush()

    teacher_a = models.Teacher(
        teacher_id=teacher_user_a.user_id,
        first_name="Alice",
        last_name="Teacher",
        degree_major="Mathematics",
        subject_code=subject_a.subject_code,
        max_hours=24,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
    )
    teacher_b = models.Teacher(
        teacher_id=teacher_user_b.user_id,
        first_name="Bella",
        last_name="Teacher",
        degree_major="Science",
        subject_code=subject_b.subject_code,
        max_hours=24,
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
    )
    db.add_all([teacher_a, teacher_b])
    db.flush()
    db.add_all(
        [
            models.TeacherSubjectAllocation(teacher_id=teacher_a.id, subject_code=subject_a.subject_code),
            models.TeacherSubjectAllocation(teacher_id=teacher_b.id, subject_code=subject_b.subject_code),
        ]
    )

    planning_a = models.PlanningSection(
        grade_level="1",
        section_name="A",
        class_status="Current",
        homeroom_teacher_id=teacher_a.id,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
    )
    planning_b = models.PlanningSection(
        grade_level="1",
        section_name="A",
        class_status="Current",
        homeroom_teacher_id=teacher_b.id,
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
    )
    db.add_all([planning_a, planning_b])
    db.flush()

    event_type_a = models.CalendarEventType(
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        name="School A Type",
        color="#0A4EA3",
        icon="calendar",
        is_active=True,
    )
    event_type_b = models.CalendarEventType(
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
        name="School B Type",
        color="#0A4EA3",
        icon="calendar",
        is_active=True,
    )
    db.add_all([event_type_a, event_type_b])
    db.flush()

    event_a = models.CalendarEvent(
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        event_type_id=event_type_a.id,
        title="School A Calendar Event",
        event_date="2026-09-01",
        end_date="2026-09-01",
        target_group="All School",
        priority="Normal",
        status="Planned",
        recurrence_rule="None",
        recurrence_interval=1,
        created_by_user_id=dev_a.user_id,
    )
    event_b = models.CalendarEvent(
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
        event_type_id=event_type_b.id,
        title="School B Calendar Event",
        event_date="2026-09-01",
        end_date="2026-09-01",
        target_group="All School",
        priority="Normal",
        status="Planned",
        recurrence_rule="None",
        recurrence_interval=1,
        created_by_user_id=admin_b.user_id,
    )
    db.add_all([event_a, event_b])
    db.flush()

    observation_a = models.Observation(
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        teacher_id=teacher_a.id,
        evaluator_user_id=admin_a.user_id,
        observation_type="Formal",
        observation_date="2026-09-02",
        subject=subject_a.subject_code,
        status="Final",
    )
    observation_b = models.Observation(
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
        teacher_id=teacher_b.id,
        evaluator_user_id=admin_b.user_id,
        observation_type="Formal",
        observation_date="2026-09-02",
        subject=subject_b.subject_code,
        status="Final",
    )
    db.add_all([observation_a, observation_b])

    notification_a = models.SystemNotification(
        school_group_id=school_a.id,
        branch_id=branch_a.id,
        academic_year_id=year_a.id,
        recipient_user_id=teacher_user_a.user_id,
        requesting_user_id=admin_a.user_id,
        request_type="Message",
        title="School A Notification",
        message="School A only",
        status="New",
        recipient_scope="User",
    )
    notification_b = models.SystemNotification(
        school_group_id=school_b.id,
        branch_id=branch_b.id,
        academic_year_id=year_b.id,
        recipient_user_id=teacher_user_b.user_id,
        requesting_user_id=admin_b.user_id,
        request_type="Message",
        title="School B Notification",
        message="School B only",
        status="New",
        recipient_scope="User",
    )
    db.add_all([notification_a, notification_b])
    db.commit()

    return {
        "school_a": school_a,
        "school_b": school_b,
        "branch_a": branch_a,
        "branch_b": branch_b,
        "year_a": year_a,
        "year_b": year_b,
        "dev_a": dev_a,
        "admin_a": admin_a,
        "teacher_user_a": teacher_user_a,
        "admin_b": admin_b,
        "teacher_user_b": teacher_user_b,
        "subject_a": subject_a,
        "subject_b": subject_b,
        "teacher_a": teacher_a,
        "teacher_b": teacher_b,
        "planning_a": planning_a,
        "planning_b": planning_b,
        "event_type_a": event_type_a,
        "event_type_b": event_type_b,
        "event_a": event_a,
        "event_b": event_b,
        "observation_a": observation_a,
        "observation_b": observation_b,
        "notification_a": notification_a,
        "notification_b": notification_b,
    }


def _assert_school_b_absent(response, *tokens):
    body = _decode_response(response)
    for token in tokens:
        _assert(token not in body, f"cross-tenant token rendered: {token}")


def _subject_import_file():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["subject_code", "subject_name", "weekly_hours", "grade", "subject_color"])
    sheet.append(["IMP201", "Imported Alpha", 3, 2, "#0A4EA3"])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _assert_subject_export_scoped(client):
    response = _get(client, "/subjects/export")
    _assert(response.status_code == 200, "subject export failed")
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    values = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values.extend(str(value) for value in row if value is not None)
    joined = "\n".join(values)
    _assert("Alpha Arithmetic" in joined, "School A subject missing from export")
    _assert("Beta Biology" not in joined, "School B subject leaked in export")


def _run_http_checks(db, data):
    with TestClient(main.app) as client:
        for key, value in _session_for(data["dev_a"], data["branch_a"], data["year_a"]).items():
            client.cookies.set(key, value)

        list_checks = [
            ("/teachers/", "Alice Teacher", "Bella Teacher"),
            ("/subjects/", "Alpha Arithmetic", "Beta Biology"),
            ("/planning/", "Grade 1", "Beta Biology"),
            ("/users", "Ada Admin", "Bea Admin"),
            ("/academic-calendar/?month=2026-09", "School A Calendar Event", "School B Calendar Event"),
            ("/observations/", "Alice Teacher", "Bella Teacher"),
            ("/notifications/compose", "Alice Teacher", "Bella Teacher"),
        ]
        for url, expected, forbidden in list_checks:
            response = _get(client, url)
            _assert(response.status_code == 200, f"{url} did not render")
            body = _decode_response(response)
            _assert(expected in body, f"{url} did not show expected scoped data")
            _assert(forbidden not in body, f"{url} leaked School B data")

        detail_checks = [
            f"/teachers/edit/{data['teacher_b'].id}",
            f"/subjects/edit/{data['subject_b'].id}",
            f"/planning/edit/{data['planning_b'].id}",
            f"/observations/{data['observation_b'].id}",
            f"/observations/{data['observation_b'].id}/edit",
            f"/observations/{data['observation_b'].id}/export/pdf",
            f"/users/edit/{data['admin_b'].id}",
        ]
        for url in detail_checks:
            response = _get(client, url)
            _assert(
                response.status_code in {200, 302, 303, 307, 403, 404},
                f"{url} failed unexpectedly status={response.status_code} body={_decode_response(response)[:300]}",
            )
            _assert_school_b_absent(response, "Bella Teacher", "Beta Biology", "Bea Admin", "School B")

        _assert_subject_export_scoped(client)
        response = _get(client, "/reports/allocation-plan.xlsx")
        _assert(response.status_code == 200, "allocation report xlsx failed")
        _assert(b"Beta Biology" not in response.content, "School B token leaked in xlsx bytes")
        response = _get(client, "/reports/allocation-plan.pdf")
        _assert(response.status_code == 200, "allocation report pdf failed")
        _assert(b"Beta Biology" not in response.content, "School B token leaked in pdf bytes")
        response = _get(client, "/academic-calendar/export.pdf?month=2026-09")
        _assert(response.status_code == 200, "calendar pdf export failed")
        _assert(b"School B Calendar Event" not in response.content, "School B calendar leaked in export")

        _post(
            client,
            "/subjects/",
            data={
                "subject_code": "NEW101",
                "subject_name": "New Alpha",
                "weekly_hours": "2",
                "grade": "1",
            },
        )
        _post(
            client,
            f"/subjects/edit/{data['subject_b'].id}",
            data={
                "subject_code": "HAX999",
                "subject_name": "Hijacked",
                "weekly_hours": "1",
                "grade": "1",
            },
        )
        _get(client, f"/subjects/delete/{data['subject_b'].id}")

        import_payload = _subject_import_file()
        _post(
            client,
            "/subjects/import",
            files={
                "subject_file": (
                    "subjects.xlsx",
                    import_payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        _post(
            client,
            "/teachers/",
            data={
                "teacher_id": "1000000099",
                "first_name": "Ari",
                "middle_name": "",
                "last_name": "Teacher",
                "qualification_keys": ["mathematics"],
                "qualification_override_subject_codes": ["NEW101"],
                "subject_codes": ["NEW101"],
                "max_hours": "24",
                "extra_hours_allowed": "",
                "extra_hours_count": "0",
                "teaches_national_section": "",
                "national_section_hours": "0",
                "is_new_teacher": "",
            },
        )
        _post(
            client,
            f"/teachers/edit/{data['teacher_b'].id}",
            data={
                "teacher_id": data["teacher_b"].teacher_id,
                "first_name": "Hijacked",
                "middle_name": "",
                "last_name": "Teacher",
                "qualification_keys": ["biology"],
                "qualification_override_subject_codes": ["BBB101"],
                "subject_codes": ["BBB101"],
                "max_hours": "24",
            },
        )
        _get(client, f"/teachers/delete/{data['teacher_b'].id}")

        _post(
            client,
            "/planning/",
            data={
                "grade_level": "2",
                "section_name": "B",
                "class_status": "Current",
                "homeroom_teacher_id": "",
            },
        )
        _post(
            client,
            f"/planning/edit/{data['planning_b'].id}",
            data={
                "grade_level": "2",
                "section_name": "C",
                "class_status": "New",
                "homeroom_teacher_id": "",
            },
        )
        _get(client, f"/planning/delete/{data['planning_b'].id}")

        _post(
            client,
            "/users",
            data={
                "user_id": "1000000098",
                "first_name": "New",
                "last_name": "User",
                "position": "Teacher",
                "role": auth.ROLE_USER,
                "password": "password123",
                "branch_id": str(data["branch_a"].id),
            },
        )
        _post(
            client,
            f"/users/edit/{data['admin_b'].id}",
            data={
                "user_id": data["admin_b"].user_id,
                "first_name": "Hijacked",
                "last_name": "Admin",
                "position": "Principal",
                "role": auth.ROLE_ADMINISTRATOR,
                "branch_id": str(data["branch_a"].id),
                "is_active": "active",
            },
        )
        _get(client, f"/users/delete/{data['admin_b'].id}")

        _post(
            client,
            "/observations/",
            data={
                "teacher_id": str(data["teacher_a"].id),
                "observation_type": "Formal",
                "observation_date": "2026-09-03",
                "term": "Term 1",
                "grade": "1",
                "section": "A",
                "period": "1",
                "subject": "AAA101",
                "evaluator_notes": "School A route create",
            },
        )
        _post(
            client,
            f"/observations/{data['observation_b'].id}/edit",
            data={
                "teacher_id": str(data["teacher_b"].id),
                "observation_type": "Formal",
                "observation_date": "2026-09-04",
                "subject": "BBB101",
                "evaluator_notes": "Hijacked",
            },
        )
        _post(client, f"/observations/{data['observation_b'].id}/delete")

        _post(
            client,
            "/academic-calendar/events",
            data={
                "title": "School A Created Calendar",
                "event_type_id": str(data["event_type_a"].id),
                "event_date": "2026-10-01",
                "end_date": "2026-10-01",
                "target_group": "All School",
                "priority": "Normal",
                "status": "Planned",
                "recurrence_rule": "None",
                "recurrence_interval": "1",
                "return_to": "/academic-calendar/",
            },
        )
        _post(
            client,
            f"/academic-calendar/events/{data['event_b'].id}",
            data={
                "title": "Hijacked Calendar",
                "event_type_id": str(data["event_type_b"].id),
                "event_date": "2026-10-02",
                "end_date": "2026-10-02",
                "target_group": "All School",
                "priority": "Normal",
                "status": "Planned",
                "recurrence_rule": "None",
                "recurrence_interval": "1",
                "return_to": "/academic-calendar/",
            },
        )
        _post(client, f"/academic-calendar/events/{data['event_b'].id}/delete")

        _post(
            client,
            "/notifications/compose",
            data={
                "title": "School A Broadcast",
                "recipient": "ALL",
                "message": "Scoped broadcast",
            },
        )
        _post(
            client,
            "/notifications/compose",
            data={
                "title": "Attempt School B Direct",
                "recipient": data["teacher_user_b"].user_id,
                "message": "Should not send",
            },
        )


def _run_database_checks():
    with engine.connect() as connection:
        _assert(connection.exec_driver_sql("PRAGMA foreign_keys").scalar() == 1, "SQLite FK pragma is disabled")
        _assert(not connection.exec_driver_sql("PRAGMA foreign_key_check").fetchall(), "SQLite foreign key orphans exist")
        inspector = inspect(connection)
        for table, column in (
            ("users", "school_group_id"),
            ("system_notifications", "school_group_id"),
            ("system_notifications", "branch_id"),
            ("system_notifications", "academic_year_id"),
        ):
            columns = {item["name"] for item in inspector.get_columns(table)}
            _assert(column in columns, f"{table}.{column} is missing")
        index_names = {
            table: {item.get("name") for item in inspector.get_indexes(table)}
            for table in ("users", "system_notifications")
        }
        _assert("ix_users_school_group_id" in index_names["users"], "users school group index missing")
        _assert(
            "ix_system_notifications_school_group_id" in index_names["system_notifications"],
            "system_notifications school group index missing",
        )
        _assert(
            "ix_system_notifications_branch_year" in index_names["system_notifications"],
            "system_notifications branch/year index missing",
        )
        foreign_keys = {
            table: {
                (
                    tuple(item.get("constrained_columns") or ()),
                    item.get("referred_table"),
                    tuple(item.get("referred_columns") or ()),
                )
                for item in inspector.get_foreign_keys(table)
            }
            for table in ("users", "system_notifications")
        }
        _assert(
            (("school_group_id",), "school_groups", ("id",)) in foreign_keys["users"],
            "users.school_group_id foreign key missing",
        )
        for column, target_table in (
            ("school_group_id", "school_groups"),
            ("branch_id", "branches"),
            ("academic_year_id", "academic_years"),
        ):
            _assert(
                ((column,), target_table, ("id",)) in foreign_keys["system_notifications"],
                f"system_notifications.{column} foreign key missing",
            )
        migrations = {
            row[0]
            for row in connection.execute(text("SELECT migration_id FROM schema_migrations")).fetchall()
        }
        _assert("20260613_002_legacy_runtime_schema_columns" in migrations, "runtime schema migration not recorded")


def _run_final_data_checks(db, data):
    subject_b = db.get(models.Subject, data["subject_b"].id)
    teacher_b = db.get(models.Teacher, data["teacher_b"].id)
    planning_b = db.get(models.PlanningSection, data["planning_b"].id)
    event_b = db.get(models.CalendarEvent, data["event_b"].id)
    observation_b = db.get(models.Observation, data["observation_b"].id)
    admin_b = db.get(models.User, data["admin_b"].id)

    _assert(subject_b is not None and subject_b.subject_code == "BBB101", "School B subject changed or deleted")
    _assert(teacher_b is not None and teacher_b.first_name == "Bella", "School B teacher changed or deleted")
    _assert(planning_b is not None and planning_b.grade_level == "1", "School B planning changed or deleted")
    _assert(event_b is not None and event_b.title == "School B Calendar Event", "School B calendar event changed or deleted")
    _assert(observation_b is not None and observation_b.evaluator_notes in {None, ""}, "School B observation changed or deleted")
    _assert(admin_b is not None and admin_b.first_name == "Bea", "School B user changed or deleted")

    school_a_notifications = db.query(models.SystemNotification).filter(
        models.SystemNotification.title == "School A Broadcast"
    ).all()
    _assert(school_a_notifications, "School A broadcast was not created")
    _assert(
        all(row.school_group_id == data["school_a"].id for row in school_a_notifications),
        "School A broadcast crossed tenant boundary",
    )
    forbidden_direct = db.query(models.SystemNotification).filter(
        models.SystemNotification.title == "Attempt School B Direct"
    ).all()
    _assert(not forbidden_direct, "School A user sent direct notification to School B")

    imported_subject = db.query(models.Subject).filter(
        models.Subject.subject_code == "IMP201"
    ).one()
    _assert(imported_subject.branch_id == data["branch_a"].id, "Imported subject used wrong branch")
    _assert(imported_subject.academic_year_id == data["year_a"].id, "Imported subject used wrong year")


def main_runner():
    _run_database_checks()
    db = SessionLocal()
    try:
        data = _seed(db)
        _run_http_checks(db, data)
        db.expire_all()
        _run_final_data_checks(db, data)
        result = {
            "status": "passed",
            "checks": [
                "list",
                "detail",
                "create",
                "edit",
                "delete",
                "export",
                "import",
                "reports",
                "notifications",
                "foreign_keys",
            ],
        }
        print(json.dumps(result, sort_keys=True))
    finally:
        db.close()


if __name__ == "__main__":
    main_runner()
