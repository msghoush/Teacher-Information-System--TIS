import json
import io
import re
import os
import subprocess
import sys
import threading
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, event as sa_event, inspect, text
from sqlalchemy.orm import sessionmaker

import models
from routers import timetable as timetable_router
from db_migrations import _smart_timetable_stage51_generator
from timetable_generation_service import (
    TimetableGenerationError,
    _minimum_difference,
    claim_run_by_public_id,
    claim_next_run,
    enqueue_generation,
    _current_input_for_run,
    heartbeat_run,
    generation_run_payload,
    mark_workflow_dispatch_failed,
    mark_run_terminal,
    persist_generated_result,
    recover_expired_runs,
    request_cancellation,
    set_run_progress,
)
from timetable_problem_builder import TimetableProblemBuilder, TimetableProblemError
from timetable_snapshot_service import build_current_snapshot_data
from timetable_solution_validator import TimetableSolutionValidator
from permission_registry import get_default_permissions_for_role
from timetable_logic import build_timetable_workspace_payload
from timetable_version_service import (
    create_manual_draft,
    resolve_operational_version,
    set_entry_lock,
    set_imported_active_pointer,
)
from test_timetable_versioning import db  # noqa: F401 - shared isolated database


def _make_ready(db):
    if not db.query(models.TeacherSectionAssignment).filter_by(
        planning_section_id=2001, subject_code="MAT"
    ).first():
        db.add(models.TeacherSectionAssignment(
            teacher_id=1001, planning_section_id=2001, subject_code="MAT"
        ))
    db.flush()


def _problem(db, *, constraints=None, locks=None):
    _make_ready(db)
    snapshot = build_current_snapshot_data(
        db,
        school_group_id=1,
        branch_id=10,
        academic_year_id=100,
        locks=locks or [],
        constraint_configuration=constraints,
    )
    return TimetableProblemBuilder().build(snapshot.canonical_json), snapshot


def _solve(problem, *, seed=7, timeout=10):
    pytest.importorskip("ortools")
    from timetable_cp_sat_solver import solve_timetable

    return solve_timetable(
        problem,
        timeout_seconds=timeout,
        seed=seed,
        search_workers=1,
    )


def test_web_dependency_graph_does_not_require_ortools():
    root = Path(__file__).resolve().parents[1]
    assert "timetable_cp_sat_solver" not in (root / "main.py").read_text(encoding="utf-8")
    assert "timetable_generation_worker" not in (root / "main.py").read_text(encoding="utf-8")
    assert "ortools" not in (root / "requirements.txt").read_text(encoding="utf-8").lower()
    assert "ortools==9.15.6755" in (
        root / "requirements-worker.txt"
    ).read_text(encoding="utf-8")
    code = """
import importlib.abc
import sys
class BlockOrtools(importlib.abc.MetaPathFinder):
    def find_spec(self, fullname, path=None, target=None):
        if fullname == 'ortools' or fullname.startswith('ortools.'):
            raise ImportError('OR-Tools deliberately unavailable to web process')
        return None
sys.meta_path.insert(0, BlockOrtools())
import main
"""
    environment = os.environ.copy()
    environment["DATABASE_URL"] = "sqlite:///:memory:"
    result = subprocess.run(
        [sys.executable, "-c", code], cwd=root, env=environment,
        capture_output=True, text=True, timeout=30,
    )
    assert result.returncode == 0, result.stderr


def test_generate_permission_defaults_are_deliberate():
    assert "timetable.generate" in get_default_permissions_for_role("administrator")
    assert "timetable.generate" not in get_default_permissions_for_role("editor")
    assert "timetable.generate" not in get_default_permissions_for_role("user")
    assert "timetable.generate" not in get_default_permissions_for_role("limited")


def test_stage51_snapshot_contract_preserves_hrt_and_excludes_display_names(db):
    db.query(models.TeacherSectionAssignment).delete()
    for section, teacher_id in zip(
        db.query(models.PlanningSection).order_by(models.PlanningSection.id),
        (1000, 1001),
    ):
        section.homeroom_teacher_id = teacher_id
    db.flush()
    snapshot = build_current_snapshot_data(
        db, school_group_id=1, branch_id=10, academic_year_id=100
    )
    problem = TimetableProblemBuilder().build(snapshot.canonical_json)
    payload = json.loads(snapshot.canonical_json)
    assert payload["schema_version"] == 3
    assert payload["scope"] == {
        "school_group_id": 1, "branch_id": 10, "academic_year_id": 100
    }
    assert {item["assignment_source"] for item in problem["demands"]} == {
        "homeroom_default"
    }
    assert {item["subject_code"] for item in problem["demands"]} == {"MAT"}
    assert "Teacher" not in snapshot.canonical_json


def test_cp_sat_feasible_exact_demand_shared_teacher_and_validator(db):
    _make_ready(db)
    db.query(models.TeacherSectionAssignment).filter_by(
        planning_section_id=2001
    ).one().teacher_id = 1000
    problem, snapshot = _problem(db)
    result = _solve(problem)
    assert result["outcome"] == "feasible"
    assert len(result["placements"]) == 8
    validation = TimetableSolutionValidator().validate(
        problem=problem,
        placements=result["placements"],
        expected_fingerprint=snapshot.full_input_fingerprint,
        current_fingerprint=snapshot.full_input_fingerprint,
    )
    assert validation["valid"] is True
    assert len({
        (item["teacher_id"], item["day_key"], item["period_index"])
        for item in result["placements"]
    }) == 8


def test_cp_sat_preserves_lock_and_excludes_non_teaching_block(db):
    _make_ready(db)
    setting = db.query(models.TimetableSetting).one()
    db.add(models.TimetableNonTeachingBlock(
        timetable_setting_id=setting.id,
        block_type="assembly",
        label="Assembly",
        day_key="monday",
        start_time="08:00",
        end_time="08:45",
        start_period=1,
        end_period=1,
    ))
    db.flush()
    lock = {
        "section_id": 2000, "subject_code": "MAT", "teacher_id": 1000,
        "day_key": "monday", "period_index": 1,
    }
    problem, snapshot = _problem(db, locks=[lock])
    result = _solve(problem)
    assert result["outcome"] == "feasible"
    assert lock in result["placements"]
    assert all(
        not (item["day_key"] == "monday" and item["period_index"] == 0)
        for item in result["placements"]
    )
    assert TimetableSolutionValidator().validate(
        problem=problem,
        placements=result["placements"],
        expected_fingerprint=snapshot.full_input_fingerprint,
        current_fingerprint=snapshot.full_input_fingerprint,
    )["valid"]


def test_problem_rejects_conflicting_and_excess_locks(db):
    first = {"section_id": 2000, "subject_code": "MAT", "teacher_id": 1000,
             "day_key": "monday", "period_index": 1}
    conflict = {"section_id": 2001, "subject_code": "MAT", "teacher_id": 1000,
                "day_key": "monday", "period_index": 1}
    _make_ready(db)
    db.query(models.TeacherSectionAssignment).filter_by(
        planning_section_id=2001
    ).one().teacher_id = 1000
    snapshot = build_current_snapshot_data(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        locks=[first, conflict],
    )
    with pytest.raises(TimetableProblemError) as exc:
        TimetableProblemBuilder().build(snapshot.canonical_json)
    assert exc.value.code == "lock_conflict"


def test_cp_sat_infeasible_global_teacher_arrangement(db):
    _make_ready(db)
    db.query(models.TeacherSectionAssignment).filter_by(
        planning_section_id=2001
    ).one().teacher_id = 1000
    setting = db.query(models.TimetableSetting).one()
    setting.working_days_csv = "monday"
    problem, _ = _problem(db)
    result = _solve(problem)
    assert result["outcome"] == "infeasible"


def test_regeneration_enforces_approved_difference_and_locks(db):
    base, _ = _problem(db)
    first = _solve(base, seed=11)
    assert first["outcome"] == "feasible"
    source = [dict(item, is_locked=False) for item in first["placements"]]
    source[0]["is_locked"] = True
    constraints = {
        "generation": {
            "request_mode": "regenerate",
            "source_version_id": 99,
            "source_edit_revision": 3,
            "source_arrangement": source,
            "minimum_difference": _minimum_difference(len(source) - 1),
        }
    }
    regenerated, snapshot = _problem(db, constraints=constraints, locks=[source[0]])
    second = _solve(regenerated, seed=13)
    assert second["outcome"] == "feasible"
    assert source[0] in [dict(item, is_locked=True) for item in second["placements"]]
    validation = TimetableSolutionValidator().validate(
        problem=regenerated,
        placements=second["placements"],
        expected_fingerprint=snapshot.full_input_fingerprint,
        current_fingerprint=snapshot.full_input_fingerprint,
        expected_source_revision=3,
        current_source_revision=3,
    )
    assert validation["valid"]
    assert _minimum_difference(0) == 0
    assert _minimum_difference(1) == 1
    assert _minimum_difference(2) == 1
    assert _minimum_difference(20) == 5
    assert _minimum_difference(100) == 25
    assert _minimum_difference(240) == 60
    assert _minimum_difference(400) == 100


def test_240_placement_regeneration_changes_at_least_twenty_five_percent():
    days = ["monday", "tuesday", "wednesday", "thursday", "friday"]
    slots = [
        {"slot_id": f"{day}:{period}", "day_key": day, "period_index": period}
        for day in days for period in range(1, 9)
    ]
    demands = [
        {
            "demand_id": f"section:{section}|subject:S{subject}|teacher:{section * 100 + subject}",
            "section_id": section,
            "subject_code": f"S{subject}",
            "teacher_id": section * 100 + subject,
            "required_weekly_periods": 5,
        }
        for section in range(1, 7) for subject in range(1, 9)
    ]
    generated_problem = {
        "demands": demands, "slots": slots, "locks": [],
        "request_mode": "generate", "source_arrangement": [],
        "minimum_difference": 0,
    }
    generated = _solve(generated_problem, seed=101, timeout=30)
    assert generated["outcome"] == "feasible"
    source = [dict(item, is_locked=False) for item in generated["placements"]]
    regenerated_problem = dict(
        generated_problem,
        request_mode="regenerate",
        source_arrangement=source,
        minimum_difference=_minimum_difference(len(source)),
    )
    regenerated = _solve(regenerated_problem, seed=202, timeout=30)
    assert regenerated["outcome"] == "feasible"

    source_keys = {
        (item["section_id"], item["subject_code"], item["teacher_id"], item["day_key"], item["period_index"])
        for item in source
    }
    result_keys = {
        (item["section_id"], item["subject_code"], item["teacher_id"], item["day_key"], item["period_index"])
        for item in regenerated["placements"]
    }
    assert len(source_keys) == len(result_keys) == 240
    assert len(source_keys - result_keys) >= 60
    assert {
        (item["section_id"], item["subject_code"], item["teacher_id"])
        for item in source
    } == {
        (item["section_id"], item["subject_code"], item["teacher_id"])
        for item in regenerated["placements"]
    }


def test_regeneration_with_no_alternative_is_infeasible():
    problem = {
        "demands": [{
            "demand_id": "section:1|subject:MAT|teacher:1",
            "section_id": 1,
            "subject_code": "MAT",
            "teacher_id": 1,
            "required_weekly_periods": 1,
        }],
        "slots": [{"slot_id": "monday:1", "day_key": "monday", "period_index": 1}],
        "locks": [],
        "request_mode": "regenerate",
        "source_arrangement": [{
            "section_id": 1, "subject_code": "MAT", "teacher_id": 1,
            "day_key": "monday", "period_index": 1, "is_locked": False,
        }],
        "minimum_difference": 1,
    }
    assert _solve(problem)["outcome"] == "infeasible"


def test_regeneration_infeasibility_reports_controlled_diversity_result(monkeypatch):
    import timetable_generation_worker as worker

    terminal = {}
    monkeypatch.setattr(
        worker,
        "_terminal",
        lambda run_id, owner, status, category, message: terminal.update(
            run_id=run_id, owner=owner, status=status, category=category, message=message
        ),
    )
    worker._mark_infeasible(7, "workflow", {"request_mode": "regenerate"})
    assert terminal == {
        "run_id": 7,
        "owner": "workflow",
        "status": "infeasible",
        "category": "regeneration_diversity_unavailable",
        "message": "No sufficiently different valid timetable could be generated while preserving all current requirements and locks.",
    }


def test_validator_rejects_extra_collision_stale_and_insufficient_diversity(db):
    problem, snapshot = _problem(db)
    result = _solve(problem)
    bad = list(result["placements"]) + [dict(result["placements"][0])]
    validation = TimetableSolutionValidator().validate(
        problem=problem,
        placements=bad,
        expected_fingerprint=snapshot.full_input_fingerprint,
        current_fingerprint="changed",
    )
    codes = {item["code"] for item in validation["errors"]}
    assert {"stale_input", "section_collision", "teacher_collision",
            "total_demand_mismatch", "demand_mismatch"} <= codes


def test_enqueue_claim_heartbeat_cancel_and_bounded_recovery(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="lifecycle-1",
    )
    db.commit()
    claimed = claim_next_run(db, lease_seconds=30, max_attempts=2, lease_owner="worker-a")
    assert claimed.id == run.id and claimed.attempt_count == 1
    assert heartbeat_run(
        db, run_id=run.id, lease_owner="worker-a", lease_seconds=30
    ) == "active"
    request_cancellation(db, run=claimed, actor_user_id="U1")
    assert heartbeat_run(
        db, run_id=run.id, lease_owner="worker-a", lease_seconds=30
    ) == "cancel_requested"
    with pytest.raises(TimetableGenerationError) as cancellation:
        set_run_progress(
            db, run_id=run.id, lease_owner="worker-a", phase="saving"
        )
    assert cancellation.value.code == "cancel_requested"
    assert db.get(models.TimetableGenerationRun, run.id).status == "cancel_requested"
    mark_run_terminal(
        db, run_id=run.id, lease_owner="worker-a", status="cancelled",
        failure_category="cancelled", safe_message="Generation was cancelled.",
    )
    assert db.get(models.TimetableGenerationRun, run.id).status == "cancelled"

    second = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="lifecycle-2",
    )
    db.flush()
    second.status = "running"
    second.attempt_count = 1
    second.lease_owner = "dead-worker"
    second.lease_expires_at = datetime.utcnow() - timedelta(seconds=1)
    db.flush()
    assert recover_expired_runs(db, max_attempts=2) == 1
    assert second.status == "queued" and second.lease_owner is None


def test_fresh_draft_generation_snapshot_matches_immediate_current_input(db):
    _make_ready(db)
    published = create_manual_draft(
        db, school_group_id=1, branch_id=10, academic_year_id=100, origin="imported"
    )
    published.lifecycle_status = "publication_ready"
    set_imported_active_pointer(db, version=published)
    fresh = create_manual_draft(
        db, school_group_id=1, branch_id=10, academic_year_id=100
    )
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="fresh-snapshot-components", draft_public_id=fresh.public_id,
    )

    current, _, _ = _current_input_for_run(db, run)
    snapshot = db.get(models.TimetableInputSnapshot, run.input_snapshot_id)

    assert snapshot.planning_fingerprint == current.planning_fingerprint
    assert snapshot.period_configuration_fingerprint == current.period_configuration_fingerprint
    assert snapshot.constraint_fingerprint == current.constraint_fingerprint
    assert snapshot.lock_fingerprint == current.lock_fingerprint
    assert snapshot.full_input_fingerprint == current.full_input_fingerprint


def test_workflow_exact_claim_is_idempotent_and_scope_safe(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="workflow-exact-claim",
    )
    db.commit()
    claimed = claim_run_by_public_id(
        db,
        public_id=run.public_id,
        lease_seconds=90,
        lease_owner="workflow-one",
        expected_school_group_id=1,
        expected_branch_id=10,
        expected_academic_year_id=100,
    )
    db.commit()
    assert claimed.id == run.id
    assert claimed.progress_phase == "building"
    assert claimed.attempt_count == 1
    assert claim_run_by_public_id(
        db, public_id=run.public_id, lease_seconds=90, lease_owner="workflow-two"
    ) is None
    with pytest.raises(TimetableGenerationError) as mismatch:
        claim_run_by_public_id(
            db,
            public_id=run.public_id,
            lease_seconds=90,
            expected_school_group_id=999,
        )
    assert mismatch.value.code == "scope_mismatch"


def test_workflow_dispatch_failure_is_terminal_only_before_claim(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="workflow-dispatch-failure",
    )
    db.commit()
    assert mark_workflow_dispatch_failed(db, run_id=run.id) is True
    db.commit()
    assert run.status == "internal_error"
    assert run.failure_category == "workflow_dispatch_failed"
    assert "Generate Again" in run.safe_failure_details

    replacement = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="workflow-dispatch-raced",
    )
    db.commit()
    claim_run_by_public_id(
        db, public_id=replacement.public_id, lease_seconds=90, lease_owner="workflow"
    )
    db.commit()
    assert mark_workflow_dispatch_failed(db, run_id=replacement.id) is False
    assert replacement.status == "running"


def test_queued_workflow_delayed_start_message(db, monkeypatch):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="workflow-delayed",
    )
    run.queued_at = datetime.utcnow() - timedelta(seconds=61)
    monkeypatch.setenv("TIS_TIMETABLE_WORKFLOW_START_WARNING_SECONDS", "60")
    payload = generation_run_payload(run)
    assert payload["phase_label"] == "Waiting for Generation Service"
    assert "waiting for compute" in payload["message"]


def test_success_is_atomic_unpublished_and_old_worker_cannot_save(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="persist-success",
    )
    db.commit()
    run = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-a")
    db.commit()
    snapshot = db.get(models.TimetableInputSnapshot, run.input_snapshot_id)
    problem = TimetableProblemBuilder().build(snapshot.canonical_snapshot_json)
    result = _solve(problem)
    set_run_progress(db, run_id=run.id, lease_owner="worker-a", phase="checking")
    set_run_progress(db, run_id=run.id, lease_owner="worker-a", phase="saving")
    version = persist_generated_result(
        db, run_id=run.id, lease_owner="worker-a", problem=problem,
        placements=result["placements"], solver_result=result,
    )
    assert version.approved_at is None
    assert version.approved_by_user_id is None
    db.commit()
    assert version.origin == "generated"
    assert version.lifecycle_status == "publication_ready"
    assert version.public_id
    assert version.quality_score is None
    assert db.query(models.TimetableEntry).filter_by(
        timetable_version_id=version.id
    ).count() == 8
    assert db.query(models.TimetableActiveVersion).count() == 0
    assert db.get(models.TimetableGenerationRun, run.id).result_version_id == version.id
    rendered = build_timetable_workspace_payload(
        db, branch_id=10, academic_year_id=100, version_id=version.id
    )
    assert rendered["version"]["is_stale"] is False
    with pytest.raises(TimetableGenerationError) as exc:
        persist_generated_result(
            db, run_id=run.id, lease_owner="old-worker", problem=problem,
            placements=result["placements"], solver_result=result,
        )
    assert exc.value.code == "lease_lost"


def test_generate_uses_explicit_fresh_draft_instead_of_active_published_version(db):
    _make_ready(db)
    published = create_manual_draft(
        db, school_group_id=1, branch_id=10, academic_year_id=100, origin="imported"
    )
    published.lifecycle_status = "publication_ready"
    set_imported_active_pointer(db, version=published)
    fresh = create_manual_draft(
        db, school_group_id=1, branch_id=10, academic_year_id=100
    )

    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="fresh-draft-context", draft_public_id=fresh.public_id,
    )

    assert run.source_version_id == fresh.id
    assert run.source_version_id != published.id


def test_cancellation_wins_immediately_before_persistence(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="cancel-before-save",
    )
    db.commit()
    run = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-a")
    set_run_progress(db, run_id=run.id, lease_owner="worker-a", phase="saving")
    request_cancellation(db, run=run, actor_user_id="U1")
    with pytest.raises(TimetableGenerationError) as exc:
        persist_generated_result(
            db, run_id=run.id, lease_owner="worker-a", problem={}, placements=[],
            solver_result={},
        )
    assert exc.value.code == "cancel_requested"
    db.commit()
    assert db.get(models.TimetableGenerationRun, run.id).status == "cancelled"
    assert db.query(models.TimetableVersion).count() == 0


def test_regeneration_persists_separate_version_and_keeps_source_and_lock(db):
    _make_ready(db)
    generate = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="regen-source",
    )
    db.commit()
    generate = claim_next_run(
        db, lease_seconds=90, max_attempts=2, lease_owner="worker-generate"
    )
    db.commit()
    snapshot = db.get(models.TimetableInputSnapshot, generate.input_snapshot_id)
    problem = TimetableProblemBuilder().build(snapshot.canonical_snapshot_json)
    first_result = _solve(problem, seed=21)
    set_run_progress(
        db, run_id=generate.id, lease_owner="worker-generate", phase="saving"
    )
    source = persist_generated_result(
        db, run_id=generate.id, lease_owner="worker-generate", problem=problem,
        placements=first_result["placements"], solver_result=first_result,
    )
    db.commit()
    locked = db.query(models.TimetableEntry).filter_by(
        timetable_version_id=source.id
    ).order_by(models.TimetableEntry.id).first()
    set_entry_lock(
        db, version=source, entry=locked, is_locked=True, actor_user_id="U1"
    )
    db.commit()
    source_before = [
        (row.planning_section_id, row.subject_code, row.teacher_id,
         row.day_key, row.period_index, row.is_locked)
        for row in db.query(models.TimetableEntry).filter_by(
            timetable_version_id=source.id
        ).order_by(models.TimetableEntry.id)
    ]

    regenerate = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="regenerate",
        idempotency_key="regen-result", source_public_id=source.public_id,
    )
    db.commit()
    regenerate = claim_next_run(
        db, lease_seconds=90, max_attempts=2, lease_owner="worker-regenerate"
    )
    db.commit()
    regen_snapshot = db.get(models.TimetableInputSnapshot, regenerate.input_snapshot_id)
    regen_problem = TimetableProblemBuilder().build(regen_snapshot.canonical_snapshot_json)
    second_result = _solve(regen_problem, seed=22)
    assert second_result["outcome"] == "feasible"
    set_run_progress(
        db, run_id=regenerate.id, lease_owner="worker-regenerate", phase="saving"
    )
    result_version = persist_generated_result(
        db, run_id=regenerate.id, lease_owner="worker-regenerate",
        problem=regen_problem, placements=second_result["placements"],
        solver_result=second_result,
    )
    db.commit()

    source_after = [
        (row.planning_section_id, row.subject_code, row.teacher_id,
         row.day_key, row.period_index, row.is_locked)
        for row in db.query(models.TimetableEntry).filter_by(
            timetable_version_id=source.id
        ).order_by(models.TimetableEntry.id)
    ]
    assert source_after == source_before
    assert result_version.id != source.id
    assert result_version.origin == "regenerated"
    assert result_version.source_version_id == source.id
    assert result_version.approved_at is None
    assert result_version.approved_by_user_id is None
    assert db.query(models.TimetableEntry).filter_by(
        timetable_version_id=result_version.id, is_locked=True
    ).count() == 1
    assert db.query(models.TimetableActiveVersion).count() == 0


def test_stale_input_and_validator_failure_create_no_version(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="persist-stale",
    )
    db.commit()
    run = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-a")
    db.commit()
    snapshot = db.get(models.TimetableInputSnapshot, run.input_snapshot_id)
    problem = TimetableProblemBuilder().build(snapshot.canonical_snapshot_json)
    result = _solve(problem)
    db.query(models.Subject).filter_by(id=3000).one().weekly_hours = 3
    set_run_progress(db, run_id=run.id, lease_owner="worker-a", phase="saving")
    with pytest.raises(TimetableGenerationError) as exc:
        persist_generated_result(
            db, run_id=run.id, lease_owner="worker-a", problem=problem,
            placements=result["placements"], solver_result=result,
        )
    assert exc.value.code == "stale_input"
    db.commit()
    assert db.query(models.TimetableVersion).count() == 0
    assert db.get(models.TimetableGenerationRun, run.id).status == "stale_input"


def test_validator_rejection_and_persistence_exception_leave_no_partial_version(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="validator-reject",
    )
    db.commit()
    run = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-a")
    db.commit()
    snapshot = db.get(models.TimetableInputSnapshot, run.input_snapshot_id)
    problem = TimetableProblemBuilder().build(snapshot.canonical_snapshot_json)
    result = _solve(problem)
    set_run_progress(db, run_id=run.id, lease_owner="worker-a", phase="saving")
    with pytest.raises(TimetableGenerationError) as exc:
        persist_generated_result(
            db, run_id=run.id, lease_owner="worker-a", problem=problem,
            placements=result["placements"][:-1], solver_result=result,
        )
    assert exc.value.code == "validator_rejected"
    db.commit()
    assert db.query(models.TimetableVersion).count() == 0


def test_nonzero_demand_rejects_zero_placements(db):
    _make_ready(db)
    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="zero-placement-reject",
    )
    db.commit()
    run = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-zero")
    db.commit()
    snapshot = db.get(models.TimetableInputSnapshot, run.input_snapshot_id)
    problem = TimetableProblemBuilder().build(snapshot.canonical_snapshot_json)
    set_run_progress(db, run_id=run.id, lease_owner="worker-zero", phase="saving")
    with pytest.raises(TimetableGenerationError) as exc:
        persist_generated_result(
            db, run_id=run.id, lease_owner="worker-zero", problem=problem,
            placements=[], solver_result={"placements": []},
        )
    assert exc.value.code == "validator_rejected"
    db.commit()
    assert db.query(models.TimetableVersion).count() == 0

    # A database failure after the candidate version is allocated must roll back all rows.
    failed = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="persistence-rollback",
    )
    db.commit()
    failed = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-b")
    db.commit()
    failed_snapshot = db.get(models.TimetableInputSnapshot, failed.input_snapshot_id)
    failed_problem = TimetableProblemBuilder().build(failed_snapshot.canonical_snapshot_json)
    failed_result = _solve(failed_problem)
    set_run_progress(db, run_id=failed.id, lease_owner="worker-b", phase="saving")

    def reject_entry(*_args, **_kwargs):
        raise RuntimeError("injected entry persistence failure")

    sa_event.listen(models.TimetableEntry, "before_insert", reject_entry)
    try:
        with pytest.raises(RuntimeError, match="injected entry"):
            persist_generated_result(
                db, run_id=failed.id, lease_owner="worker-b", problem=failed_problem,
                placements=failed_result["placements"], solver_result=failed_result,
            )
        db.rollback()
    finally:
        sa_event.remove(models.TimetableEntry, "before_insert", reject_entry)
    assert db.query(models.TimetableVersion).count() == 0
    assert db.query(models.TimetableEntry).count() == 0


def test_240_period_generation_persists_reloads_and_renders_complete_ui(db):
    pytest.importorskip("ortools")
    db.query(models.TeacherSectionAssignment).delete()
    subject = db.get(models.Subject, 3000)
    subject.weekly_hours = 40
    setting = db.get(models.TimetableSetting, 5000)
    setting.working_days_csv = "monday,tuesday,wednesday,thursday,friday"
    setting.periods_per_day = 8
    setting.school_end_time = "14:00"
    db.get(models.Teacher, 1000).max_hours = 40
    db.get(models.Teacher, 1001).max_hours = 40
    for index in range(2, 6):
        db.add(models.Teacher(
            id=1000 + index, teacher_id=f"T{index + 1}", first_name="Generated",
            last_name=f"Teacher {index + 1}", branch_id=10, academic_year_id=100,
            max_hours=40,
        ))
        db.add(models.PlanningSection(
            id=2000 + index, grade_level="1", section_name=chr(65 + index),
            class_status="Current", branch_id=10, academic_year_id=100,
        ))
    db.flush()
    for index in range(6):
        db.add(models.TeacherSectionAssignment(
            teacher_id=1000 + index,
            planning_section_id=2000 + index,
            subject_code="MAT",
        ))
    db.flush()
    published = create_manual_draft(
        db, school_group_id=1, branch_id=10, academic_year_id=100, origin="imported"
    )
    published.lifecycle_status = "publication_ready"
    set_imported_active_pointer(db, version=published)
    db.commit()

    run = enqueue_generation(
        db, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="exact-240",
    )
    db.commit()
    run = claim_next_run(db, lease_seconds=90, max_attempts=2, lease_owner="worker-240")
    db.commit()
    snapshot = db.get(models.TimetableInputSnapshot, run.input_snapshot_id)
    problem = TimetableProblemBuilder().build(snapshot.canonical_snapshot_json)
    expected_required_periods = sum(
        item["required_weekly_periods"] for item in problem["demands"]
    )
    result = _solve(problem, seed=240, timeout=30)
    validation = TimetableSolutionValidator().validate(
        problem=problem,
        placements=result["placements"],
        expected_fingerprint=snapshot.full_input_fingerprint,
        current_fingerprint=snapshot.full_input_fingerprint,
        expected_scope=problem["scope"],
        current_scope=problem["scope"],
    )
    assert expected_required_periods == 240
    assert len(result["placements"]) == 240
    assert validation["valid"] is True
    assert validation["counts"]["placements"] == 240
    set_run_progress(db, run_id=run.id, lease_owner="worker-240", phase="checking")
    set_run_progress(db, run_id=run.id, lease_owner="worker-240", phase="saving")
    generated = persist_generated_result(
        db, run_id=run.id, lease_owner="worker-240", problem=problem,
        placements=result["placements"], solver_result=result,
    )
    generated_id = generated.id
    db.commit()
    assert generated.source_version_id == published.id
    assert generated.public_id
    assert resolve_operational_version(
        db, school_group_id=1, branch_id=10, academic_year_id=100
    ).id == generated_id

    FreshSession = sessionmaker(bind=db.get_bind())
    fresh = FreshSession()
    try:
        assert fresh.query(models.TimetableEntry).filter_by(
            timetable_version_id=generated_id
        ).count() == 240
        payload = build_timetable_workspace_payload(
            fresh, branch_id=10, academic_year_id=100, version_id=generated_id
        )
        assert payload["summary"]["scheduled_hours"] == 240
        assert payload["summary"]["remaining_hours"] == 0
        assert payload["version"]["id"] == generated_id
        assert payload["version"]["origin"] == "generated"
        assert payload["version"]["is_active"] is False
        assert payload["version"]["is_logical_draft_source"] is False
        xlsx = timetable_router._build_timetable_xlsx_bytes(
            payload, "Main", "2026", logo_assets=[]
        )
        workbook = load_workbook(io.BytesIO(xlsx), data_only=True)
        assert workbook.sheetnames[0] == "By Section"
        by_section = workbook["By Section"]
        assert len(by_section._images) == 240
        assert all(image.anchor.ext.cx <= 14 * 9525 and image.anchor.ext.cy <= 14 * 9525 for image in by_section._images)
        assert all(image.anchor._from.col >= 2 and image.anchor._from.row >= 4 for image in by_section._images)
        assert by_section.column_dimensions["C"].width >= 28
        assert any((dimension.height or 0) >= 54 for dimension in by_section.row_dimensions.values())
        assert any(cell.alignment.wrap_text for row in by_section.iter_rows() for cell in row)
        overview_values = [
            str(cell.value) for row in workbook["Overview"].iter_rows() for cell in row
            if cell.value is not None
        ]
        assert "Scheduled Periods" in overview_values
        assert "240" in overview_values
        assert "Remaining Periods" in overview_values
        assert "0" in overview_values
        assert any("Working Timetable | Version" in value for value in overview_values)
        section_values = [
            str(cell.value) for row in by_section.iter_rows() for cell in row
            if cell.value is not None
        ]
        first_range = payload["time_slots"][0]["time_range"]
        assert " - " in first_range and "?" not in first_range
        assert any(first_range in value for value in section_values)
        assert any("MAT" in value and "Generated Teacher" in value for value in section_values)
        assert any("School:" in value and "Working Timetable" in value and f"Version {generated.version_number}" in value for value in section_values)
        pdf = timetable_router._build_timetable_pdf_bytes(
            payload, "Main", "2026", logo_assets=[]
        )
        assert pdf.startswith(b"%PDF")
        assert b"/Subtype /Image" in pdf
        pdf_text = b"\n".join(re.findall(rb"\((.*?)\) Tj", pdf)).decode("latin-1")
        assert "Working Timetable" in pdf_text
        assert first_range in pdf_text
        assert "?" not in pdf_text
        assert len(payload["sections"]) == 6
        assert all(section["scheduled_hours"] == 40 for section in payload["sections"])
        assert all(section["remaining_hours"] == 0 for section in payload["sections"])
        pointer = fresh.query(models.TimetableActiveVersion).one()
        assert pointer.timetable_version_id == published.id
        assert fresh.query(models.TimetableVersion).filter_by(
            generation_run_id=run.id
        ).count() == 1
        assert claim_run_by_public_id(
            fresh,
            public_id=run.public_id,
            lease_seconds=90,
            lease_owner="duplicate-workflow",
        ) is None
        fresh.commit()
        assert fresh.query(models.TimetableVersion).filter_by(
            generation_run_id=run.id
        ).count() == 1
    finally:
        fresh.close()

def test_representative_school_sized_cp_sat_problem():
    slots = [
        {"slot_id": f"{day}:{period}", "day_key": day, "period_index": period}
        for day in ("monday", "tuesday", "wednesday", "thursday", "friday")
        for period in range(1, 8)
    ]
    demands = []
    for section_id in range(1, 31):
        for subject_index in range(10):
            teacher_id = 1 + subject_index * 6 + ((section_id - 1) % 6)
            subject_code = f"S{subject_index + 1:02d}"
            demands.append({
                "demand_id": (
                    f"section:{section_id}|subject:{subject_code}|teacher:{teacher_id}"
                ),
                "section_id": section_id,
                "subject_code": subject_code,
                "teacher_id": teacher_id,
                "required_weekly_periods": 3,
            })
    problem = {
        "demands": demands,
        "slots": slots,
        "locks": [],
        "request_mode": "generate",
        "source_arrangement": [],
        "minimum_difference": 0,
    }
    started = time.perf_counter()
    result = _solve(problem, seed=20260822, timeout=30)
    elapsed = time.perf_counter() - started
    assert result["outcome"] == "feasible"
    assert len(result["placements"]) == 900
    assert elapsed < 30


POSTGRESQL_URL = os.getenv("TIS_TEST_POSTGRESQL_URL", "")


@pytest.fixture()
def pg_stage51():
    if not POSTGRESQL_URL.startswith("postgresql"):
        pytest.skip("TIS_TEST_POSTGRESQL_URL is required for Stage 5.1 PostgreSQL tests")
    schema = f"tis_timetable_s51_{uuid.uuid4().hex}"
    admin = create_engine(POSTGRESQL_URL)
    with admin.begin() as connection:
        connection.execute(text(f'CREATE SCHEMA "{schema}"'))
    engine = create_engine(
        POSTGRESQL_URL,
        connect_args={"options": f"-csearch_path={schema}"},
    )
    models.Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    session.add(models.SchoolGroup(id=1, name="Stage51 PG"))
    session.flush()
    session.add_all([
        models.Branch(id=10, school_group_id=1, name="Main"),
        models.AcademicYear(id=100, school_group_id=1, year_name="2026"),
    ])
    session.flush()
    session.add_all([
        models.User(
            user_id="U1", username="pg.u1", first_name="PG", last_name="User",
            school_group_id=1, branch_id=10, academic_year_id=100,
        ),
        models.Teacher(
            id=1000, teacher_id="PGT1", first_name="One", last_name="Teacher",
            branch_id=10, academic_year_id=100,
        ),
        models.PlanningSection(
            id=2000, grade_level="3", section_name="A", class_status="Current",
            branch_id=10, academic_year_id=100,
        ),
        models.Subject(
            id=3000, subject_code="MAT", subject_name="Mathematics",
            weekly_hours=2, grade=3, branch_id=10, academic_year_id=100,
        ),
        models.TimetableSetting(
            id=5000, branch_id=10, academic_year_id=100,
            working_days_csv="monday,tuesday", periods_per_day=2,
            period_duration_minutes=45, school_start_time="08:00",
            school_end_time="09:30",
        ),
    ])
    session.flush()
    session.add(models.TeacherSectionAssignment(
        id=4000, teacher_id=1000, planning_section_id=2000, subject_code="MAT"
    ))
    session.commit()
    session.close()
    try:
        yield engine, Session
    finally:
        engine.dispose()
        with admin.begin() as connection:
            connection.execute(text(f'DROP SCHEMA "{schema}" CASCADE'))
        admin.dispose()


def test_postgresql_stage51_indexes_and_concurrent_enqueue(pg_stage51):
    engine, Session = pg_stage51
    with engine.begin() as connection:
        _smart_timetable_stage51_generator(engine, connection)
        _smart_timetable_stage51_generator(engine, connection)
    indexes = {item["name"]: item for item in inspect(engine).get_indexes(
        "timetable_generation_runs"
    )}
    assert "ix_timetable_generation_runs_worker_claim" in indexes
    assert indexes["uq_timetable_generation_runs_active_scope"]["unique"] is True

    barrier = threading.Barrier(2)
    outcomes = []
    lock = threading.Lock()

    def enqueue(key):
        session = Session()
        try:
            barrier.wait(timeout=5)
            run = enqueue_generation(
                session, school_group_id=1, branch_id=10, academic_year_id=100,
                requested_by_user_id="U1", request_mode="generate",
                idempotency_key=key,
            )
            session.commit()
            value = ("queued", run.public_id)
        except TimetableGenerationError as exc:
            session.rollback()
            value = (exc.code, None)
        finally:
            session.close()
        with lock:
            outcomes.append(value)

    threads = [threading.Thread(target=enqueue, args=(f"pg-{index}",)) for index in (1, 2)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)
    assert sorted(item[0] for item in outcomes) == ["active_generation_exists", "queued"]
    session = Session()
    assert session.query(models.TimetableGenerationRun).filter_by(status="queued").count() == 1
    session.close()


def test_postgresql_two_worker_claim_lease_recovery_and_old_worker_guard(pg_stage51):
    _, Session = pg_stage51
    session = Session()
    run = enqueue_generation(
        session, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate", idempotency_key="claim-race",
    )
    session.commit()
    run_id = run.id
    session.close()

    barrier = threading.Barrier(2)
    claimed = []
    lock = threading.Lock()

    def claim(owner):
        worker = Session()
        barrier.wait(timeout=5)
        row = claim_next_run(
            worker, lease_seconds=30, max_attempts=2, lease_owner=owner
        )
        worker.commit()
        with lock:
            claimed.append((owner, row.id if row else None))
        worker.close()

    threads = [threading.Thread(target=claim, args=(owner,)) for owner in ("a", "b")]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)
    winners = [item for item in claimed if item[1] == run_id]
    assert len(winners) == 1
    old_owner = winners[0][0]

    session = Session()
    leased = session.get(models.TimetableGenerationRun, run_id)
    leased.lease_expires_at = datetime.utcnow() - timedelta(seconds=1)
    session.commit()
    assert recover_expired_runs(session, max_attempts=2) == 1
    session.commit()
    replacement = claim_next_run(
        session, lease_seconds=30, max_attempts=2, lease_owner="replacement"
    )
    session.commit()
    assert replacement.id == run_id
    with pytest.raises(TimetableGenerationError) as lost_save:
        persist_generated_result(
            session, run_id=run_id, lease_owner=old_owner,
            problem={}, placements=[], solver_result={},
        )
    assert lost_save.value.code == "lease_lost"
    assert heartbeat_run(
        session, run_id=run_id, lease_owner=old_owner, lease_seconds=30
    ) == "lost"
    assert heartbeat_run(
        session, run_id=run_id, lease_owner="replacement", lease_seconds=30
    ) == "active"
    request_cancellation(session, run=replacement, actor_user_id="U1")
    session.commit()
    assert heartbeat_run(
        session, run_id=run_id, lease_owner="replacement", lease_seconds=30
    ) == "cancel_requested"
    with pytest.raises(TimetableGenerationError) as cancellation:
        set_run_progress(
            session, run_id=run_id, lease_owner="replacement", phase="saving"
        )
    assert cancellation.value.code == "cancel_requested"
    assert session.get(models.TimetableGenerationRun, run_id).status == "cancel_requested"
    session.close()


def test_postgresql_duplicate_workflow_invocation_claims_exact_run_once(pg_stage51):
    _, Session = pg_stage51
    session = Session()
    run = enqueue_generation(
        session, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate",
        idempotency_key="workflow-duplicate-pg",
    )
    session.commit()
    public_id = run.public_id
    run_id = run.id
    session.close()

    barrier = threading.Barrier(2)
    outcomes = []
    outcome_lock = threading.Lock()

    def claim(owner):
        worker = Session()
        try:
            barrier.wait(timeout=5)
            row = claim_run_by_public_id(
                worker,
                public_id=public_id,
                lease_seconds=90,
                lease_owner=owner,
            )
            worker.commit()
            value = row.id if row else None
        finally:
            worker.close()
        with outcome_lock:
            outcomes.append(value)

    threads = [threading.Thread(target=claim, args=(owner,)) for owner in ("wf-a", "wf-b")]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)
    assert outcomes.count(run_id) == 1
    assert outcomes.count(None) == 1
    check = Session()
    stored = check.get(models.TimetableGenerationRun, run_id)
    assert stored.status == "running"
    assert stored.attempt_count == 1
    assert check.query(models.TimetableVersion).count() == 0
    check.close()


def test_postgresql_stale_race_and_persistence_rollback_are_atomic(pg_stage51):
    _, Session = pg_stage51
    session = Session()
    stale = enqueue_generation(
        session, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate", idempotency_key="pg-stale",
    )
    session.commit()
    stale = claim_next_run(
        session, lease_seconds=90, max_attempts=2, lease_owner="stale-worker"
    )
    session.commit()
    stale_snapshot = session.get(models.TimetableInputSnapshot, stale.input_snapshot_id)
    stale_problem = TimetableProblemBuilder().build(stale_snapshot.canonical_snapshot_json)
    placements = [
        {"section_id": 2000, "subject_code": "MAT", "teacher_id": 1000,
         "day_key": "monday", "period_index": 1},
        {"section_id": 2000, "subject_code": "MAT", "teacher_id": 1000,
         "day_key": "tuesday", "period_index": 1},
    ]
    session.query(models.TimetableSetting).filter_by(id=5000).one().period_duration_minutes = 40
    set_run_progress(
        session, run_id=stale.id, lease_owner="stale-worker", phase="saving"
    )
    with pytest.raises(TimetableGenerationError) as stale_error:
        persist_generated_result(
            session, run_id=stale.id, lease_owner="stale-worker",
            problem=stale_problem, placements=placements, solver_result={},
        )
    assert stale_error.value.code == "stale_input"
    session.commit()
    assert session.query(models.TimetableVersion).count() == 0

    # Restore current configuration before capturing the rollback candidate.
    session.query(models.TimetableSetting).filter_by(id=5000).one().period_duration_minutes = 45
    session.commit()
    rollback_run = enqueue_generation(
        session, school_group_id=1, branch_id=10, academic_year_id=100,
        requested_by_user_id="U1", request_mode="generate", idempotency_key="pg-rollback",
    )
    session.commit()
    rollback_run = claim_next_run(
        session, lease_seconds=90, max_attempts=2, lease_owner="rollback-worker"
    )
    session.commit()
    rollback_snapshot = session.get(
        models.TimetableInputSnapshot, rollback_run.input_snapshot_id
    )
    rollback_problem = TimetableProblemBuilder().build(
        rollback_snapshot.canonical_snapshot_json
    )
    set_run_progress(
        session, run_id=rollback_run.id, lease_owner="rollback-worker", phase="saving"
    )

    def reject_entry(*_args, **_kwargs):
        raise RuntimeError("injected PostgreSQL persistence failure")

    sa_event.listen(models.TimetableEntry, "before_insert", reject_entry)
    try:
        with pytest.raises(RuntimeError, match="PostgreSQL persistence"):
            persist_generated_result(
                session, run_id=rollback_run.id, lease_owner="rollback-worker",
                problem=rollback_problem, placements=placements, solver_result={},
            )
        session.rollback()
    finally:
        sa_event.remove(models.TimetableEntry, "before_insert", reject_entry)
    assert session.query(models.TimetableVersion).count() == 0
    assert session.query(models.TimetableEntry).count() == 0
    session.close()
