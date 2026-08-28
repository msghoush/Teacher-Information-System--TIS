import io

from openpyxl import load_workbook
from PIL import Image

import branding_storage
import tenant_report_branding
from routers import academic_calendar as calendar_router
from routers import timetable as timetable_router


def _logo(path, label="Academy"):
    return {
        "is_configured": True,
        "absolute_path": str(path),
        "alt": f"{label} logo",
        "label": label,
    }


def _workspace():
    return {
        "days": [{"key": "monday", "label": "Monday"}],
        "timeline_rows": [{
            "row_type": "period",
            "slot": {"period_index": 1, "label": "Period 1"},
            "items_by_day": {
                "monday": {
                    "row_type": "period", "period_index": 1,
                    "start_time": "08:00", "end_time": "08:45",
                }
            },
        }],
        "time_slots": [{"period_index": 1, "label": "Period 1"}],
        "sections": [{
            "id": 1, "section_label": "Grade 3 - A", "class_status": "active",
            "scheduled_hours": 0, "total_required_hours": 0, "remaining_hours": 0,
        }],
        "teachers": [], "entries": [], "blocked_slots": [],
        "settings": {"blocks": []}, "warnings": [], "readiness": {},
    }


def test_central_tenant_report_branding_returns_configured_logo_only(monkeypatch, tmp_path):
    organization_logo = tmp_path / "organization.png"
    Image.new("RGB", (48, 24), "navy").save(organization_logo)
    calls = []

    def scoped_slots(request, db, branch_id, school_group_id, *, include_empty):
        calls.append((branch_id, school_group_id, include_empty))
        return [_logo(organization_logo)] if (branch_id, school_group_id) == (10, 1) else []

    monkeypatch.setattr(tenant_report_branding, "get_school_logo_slots", scoped_slots)

    tenant_one = tenant_report_branding.get_tenant_report_logos(None, None, 10, 1)
    tenant_two = tenant_report_branding.get_tenant_report_logos(None, None, 20, 2)

    assert [item["absolute_path"] for item in tenant_one] == [str(organization_logo)]
    assert tenant_two == []
    assert calls == [(10, 1, False), (20, 2, False)]


def test_timetable_exports_use_org_logo_and_never_inject_tis_fallback(monkeypatch, tmp_path):
    organization_logo = tmp_path / "organization.png"
    Image.new("RGB", (96, 40), "navy").save(organization_logo)
    monkeypatch.setattr(
        timetable_router,
        "get_tenant_report_logos",
        lambda request, db, branch_id: [_logo(organization_logo)] if branch_id == 10 else [],
    )

    branded_assets = timetable_router._build_export_logo_assets(None, None, 10)
    unbranded_assets = timetable_router._build_export_logo_assets(None, None, 20)
    product_logo = branding_storage.tis_logo_absolute_path(theme="light", layout="horizontal")

    assert [item["path"] for item in branded_assets] == [organization_logo]
    assert all(item["path"] != product_logo for item in branded_assets)
    assert unbranded_assets == []
    assert not hasattr(timetable_router, "PLATFORM_EXPORT_LOGO_ASSET")

    branded_xlsx = timetable_router._build_timetable_xlsx_bytes(
        _workspace(), "Academy Branch", "2026-2027", logo_assets=branded_assets,
    )
    plain_xlsx = timetable_router._build_timetable_xlsx_bytes(
        _workspace(), "No Logo Branch", "2026-2027", logo_assets=[],
    )
    assert len(load_workbook(io.BytesIO(branded_xlsx))["Overview"]._images) == 1
    plain_overview = load_workbook(io.BytesIO(plain_xlsx))["Overview"]
    assert len(plain_overview._images) == 0
    assert plain_overview["A1"].value == "Weekly Timetable Export"

    assert timetable_router._build_timetable_pdf_bytes(
        _workspace(), "No Logo Branch", "2026-2027", logo_assets=[]
    ).startswith(b"%PDF")


def test_academic_calendar_has_no_product_logo_default(monkeypatch):
    monkeypatch.setattr(
        calendar_router,
        "get_tenant_report_logo_paths",
        lambda request, db, branch_id: (),
    )
    assert calendar_router._calendar_pdf_logos(None, None, 10) == ()
    assert not hasattr(calendar_router, "PLATFORM_CALENDAR_PDF_LOGO")
